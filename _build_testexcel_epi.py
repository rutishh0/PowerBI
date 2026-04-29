"""
Build a self-contained HTML visualization for EPI 16.02.xlsx (INVOICE_LIST).

Runs independently of V6/parser.py - reads the xlsx directly with openpyxl.
Emits V6/TESTEXCEL/EPI_16_02.html embedding all data as a JSON constant,
rendered with ApexCharts 3.49.0 via CDN.
"""

from __future__ import annotations

import json
import os
from datetime import datetime, date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SRC_PATH = Path(
    r"C:\Users\Rutishkrishna\Desktop\RR\RR Powerbi\V6\New info\EPI 16.02.xlsx"
)
OUT_PATH = Path(
    r"C:\Users\Rutishkrishna\Desktop\RR\RR Powerbi\V6\TESTEXCEL\EPI_16_02.html"
)

# Today's date for aging computation (fixed to the conversation's "today")
TODAY = datetime(2026, 4, 22)

AGING_BUCKETS = [
    "Not due",
    "0-30",
    "31-60",
    "61-90",
    "91-180",
    "181-365",
    "365+",
]


def aging_bucket(days_late: int) -> str:
    if days_late <= 0:
        return "Not due"
    if days_late <= 30:
        return "0-30"
    if days_late <= 60:
        return "31-60"
    if days_late <= 90:
        return "61-90"
    if days_late <= 180:
        return "91-180"
    if days_late <= 365:
        return "181-365"
    return "365+"


def normalize_status(text: str | None, days_late: int, amount: float | None) -> str:
    """
    Status inference for EPI file - no explicit Status column.
    Heuristic:
      - Negative amount -> 'Credit'
      - 'disputed' / 'dispute' in text -> 'Disputed'
      - 'cancel' / 'canx' / 'credit' in text -> 'Disputed' (offsetting/credit notes)
      - days_late > 0 -> 'Overdue'
      - days_late <= 0 -> 'Outstanding'
    """
    t = (text or "").lower()
    if isinstance(amount, (int, float)) and amount < 0:
        return "Credit"
    if "disput" in t:
        return "Disputed"
    if "canx" in t or "cancel" in t or "credited" in t or "credit" in t:
        return "Disputed"
    if days_late > 0:
        return "Overdue"
    return "Outstanding"


def to_iso(dt: Any) -> str | None:
    if dt is None or dt == "":
        return None
    if isinstance(dt, (datetime, date)):
        return dt.strftime("%Y-%m-%d")
    s = str(dt).strip()
    # Attempt common formats
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s  # fallback raw


def ref_to_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        # preserve as plain integer-like string (no scientific)
        if float(v).is_integer():
            return str(int(v))
        return str(v)
    return str(v)


def extract() -> dict[str, Any]:
    wb = load_workbook(SRC_PATH, data_only=True)
    wb_formulas = load_workbook(SRC_PATH, data_only=False)
    ws = wb["Sheet1"]
    ws_f = wb_formulas["Sheet1"]

    header = [
        (ws.cell(row=1, column=c).value or "").strip()
        for c in range(1, ws.max_column + 1)
    ]

    invoices: list[dict[str, Any]] = []
    subtotals: list[dict[str, Any]] = []

    for r in range(2, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        row_formulas = [
            ws_f.cell(row=r, column=c).value for c in range(1, ws_f.max_column + 1)
        ]

        # Skip fully blank rows
        if all(v is None or v == "" for v in row_vals):
            continue

        # Detect subtotal row: amount cell (col E, idx 4) contains a SUM formula
        amt_formula = row_formulas[4] if len(row_formulas) > 4 else None
        is_sum = isinstance(amt_formula, str) and amt_formula.lstrip().upper().startswith("=SUM")

        if is_sum:
            subtotals.append(
                {
                    "row": r,
                    "formula": amt_formula,
                    "value": row_vals[4],
                }
            )
            continue

        ref = ref_to_str(row_vals[0])
        doc_date = row_vals[1]
        net_due = row_vals[2]
        currency = row_vals[3] or ""
        amount = row_vals[4]
        text = row_vals[6] or ""
        assignment = row_vals[7] or ""

        if not isinstance(amount, (int, float)):
            # nothing useful to chart on
            continue

        net_due_iso = to_iso(net_due)
        days_late = 0
        if isinstance(net_due, (datetime, date)):
            nd = net_due if isinstance(net_due, datetime) else datetime.combine(net_due, datetime.min.time())
            days_late = (TODAY - nd).days

        bucket = aging_bucket(days_late)
        status = normalize_status(str(text), days_late, amount)

        invoices.append(
            {
                "row": r,
                "reference": ref,
                "doc_date": to_iso(doc_date),
                "net_due_date": net_due_iso,
                "currency": currency,
                "amount": float(amount),
                "text": str(text),
                "assignment": str(assignment),
                "days_late": days_late,
                "aging_bucket": bucket,
                "status": status,
            }
        )

    # Aggregations
    total_count = len(invoices)
    total_amount = sum(i["amount"] for i in invoices)
    total_outstanding = sum(i["amount"] for i in invoices if i["status"] == "Outstanding")
    total_overdue = sum(i["amount"] for i in invoices if i["status"] == "Overdue")
    total_disputed = sum(i["amount"] for i in invoices if i["status"] == "Disputed")
    total_credit = sum(i["amount"] for i in invoices if i["status"] == "Credit")
    overdue_items = [i for i in invoices if i["status"] == "Overdue"]
    avg_days_late = (
        sum(i["days_late"] for i in overdue_items) / len(overdue_items)
        if overdue_items
        else 0
    )

    # Aging distribution (in canonical order)
    aging_counts = {b: 0 for b in AGING_BUCKETS}
    aging_amounts = {b: 0.0 for b in AGING_BUCKETS}
    for i in invoices:
        aging_counts[i["aging_bucket"]] += 1
        aging_amounts[i["aging_bucket"]] += i["amount"]

    # Status distribution
    status_counts: dict[str, int] = {}
    status_amounts: dict[str, float] = {}
    for i in invoices:
        status_counts[i["status"]] = status_counts.get(i["status"], 0) + 1
        status_amounts[i["status"]] = status_amounts.get(i["status"], 0.0) + i["amount"]

    # Monthly distribution by doc_date
    monthly: dict[str, dict[str, float]] = {}
    for i in invoices:
        d = i["doc_date"]
        if not d:
            continue
        key = d[:7]  # YYYY-MM
        entry = monthly.setdefault(key, {"count": 0, "amount": 0.0})
        entry["count"] += 1
        entry["amount"] += i["amount"]
    monthly_sorted = sorted(monthly.items())

    # Currency distribution
    currency_counts: dict[str, int] = {}
    for i in invoices:
        c = i["currency"] or "?"
        currency_counts[c] = currency_counts.get(c, 0) + 1

    # Top 10 by abs(amount) - show actual amount but rank by magnitude
    top10 = sorted(invoices, key=lambda x: abs(x["amount"]), reverse=True)[:10]

    # Subtotal groupings: use row ranges between subtotals as logical groups
    groups = []
    prev_end = 1
    for s in subtotals:
        group_rows = [i for i in invoices if prev_end < i["row"] < s["row"]]
        if group_rows:
            groups.append(
                {
                    "label": f"Rows {group_rows[0]['row']}-{group_rows[-1]['row']}",
                    "count": len(group_rows),
                    "amount": sum(g["amount"] for g in group_rows),
                    "subtotal_row": s["row"],
                    "subtotal_value": s["value"],
                    "formula": s["formula"],
                }
            )
        prev_end = s["row"]

    return {
        "meta": {
            "source_path": str(SRC_PATH),
            "sheet": "Sheet1",
            "header": header,
            "row_count": total_count,
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "today": TODAY.strftime("%Y-%m-%d"),
        },
        "kpis": {
            "total_count": total_count,
            "total_amount": total_amount,
            "total_outstanding": total_outstanding,
            "total_overdue": total_overdue,
            "total_disputed": total_disputed,
            "total_credit": total_credit,
            "avg_days_late_overdue": avg_days_late,
        },
        "aging_counts": aging_counts,
        "aging_amounts": aging_amounts,
        "status_counts": status_counts,
        "status_amounts": status_amounts,
        "monthly": monthly_sorted,
        "currency_counts": currency_counts,
        "top10": top10,
        "invoices": invoices,
        "subtotals": subtotals,
        "groups": groups,
    }


HTML_TEMPLATE = r"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8" />
<meta name="viewport" content="width=device-width, initial-scale=1" />
<title>EPI Invoice List — Benchmark</title>
<script src="https://cdn.jsdelivr.net/npm/apexcharts@3.49.0/dist/apexcharts.min.js"></script>
<style>
  :root {
    --bg: #0b1220;
    --panel: #111a2e;
    --panel-2: #162241;
    --border: #24324f;
    --text: #e6ecf7;
    --muted: #8b99b8;
    --accent: #5aa2ff;
    --ok: #22c55e;
    --warn: #f59e0b;
    --bad: #ef4444;
    --purple: #a78bfa;
  }
  * { box-sizing: border-box; }
  html, body { margin: 0; padding: 0; background: var(--bg); color: var(--text);
               font: 14px/1.45 "Segoe UI", Roboto, system-ui, -apple-system, Arial, sans-serif; }
  a { color: var(--accent); }
  header { padding: 22px 28px; border-bottom: 1px solid var(--border);
           background: linear-gradient(180deg, #0f1a35, #0b1220); }
  header h1 { margin: 0 0 6px; font-size: 22px; font-weight: 600; letter-spacing: .2px; }
  header .sub { color: var(--muted); font-size: 12.5px; }
  main { padding: 22px 28px 60px; max-width: 1400px; margin: 0 auto; }
  .kpis { display: grid; grid-template-columns: repeat(6, minmax(0,1fr)); gap: 14px; margin: 18px 0 22px; }
  .kpi { background: var(--panel); border: 1px solid var(--border); border-radius: 10px; padding: 14px 16px; }
  .kpi .label { color: var(--muted); font-size: 11.5px; text-transform: uppercase; letter-spacing: .8px; }
  .kpi .val { font-size: 20px; font-weight: 600; margin-top: 6px; }
  .kpi .val.small { font-size: 16px; }
  .kpi.ok .val { color: var(--ok); }
  .kpi.warn .val { color: var(--warn); }
  .kpi.bad .val { color: var(--bad); }
  .charts { display: grid; grid-template-columns: 2fr 1fr; gap: 16px; margin-bottom: 18px; }
  .charts-2 { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 18px; }
  .card { background: var(--panel); border: 1px solid var(--border); border-radius: 10px; padding: 14px 16px; }
  .card h2 { margin: 0 0 12px; font-size: 14px; font-weight: 600; color: #cfd8ec; letter-spacing: .3px; }
  .card h2 .hint { color: var(--muted); font-weight: 400; font-size: 12px; margin-left: 8px; }
  .groups { display: grid; grid-template-columns: repeat(3, 1fr); gap: 12px; }
  .group { background: var(--panel-2); border: 1px solid var(--border); border-radius: 8px; padding: 10px 12px; font-size: 12.5px; }
  .group .lbl { color: var(--muted); font-size: 11px; text-transform: uppercase; letter-spacing: .6px; }
  .group .amt { font-size: 16px; font-weight: 600; margin-top: 4px; }
  .group .frm { color: var(--muted); font-size: 11px; margin-top: 4px; font-family: ui-monospace, Consolas, monospace; }
  .filters { display: flex; flex-wrap: wrap; gap: 10px; margin: 4px 0 14px; align-items: center; }
  .filters input, .filters select {
    background: var(--panel-2); color: var(--text); border: 1px solid var(--border);
    border-radius: 8px; padding: 8px 10px; font-size: 13px; min-width: 140px;
  }
  .filters input[type=search] { min-width: 260px; }
  .filters .count { color: var(--muted); margin-left: auto; font-size: 12.5px; }
  table { width: 100%; border-collapse: collapse; font-size: 12.5px; }
  th, td { padding: 8px 10px; text-align: left; border-bottom: 1px solid var(--border); }
  thead th { background: var(--panel-2); color: #cfd8ec; font-weight: 600; cursor: pointer; user-select: none; position: sticky; top: 0; }
  thead th:hover { background: #1c2a4e; }
  tbody tr:hover { background: #16213e; }
  td.num { text-align: right; font-variant-numeric: tabular-nums; }
  td.ref { font-family: ui-monospace, Consolas, monospace; }
  .pill { display: inline-block; padding: 2px 8px; border-radius: 999px; font-size: 11px; font-weight: 600; }
  .pill.Paid { background: #0f3b22; color: #5ee89a; }
  .pill.Outstanding { background: #0f2a4a; color: #7ab5ff; }
  .pill.Overdue { background: #3a1414; color: #ff8b8b; }
  .pill.Disputed { background: #3a2a14; color: #f5c26b; }
  .pill.Credit { background: #28124a; color: #c7a3ff; }
  .pill.Unknown { background: #222; color: #aaa; }
  .table-wrap { max-height: 560px; overflow: auto; border: 1px solid var(--border); border-radius: 10px; }
  footer { padding: 20px 28px; color: var(--muted); font-size: 12px; border-top: 1px solid var(--border); }
  .grid2 { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }
  .desc-trunc { max-width: 340px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; display: inline-block; }
  @media (max-width: 1100px) {
    .kpis { grid-template-columns: repeat(3, 1fr); }
    .charts { grid-template-columns: 1fr; }
    .charts-2 { grid-template-columns: 1fr; }
  }
</style>
</head>
<body>
<header>
  <h1>EPI Invoice List</h1>
  <div class="sub">
    <span id="meta-file"></span> · <span id="meta-sheet"></span> ·
    <span id="meta-rows"></span> invoices · <span id="meta-currency"></span>
  </div>
</header>
<main>
  <section class="kpis" id="kpis"></section>

  <section class="charts">
    <div class="card">
      <h2>Aging distribution <span class="hint">count per bucket</span></h2>
      <div id="chartAging"></div>
    </div>
    <div class="card">
      <h2>Status <span class="hint">by invoice count</span></h2>
      <div id="chartStatus"></div>
    </div>
  </section>

  <section class="charts-2">
    <div class="card">
      <h2>Monthly invoice count <span class="hint">by document date</span></h2>
      <div id="chartMonthly"></div>
    </div>
    <div class="card">
      <h2>Top 10 by |amount|</h2>
      <div id="chartTop10"></div>
    </div>
  </section>

  <section class="card" style="margin-bottom:18px">
    <h2>Source subtotals <span class="hint">SUM formulas preserved from Sheet1</span></h2>
    <div class="groups" id="groups"></div>
  </section>

  <section class="card">
    <h2>Invoice table</h2>
    <div class="filters">
      <input id="q" type="search" placeholder="Search text, reference, assignment..." />
      <select id="fStatus"><option value="">All statuses</option></select>
      <select id="fCurrency"><option value="">All currencies</option></select>
      <select id="fBucket"><option value="">All aging buckets</option></select>
      <span class="count" id="tableCount"></span>
    </div>
    <div class="table-wrap">
      <table id="tbl">
        <thead>
          <tr>
            <th data-k="reference">Reference</th>
            <th data-k="doc_date">Doc Date</th>
            <th data-k="text">Description</th>
            <th data-k="amount" class="num">Amount</th>
            <th data-k="currency">Cur</th>
            <th data-k="net_due_date">Net Due</th>
            <th data-k="days_late" class="num">Days Late</th>
            <th data-k="status">Status</th>
            <th data-k="aging_bucket">Aging</th>
            <th data-k="assignment">Assignment</th>
          </tr>
        </thead>
        <tbody id="tbody"></tbody>
      </table>
    </div>
  </section>
</main>
<footer>
  Generated <span id="gen-ts"></span> · Source <span id="f-src"></span> · Today anchor <span id="f-today"></span>
</footer>

<script id="epi-data" type="application/json">__DATA_JSON__</script>
<script>
(function(){
  const D = JSON.parse(document.getElementById('epi-data').textContent);
  const fmtMoney = (v) => {
    const n = Number(v || 0);
    const sign = n < 0 ? '-' : '';
    return sign + '$' + Math.abs(n).toLocaleString(undefined, {maximumFractionDigits: 2});
  };
  const fmtInt = (v) => Number(v || 0).toLocaleString();

  // Meta
  document.getElementById('meta-file').textContent = D.meta.source_path;
  document.getElementById('meta-sheet').textContent = 'Sheet1';
  document.getElementById('meta-rows').textContent = D.meta.row_count;
  const currencies = Object.keys(D.currency_counts);
  document.getElementById('meta-currency').textContent = currencies.join(', ') || '-';
  document.getElementById('gen-ts').textContent = D.meta.generated_at;
  document.getElementById('f-src').textContent = D.meta.source_path;
  document.getElementById('f-today').textContent = D.meta.today;

  // KPIs
  const kpis = [
    {label:'Invoices', val:fmtInt(D.kpis.total_count)},
    {label:'Net Total', val:fmtMoney(D.kpis.total_amount), cls:'small'},
    {label:'Outstanding', val:fmtMoney(D.kpis.total_outstanding), cls:'ok'},
    {label:'Overdue', val:fmtMoney(D.kpis.total_overdue), cls:'bad'},
    {label:'Disputed', val:fmtMoney(D.kpis.total_disputed), cls:'warn'},
    {label:'Avg days late (overdue)', val:Math.round(D.kpis.avg_days_late_overdue) + 'd'},
  ];
  document.getElementById('kpis').innerHTML = kpis.map(k =>
    `<div class="kpi ${k.cls||''}"><div class="label">${k.label}</div><div class="val ${k.cls==='small'?'small':''}">${k.val}</div></div>`
  ).join('');

  // Aging chart - horizontal bar
  const agingOrder = ["Not due","0-30","31-60","61-90","91-180","181-365","365+"];
  new ApexCharts(document.getElementById('chartAging'), {
    chart: {type:'bar', height: 300, toolbar:{show:false}, background:'transparent', foreColor:'#cfd8ec'},
    theme: {mode:'dark'},
    series: [{name:'Invoices', data: agingOrder.map(b => D.aging_counts[b] || 0)}],
    plotOptions: {bar: {horizontal: true, borderRadius: 3, distributed: true}},
    colors: ['#4ade80','#60a5fa','#fbbf24','#f97316','#ef4444','#b91c1c','#7f1d1d'],
    xaxis: {categories: agingOrder},
    dataLabels: {enabled: true, style: {colors:['#fff']}},
    legend: {show: false},
    grid: {borderColor: '#24324f'},
  }).render();

  // Status donut
  const statusLabels = Object.keys(D.status_counts);
  const statusValues = statusLabels.map(l => D.status_counts[l]);
  const statusColorMap = {
    Paid:'#22c55e', Outstanding:'#60a5fa', Overdue:'#ef4444',
    Disputed:'#f59e0b', Credit:'#a78bfa', Unknown:'#64748b'
  };
  new ApexCharts(document.getElementById('chartStatus'), {
    chart: {type:'donut', height: 300, background:'transparent', foreColor:'#cfd8ec'},
    theme: {mode:'dark'},
    series: statusValues,
    labels: statusLabels,
    colors: statusLabels.map(l => statusColorMap[l] || '#64748b'),
    legend: {position: 'bottom'},
    dataLabels: {enabled: true},
    plotOptions: {pie: {donut: {size:'60%'}}},
  }).render();

  // Monthly line
  new ApexCharts(document.getElementById('chartMonthly'), {
    chart: {type:'area', height: 300, toolbar:{show:false}, background:'transparent', foreColor:'#cfd8ec'},
    theme: {mode:'dark'},
    series: [{name: 'Invoices', data: D.monthly.map(([k,v]) => v.count)}],
    xaxis: {categories: D.monthly.map(([k]) => k), labels: {rotate: -45}},
    colors: ['#60a5fa'],
    stroke: {curve:'smooth', width: 2},
    fill: {type:'gradient', gradient:{shadeIntensity:1, opacityFrom:0.4, opacityTo:0.05}},
    dataLabels: {enabled: false},
    grid: {borderColor: '#24324f'},
  }).render();

  // Top10 horizontal bar
  new ApexCharts(document.getElementById('chartTop10'), {
    chart: {type:'bar', height: 360, toolbar:{show:false}, background:'transparent', foreColor:'#cfd8ec'},
    theme: {mode:'dark'},
    series: [{name: 'Amount', data: D.top10.map(i => i.amount)}],
    plotOptions: {bar: {horizontal: true, borderRadius: 2, distributed: true,
                         colors: {ranges: [{from: -1e18, to: 0, color: '#ef4444'},
                                           {from: 0, to: 1e18, color: '#60a5fa'}]}}},
    xaxis: {categories: D.top10.map(i => (i.reference || '(blank)') + ' · r' + i.row),
            labels: {formatter: (v) => fmtMoney(v)}},
    dataLabels: {enabled: true, formatter: (v) => fmtMoney(v), style:{colors:['#fff']}},
    legend: {show: false},
    grid: {borderColor: '#24324f'},
    tooltip: {y: {formatter: (v) => fmtMoney(v)}},
  }).render();

  // Groups / subtotals
  document.getElementById('groups').innerHTML = D.groups.map(g => `
    <div class="group">
      <div class="lbl">${g.label} · ${g.count} rows</div>
      <div class="amt">${fmtMoney(g.amount)}</div>
      <div class="lbl" style="margin-top:6px">Source subtotal (row ${g.subtotal_row})</div>
      <div class="amt">${fmtMoney(g.subtotal_value)}</div>
      <div class="frm">${g.formula}</div>
    </div>
  `).join('');

  // Table + filters
  const tbody = document.getElementById('tbody');
  const qEl = document.getElementById('q');
  const sEl = document.getElementById('fStatus');
  const cEl = document.getElementById('fCurrency');
  const bEl = document.getElementById('fBucket');
  const countEl = document.getElementById('tableCount');

  // Populate selects
  const uniq = (arr) => Array.from(new Set(arr)).filter(x => x !== undefined && x !== null);
  uniq(D.invoices.map(i => i.status)).sort().forEach(s => {
    const o = document.createElement('option'); o.value = s; o.textContent = s; sEl.appendChild(o);
  });
  uniq(D.invoices.map(i => i.currency)).sort().forEach(c => {
    const o = document.createElement('option'); o.value = c; o.textContent = c || '(none)'; cEl.appendChild(o);
  });
  ["Not due","0-30","31-60","61-90","91-180","181-365","365+"].forEach(b => {
    const o = document.createElement('option'); o.value = b; o.textContent = b; bEl.appendChild(o);
  });

  let sortKey = 'amount';
  let sortDir = -1; // desc by default

  function truncate(s, n) {
    s = s || '';
    return s.length > n ? s.slice(0, n-1) + '…' : s;
  }

  function render() {
    const q = qEl.value.trim().toLowerCase();
    const s = sEl.value;
    const c = cEl.value;
    const b = bEl.value;
    let rows = D.invoices.filter(i => {
      if (s && i.status !== s) return false;
      if (c && i.currency !== c) return false;
      if (b && i.aging_bucket !== b) return false;
      if (q) {
        const hay = (i.reference + ' ' + (i.text||'') + ' ' + (i.assignment||'') + ' ' + i.status).toLowerCase();
        if (!hay.includes(q)) return false;
      }
      return true;
    });
    rows.sort((a, b2) => {
      const va = a[sortKey], vb = b2[sortKey];
      if (va == null && vb == null) return 0;
      if (va == null) return 1;
      if (vb == null) return -1;
      if (typeof va === 'number' && typeof vb === 'number') return (va - vb) * sortDir;
      return String(va).localeCompare(String(vb)) * sortDir;
    });
    countEl.textContent = rows.length + ' of ' + D.invoices.length + ' invoices';
    tbody.innerHTML = rows.map(i => `
      <tr>
        <td class="ref">${i.reference || '<span style="color:var(--muted)">(blank)</span>'}</td>
        <td>${i.doc_date || ''}</td>
        <td><span class="desc-trunc" title="${(i.text||'').replace(/"/g,'&quot;')}">${truncate(i.text || '', 60)}</span></td>
        <td class="num">${fmtMoney(i.amount)}</td>
        <td>${i.currency || ''}</td>
        <td>${i.net_due_date || ''}</td>
        <td class="num">${i.days_late}</td>
        <td><span class="pill ${i.status}">${i.status}</span></td>
        <td>${i.aging_bucket}</td>
        <td>${i.assignment || ''}</td>
      </tr>
    `).join('');
  }

  document.querySelectorAll('thead th').forEach(th => {
    th.addEventListener('click', () => {
      const k = th.dataset.k;
      if (!k) return;
      if (sortKey === k) sortDir = -sortDir;
      else { sortKey = k; sortDir = (k === 'amount' || k === 'days_late') ? -1 : 1; }
      render();
    });
  });
  [qEl, sEl, cEl, bEl].forEach(el => el.addEventListener('input', render));
  render();
})();
</script>
</body>
</html>
"""


def build() -> None:
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    data = extract()
    # Safe JSON embedding - avoid </script> injection and ensure valid inside <script>
    payload = json.dumps(data, default=str, ensure_ascii=False).replace("</", "<\\/")
    html = HTML_TEMPLATE.replace("__DATA_JSON__", payload)
    OUT_PATH.write_text(html, encoding="utf-8")

    # Console report
    kpis = data["kpis"]
    print(f"[OK] Invoices extracted : {kpis['total_count']}")
    print(f"[OK] Subtotal rows      : {len(data['subtotals'])} ({[s['row'] for s in data['subtotals']]})")
    print(f"[OK] Currencies         : {data['currency_counts']}")
    print(f"[OK] Status distribution: {data['status_counts']}")
    print(f"[OK] Aging distribution : {data['aging_counts']}")
    print(f"[OK] Total amount       : {kpis['total_amount']:.2f}")
    print(f"[OK] Output             : {OUT_PATH}")
    print(f"[OK] Size               : {OUT_PATH.stat().st_size:,} bytes")


if __name__ == "__main__":
    build()
