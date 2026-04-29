"""
Ground-truth SOA visualization builder for ETH SOA 30.1.26.xlsx.

Reads the workbook DIRECTLY with openpyxl (does NOT use V6/parser.py) and emits
a single self-contained HTML file with embedded JSON + ApexCharts rendering.

This is the benchmark the production parser/dashboard must match.
"""

from __future__ import annotations

import json
import os
import re
import sys
from dataclasses import dataclass, asdict, field
from datetime import datetime, date
from typing import Any

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "New info", "ETH SOA 30.1.26.xlsx")
OUT_DIR = os.path.join(HERE, "TESTEXCEL")
OUT = os.path.join(OUT_DIR, "ETH_SOA_30_1_26.html")

SHEET = "SoA 26.1.26"
HEADER_ROW = 7

# Canonical section banner -> normalized key
SECTION_MAP = [
    ("credits usable", "Credits Usable"),
    ("totalcare charges", "TotalCare Charges"),
    ("customer responsible charges", "Customer Responsible Charges"),
    ("spare parts charges", "Spare Parts Charges"),
    ("late payment interest", "Late Payment Interest"),
]

AGING_BUCKETS = [
    ("Current", lambda d: d is not None and d <= 0),
    ("1-30 Days", lambda d: d is not None and 1 <= d <= 30),
    ("31-60 Days", lambda d: d is not None and 31 <= d <= 60),
    ("61-90 Days", lambda d: d is not None and 61 <= d <= 90),
    ("91-180 Days", lambda d: d is not None and 91 <= d <= 180),
    ("180+ Days", lambda d: d is not None and d > 180),
    ("Unknown", lambda d: d is None),
]


def parse_date(v: Any) -> datetime | None:
    """Accept datetime, date, or dd/mm/yyyy / yyyy-mm-dd string."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        # dd/mm/yyyy
        m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
        if m:
            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            try:
                return datetime(y, mo, d)
            except ValueError:
                return None
        # yyyy-mm-dd [hh:mm:ss]
        m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", s)
        if m:
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            try:
                return datetime(y, mo, d)
            except ValueError:
                return None
    return None


def to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.replace(",", "").replace("$", "").replace("USD", "").strip()
        try:
            return float(s)
        except ValueError:
            return None
    return None


def iso(d: datetime | None) -> str | None:
    return d.strftime("%Y-%m-%d") if d else None


def aging_bucket(days_late: float | None) -> str:
    for name, pred in AGING_BUCKETS:
        if pred(days_late):
            return name
    return "Unknown"


def classify_status(action: str | None, days_late: float | None, amount: float | None) -> str:
    """Derive a coarse status label from Action/Awaiting text + days_late."""
    a = (action or "").lower().strip()
    # Explicit keywords win
    if "disput" in a:
        return "Disputed"
    if "paid" in a:
        return "Paid"
    if "ready for payment" in a or "ready" in a:
        return "Ready for Payment"
    if "ongoing" in a:
        return "Disputed"
    # Credit notes (negative amounts in credit sections)
    if amount is not None and amount < 0:
        return "Credit"
    # Fall back on days_late
    if days_late is None:
        return "Unknown"
    if days_late > 0:
        return "Overdue"
    return "Current"


@dataclass
class LineItem:
    row: int
    section: str
    company: str | None
    account: str | None
    reference: str | None
    document_date: str | None
    net_due_date: str | None
    amount: float | None
    currency: str | None
    text: str | None
    assignment: str | None
    rr_comments: str | None
    action_owner: str | None
    days_late_sheet: float | None   # col L value (from sheet)
    eth_comments: str | None
    eth_po_ref: str | None
    lpi_cumulated: float | None
    # derived
    days_late: float | None = None
    aging: str = "Unknown"
    status: str = "Unknown"


def norm_str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def main() -> None:
    os.makedirs(OUT_DIR, exist_ok=True)
    wb = openpyxl.load_workbook(SRC, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"FATAL: sheet {SHEET!r} not found. Found: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[SHEET]

    # Metadata (rows 1-6)
    meta: dict[str, Any] = {
        "title": norm_str(ws.cell(1, 1).value) or "Ethiopian Statement of Account",
        "customer_name": norm_str(ws.cell(2, 3).value),
        "customer_number": norm_str(ws.cell(3, 3).value),
        "contact_email": norm_str(ws.cell(4, 3).value),
        "lpi_rate_label": norm_str(ws.cell(2, 11).value),
        "lpi_rate": to_float(ws.cell(2, 12).value),
        "today_label": norm_str(ws.cell(4, 11).value),
        "today": iso(parse_date(ws.cell(4, 12).value)),
        "avg_days_late_label": norm_str(ws.cell(5, 11).value),
        "avg_days_late": to_float(ws.cell(5, 12).value),
        "source_file": os.path.abspath(SRC),
        "sheet": SHEET,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }

    # Reference "today" for days_late computation — prefer sheet's cell L4
    today_ref = parse_date(ws.cell(4, 12).value) or datetime.now()

    # Walk rows, tracking section
    sections: list[dict[str, Any]] = []
    current_section: str | None = None
    current_items: list[LineItem] = []
    section_totals_from_sheet: dict[str, float | None] = {}
    warnings: list[str] = []

    # Also capture totals announced in sheet (rows where E='Total' or 'Overdue')
    sheet_totals: list[dict[str, Any]] = []

    def close_section() -> None:
        nonlocal current_section, current_items
        if current_section is not None:
            total = sum(i.amount for i in current_items if i.amount is not None)
            sections.append({
                "name": current_section,
                "items": [asdict(i) for i in current_items],
                "computed_total": total,
                "item_count": len(current_items),
            })
        current_section = None
        current_items = []

    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        d = ws.cell(r, 4).value
        e = ws.cell(r, 5).value
        f = ws.cell(r, 6).value
        g = ws.cell(r, 7).value
        h = ws.cell(r, 8).value
        i_ = ws.cell(r, 9).value
        j = ws.cell(r, 10).value
        k = ws.cell(r, 11).value
        l = ws.cell(r, 12).value
        m = ws.cell(r, 13).value
        n = ws.cell(r, 14).value
        o = ws.cell(r, 15).value

        a_s = norm_str(a)

        # Section banner?
        if a_s:
            al = a_s.lower()
            matched_sec = None
            for key, canon in SECTION_MAP:
                if al.startswith(key[:15]):
                    matched_sec = canon
                    break
            if matched_sec and not str(a_s).strip().isdigit():
                # new section
                close_section()
                current_section = matched_sec
                continue

        # Total / Overdue / Available Credit row?
        e_s = norm_str(e)
        if e_s and e_s.lower() in ("total", "overdue", "available credit", "total overdue"):
            sheet_totals.append({
                "row": r,
                "section": current_section,
                "label": e_s,
                "amount": to_float(f),
            })
            if e_s.lower() == "total" and current_section is not None:
                section_totals_from_sheet[current_section] = to_float(f)
            continue

        # Detail row: must have a reference AND amount
        ref = norm_str(c)
        amt = to_float(f)
        if ref is None or amt is None:
            # fully blank or metadata row — skip silently
            if any(v not in (None, "") for v in (a, b, c, d, e, f, g, h, i_, j, k, l, m, n, o)):
                # non-empty but unparseable
                if current_section is not None and (c is not None or f is not None):
                    warnings.append(f"row {r}: partial/unparseable row dropped (ref={c!r}, amount={f!r})")
            continue

        if current_section is None:
            warnings.append(f"row {r}: detail row before any section banner — skipped")
            continue

        net_due = parse_date(e)
        days_late_calc: float | None = None
        if net_due is not None:
            delta = (today_ref - net_due).days
            days_late_calc = float(delta) if delta > 0 else 0.0
        else:
            # fallback to sheet-provided col L
            sheet_dl = to_float(l)
            days_late_calc = sheet_dl if sheet_dl is not None else None

        item = LineItem(
            row=r,
            section=current_section,
            company=norm_str(a),
            account=norm_str(b),
            reference=ref,
            document_date=iso(parse_date(d)),
            net_due_date=iso(net_due),
            amount=amt,
            currency=norm_str(g),
            text=norm_str(h),
            assignment=norm_str(i_),
            rr_comments=norm_str(j),
            action_owner=norm_str(k),
            days_late_sheet=to_float(l),
            eth_comments=norm_str(m),
            eth_po_ref=norm_str(n),
            lpi_cumulated=to_float(o),
            days_late=days_late_calc,
            aging=aging_bucket(days_late_calc),
        )
        item.status = classify_status(item.action_owner, item.days_late, item.amount)
        current_items.append(item)

    close_section()

    # Flat item list
    flat = [it for sec in sections for it in sec["items"]]

    # Aging breakdown
    aging_breakdown = []
    for name, pred in AGING_BUCKETS:
        bucket_items = [it for it in flat if it["aging"] == name]
        aging_breakdown.append({
            "bucket": name,
            "count": len(bucket_items),
            "total_amount": sum((it["amount"] or 0) for it in bucket_items),
            "overdue_amount": sum(
                (it["amount"] or 0) for it in bucket_items
                if (it["amount"] or 0) > 0 and name not in ("Current", "Unknown")
            ),
        })

    # Status breakdown
    status_counts: dict[str, dict[str, float]] = {}
    for it in flat:
        s = it["status"]
        entry = status_counts.setdefault(s, {"count": 0, "total_amount": 0.0})
        entry["count"] += 1
        entry["total_amount"] += it["amount"] or 0
    status_breakdown = [
        {"status": k, "count": v["count"], "total_amount": v["total_amount"]}
        for k, v in sorted(status_counts.items(), key=lambda kv: -kv[1]["count"])
    ]

    # Top-10 overdue (positive amount, days_late > 0, sort desc by amount)
    overdue = sorted(
        [it for it in flat if (it["amount"] or 0) > 0 and (it["days_late"] or 0) > 0],
        key=lambda it: -(it["amount"] or 0),
    )[:10]

    # Summary
    total_debit = sum((it["amount"] or 0) for it in flat if (it["amount"] or 0) > 0)
    total_credit = sum((it["amount"] or 0) for it in flat if (it["amount"] or 0) < 0)
    net_balance = total_debit + total_credit
    total_overdue = sum(
        (it["amount"] or 0) for it in flat
        if (it["amount"] or 0) > 0 and (it["days_late"] or 0) > 0
    )
    days_late_vals = [it["days_late"] for it in flat if it["days_late"] is not None and it["days_late"] > 0]
    avg_days_late_calc = (sum(days_late_vals) / len(days_late_vals)) if days_late_vals else 0

    credits_usable_total = next(
        (s["computed_total"] for s in sections if s["name"] == "Credits Usable"), 0
    )

    summary = {
        "total_items": len(flat),
        "total_sections": len(sections),
        "total_debit": total_debit,
        "total_credit": total_credit,
        "net_balance": net_balance,
        "total_overdue": total_overdue,
        "avg_days_late_calc": avg_days_late_calc,
        "credits_usable_total": credits_usable_total,
        "today_ref": iso(today_ref),
    }

    data = {
        "metadata": meta,
        "sections": sections,
        "flat_items": flat,
        "aging_breakdown": aging_breakdown,
        "status_breakdown": status_breakdown,
        "top_overdue": overdue,
        "sheet_totals": sheet_totals,
        "section_totals_from_sheet": section_totals_from_sheet,
        "summary": summary,
        "warnings": warnings,
    }

    # Sanity/audit comparison vs sheet totals.
    # Observed quirk in ETH SOA 30.1.26:
    #  - Credits Usable / TotalCare / LPI: sheet 'Total' == sum of all rows in section.
    #  - CRC: sheet 'Total' == sum of POSITIVE rows only (debits), not net.
    #        Credit notes are split out at row 66 as 'Available Credit'.
    #  - Spare Parts: sheet 'Total' (581,179) appears to be a hand-entered
    #        overdue figure that does not match any clean derivation from
    #        the row-level due dates. Documented as a data-quality note.
    audit_lines: list[str] = []
    for sec in sections:
        sheet_t = section_totals_from_sheet.get(sec["name"])
        computed = sec["computed_total"]
        pos_sum = sum(it["amount"] for it in sec["items"] if (it["amount"] or 0) > 0)
        neg_sum = sum(it["amount"] for it in sec["items"] if (it["amount"] or 0) < 0)
        if sheet_t is None:
            audit_lines.append(f"NO SHEET TOTAL {sec['name']}: computed={computed:,.2f}")
            continue
        if abs(sheet_t - computed) <= 0.01:
            audit_lines.append(f"OK  {sec['name']}: computed={computed:,.2f} matches sheet Total")
        elif abs(sheet_t - pos_sum) <= 0.01:
            audit_lines.append(
                f"NOTE {sec['name']}: sheet Total ({sheet_t:,.2f}) == debits only; "
                f"net computed = {computed:,.2f} (credits {neg_sum:,.2f})"
            )
        else:
            audit_lines.append(
                f"DATA-QUALITY {sec['name']}: sheet Total {sheet_t:,.2f} does not match "
                f"computed net {computed:,.2f} or debits-only {pos_sum:,.2f} — "
                f"sheet value may be a manually entered overdue figure; "
                f"using computed net for visualizations."
            )
    data["audit"] = audit_lines

    # Print for console
    print("=== BUILD REPORT ===")
    print(f"Items: {len(flat)}  Sections: {len(sections)}")
    for sec in sections:
        print(f"  - {sec['name']}: {sec['item_count']} items, total {sec['computed_total']:,.2f}")
    print(f"Debit total:   {total_debit:,.2f}")
    print(f"Credit total:  {total_credit:,.2f}")
    print(f"Net balance:   {net_balance:,.2f}")
    print(f"Total overdue: {total_overdue:,.2f}")
    for line in audit_lines:
        print("  " + line)
    for w in warnings:
        print("  WARN:", w)

    html = build_html(data)
    with open(OUT, "w", encoding="utf-8") as fh:
        fh.write(html)
    print(f"\nWrote {OUT} ({os.path.getsize(OUT):,} bytes)")


def build_html(data: dict[str, Any]) -> str:
    # JSON blob (keep compact)
    blob = json.dumps(data, default=str, separators=(",", ":"))
    return HTML_TEMPLATE.replace("__DATA_JSON__", blob)


HTML_TEMPLATE = r"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8" />
<title>ETH SOA 30.1.26 — Ground Truth</title>
<meta name="viewport" content="width=device-width,initial-scale=1" />
<script src="https://cdn.jsdelivr.net/npm/apexcharts@3.49.0/dist/apexcharts.min.js"></script>
<style>
  :root{
    --bg:#0f1115; --panel:#171a21; --panel2:#1c2029; --border:#262b36;
    --ink:#e7ecf3; --muted:#8a93a5; --accent:#4aa8ff; --good:#36d399;
    --warn:#f7b955; --bad:#ef4a6a; --credit:#8b5cf6;
    --radius:10px;
  }
  *{box-sizing:border-box}
  html,body{margin:0;background:var(--bg);color:var(--ink);
    font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Inter,"Helvetica Neue",Arial,sans-serif;
    font-size:14px;line-height:1.4}
  a{color:var(--accent)}
  h1,h2,h3{margin:0 0 .5em;font-weight:600;letter-spacing:.2px}
  .wrap{max-width:1400px;margin:0 auto;padding:24px}
  header.top{padding:24px 24px 16px;background:linear-gradient(180deg,#1b2130,#14171e);
    border:1px solid var(--border);border-radius:var(--radius);margin-bottom:20px}
  .title{font-size:22px;font-weight:700}
  .subtitle{color:var(--muted);margin-top:4px;font-size:13px}
  .meta-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:10px;margin-top:14px}
  .meta-grid .cell{background:var(--panel2);border:1px solid var(--border);border-radius:8px;padding:10px 12px}
  .meta-grid .k{color:var(--muted);font-size:11px;text-transform:uppercase;letter-spacing:.8px}
  .meta-grid .v{font-size:14px;margin-top:3px;word-break:break-word}
  .kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:12px;margin-bottom:20px}
  .kpi{background:var(--panel);border:1px solid var(--border);border-radius:var(--radius);padding:16px}
  .kpi .label{color:var(--muted);font-size:11px;text-transform:uppercase;letter-spacing:.8px}
  .kpi .value{font-size:24px;font-weight:700;margin-top:6px}
  .kpi .sub{color:var(--muted);font-size:11px;margin-top:3px}
  .kpi.good .value{color:var(--good)} .kpi.bad .value{color:var(--bad)}
  .kpi.warn .value{color:var(--warn)} .kpi.credit .value{color:var(--credit)}
  .grid-2{display:grid;grid-template-columns:1fr 1fr;gap:16px}
  .grid-1{display:grid;grid-template-columns:1fr;gap:16px}
  .card{background:var(--panel);border:1px solid var(--border);border-radius:var(--radius);padding:16px;margin-bottom:16px}
  .card h3{font-size:14px;color:var(--muted);margin-bottom:12px;text-transform:uppercase;letter-spacing:.6px}
  .controls{display:flex;flex-wrap:wrap;gap:10px;margin-bottom:12px;align-items:center}
  .controls input,.controls select{background:var(--panel2);border:1px solid var(--border);color:var(--ink);
    padding:8px 10px;border-radius:6px;font:inherit}
  .controls input{flex:1;min-width:200px}
  .controls label{font-size:12px;color:var(--muted)}
  details{background:var(--panel);border:1px solid var(--border);border-radius:var(--radius);margin-bottom:12px}
  details summary{padding:14px 16px;cursor:pointer;font-weight:600;list-style:none;
    display:flex;justify-content:space-between;align-items:center}
  details summary::-webkit-details-marker{display:none}
  details summary::after{content:"+";color:var(--muted);font-size:18px}
  details[open] summary::after{content:"–"}
  details[open] summary{border-bottom:1px solid var(--border)}
  table{width:100%;border-collapse:collapse;font-size:13px}
  th,td{text-align:left;padding:8px 10px;border-bottom:1px solid var(--border);vertical-align:top}
  th{background:var(--panel2);position:sticky;top:0;color:var(--muted);font-weight:600;
    text-transform:uppercase;font-size:11px;letter-spacing:.6px;cursor:pointer;user-select:none}
  th:hover{color:var(--ink)}
  td.num{text-align:right;font-variant-numeric:tabular-nums;font-family:"JetBrains Mono","SF Mono",Menlo,monospace}
  td.neg{color:var(--bad)} td.pos{color:var(--ink)}
  tr.credit td{color:var(--credit)}
  .tag{display:inline-block;padding:2px 8px;border-radius:100px;font-size:11px;font-weight:600;
    background:var(--panel2);border:1px solid var(--border)}
  .tag.overdue{background:rgba(239,74,106,.15);color:var(--bad);border-color:rgba(239,74,106,.4)}
  .tag.current{background:rgba(54,211,153,.15);color:var(--good);border-color:rgba(54,211,153,.4)}
  .tag.credit{background:rgba(139,92,246,.15);color:var(--credit);border-color:rgba(139,92,246,.4)}
  .tag.disputed{background:rgba(247,185,85,.15);color:var(--warn);border-color:rgba(247,185,85,.4)}
  .tag.ready{background:rgba(74,168,255,.15);color:var(--accent);border-color:rgba(74,168,255,.4)}
  .tag.paid{background:rgba(54,211,153,.15);color:var(--good);border-color:rgba(54,211,153,.4)}
  .tag.unknown{background:rgba(138,147,165,.15);color:var(--muted)}
  .section-meta{display:flex;gap:18px;font-size:12px;color:var(--muted)}
  .warning-banner{background:rgba(247,185,85,.1);border:1px solid rgba(247,185,85,.4);
    color:var(--warn);padding:10px 14px;border-radius:8px;margin-bottom:12px;font-size:13px}
  .footer{color:var(--muted);font-size:12px;padding:20px 0;border-top:1px solid var(--border);margin-top:24px}
  .footer code{background:var(--panel2);padding:2px 6px;border-radius:4px;font-size:11px}
  .scroll-x{overflow-x:auto}
  @media print {
    :root{--bg:#fff;--panel:#fff;--panel2:#f5f5f5;--border:#ddd;--ink:#000;--muted:#555}
    body{font-size:11px}
    details{break-inside:avoid} .card,.kpi{break-inside:avoid}
  }
  @media (max-width:860px){ .grid-2{grid-template-columns:1fr} }
</style>
</head>
<body>
<div class="wrap">

  <header class="top">
    <div class="title" id="hdr-title">Loading…</div>
    <div class="subtitle" id="hdr-subtitle"></div>
    <div class="meta-grid" id="meta-grid"></div>
  </header>

  <div class="kpis" id="kpis"></div>

  <div class="grid-2">
    <div class="card"><h3>Aging Buckets</h3><div id="chart-aging"></div></div>
    <div class="card"><h3>Payment Status</h3><div id="chart-status"></div></div>
  </div>

  <div class="grid-2">
    <div class="card"><h3>Debit / Credit / Net</h3><div id="chart-balance"></div></div>
    <div class="card"><h3>Section Totals</h3><div id="chart-sections"></div></div>
  </div>

  <div class="card">
    <h3>Top 10 Overdue Invoices (by amount)</h3>
    <div id="chart-top"></div>
  </div>

  <div class="card">
    <h3>Filter Line Items</h3>
    <div class="controls">
      <input id="f-search" placeholder="Search reference / document / text…" />
      <label>Status
        <select id="f-status"><option value="">All</option></select>
      </label>
      <label>Section
        <select id="f-section"><option value="">All</option></select>
      </label>
      <label>Aging
        <select id="f-aging"><option value="">All</option></select>
      </label>
      <span id="f-count" style="color:var(--muted);font-size:12px"></span>
    </div>
  </div>

  <div id="sections-container"></div>

  <div class="card">
    <h3>Build Audit</h3>
    <pre id="audit-log" style="white-space:pre-wrap;color:var(--muted);font-family:monospace;font-size:12px;margin:0"></pre>
  </div>

  <div class="footer" id="footer"></div>

</div>

<script>
const DATA = __DATA_JSON__;

const fmt = {
  money(n){
    if(n === null || n === undefined || isNaN(n)) return '—';
    const sign = n < 0 ? '-' : '';
    const abs = Math.abs(n);
    return sign + abs.toLocaleString('en-US', {minimumFractionDigits:2, maximumFractionDigits:2});
  },
  int(n){ return (n==null) ? '—' : Number(n).toLocaleString('en-US'); },
  date(s){ return s || '—'; },
  days(n){ return (n==null) ? '—' : Math.round(n).toString(); },
};

function statusClass(s){
  const k=(s||'').toLowerCase();
  if(k.includes('overdue')) return 'overdue';
  if(k.includes('current')) return 'current';
  if(k.includes('credit')) return 'credit';
  if(k.includes('disput')) return 'disputed';
  if(k.includes('ready')) return 'ready';
  if(k.includes('paid')) return 'paid';
  return 'unknown';
}

// ===== Header =====
(function renderHeader(){
  const m = DATA.metadata;
  document.getElementById('hdr-title').textContent = m.customer_name || 'Statement of Account';
  document.getElementById('hdr-subtitle').textContent =
    `Customer #${m.customer_number || '—'} · Sheet "${m.sheet}" · ${DATA.summary.total_items} line items across ${DATA.summary.total_sections} sections`;
  const cells = [
    ['Customer', m.customer_name],
    ['Customer #', m.customer_number],
    ['Contact', m.contact_email],
    ['LPI Rate', m.lpi_rate != null ? m.lpi_rate.toFixed(5) : '—'],
    ['Report "Today"', m.today],
    ['Avg Days Late (sheet)', m.avg_days_late != null ? m.avg_days_late.toFixed(2) : '—'],
  ];
  const g = document.getElementById('meta-grid');
  g.innerHTML = cells.map(([k,v])=>`<div class="cell"><div class="k">${k}</div><div class="v">${v ?? '—'}</div></div>`).join('');
})();

// ===== KPIs =====
(function renderKPIs(){
  const s = DATA.summary;
  const kpis = [
    {label:'Total Overdue',        value: fmt.money(s.total_overdue),       sub:'positive amounts past due',             cls:'bad'},
    {label:'Avg Days Late',        value: s.avg_days_late_calc.toFixed(1),  sub:'across overdue items',                  cls:'warn'},
    {label:'LPI Rate',             value: DATA.metadata.lpi_rate != null ? DATA.metadata.lpi_rate.toFixed(5) : '—', sub:'per sheet', cls:''},
    {label:'Credits Usable',       value: fmt.money(s.credits_usable_total), sub:'available offset',                     cls:'credit'},
    {label:'Net Balance',          value: fmt.money(s.net_balance),          sub:'debits + credits',                     cls: s.net_balance >= 0 ? 'bad' : 'good'},
  ];
  document.getElementById('kpis').innerHTML = kpis.map(k=>
    `<div class="kpi ${k.cls}"><div class="label">${k.label}</div><div class="value">${k.value}</div><div class="sub">${k.sub}</div></div>`
  ).join('');
})();

// ===== Charts =====
const COMMON = {
  chart:{ foreColor:'#e7ecf3', toolbar:{show:false}, animations:{enabled:true,speed:400} },
  tooltip:{ theme:'dark' },
  grid:{ borderColor:'#262b36' },
  dataLabels:{ enabled:false },
  legend:{ labels:{ colors:'#e7ecf3' } },
};

// 1. Aging
(function(){
  const b = DATA.aging_breakdown;
  const opts = {
    ...COMMON,
    chart:{...COMMON.chart, type:'bar', height:300},
    series:[
      {name:'Item Count',    data: b.map(x=>x.count)},
      {name:'Total (abs)',   data: b.map(x=>Math.abs(x.total_amount))},
    ],
    xaxis:{ categories: b.map(x=>x.bucket) },
    plotOptions:{ bar:{ horizontal:true, borderRadius:3 } },
    colors:['#4aa8ff','#ef4a6a'],
    yaxis:[
      {labels:{formatter:v=>Math.round(v)}},
    ],
  };
  new ApexCharts(document.getElementById('chart-aging'), opts).render();
})();

// 2. Status
(function(){
  const s = DATA.status_breakdown;
  const opts = {
    ...COMMON,
    chart:{...COMMON.chart, type:'donut', height:300},
    series: s.map(x=>x.count),
    labels: s.map(x=>x.status),
    colors:['#ef4a6a','#36d399','#8b5cf6','#f7b955','#4aa8ff','#8a93a5'],
    legend:{position:'bottom', labels:{colors:'#e7ecf3'}},
    plotOptions:{ pie:{ donut:{ size:'60%', labels:{show:true, total:{show:true, label:'Total Items', color:'#8a93a5'} } } } },
  };
  new ApexCharts(document.getElementById('chart-status'), opts).render();
})();

// 3. Balance overview
(function(){
  const s = DATA.summary;
  const opts = {
    ...COMMON,
    chart:{...COMMON.chart, type:'bar', height:300},
    series:[{name:'USD', data:[
      {x:'Total Debit',  y: s.total_debit,  fillColor:'#ef4a6a'},
      {x:'Total Credit', y: s.total_credit, fillColor:'#8b5cf6'},
      {x:'Net Balance',  y: s.net_balance,  fillColor: s.net_balance>=0?'#ef4a6a':'#36d399'},
    ]}],
    plotOptions:{ bar:{ horizontal:true, borderRadius:3, distributed:false, colors:{ranges:[]} } },
    yaxis:{ labels:{ formatter: v=>fmt.money(v) } },
    xaxis:{ labels:{ formatter: v=>fmt.money(v) } },
  };
  new ApexCharts(document.getElementById('chart-balance'), opts).render();
})();

// 4. Section totals
(function(){
  const s = DATA.sections;
  const opts = {
    ...COMMON,
    chart:{...COMMON.chart, type:'bar', height:300},
    series:[{name:'USD', data: s.map(x=>Number(x.computed_total.toFixed(2))) }],
    xaxis:{ categories: s.map(x=>x.name) },
    plotOptions:{ bar:{ horizontal:true, borderRadius:3, colors:{ranges:[
      {from:-Infinity,to:-0.01,color:'#8b5cf6'},
      {from:0,to:Infinity,color:'#4aa8ff'}
    ]} } },
    yaxis:{ labels:{ formatter: v=>fmt.money(v) } },
  };
  new ApexCharts(document.getElementById('chart-sections'), opts).render();
})();

// 5. Top 10 overdue
(function(){
  const t = DATA.top_overdue;
  const opts = {
    ...COMMON,
    chart:{...COMMON.chart, type:'bar', height:360},
    series:[{name:'Amount (USD)', data: t.map(x=>Number((x.amount||0).toFixed(2))) }],
    xaxis:{ categories: t.map(x=>`${x.reference} (${Math.round(x.days_late||0)}d)`) },
    plotOptions:{ bar:{ horizontal:true, borderRadius:3, distributed:true } },
    colors:['#ef4a6a'],
    tooltip:{ theme:'dark', y:{ formatter: v=>fmt.money(v) },
      custom: ({dataPointIndex})=>{
        const it = t[dataPointIndex];
        return `<div style="padding:8px 10px;background:#171a21;border:1px solid #262b36">
          <div style="font-weight:600;margin-bottom:4px">${it.reference}</div>
          <div style="color:#8a93a5;font-size:11px;margin-bottom:4px">${it.text||''}</div>
          <div><b>${fmt.money(it.amount)} ${it.currency||''}</b></div>
          <div style="color:#8a93a5">Due ${it.net_due_date||'—'} · ${Math.round(it.days_late||0)} days late</div>
          <div style="color:#8a93a5">Section: ${it.section}</div>
        </div>`;
      }
    },
  };
  new ApexCharts(document.getElementById('chart-top'), opts).render();
})();

// ===== Filters & Tables =====
const FILTERS = { search:'', status:'', section:'', aging:'' };

function buildFilterOptions(){
  const statuses = [...new Set(DATA.flat_items.map(i=>i.status))].sort();
  const sections = DATA.sections.map(s=>s.name);
  const agings = ['Current','1-30 Days','31-60 Days','61-90 Days','91-180 Days','180+ Days','Unknown'];
  document.getElementById('f-status').innerHTML  = '<option value="">All</option>'+statuses.map(s=>`<option>${s}</option>`).join('');
  document.getElementById('f-section').innerHTML = '<option value="">All</option>'+sections.map(s=>`<option>${s}</option>`).join('');
  document.getElementById('f-aging').innerHTML   = '<option value="">All</option>'+agings.map(s=>`<option>${s}</option>`).join('');
}
buildFilterOptions();

function itemMatches(it){
  if(FILTERS.status && it.status !== FILTERS.status) return false;
  if(FILTERS.section && it.section !== FILTERS.section) return false;
  if(FILTERS.aging && it.aging !== FILTERS.aging) return false;
  if(FILTERS.search){
    const q = FILTERS.search.toLowerCase();
    const hay = [it.reference, it.text, it.assignment, it.rr_comments, it.eth_comments, it.eth_po_ref]
      .filter(Boolean).join(' ').toLowerCase();
    if(!hay.includes(q)) return false;
  }
  return true;
}

const SORT = {}; // per-section sort state: {section: {key, dir}}

function renderSections(){
  const c = document.getElementById('sections-container');
  c.innerHTML = '';
  let totalShown = 0;
  DATA.sections.forEach(sec=>{
    const items = sec.items.filter(itemMatches);
    totalShown += items.length;
    const sortState = SORT[sec.name] || {key:'row', dir:1};
    items.sort((a,b)=>{
      let va=a[sortState.key], vb=b[sortState.key];
      if(va==null) va = sortState.dir<0 ? Infinity : -Infinity;
      if(vb==null) vb = sortState.dir<0 ? Infinity : -Infinity;
      if(typeof va==='string') va = va.toLowerCase();
      if(typeof vb==='string') vb = vb.toLowerCase();
      return va<vb ? -sortState.dir : va>vb ? sortState.dir : 0;
    });

    const rows = items.map(it=>{
      const amt = it.amount;
      const amtClass = (amt||0) < 0 ? 'neg' : 'pos';
      const tr = (amt||0) < 0 ? 'class="credit"' : '';
      return `<tr ${tr}>
        <td>${it.reference||'—'}</td>
        <td>${it.assignment||'—'}</td>
        <td>${fmt.date(it.net_due_date)}</td>
        <td class="num ${amtClass}">${fmt.money(amt)}</td>
        <td>${it.currency||''}</td>
        <td class="num">${fmt.days(it.days_late)}</td>
        <td><span class="tag ${statusClass(it.status)}">${it.status}</span></td>
        <td>${escapeHtml(it.text||'')}</td>
      </tr>`;
    }).join('');
    const secTotal = items.reduce((a,b)=>a+(b.amount||0),0);

    const det = document.createElement('details');
    det.open = items.length > 0 && items.length <= 40;
    det.innerHTML = `
      <summary>
        <div>
          <span style="font-size:15px">${sec.name}</span>
          <span class="section-meta" style="display:inline-flex;margin-left:12px">
            <span>${items.length}/${sec.item_count} items</span>
            <span>Filtered total: <b>${fmt.money(secTotal)}</b></span>
            <span>Section total: <b>${fmt.money(sec.computed_total)}</b></span>
          </span>
        </div>
      </summary>
      <div class="scroll-x">
        <table data-section="${sec.name}">
          <thead><tr>
            <th data-k="reference">Reference</th>
            <th data-k="assignment">Assignment</th>
            <th data-k="net_due_date">Net Due</th>
            <th data-k="amount" style="text-align:right">Amount</th>
            <th data-k="currency">Curr</th>
            <th data-k="days_late" style="text-align:right">Days Late</th>
            <th data-k="status">Status</th>
            <th data-k="text">Text</th>
          </tr></thead>
          <tbody>${rows || '<tr><td colspan="8" style="color:var(--muted);text-align:center;padding:20px">No items match filters.</td></tr>'}</tbody>
        </table>
      </div>
    `;
    det.querySelectorAll('th[data-k]').forEach(th=>{
      th.addEventListener('click', ()=>{
        const k = th.dataset.k;
        const cur = SORT[sec.name] || {key:'row', dir:1};
        SORT[sec.name] = { key:k, dir: cur.key===k ? -cur.dir : 1 };
        renderSections();
      });
    });
    c.appendChild(det);
  });
  document.getElementById('f-count').textContent =
    `Showing ${totalShown} of ${DATA.flat_items.length} items`;
}

function escapeHtml(s){ return String(s).replace(/[&<>"']/g, c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c])); }

['f-search','f-status','f-section','f-aging'].forEach(id=>{
  document.getElementById(id).addEventListener('input', e=>{
    const key = id.slice(2);
    FILTERS[key] = e.target.value;
    renderSections();
  });
});
renderSections();

// Audit
document.getElementById('audit-log').textContent = (DATA.audit||[]).join('\n') +
  (DATA.warnings && DATA.warnings.length ? '\n\nWARNINGS:\n' + DATA.warnings.join('\n') : '');

// Footer
document.getElementById('footer').innerHTML = `
  <div><b>Source:</b> <code>${DATA.metadata.source_file}</code></div>
  <div><b>Sheet:</b> <code>${DATA.metadata.sheet}</code> · <b>Generated:</b> <code>${DATA.metadata.generated_at}</code></div>
  <div style="margin-top:8px">Ground-truth benchmark · parsed directly via openpyxl · independent of V6/parser.py</div>
`;
</script>
</body>
</html>
"""


if __name__ == "__main__":
    main()
