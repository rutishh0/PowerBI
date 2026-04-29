"""
Build a self-contained HTML visualization benchmark for 2026_PLAN.xlsx.

Independent of V6/parser.py. Reads xlsx via openpyxl directly.
Emits a canonical JSON payload + ApexCharts visualisation as a single HTML file.

Sheets:
  1YP                  - One Year Plan (action log, category status, weekly activity)
  5YP SPE SALES        - Five Year Plan SPE Sales pipeline
  SPE SALES PER YEAR   - Annual pivot of SPE sales by customer and engine
"""
from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from datetime import datetime, date
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

SRC = Path(r"C:\Users\Rutishkrishna\Desktop\RR\RR Powerbi\V6\New info\2026_PLAN.xlsx")
OUT_HTML = Path(r"C:\Users\Rutishkrishna\Desktop\RR\RR Powerbi\V6\TESTEXCEL\2026_PLAN.html")
OUT_JSON = Path(r"C:\Users\Rutishkrishna\Desktop\RR\RR Powerbi\V6\TESTEXCEL\_payloads\2026_PLAN.json")

CATEGORY_COLUMNS = ["COMMERCIAL", "AM", "SALES", "CUSTOMER OPS"]
ENGINE_FAMILIES = ["XWB-84", "XWB-97", "Trent 7000", "TRENT7000", "TRENT1000"]


# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------
def _s(v):
    """Safe string or None."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.isoformat() if isinstance(v, datetime) else v.isoformat()
    s = str(v).strip()
    if not s or s in {".", "\xa0"}:
        return None
    return s


def _iso_date(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except Exception:
        return None


def _clean_text(v):
    s = _s(v)
    if s is None:
        return None
    # Normalise newlines
    return s.replace("\r", "").strip()


# ------------------------------------------------------------------
# Sheet 1: 1YP
# ------------------------------------------------------------------
def parse_1yp(ws):
    """
    Row 1  : G='Commercial', H='AM', I='Sales', J='Customer Ops'  (category labels on hidden cols G-J)
    Row 3  : H..S numeric week numbers (1..12)
    Row 4  : G4:S4 merged 'Week Beginning' banner
    Row 5  : A='Blue Chip', B='Customer', C='Issue', D='Description', E='Owner', F='Latest Update',
             G..S = 2024-10-07 .. 2024-12-30 (historical hidden weekly dates),
             T='Description', U='Owner', V..='Week 1','Week 2',...
    Row 6  : V.. = 2026-01-01, 2026-01-08, ... (real 2026 weekly dates)
    Row 7+ : action items
    """
    # Map category columns from row 1 G..J (or scan wider to be safe)
    category_col = {}
    for c in range(7, 11):  # G(7)..J(10)
        v = _s(ws.cell(row=1, column=c).value)
        if v:
            category_col[v.upper()] = c
    # Keep order COMMERCIAL, AM, SALES, CUSTOMER OPS
    ordered_categories = []
    for label in CATEGORY_COLUMNS:
        if label in category_col:
            ordered_categories.append(label)

    # 2024 weekly dates sit on cols G..S via row 5
    historical_weeks = {}  # col_idx -> iso date
    for c in range(7, 20):  # G..S
        v = ws.cell(row=5, column=c).value
        d = _iso_date(v)
        if d:
            historical_weeks[c] = d

    # Secondary campaign description/owner at T=20, U=21
    campaign_desc_col = 20
    campaign_owner_col = 21

    # 2026 weekly dates from row 6, columns V(22) onwards
    weekly_cols = {}  # col_idx -> iso date (YYYY-MM-DD)
    for c in range(22, ws.max_column + 1):
        v = ws.cell(row=6, column=c).value
        d = _iso_date(v)
        if d:
            weekly_cols[c] = d
    week_columns_iso = [weekly_cols[c] for c in sorted(weekly_cols.keys())]

    # Forward-fill helpers for merged cells (Blue Chip, Customer, Issue, Description, Owner)
    last_blue_chip = None
    last_customer = None

    items = []
    for r in range(7, ws.max_row + 1):
        raw_blue_chip = _clean_text(ws.cell(row=r, column=1).value)
        raw_customer = _clean_text(ws.cell(row=r, column=2).value)
        issue = _clean_text(ws.cell(row=r, column=3).value)
        description = _clean_text(ws.cell(row=r, column=4).value)
        owner = _clean_text(ws.cell(row=r, column=5).value)
        latest_update = _clean_text(ws.cell(row=r, column=6).value)
        campaign_desc = _clean_text(ws.cell(row=r, column=campaign_desc_col).value)
        campaign_owner = _clean_text(ws.cell(row=r, column=campaign_owner_col).value)

        # Category status (G,H,I,J)
        category_status = {}
        for label, col_idx in category_col.items():
            val = _clean_text(ws.cell(row=r, column=col_idx).value)
            category_status[label] = val

        # Historical 2024 weekly activity
        historical_status = {}
        for col_idx, iso_d in historical_weeks.items():
            val = _clean_text(ws.cell(row=r, column=col_idx).value)
            if val:
                historical_status[iso_d] = val

        # 2026 weekly activity
        weekly_status = {}
        for col_idx, iso_d in weekly_cols.items():
            val = _clean_text(ws.cell(row=r, column=col_idx).value)
            if val:
                weekly_status[iso_d] = val

        # Decide whether this row has any REAL content (do NOT count inherited-only customer name).
        # A row is real if it has any of: own issue/description/owner/latest_update,
        # campaign fields, category status, 2026 weekly status, or historical 2024 status.
        has_real_content = any([
            issue, description, owner, latest_update,
            campaign_desc, campaign_owner,
            any(v for v in category_status.values()),
            weekly_status, historical_status,
        ])
        # Also accept if the row itself carries a fresh customer/blue-chip identity
        # (i.e., it's a new record with no per-week content yet).
        carries_own_identity = bool(raw_blue_chip or raw_customer)

        # Update forward-fill state based on RAW values
        if raw_blue_chip:
            last_blue_chip = raw_blue_chip
        if raw_customer:
            last_customer = raw_customer

        # Skip rows that are purely empty shells below the data region
        if not has_real_content and not carries_own_identity:
            continue

        blue_chip = raw_blue_chip or last_blue_chip
        customer = raw_customer or last_customer

        # Derive status summary
        total_weeks = len(weekly_status)
        cat_filled = sum(1 for v in category_status.values() if v)
        has_latest = bool(latest_update)
        if total_weeks == 0 and cat_filled == 0 and not has_latest:
            summary = "Unassigned"
        elif total_weeks == 0 and cat_filled > 0:
            summary = "Planned"
        elif total_weeks >= 1 and owner:
            summary = "Active"
        else:
            summary = "Active" if (total_weeks or cat_filled) else "Unassigned"

        items.append({
            "row_index": r,
            "blue_chip": blue_chip,
            "customer": customer,
            "issue": issue,
            "description": description,
            "owner": owner or campaign_owner,  # fall back to campaign owner
            "primary_owner": owner,
            "campaign_owner": campaign_owner,
            "latest_update": latest_update,
            "campaign_description": campaign_desc,
            "category_status": category_status,
            "historical_status": historical_status,
            "weekly_status": weekly_status,
            "weeks_active": total_weeks,
            "status_summary": summary,
        })

    return {
        "week_columns": week_columns_iso,
        "historical_week_columns": sorted(historical_weeks.values()),
        "category_columns": ordered_categories,
        "items": items,
    }


# ------------------------------------------------------------------
# Sheet 2: 5YP SPE SALES
# ------------------------------------------------------------------
def parse_5yp(ws):
    """
    Row 1 header: Customer | Engine Type | Year | Quarter | Amount | Comments
    Rows 2+   : Customer may be blank (forward-fill from above)
    """
    items = []
    last_customer = None
    for r in range(2, ws.max_row + 1):
        cust = _clean_text(ws.cell(row=r, column=1).value)
        engine = _clean_text(ws.cell(row=r, column=2).value)
        year = ws.cell(row=r, column=3).value
        quarter = _clean_text(ws.cell(row=r, column=4).value)
        amount = _num(ws.cell(row=r, column=5).value)
        comments = _clean_text(ws.cell(row=r, column=6).value)

        if cust:
            last_customer = cust
        effective_customer = cust or last_customer
        year_int = None
        if isinstance(year, (int, float)):
            year_int = int(year)
        else:
            try:
                year_int = int(str(year)) if year else None
            except Exception:
                year_int = None

        if not effective_customer and not engine and year_int is None:
            continue

        items.append({
            "row_index": r,
            "customer": effective_customer,
            "engine_type": engine,
            "year": year_int,
            "quarter": quarter,
            "amount": amount,
            "comments": comments,
        })

    # Totals
    by_year = Counter()
    by_engine = Counter()
    by_customer = Counter()
    total_amount = 0.0
    for it in items:
        if it["year"] is not None:
            by_year[str(it["year"])] += 1
        if it["engine_type"]:
            by_engine[it["engine_type"]] += 1
        if it["customer"]:
            by_customer[it["customer"]] += 1
        if it["amount"]:
            total_amount += it["amount"]

    return {
        "items": items,
        "totals": {
            "by_year": dict(sorted(by_year.items())),
            "by_engine": dict(by_engine.most_common()),
            "by_customer": dict(by_customer.most_common()),
            "total_opportunities": len(items),
            "total_amount": round(total_amount, 2),
        },
    }


# ------------------------------------------------------------------
# Sheet 3: SPE SALES PER YEAR (pivot)
# ------------------------------------------------------------------
def parse_spe_per_year(ws):
    """
    Row 1 year banners on cols A, D, G, J, M (years 2026..2030)
    Row 3: 'Customer' | 'Engines' pairs under each year
    Rows 4+: alternating Customer row / Engine-type row, each with a numeric count.
             Pattern per year block: [Customer name] | count, then next row [Engine type] | count.
             Ends with 'Grand Total' row.
    """
    # Locate year blocks
    year_blocks = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, (int, float)) and 2020 <= int(v) <= 2040:
            year_blocks.append({"year": int(v), "customer_col": c, "engine_count_col": c + 1})

    by_year = {}
    for blk in year_blocks:
        year = blk["year"]
        cc = blk["customer_col"]
        ec = blk["engine_count_col"]

        customers = []
        grand_total = None
        current_customer = None  # being built
        r = 4
        while r <= ws.max_row:
            label = _clean_text(ws.cell(row=r, column=cc).value)
            count = _num(ws.cell(row=r, column=ec).value)
            if label is None and count is None:
                # Possible end of block if we've seen Grand Total, OR skip
                if grand_total is not None:
                    break
                r += 1
                continue

            if label and label.lower().startswith("grand total"):
                grand_total = int(count) if count is not None else None
                break

            # Heuristic: if label is a known engine family OR matches engine pattern
            is_engine = False
            if label:
                up = label.upper()
                if any(up.startswith(fam.upper()) for fam in ENGINE_FAMILIES):
                    is_engine = True
                elif re.match(r"^(XWB|TRENT|CFM|GEnx|PW\d)", up):
                    is_engine = True

            if is_engine and current_customer is not None:
                # attach engine line to current customer
                current_customer["engines"].append({
                    "type": label,
                    "count": int(count) if count is not None else 0,
                })
            else:
                # new customer row
                if current_customer is not None:
                    customers.append(current_customer)
                current_customer = {
                    "name": label,
                    "total": int(count) if count is not None else 0,
                    "engines": [],
                }
            r += 1

        # Flush last customer if not already
        if current_customer is not None and current_customer not in customers:
            # Add only if not grand total
            if current_customer["name"] and not current_customer["name"].lower().startswith("grand total"):
                customers.append(current_customer)

        by_year[str(year)] = {
            "year": year,
            "customers": customers,
            "grand_total": grand_total if grand_total is not None else sum(c["total"] for c in customers),
        }

    return {"by_year": by_year, "year_list": [b["year"] for b in year_blocks]}


# ------------------------------------------------------------------
# Main
# ------------------------------------------------------------------
def build_payload():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    one = parse_1yp(wb["1YP"])
    five = parse_5yp(wb["5YP SPE SALES"])
    annual = parse_spe_per_year(wb["SPE SALES PER YEAR"])

    payload = {
        "file_type": "COMMERCIAL_PLAN",
        "metadata": {
            "source_file": SRC.name,
            "sheets_parsed": ["1YP", "5YP SPE SALES", "SPE SALES PER YEAR"],
            "plan_year": 2026,
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "counts": {
                "one_year_plan_items": len(one["items"]),
                "five_year_spe_items": len(five["items"]),
                "annual_summary_years": len(annual["by_year"]),
            },
        },
        "one_year_plan": one,
        "five_year_spe_sales": five,
        "annual_summary": annual,
    }
    return payload


# ------------------------------------------------------------------
# HTML rendering
# ------------------------------------------------------------------
HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8" />
<title>2026 Commercial Plan — Benchmark</title>
<meta name="viewport" content="width=device-width,initial-scale=1" />
<script src="https://cdn.jsdelivr.net/npm/apexcharts@3.49.0/dist/apexcharts.min.js"></script>
<style>
:root{
  --bg: #0b0f14;
  --panel: #131a22;
  --panel-2: #18212c;
  --ink: #e7eef7;
  --ink-dim: #9eb0c4;
  --accent: #7bd3ff;
  --accent-2: #b6a1ff;
  --ok: #5ee7a7;
  --warn: #ffb547;
  --hot: #ff6b6b;
  --line: #1f2b38;
  --chip: #22313f;
  --shadow: 0 6px 24px rgba(0,0,0,.35);
}
*{box-sizing:border-box}
html,body{margin:0;padding:0;background:linear-gradient(180deg,#0a0e13,#0b0f14 280px);color:var(--ink);font-family:'Inter','Segoe UI',system-ui,-apple-system,sans-serif;}
a{color:var(--accent)}
header.top{
  padding:28px 32px 16px 32px;
  border-bottom:1px solid var(--line);
  background:linear-gradient(135deg,rgba(123,211,255,.08),rgba(182,161,255,.04));
}
header.top h1{margin:0;font-size:28px;letter-spacing:-.5px}
header.top .sub{color:var(--ink-dim);font-size:13px;margin-top:6px}
.badge{display:inline-block;padding:2px 10px;border-radius:999px;background:var(--chip);color:var(--ink);font-size:11px;margin-right:6px;border:1px solid var(--line)}
main{padding:18px 24px 80px}
.tabs{display:flex;gap:8px;margin:18px 8px 22px 8px;flex-wrap:wrap}
.tab{
  padding:10px 18px;border-radius:10px;background:var(--panel);
  border:1px solid var(--line);color:var(--ink-dim);cursor:pointer;
  font-weight:600;font-size:13px;transition:all .18s;
}
.tab:hover{color:var(--ink);border-color:#2a3a4c}
.tab.active{background:linear-gradient(135deg,var(--accent),var(--accent-2));color:#0a0e13;border-color:transparent;box-shadow:var(--shadow)}
.view{display:none}
.view.active{display:block;animation:fade .25s ease}
@keyframes fade{from{opacity:0;transform:translateY(4px)}to{opacity:1;transform:none}}
.grid{display:grid;gap:16px}
.grid-6{grid-template-columns:repeat(6,minmax(0,1fr))}
.grid-3{grid-template-columns:repeat(3,minmax(0,1fr))}
.grid-2{grid-template-columns:1fr 1fr}
.grid-5{grid-template-columns:repeat(5,minmax(0,1fr))}
@media (max-width:1100px){.grid-6{grid-template-columns:repeat(3,1fr)}.grid-3,.grid-2,.grid-5{grid-template-columns:1fr}}
.card{
  background:var(--panel);border:1px solid var(--line);border-radius:14px;
  padding:18px;box-shadow:var(--shadow)
}
.card h3{margin:0 0 14px 0;font-size:14px;letter-spacing:.5px;color:var(--ink-dim);text-transform:uppercase}
.kpi{padding:18px;background:linear-gradient(135deg,var(--panel),var(--panel-2));border:1px solid var(--line);border-radius:14px}
.kpi .label{font-size:11px;color:var(--ink-dim);text-transform:uppercase;letter-spacing:1px}
.kpi .value{font-size:28px;font-weight:700;margin-top:6px}
.kpi .meta{font-size:11px;color:var(--ink-dim);margin-top:4px}
.filterbar{display:flex;gap:10px;flex-wrap:wrap;align-items:center;margin-bottom:14px}
.filterbar select,.filterbar input[type=text]{
  background:var(--panel-2);border:1px solid var(--line);color:var(--ink);
  padding:8px 10px;border-radius:8px;font-size:13px;outline:none;min-width:150px
}
.filterbar button{
  background:var(--chip);color:var(--ink);border:1px solid var(--line);
  padding:8px 14px;border-radius:8px;cursor:pointer;font-size:12px;font-weight:600
}
.filterbar button:hover{background:#2a3a4c}
.filterbar .btn-export{background:linear-gradient(135deg,var(--accent),var(--accent-2));color:#0a0e13;border:0}
table{width:100%;border-collapse:collapse;font-size:12.5px}
thead th{
  text-align:left;padding:10px 10px;border-bottom:1px solid var(--line);
  background:var(--panel-2);color:var(--ink-dim);font-weight:600;letter-spacing:.3px;
  position:sticky;top:0;cursor:pointer;user-select:none
}
thead th:hover{color:var(--ink)}
tbody td{padding:9px 10px;border-bottom:1px solid var(--line);vertical-align:top}
tbody tr:hover{background:rgba(123,211,255,.04)}
.table-wrap{max-height:640px;overflow:auto;border:1px solid var(--line);border-radius:10px;background:var(--panel)}
.chip{display:inline-block;padding:2px 8px;border-radius:999px;font-size:10px;margin:1px 2px;border:1px solid var(--line);background:var(--panel-2);color:var(--ink)}
.chip.L1{background:rgba(94,231,167,.12);color:var(--ok);border-color:rgba(94,231,167,.3)}
.chip.L2{background:rgba(255,181,71,.12);color:var(--warn);border-color:rgba(255,181,71,.3)}
.chip.L3{background:rgba(255,140,97,.12);color:#ff8c61;border-color:rgba(255,140,97,.3)}
.chip.L4{background:rgba(255,107,107,.14);color:var(--hot);border-color:rgba(255,107,107,.3)}
.chip.status-Active{background:rgba(123,211,255,.12);color:var(--accent);border-color:rgba(123,211,255,.3)}
.chip.status-Planned{background:rgba(255,181,71,.12);color:var(--warn);border-color:rgba(255,181,71,.3)}
.chip.status-Unassigned{background:rgba(158,176,196,.10);color:var(--ink-dim);border-color:var(--line)}
.row-detail{background:var(--panel-2);font-size:11.5px;color:var(--ink-dim)}
.row-detail .mini-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(170px,1fr));gap:8px;padding:8px 0}
.row-detail .mini-grid .w{background:var(--panel);border:1px solid var(--line);border-radius:6px;padding:6px 8px}
.row-detail .w .d{color:var(--ink);font-size:10.5px;font-weight:600}
.row-detail .w .v{margin-top:3px;color:var(--ink-dim)}
.expand-btn{background:transparent;border:0;color:var(--accent);cursor:pointer;font-size:12px}
.year-card{background:linear-gradient(150deg,var(--panel),var(--panel-2));border:1px solid var(--line);border-radius:14px;padding:16px;box-shadow:var(--shadow)}
.year-card h4{margin:0;font-size:22px;letter-spacing:-.5px}
.year-card .total{font-size:12px;color:var(--ink-dim);margin-top:4px}
.year-card .cust-list{margin-top:10px;max-height:180px;overflow:auto}
.year-card .cust-list .c{display:flex;justify-content:space-between;padding:4px 0;border-bottom:1px dashed var(--line);font-size:12px}
.year-card .cust-list .c:last-child{border-bottom:0}
.year-card .c .n{color:var(--ink)}
.year-card .c .e{color:var(--ink-dim);font-size:11px;margin-left:6px}
.legend{display:flex;gap:12px;flex-wrap:wrap;font-size:11px;color:var(--ink-dim);margin-top:8px}
.legend .sw{display:inline-block;width:10px;height:10px;border-radius:3px;margin-right:4px;vertical-align:middle}
footer{color:var(--ink-dim);font-size:11px;padding:16px 24px;border-top:1px solid var(--line);margin-top:24px}
.tag{display:inline-block;background:var(--chip);color:var(--ink);padding:1px 7px;border-radius:5px;font-size:10px;border:1px solid var(--line);margin-right:3px}
</style>
</head>
<body>
<header class="top">
  <h1 id="h-title">2026 Commercial Plan</h1>
  <div class="sub" id="h-sub">Loading…</div>
</header>

<main>
  <div class="filterbar" style="margin-left:8px">
    <label style="font-size:12px;color:var(--ink-dim)">Year filter (cascades):</label>
    <select id="globalYear">
      <option value="ALL">All years</option>
    </select>
  </div>

  <div class="tabs">
    <div class="tab active" data-view="overview">Overview</div>
    <div class="tab" data-view="actions">Action Log (1YP)</div>
    <div class="tab" data-view="pipeline">Pipeline (5YP)</div>
    <div class="tab" data-view="yearly">Yearly Breakdown</div>
  </div>

  <!-- ==================== OVERVIEW ==================== -->
  <section class="view active" id="view-overview">
    <div class="grid grid-6" id="kpiRow"></div>
    <div class="grid grid-2" style="margin-top:16px">
      <div class="card"><h3>Annual Engine Forecast (2026–2030)</h3><div id="chartAnnual"></div></div>
      <div class="card"><h3>Engine Family Mix by Year</h3><div id="chartFamilyMix"></div></div>
    </div>
    <div class="grid grid-2" style="margin-top:16px">
      <div class="card"><h3>Top 10 Customers — Total Engines</h3><div id="chartTopCust"></div></div>
      <div class="card"><h3>1YP Actions by Owner</h3><div id="chartOwners"></div></div>
    </div>
  </section>

  <!-- ==================== ACTION LOG ==================== -->
  <section class="view" id="view-actions">
    <div class="card">
      <div class="filterbar">
        <select id="flt-cust"><option value="">Customer (all)</option></select>
        <select id="flt-owner"><option value="">Owner (all)</option></select>
        <select id="flt-cat"><option value="">Category (all)</option></select>
        <select id="flt-status"><option value="">Status (all)</option></select>
        <input type="text" id="flt-search" placeholder="Search issue / description / update…" style="flex:1;min-width:220px" />
        <button id="btn-reset">Reset</button>
        <button id="btn-export-actions" class="btn-export">Export CSV</button>
      </div>
      <h3>Weekly Activity Heatmap (top 15 customers by activity)</h3>
      <div id="chartHeatmap"></div>
      <div class="legend">
        <span><span class="sw" style="background:#5ee7a7"></span>L1</span>
        <span><span class="sw" style="background:#ffb547"></span>L2</span>
        <span><span class="sw" style="background:#ff8c61"></span>L3</span>
        <span><span class="sw" style="background:#ff6b6b"></span>L4</span>
        <span><span class="sw" style="background:#7bd3ff"></span>Other activity</span>
      </div>
    </div>
    <div class="card" style="margin-top:16px">
      <h3>Action Items</h3>
      <div class="table-wrap">
        <table id="tbl-actions">
          <thead><tr>
            <th data-k="customer">Customer</th>
            <th data-k="issue">Issue</th>
            <th data-k="description">Description</th>
            <th data-k="owner">Owner</th>
            <th data-k="latest_update">Latest Update</th>
            <th>Category</th>
            <th data-k="weeks_active">Weeks</th>
            <th data-k="status_summary">Status</th>
            <th></th>
          </tr></thead>
          <tbody></tbody>
        </table>
      </div>
    </div>
  </section>

  <!-- ==================== PIPELINE ==================== -->
  <section class="view" id="view-pipeline">
    <div class="card">
      <div class="filterbar">
        <select id="p-year"><option value="">Year (all)</option></select>
        <select id="p-engine"><option value="">Engine (all)</option></select>
        <select id="p-cust"><option value="">Customer (all)</option></select>
        <select id="p-qtr"><option value="">Quarter (all)</option></select>
        <button id="p-reset">Reset</button>
        <button id="btn-export-pipeline" class="btn-export">Export CSV</button>
      </div>
    </div>
    <div class="grid grid-2" style="margin-top:16px">
      <div class="card"><h3>Pipeline Funnel — Year × Quarter × Engine count</h3><div id="chartFunnel"></div></div>
      <div class="card"><h3>Customer × Year Heatmap</h3><div id="chartPipeHeat"></div></div>
    </div>
    <div class="grid grid-2" style="margin-top:16px">
      <div class="card"><h3>Amount by Year (£m — where captured)</h3><div id="chartAmount"></div></div>
      <div class="card">
        <h3>Pipeline Items</h3>
        <div class="table-wrap" style="max-height:360px">
          <table id="tbl-pipeline">
            <thead><tr>
              <th data-k="customer">Customer</th>
              <th data-k="engine_type">Engine</th>
              <th data-k="year">Year</th>
              <th data-k="quarter">Qtr</th>
              <th data-k="amount">Amount</th>
              <th data-k="comments">Comments</th>
            </tr></thead>
            <tbody></tbody>
          </table>
        </div>
      </div>
    </div>
  </section>

  <!-- ==================== YEARLY ==================== -->
  <section class="view" id="view-yearly">
    <div class="grid grid-5" id="yearCards"></div>
    <div class="card" style="margin-top:16px">
      <h3>SPE Sales Per Year — Full pivot</h3>
      <div class="table-wrap">
        <table id="tbl-yearly">
          <thead id="tbl-yearly-head"></thead>
          <tbody id="tbl-yearly-body"></tbody>
        </table>
      </div>
    </div>
  </section>
</main>

<footer>
  Benchmark reference visualisation. Parsed directly from <code>2026_PLAN.xlsx</code> via openpyxl — independent of V6/parser.py.
  <span class="tag">ApexCharts 3.49.0</span>
  <span class="tag" id="genTag"></span>
</footer>

<script id="payload" type="application/json">__PAYLOAD__</script>
<script>
const DATA = JSON.parse(document.getElementById('payload').textContent);

// -------------- helpers --------------
const $ = (s,root=document)=>root.querySelector(s);
const $$ = (s,root=document)=>Array.from(root.querySelectorAll(s));
const esc = s => s==null ? '' : String(s).replace(/[&<>"']/g, c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));
const by = (arr, k) => arr.reduce((o,x)=>(o[x[k]]=(o[x[k]]||0)+1,o),{});
const uniq = arr => Array.from(new Set(arr.filter(Boolean))).sort();
function downloadCSV(rows, name){
  const csv = rows.map(r=>r.map(v=>{
    if(v==null) return '';
    v = String(v).replace(/"/g,'""');
    return /[",\n]/.test(v) ? `"${v}"` : v;
  }).join(',')).join('\n');
  const blob = new Blob([csv], {type:'text/csv;charset=utf-8'});
  const url = URL.createObjectURL(blob);
  const a = Object.assign(document.createElement('a'),{href:url,download:name});
  document.body.appendChild(a); a.click(); a.remove(); URL.revokeObjectURL(url);
}

// -------------- header --------------
const meta = DATA.metadata;
document.getElementById('h-title').textContent = `${meta.plan_year} Commercial Plan — Benchmark`;
document.getElementById('h-sub').innerHTML =
  `<span class="badge">${esc(meta.source_file)}</span>` +
  meta.sheets_parsed.map(s=>`<span class="badge">${esc(s)}</span>`).join('') +
  `<span class="badge">${meta.counts.one_year_plan_items} 1YP items</span>` +
  `<span class="badge">${meta.counts.five_year_spe_items} 5YP items</span>` +
  `<span class="badge">${meta.counts.annual_summary_years} years summarised</span>`;
document.getElementById('genTag').textContent = 'generated ' + meta.generated_at;

// -------------- tabs --------------
$$('.tab').forEach(t=>t.addEventListener('click',()=>{
  $$('.tab').forEach(x=>x.classList.remove('active'));
  $$('.view').forEach(x=>x.classList.remove('active'));
  t.classList.add('active');
  $('#view-'+t.dataset.view).classList.add('active');
  // Re-render heavy charts on first show to avoid apex zero-size issue
  if(t.dataset.view==='actions') renderHeatmap();
  if(t.dataset.view==='pipeline') { renderFunnel(); renderPipeHeat(); renderAmount(); }
}));

// -------------- global year filter --------------
const allYears = Array.from(new Set(
  DATA.five_year_spe_sales.items.map(x=>x.year).concat(
    Object.keys(DATA.annual_summary.by_year).map(Number)
  ).filter(Boolean)
)).sort();
const gy = $('#globalYear');
allYears.forEach(y=>{ const o=document.createElement('option'); o.value=y; o.textContent=y; gy.appendChild(o); });
gy.addEventListener('change',()=>{ renderOverview(); renderPipeline(); renderYearly(); });

function getYearFilter(){ const v = gy.value; return v==='ALL' ? null : Number(v); }

// -------------- OVERVIEW --------------
function renderKpis(){
  const yf = getYearFilter();
  const items5 = DATA.five_year_spe_sales.items.filter(i=>!yf || i.year===yf);
  const actions = DATA.one_year_plan.items;
  const yearTotals = DATA.annual_summary.by_year;
  const yearSum = Object.values(yearTotals).reduce((s,y)=>s+(y.grand_total||0),0);
  const uniqueCust = new Set(actions.map(a=>a.customer).filter(Boolean).concat(items5.map(i=>i.customer).filter(Boolean))).size;
  const uniqueOwners = new Set(actions.map(a=>a.owner).filter(Boolean)).size;
  const kpis = [
    ['1YP action items', actions.length, '= open planning rows'],
    ['5YP SPE opportunities', items5.length, yf? `year ${yf}` : 'across 2026-2031'],
    ['5-year engine total', yearSum, 'across annual pivot'],
    ['Active customers', uniqueCust, '1YP + 5YP union'],
    ['Unique owners (1YP)', uniqueOwners, 'action-log owners'],
    ['Plan year', meta.plan_year, meta.sheets_parsed.length+' sheets'],
  ];
  $('#kpiRow').innerHTML = kpis.map(k=>`
    <div class="kpi"><div class="label">${esc(k[0])}</div><div class="value">${esc(k[1])}</div><div class="meta">${esc(k[2])}</div></div>
  `).join('');
}

let chAnnual, chFamily, chTopCust, chOwners;
function renderOverview(){
  renderKpis();
  const yearTotals = DATA.annual_summary.by_year;
  const yearsList = Object.keys(yearTotals).sort();
  // 1. Annual engine forecast
  const annualData = yearsList.map(y=>yearTotals[y].grand_total||0);
  if(chAnnual) chAnnual.destroy();
  chAnnual = new ApexCharts($('#chartAnnual'), {
    chart:{type:'bar',height:280,toolbar:{show:false},background:'transparent'},
    theme:{mode:'dark'},
    series:[{name:'Engines',data:annualData}],
    xaxis:{categories:yearsList},
    colors:['#7bd3ff'],
    plotOptions:{bar:{columnWidth:'45%',borderRadius:6,dataLabels:{position:'top'}}},
    dataLabels:{enabled:true,offsetY:-18,style:{colors:['#e7eef7'],fontSize:'11px'}},
    grid:{borderColor:'#1f2b38'}, tooltip:{theme:'dark'}
  });
  chAnnual.render();

  // 2. Engine family mix by year
  const families = ['XWB-84','XWB-97','Trent 7000','TRENT7000','TRENT1000'];
  const familySeries = families.map(f=>({
    name:f,
    data: yearsList.map(y=>{
      const custs = yearTotals[y].customers||[];
      let total=0;
      custs.forEach(c=>c.engines.forEach(e=>{ if((e.type||'').toUpperCase().replace(/\s+/g,'')===f.toUpperCase().replace(/\s+/g,'')) total+=e.count; }));
      return total;
    })
  }));
  if(chFamily) chFamily.destroy();
  chFamily = new ApexCharts($('#chartFamilyMix'), {
    chart:{type:'bar',stacked:true,height:280,toolbar:{show:false},background:'transparent'},
    theme:{mode:'dark'},
    series:familySeries,
    xaxis:{categories:yearsList},
    colors:['#7bd3ff','#b6a1ff','#5ee7a7','#ffb547','#ff6b6b'],
    plotOptions:{bar:{columnWidth:'60%',borderRadius:4}},
    legend:{position:'top',labels:{colors:'#e7eef7'}},
    grid:{borderColor:'#1f2b38'}, tooltip:{theme:'dark'}
  });
  chFamily.render();

  // 3. Top 10 customers by total engines (across annual pivot)
  const custTotals = {};
  Object.values(yearTotals).forEach(y=>{
    (y.customers||[]).forEach(c=>{
      if(!c.name) return;
      custTotals[c.name] = (custTotals[c.name]||0) + (c.total||0);
    });
  });
  const topCust = Object.entries(custTotals).sort((a,b)=>b[1]-a[1]).slice(0,10);
  if(chTopCust) chTopCust.destroy();
  chTopCust = new ApexCharts($('#chartTopCust'), {
    chart:{type:'bar',height:320,toolbar:{show:false},background:'transparent'},
    theme:{mode:'dark'},
    series:[{name:'Engines',data:topCust.map(x=>x[1])}],
    xaxis:{categories:topCust.map(x=>x[0])},
    plotOptions:{bar:{horizontal:true,borderRadius:4,dataLabels:{position:'top'}}},
    dataLabels:{enabled:true,offsetX:24,style:{colors:['#e7eef7'],fontSize:'11px'}},
    colors:['#b6a1ff'],
    grid:{borderColor:'#1f2b38'}, tooltip:{theme:'dark'}
  });
  chTopCust.render();

  // 4. 1YP actions by owner
  const owners = {};
  DATA.one_year_plan.items.forEach(i=>{ if(i.owner) owners[i.owner]=(owners[i.owner]||0)+1; });
  const sortedOwners = Object.entries(owners).sort((a,b)=>b[1]-a[1]);
  if(chOwners) chOwners.destroy();
  chOwners = new ApexCharts($('#chartOwners'), {
    chart:{type:'bar',height:320,toolbar:{show:false},background:'transparent'},
    theme:{mode:'dark'},
    series:[{name:'Actions',data:sortedOwners.map(x=>x[1])}],
    xaxis:{categories:sortedOwners.map(x=>x[0])},
    plotOptions:{bar:{horizontal:true,borderRadius:4,dataLabels:{position:'top'}}},
    dataLabels:{enabled:true,offsetX:20,style:{colors:['#e7eef7'],fontSize:'11px'}},
    colors:['#5ee7a7'],
    grid:{borderColor:'#1f2b38'}, tooltip:{theme:'dark'}
  });
  chOwners.render();
}

// -------------- ACTION LOG --------------
function levelColor(v){
  if(!v) return null;
  const s = String(v).toUpperCase();
  if(/L4/.test(s)) return {cls:'L4', color:'#ff6b6b', value:4};
  if(/L3/.test(s)) return {cls:'L3', color:'#ff8c61', value:3};
  if(/L2/.test(s)) return {cls:'L2', color:'#ffb547', value:2};
  if(/L1/.test(s)) return {cls:'L1', color:'#5ee7a7', value:1};
  return {cls:'', color:'#7bd3ff', value:0.6};
}

function populateActionFilters(){
  const items = DATA.one_year_plan.items;
  const fillSel = (sel, values) => {
    values.forEach(v=>{ const o=document.createElement('option'); o.value=v; o.textContent=v; sel.appendChild(o); });
  };
  fillSel($('#flt-cust'), uniq(items.map(i=>i.customer)));
  fillSel($('#flt-owner'), uniq(items.map(i=>i.owner)));
  fillSel($('#flt-cat'), DATA.one_year_plan.category_columns);
  fillSel($('#flt-status'), uniq(items.map(i=>i.status_summary)));
}

function actionFiltered(){
  const c = $('#flt-cust').value;
  const o = $('#flt-owner').value;
  const cat = $('#flt-cat').value;
  const st = $('#flt-status').value;
  const q = $('#flt-search').value.trim().toLowerCase();
  return DATA.one_year_plan.items.filter(i=>{
    if(c && i.customer!==c) return false;
    if(o && i.owner!==o) return false;
    if(cat){
      const v = i.category_status && i.category_status[cat];
      if(!v) return false;
    }
    if(st && i.status_summary!==st) return false;
    if(q){
      const hay = [i.issue,i.description,i.latest_update,i.campaign_description].filter(Boolean).join(' ').toLowerCase();
      if(!hay.includes(q)) return false;
    }
    return true;
  });
}

let sortState = {key:null, dir:1};
function renderActionTable(){
  const tbody = $('#tbl-actions tbody');
  let data = actionFiltered();
  if(sortState.key){
    const k = sortState.key;
    data = data.slice().sort((a,b)=>{
      const av = a[k]??'', bv = b[k]??'';
      if(av<bv) return -sortState.dir;
      if(av>bv) return sortState.dir;
      return 0;
    });
  }
  tbody.innerHTML = data.map((i,idx)=>{
    const cats = (i.category_status && DATA.one_year_plan.category_columns.map(cn=>{
      const v = i.category_status[cn];
      if(!v) return '';
      const lc = levelColor(v);
      return `<span class="chip ${lc?lc.cls:''}" title="${esc(cn)}: ${esc(v)}">${esc(cn.slice(0,3))}: ${esc(v)}</span>`;
    }).join('')) || '';
    return `
      <tr>
        <td><strong>${esc(i.customer||'—')}</strong>${i.blue_chip?`<br><span class="tag">${esc(i.blue_chip)}</span>`:''}</td>
        <td>${esc(i.issue||'')}</td>
        <td>${esc(i.description||i.campaign_description||'')}</td>
        <td>${esc(i.owner||'')}</td>
        <td style="max-width:260px">${esc(i.latest_update||'')}</td>
        <td>${cats}</td>
        <td style="text-align:center">${i.weeks_active||0}</td>
        <td><span class="chip status-${esc(i.status_summary)}">${esc(i.status_summary)}</span></td>
        <td><button class="expand-btn" data-idx="${idx}">Details</button></td>
      </tr>
      <tr class="row-detail" style="display:none" data-detail="${idx}">
        <td colspan="9">
          <strong>Campaign:</strong> ${esc(i.campaign_description||'—')}
          <div class="mini-grid">
            ${Object.keys(i.weekly_status).sort().map(d=>`<div class="w"><div class="d">${esc(d)}</div><div class="v">${esc(i.weekly_status[d])}</div></div>`).join('') || '<em>No 2026 weekly activity recorded.</em>'}
          </div>
          ${Object.keys(i.historical_status).length ? `<div style="margin-top:6px"><strong>Historical (2024):</strong>
             <div class="mini-grid">${Object.keys(i.historical_status).sort().map(d=>`<div class="w"><div class="d">${esc(d)}</div><div class="v">${esc(i.historical_status[d])}</div></div>`).join('')}</div></div>`:''}
        </td>
      </tr>
    `;
  }).join('');
  tbody.querySelectorAll('.expand-btn').forEach(b=>{
    b.addEventListener('click',()=>{
      const idx = b.dataset.idx;
      const det = tbody.querySelector(`[data-detail="${idx}"]`);
      det.style.display = det.style.display==='none' ? '' : 'none';
    });
  });
}

$$('#tbl-actions thead th').forEach(th=>{
  if(!th.dataset.k) return;
  th.addEventListener('click',()=>{
    const k = th.dataset.k;
    if(sortState.key===k) sortState.dir*=-1;
    else { sortState.key=k; sortState.dir=1; }
    renderActionTable();
  });
});

let chHeat;
function renderHeatmap(){
  const items = actionFiltered();
  const weeks = DATA.one_year_plan.week_columns;
  // top 15 customers by weeks_active
  const custTotals = {};
  items.forEach(i=>{ if(i.customer) custTotals[i.customer] = (custTotals[i.customer]||0) + (i.weeks_active||0); });
  const topCust = Object.entries(custTotals).sort((a,b)=>b[1]-a[1]).slice(0,15).map(x=>x[0]);

  const series = topCust.map(cust=>({
    name: cust,
    data: weeks.map(w=>{
      // aggregate any activity at that week for that customer
      const rows = items.filter(i=>i.customer===cust);
      let best = 0;
      rows.forEach(r=>{
        const v = r.weekly_status && r.weekly_status[w];
        if(v){
          const lc = levelColor(v);
          best = Math.max(best, lc?lc.value:0.6);
        }
      });
      return {x:w, y:best};
    })
  }));

  if(chHeat) chHeat.destroy();
  chHeat = new ApexCharts($('#chartHeatmap'), {
    chart:{type:'heatmap',height:420,toolbar:{show:false},background:'transparent'},
    theme:{mode:'dark'},
    series,
    dataLabels:{enabled:false},
    xaxis:{categories:weeks,labels:{rotate:-45,style:{fontSize:'9px'},formatter:v=>v.slice(5)}},
    colors:['#7bd3ff'],
    plotOptions:{
      heatmap:{
        enableShades:false,
        colorScale:{ranges:[
          {from:0,to:0,color:'#131a22',name:'none'},
          {from:0.5,to:0.9,color:'#7bd3ff',name:'note'},
          {from:1,to:1.4,color:'#5ee7a7',name:'L1'},
          {from:1.5,to:2.4,color:'#ffb547',name:'L2'},
          {from:2.5,to:3.4,color:'#ff8c61',name:'L3'},
          {from:3.5,to:4,color:'#ff6b6b',name:'L4'},
        ]}
      }
    },
    grid:{borderColor:'#1f2b38'}, tooltip:{theme:'dark'}
  });
  chHeat.render();
}

['flt-cust','flt-owner','flt-cat','flt-status','flt-search'].forEach(id=>{
  $('#'+id).addEventListener('input',()=>{ renderActionTable(); renderHeatmap(); });
});
$('#btn-reset').addEventListener('click',()=>{
  ['flt-cust','flt-owner','flt-cat','flt-status','flt-search'].forEach(id=>$('#'+id).value='');
  renderActionTable(); renderHeatmap();
});
$('#btn-export-actions').addEventListener('click',()=>{
  const rows = actionFiltered();
  const header = ['Customer','Blue Chip','Issue','Description','Owner','Latest Update','Commercial','AM','Sales','Customer Ops','Weeks Active','Status','Weekly Activity'];
  const out = [header];
  rows.forEach(i=>{
    out.push([
      i.customer,i.blue_chip,i.issue,i.description||i.campaign_description,i.owner,i.latest_update,
      (i.category_status||{}).COMMERCIAL, (i.category_status||{}).AM, (i.category_status||{}).SALES, (i.category_status||{})['CUSTOMER OPS'],
      i.weeks_active, i.status_summary,
      Object.entries(i.weekly_status||{}).map(([d,v])=>`${d}: ${v}`).join(' | ')
    ]);
  });
  downloadCSV(out,'2026_PLAN_action_log.csv');
});

// -------------- PIPELINE --------------
function populatePipeFilters(){
  const items = DATA.five_year_spe_sales.items;
  const fill = (sel, values) => values.forEach(v=>{ const o=document.createElement('option'); o.value=v; o.textContent=v; sel.appendChild(o); });
  fill($('#p-year'), uniq(items.map(i=>String(i.year)).filter(x=>x!=='null')));
  fill($('#p-engine'), uniq(items.map(i=>i.engine_type)));
  fill($('#p-cust'), uniq(items.map(i=>i.customer)));
  fill($('#p-qtr'), uniq(items.map(i=>i.quarter)));
}

function pipeFiltered(){
  const yf = getYearFilter();
  const y = $('#p-year').value;
  const e = $('#p-engine').value;
  const c = $('#p-cust').value;
  const q = $('#p-qtr').value;
  return DATA.five_year_spe_sales.items.filter(i=>{
    if(yf && i.year!==yf) return false;
    if(y && String(i.year)!==y) return false;
    if(e && i.engine_type!==e) return false;
    if(c && i.customer!==c) return false;
    if(q && i.quarter!==q) return false;
    return true;
  });
}

let chFunnel, chPipeHeat, chAmount;
function renderFunnel(){
  const items = pipeFiltered();
  const years = uniq(items.map(i=>i.year).filter(Boolean)).sort();
  const quarters = ['Q1','Q2','Q3','Q4','Q3/Q4','—'];
  const qSet = uniq(items.map(i=>i.quarter || '—'));
  const qOrdered = quarters.filter(q=>qSet.includes(q)).concat(qSet.filter(q=>!quarters.includes(q)));
  const series = qOrdered.map(q=>({
    name: q,
    data: years.map(y=>items.filter(i=>i.year===y && (i.quarter||'—')===q).length)
  }));
  if(chFunnel) chFunnel.destroy();
  chFunnel = new ApexCharts($('#chartFunnel'), {
    chart:{type:'bar',stacked:true,height:340,toolbar:{show:false},background:'transparent'},
    theme:{mode:'dark'}, series,
    xaxis:{categories:years},
    colors:['#5ee7a7','#7bd3ff','#b6a1ff','#ffb547','#ff8c61','#9eb0c4'],
    plotOptions:{bar:{columnWidth:'45%',borderRadius:4}},
    legend:{position:'top',labels:{colors:'#e7eef7'}},
    grid:{borderColor:'#1f2b38'}, tooltip:{theme:'dark'}
  });
  chFunnel.render();
}

function renderPipeHeat(){
  const items = pipeFiltered();
  const years = uniq(items.map(i=>i.year).filter(Boolean)).sort();
  const custs = uniq(items.map(i=>i.customer).filter(Boolean));
  const series = custs.map(c=>({
    name:c, data: years.map(y=>({x:String(y), y:items.filter(i=>i.customer===c && i.year===y).length}))
  }));
  if(chPipeHeat) chPipeHeat.destroy();
  chPipeHeat = new ApexCharts($('#chartPipeHeat'), {
    chart:{type:'heatmap',height:340,toolbar:{show:false},background:'transparent'},
    theme:{mode:'dark'}, series,
    dataLabels:{enabled:true,style:{colors:['#0a0e13'],fontSize:'10px'}},
    colors:['#7bd3ff'],
    plotOptions:{heatmap:{enableShades:true,shadeIntensity:0.6}},
    grid:{borderColor:'#1f2b38'}, tooltip:{theme:'dark'}
  });
  chPipeHeat.render();
}

function renderAmount(){
  const items = pipeFiltered();
  const years = uniq(items.map(i=>i.year).filter(Boolean)).sort();
  const data = years.map(y=>{
    const s = items.filter(i=>i.year===y).reduce((a,i)=>a+(i.amount||0),0);
    return +s.toFixed(2);
  });
  if(chAmount) chAmount.destroy();
  chAmount = new ApexCharts($('#chartAmount'), {
    chart:{type:'line',height:300,toolbar:{show:false},background:'transparent'},
    theme:{mode:'dark'},
    series:[{name:'Amount',data}],
    xaxis:{categories:years},
    stroke:{curve:'smooth',width:3},
    markers:{size:6},
    colors:['#b6a1ff'],
    grid:{borderColor:'#1f2b38'}, tooltip:{theme:'dark'},
    dataLabels:{enabled:true,style:{colors:['#e7eef7'],fontSize:'11px'},background:{enabled:false}}
  });
  chAmount.render();
}

function renderPipelineTable(){
  const tbody = $('#tbl-pipeline tbody');
  const items = pipeFiltered();
  tbody.innerHTML = items.map(i=>`
    <tr>
      <td><strong>${esc(i.customer||'')}</strong></td>
      <td><span class="tag">${esc(i.engine_type||'')}</span></td>
      <td>${esc(i.year||'')}</td>
      <td>${esc(i.quarter||'')}</td>
      <td>${i.amount!=null?`£${i.amount.toFixed(1)}m`:''}</td>
      <td style="max-width:260px">${esc(i.comments||'')}</td>
    </tr>
  `).join('');
}

function renderPipeline(){ renderFunnel(); renderPipeHeat(); renderAmount(); renderPipelineTable(); }
['p-year','p-engine','p-cust','p-qtr'].forEach(id=>$('#'+id).addEventListener('change',renderPipeline));
$('#p-reset').addEventListener('click',()=>{ ['p-year','p-engine','p-cust','p-qtr'].forEach(id=>$('#'+id).value=''); renderPipeline(); });
$('#btn-export-pipeline').addEventListener('click',()=>{
  const rows = pipeFiltered();
  const out = [['Customer','Engine','Year','Quarter','Amount (£m)','Comments']];
  rows.forEach(i=>out.push([i.customer,i.engine_type,i.year,i.quarter,i.amount,i.comments]));
  downloadCSV(out,'2026_PLAN_pipeline.csv');
});

// -------------- YEARLY --------------
function renderYearly(){
  const yt = DATA.annual_summary.by_year;
  const years = Object.keys(yt).sort();
  const host = $('#yearCards');
  host.innerHTML = '';
  years.forEach((y,idx)=>{
    const d = yt[y];
    const chartId = `yc-${y}`;
    const card = document.createElement('div');
    card.className='year-card';
    card.innerHTML = `
      <h4>${esc(y)}</h4>
      <div class="total"><strong style="color:var(--accent);font-size:16px">${d.grand_total}</strong> engines</div>
      <div id="${chartId}" style="margin-top:8px"></div>
      <div class="cust-list">
        ${d.customers.map(c=>`<div class="c"><span class="n">${esc(c.name)} <span class="e">${c.engines.map(e=>e.type+' ('+e.count+')').join(', ')}</span></span><span>${c.total}</span></div>`).join('')}
      </div>
    `;
    host.appendChild(card);

    // per-engine bar chart per year card
    const engineAgg = {};
    d.customers.forEach(c=>c.engines.forEach(e=>{ engineAgg[e.type] = (engineAgg[e.type]||0)+e.count; }));
    const keys = Object.keys(engineAgg);
    const vals = keys.map(k=>engineAgg[k]);
    if(keys.length){
      new ApexCharts(document.getElementById(chartId),{
        chart:{type:'bar',height:140,toolbar:{show:false},sparkline:{enabled:false},background:'transparent'},
        theme:{mode:'dark'},
        series:[{name:'Engines',data:vals}],
        xaxis:{categories:keys,labels:{style:{fontSize:'9px'}}},
        plotOptions:{bar:{borderRadius:3,columnWidth:'60%',dataLabels:{position:'top'}}},
        colors:['#7bd3ff'],
        dataLabels:{enabled:true,offsetY:-14,style:{fontSize:'10px',colors:['#e7eef7']}},
        grid:{borderColor:'#1f2b38',padding:{left:0,right:0,top:-14}},
        yaxis:{show:false},tooltip:{theme:'dark'}
      }).render();
    }
  });

  // Full pivot table
  const headRow = ['<tr>'];
  const cols = [];
  years.forEach(y=>{
    headRow.push(`<th colspan="2" style="text-align:center;background:linear-gradient(135deg,rgba(123,211,255,.1),rgba(182,161,255,.05))">${esc(y)}</th>`);
    cols.push([y+' Customer', y+' Engines']);
  });
  headRow.push('</tr>');
  const subHead = '<tr>'+years.map(()=>'<th>Customer/Engine</th><th>Count</th>').join('')+'</tr>';
  $('#tbl-yearly-head').innerHTML = headRow.join('')+subHead;

  const body = [];
  const maxRows = Math.max(...years.map(y=>{
    return (yt[y].customers||[]).reduce((s,c)=>s+1+c.engines.length,0) + 1;
  }));
  const yRows = years.map(y=>{
    const rows = [];
    (yt[y].customers||[]).forEach(c=>{
      rows.push([`<strong>${esc(c.name)}</strong>`, c.total]);
      c.engines.forEach(e=>rows.push([`<span class="tag">${esc(e.type)}</span>`, e.count]));
    });
    rows.push([`<strong style="color:var(--accent)">Grand Total</strong>`, yt[y].grand_total]);
    return rows;
  });
  for(let r=0;r<maxRows;r++){
    const cells = [];
    yRows.forEach(rows=>{
      const cell = rows[r];
      if(cell){
        cells.push(`<td>${cell[0]}</td><td>${cell[1]}</td>`);
      } else {
        cells.push(`<td></td><td></td>`);
      }
    });
    body.push('<tr>'+cells.join('')+'</tr>');
  }
  $('#tbl-yearly-body').innerHTML = body.join('');
}

// -------------- Init --------------
populateActionFilters();
populatePipeFilters();
renderOverview();
renderActionTable();
renderPipelineTable();
renderYearly();
// Heavy charts on their tab open, but also warm the first pipeline/heatmap
setTimeout(()=>{ renderHeatmap(); renderPipeline(); }, 150);
</script>
</body>
</html>
"""


def main():
    OUT_HTML.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    payload = build_payload()

    # Save raw JSON payload for parity
    OUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

    html = HTML_TEMPLATE.replace("__PAYLOAD__", json.dumps(payload, ensure_ascii=False, default=str))
    OUT_HTML.write_text(html, encoding="utf-8")

    # Print a quick summary
    print(f"Wrote: {OUT_HTML}")
    print(f"Wrote: {OUT_JSON}")
    print("Counts:", payload["metadata"]["counts"])
    print("5YP by_year:", payload["five_year_spe_sales"]["totals"]["by_year"])
    print("5YP by_engine:", payload["five_year_spe_sales"]["totals"]["by_engine"])
    print("5YP top customers:", list(payload["five_year_spe_sales"]["totals"]["by_customer"].items())[:5])
    print("Annual grand totals:", {y: v["grand_total"] for y, v in payload["annual_summary"]["by_year"].items()})
    owners = Counter([i["owner"] for i in payload["one_year_plan"]["items"] if i["owner"]])
    print("1YP owners:", owners.most_common())


if __name__ == "__main__":
    main()
