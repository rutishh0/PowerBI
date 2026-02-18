"""
Rolls-Royce Civil Aerospace — Statement of Account Dashboard
=============================================================
A professional, adaptive dashboard that ingests any RR-style SOA
Excel workbook and visualises CRC payments, TotalCare, Spare Parts,
Late Payment Interest, credits, balances, ageing, and more.

NOTHING is hard-coded to a particular customer or column position.
The parser detects sections, headers, and amounts dynamically.
"""

import io
import re
import math
import json
from datetime import datetime, timedelta
from collections import OrderedDict

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook
from pdf_export import generate_pdf_report

# ─────────────────────────────────────────────────────────────
# 0.  PAGE CONFIG & BRANDING
# ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Rolls-Royce Civil Aerospace · SOA Dashboard",
    page_icon="✈",
    layout="wide",
    initial_sidebar_state="expanded",
)

# -- Rolls-Royce Civil Aerospace colour palette --
RR_NAVY   = "#10069F"
RR_DARK   = "#0C0033"
RR_SILVER = "#C0C0C0"
RR_LIGHT  = "#E8E8EE"
RR_WHITE  = "#FFFFFF"
RR_GOLD   = "#B8860B"
RR_RED    = "#D32F2F"
RR_GREEN  = "#2E7D32"
RR_BLUE2  = "#1565C0"
RR_AMBER  = "#F9A825"

SECTION_COLOURS = [RR_NAVY, "#1565C0", "#5E35B1", "#00838F", "#C62828", "#EF6C00", "#2E7D32", "#6A1B9A"]

# Inject CSS for RR look-and-feel
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

/* ---------- global ---------- */
html, body, [class*="css"] { font-family: 'Inter', sans-serif !important; }
.stApp { background: #F0F1F5; }

/* Force dark readable text everywhere in main area */
.stApp p, .stApp span, .stApp label, .stApp div { color: #1a1a2e; }
.stMarkdown p { color: #1a1a2e !important; font-size: 0.95rem; }
.stMarkdown strong { color: #0C0033 !important; }

/* ---------- header bar ---------- */
.rr-header {
    background: linear-gradient(135deg, #10069F 0%, #0C0033 100%);
    padding: 1.8rem 2.2rem;
    border-radius: 14px;
    margin-bottom: 1.4rem;
    display: flex; align-items: center; justify-content: space-between;
    box-shadow: 0 6px 24px rgba(16,6,159,.30);
}
.rr-header h1 {
    color: #FFFFFF !important; margin: 0; font-size: 1.65rem; font-weight: 700; letter-spacing: .5px;
}
.rr-header .subtitle {
    color: #D0D0E0 !important; font-size: .92rem; margin-top: .2rem; font-weight: 400;
}
.rr-logo-text {
    color: #FFFFFF !important; font-size: 1.15rem; font-weight: 700; letter-spacing: 3px; text-transform: uppercase;
    border: 2px solid rgba(255,255,255,.6); padding: .4rem 1rem; border-radius: 4px;
}

/* ---------- metric cards ---------- */
.metric-card {
    background: #FFFFFF;
    border-radius: 12px;
    padding: 1.2rem 1rem;
    text-align: center;
    border-left: 5px solid #10069F;
    box-shadow: 0 2px 12px rgba(0,0,0,.08);
    transition: transform .15s, box-shadow .15s;
}
.metric-card:hover { transform: translateY(-3px); box-shadow: 0 6px 20px rgba(0,0,0,.12); }
.metric-card .label {
    color: #444 !important;
    font-size: .78rem !important;
    font-weight: 700 !important;
    text-transform: uppercase;
    letter-spacing: .9px;
    margin-bottom: .35rem;
}
.metric-card .value {
    color: #0C0033 !important;
    font-size: 1.45rem !important;
    font-weight: 800 !important;
}
.metric-card .value.negative { color: #C62828 !important; }
.metric-card .value.positive { color: #1B5E20 !important; }

/* ---------- section header ---------- */
.section-hdr {
    background: linear-gradient(90deg, #10069F, #1565C0);
    color: #FFFFFF !important;
    padding: .7rem 1.4rem;
    border-radius: 8px;
    font-weight: 700;
    font-size: 1.05rem;
    margin: 1.8rem 0 1rem 0;
    letter-spacing: .5px;
    box-shadow: 0 2px 8px rgba(16,6,159,.18);
}

/* ---------- tables ---------- */
div[data-testid="stDataFrame"] {
    border-radius: 10px;
    overflow: hidden;
    box-shadow: 0 2px 10px rgba(0,0,0,.08);
    font-size: .88rem !important;
}
div[data-testid="stDataFrame"] th {
    background: #10069F !important;
    color: #FFF !important;
    font-weight: 600 !important;
    font-size: .82rem !important;
}
div[data-testid="stDataFrame"] td {
    color: #1a1a2e !important;
    font-size: .84rem !important;
}

/* ---------- sidebar ---------- */
section[data-testid="stSidebar"] { background: #0C0033 !important; }
section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3,
section[data-testid="stSidebar"] h4,
section[data-testid="stSidebar"] h5 { color: #FFFFFF !important; font-weight: 700 !important; }
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] span,
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] div { color: #E0E0F0 !important; }
section[data-testid="stSidebar"] .stSelectbox label,
section[data-testid="stSidebar"] .stMultiSelect label,
section[data-testid="stSidebar"] .stCheckbox label { color: #FFFFFF !important; font-weight: 500 !important; }
section[data-testid="stSidebar"] hr { border-color: rgba(255,255,255,.15) !important; }
section[data-testid="stSidebar"] .stFileUploader label { color: #FFFFFF !important; }
section[data-testid="stSidebar"] .stFileUploader div { color: #D0D0E0 !important; }

/* ---------- tabs ---------- */
.stTabs [data-baseweb="tab-list"] { gap: 6px; }
.stTabs [data-baseweb="tab"] {
    background: #FFFFFF; border-radius: 8px 8px 0 0; padding: .65rem 1.3rem;
    font-weight: 600; color: #333 !important; border: 1px solid #CCC; border-bottom: none;
    font-size: .9rem;
}
.stTabs [aria-selected="true"] {
    background: #10069F !important; color: #FFFFFF !important; border-color: #10069F !important;
}
.stTabs [aria-selected="true"] p,
.stTabs [aria-selected="true"] span,
.stTabs [aria-selected="true"] div {
    color: #FFFFFF !important;
}

/* ---------- selectbox / multiselect in main area ---------- */
.stSelectbox label, .stMultiSelect label, .stSlider label, .stCheckbox label {
    color: #1a1a2e !important;
    font-weight: 600 !important;
    font-size: .88rem !important;
}

/* ---------- metrics (st.metric) ---------- */
div[data-testid="stMetric"] label { color: #555 !important; font-weight: 600 !important; font-size: .82rem !important; }
div[data-testid="stMetric"] div[data-testid="stMetricValue"] { color: #0C0033 !important; font-weight: 700 !important; }

/* ---------- info chip ---------- */
.info-chip {
    display: inline-block;
    background: #E0E1EC;
    padding: .3rem .75rem;
    border-radius: 6px;
    font-size: .82rem;
    color: #1a1a2e !important;
    font-weight: 500;
    margin-right: .5rem;
    margin-bottom: .35rem;
    border: 1px solid #D0D1DC;
}
.info-chip b { color: #0C0033 !important; }

/* ---------- plotly chart text ---------- */
.js-plotly-plot .plotly text { fill: #333 !important; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────
# 1.  SMART EXCEL PARSER
# ─────────────────────────────────────────────────────────────

# Keywords that signal a section header row
SECTION_KEYWORDS = [
    "charges", "credits", "credit", "totalcare", "familycare", "missioncare",
    "spare parts", "late payment", "interest", "customer respon",
    "customer responsibility", "usable", "offset",
]

SUMMARY_KEYWORDS = ["total", "overdue", "available credit", "total overdue", "net balance"]

HEADER_KEYWORDS = [
    "company", "account", "reference", "document", "date", "amount", "curr",
    "text", "assignment", "arrangement", "comments", "status", "action",
    "days", "late", "lpi", "invoice", "type", "interest", "net due",
]


def _is_section_header(row_values: list, col_count: int) -> bool:
    """Return True if this row looks like a section heading."""
    non_empty = [(i, v) for i, v in enumerate(row_values) if v is not None]
    if not non_empty:
        return False
    # Typically section headers only populate col A (and sometimes 1-2 more)
    if len(non_empty) > 3:
        return False
    text = str(non_empty[0][1]).strip().lower()
    # Must not be purely numeric
    try:
        float(text.replace(",", ""))
        return False
    except ValueError:
        pass
    # Exclude summary-row patterns (e.g. "Available Credit:", "Total Overdue")
    text_clean = text.rstrip(":")
    if text_clean in ("total", "overdue", "available credit", "total overdue", "net balance"):
        return False
    # Also exclude if the SECOND cell is a number (looks like a summary: "Total  | 12345.00")
    if len(non_empty) == 2:
        try:
            float(str(non_empty[1][1]).replace(",", "").replace("$", "").strip())
            # Row has a label + number → summary, not section header
            if any(sw in text_clean for sw in ("total", "overdue", "credit", "balance")):
                return False
        except (ValueError, TypeError):
            pass
    return any(kw in text for kw in SECTION_KEYWORDS)


def _is_header_row(row_values: list) -> bool:
    """Return True if row looks like a column header row.

    Strict checks:
      1. At least 4 non-empty cells
      2. Most values must be short text (< 35 chars) — rules out data rows
      3. No large numeric values (> 100) — header cells are labels, not amounts
      4. At least 3 keyword hits among the SHORT text values
    """
    non_empty = [v for v in row_values if v is not None]
    if len(non_empty) < 4:
        return False
    # Reject if any cell is a large number (data rows have amounts)
    for v in non_empty:
        try:
            n = float(str(v).replace(",", "").replace("$", "").strip())
            if abs(n) > 100:
                return False  # Likely a data row
        except (ValueError, TypeError):
            pass
    # Only count keywords in SHORT text cells (column names are concise)
    short_texts = [str(v).strip().lower() for v in non_empty if len(str(v).strip()) < 35]
    if len(short_texts) < 3:
        return False
    hits = sum(1 for t in short_texts for kw in HEADER_KEYWORDS if kw in t)
    return hits >= 3


def _is_summary_row(row_values: list) -> str | None:
    """Return the summary type if this looks like a Total/Overdue row, else None.

    The match must be a STANDALONE label (short text, essentially the keyword itself),
    not a substring of a description like 'TotalCare' or 'Late Payment Invoice'.
    """
    for v in row_values:
        if v is None:
            continue
        t = str(v).strip().lower().rstrip(":")
        # Must be a short cell value (labels like "Total" or "Overdue")
        if len(t) > 25:
            continue
        if t in ("total", "overdue", "available credit", "total overdue",
                 "net balance", "total:", "overdue:"):
            return t.rstrip(":")
        # Also match patterns like "Total  " with trailing spaces already stripped
    return None


def _find_amount_col(header: list) -> int | None:
    """Find the column index that holds amounts."""
    for i, h in enumerate(header):
        if h is None:
            continue
        hl = str(h).lower()
        if "amount" in hl:
            return i
    return None


def _coerce_amount(val) -> float | None:
    """Convert a cell value to a float, handling $, commas, etc."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace("$", "").replace(" ", "")
    if s in ("", "-", "$ -", "$-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _coerce_date(val) -> pd.Timestamp | None:
    """Try to parse a date from various formats."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return pd.Timestamp(val)
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return pd.Timestamp(datetime.strptime(s, fmt))
        except ValueError:
            continue
    return None


def _coerce_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _normalise_header(raw: list) -> list:
    """Clean header names so they are consistent."""
    result = []
    for h in raw:
        if h is None:
            result.append(None)
            continue
        s = str(h).strip()
        # Collapse whitespace / newlines
        s = re.sub(r"\s+", " ", s)
        result.append(s)
    return result


def _map_columns(header: list) -> dict:
    """Map semantic roles to column indices from the header row.
    Returns dict with keys like 'amount', 'date', 'due_date', 'reference', etc."""
    mapping = {}
    hl = [str(h).lower() if h else "" for h in header]
    for i, h in enumerate(hl):
        if not h:
            continue
        if "amount" in h:
            mapping["amount"] = i
        elif h in ("curr", "currency"):
            mapping["currency"] = i
        elif "net due" in h or "due date" in h:
            mapping["due_date"] = i
        elif "document" in h and "date" in h:
            mapping["doc_date"] = i
        elif "document" in h and "no" in h:
            mapping["doc_no"] = i
        elif "invoice date" in h:
            mapping["doc_date"] = i
        elif h == "reference" or "reference" in h:
            mapping["reference"] = i
        elif h == "company" or "company" in h:
            mapping["company"] = i
        elif h == "account" or "account" in h:
            mapping["account"] = i
        elif "text" == h:
            mapping["text"] = i
        elif "assignment" in h or "arrangement" in h:
            mapping["assignment"] = i
        elif "r-r comment" in h or "rr comment" in h:
            mapping["rr_comments"] = i
        elif "action" in h or "reqd" in h:
            mapping["action_owner"] = i
        elif "days" in h and "late" in h:
            mapping["days_late"] = i
        elif "rata" in h:
            mapping["rata_date"] = i          # RATA Date is a date, not days-late int
        elif "comment" in h and "r-r" not in h and "rr" not in h:
            mapping["customer_comments"] = i
        elif "status" in h:
            mapping["status"] = i
        elif "customer" in h and "comment" not in h and "name" not in h and "n" not in h and "respon" not in h:
            mapping["customer_name"] = i
        elif "lpi" in h:
            mapping["lpi_cumulated"] = i
        elif "etr" in h or "po" in h or "pr" in h:
            mapping["po_reference"] = i
        elif "type" in h:
            mapping["type"] = i
        elif "interest" in h or "calc" in h:
            mapping["interest_method"] = i

    # If no explicit date columns, try to find 'date' generically
    if "doc_date" not in mapping:
        for i, h in enumerate(hl):
            if "date" in h and i not in mapping.values():
                mapping["doc_date"] = i
                break
    if "due_date" not in mapping:
        for i, h in enumerate(hl):
            if "due" in h and i not in mapping.values():
                mapping["due_date"] = i
                break
    return mapping


# ---------- Main parse function ----------

def parse_soa_workbook(file) -> dict:
    """Parse a Rolls-Royce Statement of Account workbook (all sheets).

    Returns a dict:
        metadata  : dict of customer info, LPI rate, avg days late, etc.
        sections  : OrderedDict  section_name -> { header, colmap, rows (list[dict]), totals }
        all_items : pd.DataFrame  flattened across all sections
        grand_totals : dict
    """
    wb = load_workbook(file, data_only=True)
    
    # Global accumulators
    all_metadata = {}
    all_sections = OrderedDict()
    all_items_list = []

    # Iterate over all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_col = ws.max_column or 20
        all_rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_col, values_only=True):
            all_rows.append(list(row))

        if not all_rows:
            continue

        # ---- PASS 1: Metadata ----
        local_metadata = {}
        for idx, row in enumerate(all_rows[:15]):
            joined = " ".join(str(v) for v in row if v is not None).lower()
            for v in row:
                if v is None:
                    continue
                s = str(v).strip()
                sl = s.lower()
                if "statement of account" in sl:
                    local_metadata["title"] = s
                if "customer" in sl and ("name" in sl or ":" in sl) and "customer_name" not in local_metadata:
                    vi = row.index(v)
                    for nv in row[vi+1:]:
                        if nv is not None:
                            local_metadata["customer_name"] = str(nv).strip()
                            break
                if ("customer" in sl and ("#" in sl or "n" in sl and ":" in sl)) or "customer n" in sl:
                    vi = row.index(v)
                    for nv in row[vi+1:]:
                        if nv is not None:
                            local_metadata["customer_id"] = str(nv).strip()
                            break
                if "contact" in sl:
                    vi = row.index(v)
                    for nv in row[vi+1:]:
                        if nv is not None:
                            local_metadata["contact"] = str(nv).strip()
                            break
                if "lpi" in sl or "lp ratio" in sl or "lp rate" in sl:
                    for nv in row:
                        if nv is None:
                            continue
                        nv_str = str(nv).strip()
                        if "%" in nv_str:
                            try:
                                local_metadata["lpi_rate"] = float(nv_str.replace("%", "")) / 100.0
                                break
                            except ValueError:
                                pass
                        amt = _coerce_amount(nv)
                        if amt is not None and 0 < abs(amt) < 1:
                            local_metadata["lpi_rate"] = amt
                            break
                if "average days late" in sl or "avg days late" in sl or "average days late" in joined:
                    for nv in row:
                        if nv is None:
                            continue
                        val = _coerce_int(nv)
                        if val is not None and val > 0:
                            local_metadata["avg_days_late"] = val
                            break
                if "today" in sl:
                    for nv in row:
                        d = _coerce_date(nv)
                        if d is not None:
                            local_metadata["report_date"] = d
                            break

        # Merge local metadata
        for k, v in local_metadata.items():
            if k not in all_metadata or all_metadata[k] is None:
                all_metadata[k] = v

        # ---- PASS 2: Boundaries ----
        master_header = None
        master_header_idx = None
        sections_info = [] 

        for idx, row in enumerate(all_rows):
            if _is_header_row(row) and master_header is None:
                master_header = _normalise_header(row)
                master_header_idx = idx
                continue
            if _is_section_header(row, max_col):
                name = str([v for v in row if v is not None][0]).strip()
                sections_info.append({"name": name, "start": idx})

        for i, sec in enumerate(sections_info):
            if i + 1 < len(sections_info):
                sec["end"] = sections_info[i + 1]["start"]
            else:
                sec["end"] = len(all_rows)

        # ---- PASS 3: Parse Sections ----
        for sec in sections_info:
            sec_name = sec["name"]
            start = sec["start"]
            end = sec["end"]

            header = master_header
            header_idx = master_header_idx
            col_map = None

            for offset in range(1, 4):
                ri = start + offset
                if ri >= end:
                    break
                if _is_header_row(all_rows[ri]):
                    header = _normalise_header(all_rows[ri])
                    header_idx = ri
                    break

            if header:
                col_map = _map_columns(header)
            else:
                col_map = {}

            amt_idx = col_map.get("amount")
            if amt_idx is None and header:
                amt_idx = _find_amount_col(header)
                if amt_idx is not None:
                    col_map["amount"] = amt_idx

            data_rows = []
            totals = {}
            data_start = (header_idx + 1) if header_idx and header_idx >= start else start + 1

            for ri in range(data_start, end):
                row = all_rows[ri]
                summary_type = _is_summary_row(row)
                if summary_type:
                    for v in row:
                        amt = _coerce_amount(v)
                        if amt is not None:
                            totals[summary_type] = amt
                            break
                    continue

                if _is_section_header(row, max_col):
                    continue
                if _is_header_row(row):
                    header = _normalise_header(row)
                    col_map = _map_columns(header)
                    amt_idx = col_map.get("amount")
                    if amt_idx is None and header:
                        amt_idx = _find_amount_col(header)
                        if amt_idx is not None:
                            col_map["amount"] = amt_idx
                    continue

                amt_val = None
                if amt_idx is not None and amt_idx < len(row):
                    amt_val = _coerce_amount(row[amt_idx])
                if amt_val is None:
                    for ci, cv in enumerate(row):
                        a = _coerce_amount(cv)
                        if a is not None and abs(a) > 0.01:
                            if abs(a) > 100 or (col_map.get("days_late") is not None and ci != col_map.get("days_late")):
                                amt_val = a
                                break

                if amt_val is None:
                    continue

                record = {
                    "Section": sec_name,
                    "Amount": amt_val,
                }

                def _get(key, coerce=str):
                    ci = col_map.get(key)
                    if ci is None or ci >= len(row):
                        return None
                    v = row[ci]
                    if v is None:
                        return None
                    if coerce == float:
                        return _coerce_amount(v)
                    if coerce == "date":
                        return _coerce_date(v)
                    if coerce == int:
                        return _coerce_int(v)
                    return str(v).strip()

                record["Company"]         = _get("company")
                record["Account"]         = _get("account")
                record["Reference"]       = _get("reference")
                record["Document Date"]   = _get("doc_date", "date")
                record["Due Date"]        = _get("due_date", "date")
                record["Currency"]        = _get("currency")
                record["Text"]            = _get("text")
                record["Assignment"]      = _get("assignment")
                record["R-R Comments"]    = _get("rr_comments")
                record["Action Owner"]    = _get("action_owner")
                record["Days Late"]       = _get("days_late", int)
                record["Customer Comments"] = _get("customer_comments")
                record["Status"]          = _get("status")
                record["PO Reference"]    = _get("po_reference")
                record["LPI Cumulated"]   = _get("lpi_cumulated")
                record["Type"]            = _get("type")
                record["Document No"]     = _get("doc_no")
                record["Interest Method"] = _get("interest_method")
                record["Customer Name"]   = _get("customer_name")

                # Auto-compute Days Late from Due Date when not explicitly available
                if record["Days Late"] is None and record["Due Date"] is not None:
                    try:
                        due = record["Due Date"]
                        today = pd.Timestamp.now().normalize()
                        if due < today:
                            record["Days Late"] = (today - due).days
                        else:
                            record["Days Late"] = 0
                    except Exception:
                        pass

                # Derive a unified Status field
                if not record.get("Status"):
                    for field in ["R-R Comments", "Action Owner", "Customer Comments"]:
                        v = record.get(field, "")
                        if v and any(kw in v.lower() for kw in ["ready for payment", "under approval", "under review",
                                                                 "dispute", "ongoing", "et to process", "payment pending",
                                                                 "invoice sent", "credit note", "approved",
                                                                 "transfer", "invoice approved", "pending for payment"]):
                            record["Status"] = v
                            break
                if not record.get("Status"):
                    rrc = record.get("R-R Comments", "")
                    if rrc:
                        record["Status"] = rrc

                # Determine if debit or credit
                record["Entry Type"] = "Credit" if amt_val < 0 else "Charge"

                data_rows.append(record)
                all_items_list.append(record)

            # --- Merge logic ---
            if sec_name not in all_sections:
                all_sections[sec_name] = {
                    "header": header,
                    "colmap": col_map,
                    "rows": data_rows,
                    "totals": totals,
                }
            else:
                all_sections[sec_name]["rows"].extend(data_rows)
                # Aggregate totals
                existing_totals = all_sections[sec_name]["totals"]
                for k, v in totals.items():
                    if k in existing_totals:
                        existing_totals[k] += v
                    else:
                        existing_totals[k] = v

    # Build DataFrame
    df = pd.DataFrame(all_items_list)
    if df.empty:
        df = pd.DataFrame(columns=["Section", "Amount", "Entry Type"])

    # ---- Grand totals ----
    grand = {}
    for sec_name, sec_data in all_sections.items():
        for k, v in sec_data["totals"].items():
            if "total overdue" in k:
                grand["total_overdue"] = grand.get("total_overdue", 0) + v
            elif "overdue" in k:
                grand.setdefault("section_overdue", {})[sec_name] = grand.setdefault("section_overdue", {}).get(sec_name, 0) + v
            elif "available credit" in k:
                grand.setdefault("available_credits", {})[sec_name] = grand.setdefault("available_credits", {}).get(sec_name, 0) + v
            elif "total" in k:
                grand.setdefault("section_totals", {})[sec_name] = grand.setdefault("section_totals", {}).get(sec_name, 0) + v

    if not df.empty:
        grand["total_charges"]  = df.loc[df["Amount"] > 0, "Amount"].sum()
        grand["total_credits"]  = df.loc[df["Amount"] < 0, "Amount"].sum()
        grand["net_balance"]    = df["Amount"].sum()
        grand["item_count"]     = len(df)
        if "total_overdue" not in grand:
            overdue_sum = sum(grand.get("section_overdue", {}).values())
            if overdue_sum:
                grand["total_overdue"] = overdue_sum
            else:
                if "section_overdue" in grand:
                    grand["total_overdue"] = sum(grand["section_overdue"].values())

    return {
        "metadata": all_metadata,
        "sections": all_sections,
        "all_items": df,
        "grand_totals": grand,
    }


# ─────────────────────────────────────────────────────────────
# 2.  HELPER VIS FUNCTIONS
# ─────────────────────────────────────────────────────────────

def fmt_currency(val, short=False):
    """Format a number as USD currency string."""
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return "—"
    neg = val < 0
    av = abs(val)
    if short and av >= 1_000_000:
        s = f"${av/1_000_000:,.2f}M"
    elif short and av >= 1_000:
        s = f"${av/1_000:,.1f}K"
    else:
        s = f"${av:,.2f}"
    return f"-{s}" if neg else s


def metric_card(label, value, color_class=""):
    return f"""
    <div class="metric-card">
        <div class="label">{label}</div>
        <div class="value {color_class}">{value}</div>
    </div>"""


def make_donut(labels, values, title, colors=None):
    if colors is None:
        colors = SECTION_COLOURS[:len(labels)]
    fig = go.Figure(go.Pie(
        labels=labels, values=values, hole=.55,
        marker=dict(colors=colors, line=dict(color="#FFF", width=2)),
        textinfo="label+percent", textposition="outside",
        textfont=dict(size=12, color="#222"),
    ))
    fig.update_layout(
        title=dict(text=f"<b>{title}</b>", font=dict(size=15, color="#1a1a2e")),
        showlegend=False, margin=dict(t=55, b=25, l=25, r=25),
        height=360, paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
    )
    return fig


# Shared chart layout defaults for readable text
CHART_LAYOUT = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    font=dict(size=12, color="#1a1a2e", family="Inter, sans-serif"),
    xaxis=dict(gridcolor="#E0E0E0", tickfont=dict(size=11, color="#333")),
    yaxis=dict(gridcolor="#E0E0E0", tickfont=dict(size=11, color="#333")),
)


def make_bar(df_plot, x, y, title, color=RR_NAVY, horizontal=False, color_col=None, text_auto=True):
    if horizontal:
        fig = px.bar(df_plot, y=x, x=y, orientation="h", title=title, text_auto=".2s" if text_auto else False,
                     color=color_col, color_discrete_sequence=SECTION_COLOURS if color_col else [color])
    else:
        fig = px.bar(df_plot, x=x, y=y, title=title, text_auto=".2s" if text_auto else False,
                     color=color_col, color_discrete_sequence=SECTION_COLOURS if color_col else [color])
    fig.update_layout(
        height=360, margin=dict(t=55, b=45, l=60, r=20),
        showlegend=bool(color_col),
        title=dict(font=dict(size=15, color="#1a1a2e")),
        **CHART_LAYOUT,
    )
    return fig


def make_chartjs_pie(labels: list, values: list, title: str, colors: list = None, height: int = 350) -> str:
    """Return an HTML string containing a Chart.js pie chart (rendered via st.components.v1.html)."""
    if colors is None:
        colors = SECTION_COLOURS[:len(labels)]
    labels_js = json.dumps(labels)
    values_js = json.dumps(values)
    colors_js = json.dumps(colors)
    return f"""
    <html><head>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
    <style>
      body {{ margin:0; padding:0; background:transparent; font-family:Inter,sans-serif; }}
      .chart-wrap {{ position:relative; width:100%; height:{height - 10}px; }}
    </style>
    </head><body>
    <div class="chart-wrap"><canvas id="pie"></canvas></div>
    <script>
    new Chart(document.getElementById('pie'), {{
      type: 'pie',
      data: {{
        labels: {labels_js},
        datasets: [{{
          data: {values_js},
          backgroundColor: {colors_js},
          borderColor: '#FFF',
          borderWidth: 2
        }}]
      }},
      options: {{
        responsive: true,
        maintainAspectRatio: false,
        animation: {{ duration: 500, easing: 'easeInOutQuart' }},
        plugins: {{
          title: {{
            display: true,
            text: '{title}',
            font: {{ size: 15, weight: 'bold', family: 'Inter, sans-serif' }},
            color: '#1a1a2e'
          }},
          legend: {{
            position: 'right',
            labels: {{
              font: {{ size: 11, family: 'Inter, sans-serif' }},
              color: '#333',
              padding: 12
            }}
          }}
        }}
      }}
    }});
    </script>
    </body></html>"""


def aging_bucket(days) -> str:
    if days is None or (isinstance(days, float) and math.isnan(days)):
        return "Unknown"
    d = int(days)
    if d <= 0:
        return "Current"
    elif d <= 30:
        return "1-30 Days"
    elif d <= 60:
        return "31-60 Days"
    elif d <= 90:
        return "61-90 Days"
    elif d <= 180:
        return "91-180 Days"
    else:
        return "180+ Days"


AGING_ORDER = ["Current", "1-30 Days", "31-60 Days", "61-90 Days", "91-180 Days", "180+ Days", "Unknown"]
AGING_COLORS = {"Current": "#2E7D32", "1-30 Days": "#66BB6A", "31-60 Days": "#F9A825",
                "61-90 Days": "#EF6C00", "91-180 Days": "#D32F2F", "180+ Days": "#B71C1C", "Unknown": "#9E9E9E"}


# ─────────────────────────────────────────────────────────────
# 3.  CACHED PARSE WRAPPER
# ─────────────────────────────────────────────────────────────

@st.cache_data(show_spinner=False)
def _cached_parse(file_bytes: bytes, file_name: str) -> dict:
    """Cache-friendly wrapper — accepts bytes so Streamlit can hash them."""
    buf = io.BytesIO(file_bytes)
    return parse_soa_workbook(buf)


# ─────────────────────────────────────────────────────────────
# 3a. DATA MERGING HELPERS
# ─────────────────────────────────────────────────────────────

def _merge_parsed_files(parsed_files: dict, selected_sources: list) -> tuple:
    """Merge parsed data from multiple files into combined structures.

    Returns (df_all, merged_sections, merged_grand, all_metadata).
    """
    frames = []
    merged_sections = OrderedDict()
    all_metadata = []
    multi = len(selected_sources) > 1

    for fname in selected_sources:
        data = parsed_files[fname]
        meta = data["metadata"]
        all_metadata.append(meta)
        df = data["all_items"].copy()
        if not df.empty:
            df["Source File"] = fname
            frames.append(df)
        for sec_name, sec_data in data["sections"].items():
            key = f"{fname} — {sec_name}" if multi else sec_name
            # Tag each row with Source File
            tagged_rows = []
            for r in sec_data["rows"]:
                rc = dict(r)
                rc["Source File"] = fname
                tagged_rows.append(rc)
            merged_sections[key] = {
                "header": sec_data["header"],
                "colmap": sec_data["colmap"],
                "rows": tagged_rows,
                "totals": sec_data["totals"],
            }

    if frames:
        df_all = pd.concat(frames, ignore_index=True)
    else:
        df_all = pd.DataFrame(columns=["Section", "Amount", "Entry Type", "Source File"])

    # Build combined grand totals
    merged_grand = {}
    if not df_all.empty:
        merged_grand["total_charges"] = df_all.loc[df_all["Amount"] > 0, "Amount"].sum()
        merged_grand["total_credits"] = df_all.loc[df_all["Amount"] < 0, "Amount"].sum()
        merged_grand["net_balance"] = df_all["Amount"].sum()
        merged_grand["item_count"] = len(df_all)
        # Aggregate section totals/overdue from individual files
        section_totals = {}
        section_overdue = {}
        total_overdue = 0.0
        for fname in selected_sources:
            g = parsed_files[fname]["grand_totals"]
            for sn, sv in g.get("section_totals", {}).items():
                k = f"{fname} — {sn}" if multi else sn
                section_totals[k] = section_totals.get(k, 0) + sv
            for sn, sv in g.get("section_overdue", {}).items():
                k = f"{fname} — {sn}" if multi else sn
                section_overdue[k] = section_overdue.get(k, 0) + sv
            total_overdue += g.get("total_overdue", 0)
        merged_grand["section_totals"] = section_totals
        merged_grand["section_overdue"] = section_overdue
        merged_grand["total_overdue"] = total_overdue if total_overdue else merged_grand["net_balance"]

    return df_all, merged_sections, merged_grand, all_metadata


def _apply_global_filters(df: pd.DataFrame, selected_sections: list,
                          selected_types: list, selected_statuses: list,
                          selected_currencies: list, selected_customers: list,
                          parsed_files: dict, selected_sources: list,
                          selected_overdue_status: list = None) -> pd.DataFrame:
    """Apply global sidebar filters to the combined DataFrame."""
    if df.empty:
        return df

    filtered = df.copy()

    # Filter by source file (customer filter maps to source files via metadata)
    if selected_customers:
        matching_sources = []
        for fname in selected_sources:
            meta = parsed_files[fname]["metadata"]
            cname = meta.get("customer_name", "Unknown")
            if cname in selected_customers:
                matching_sources.append(fname)
        if "Source File" in filtered.columns:
            filtered = filtered[filtered["Source File"].isin(matching_sources)]

    if selected_sections:
        filtered = filtered[filtered["Section"].isin(selected_sections)]

    if selected_types:
        filtered = filtered[filtered["Entry Type"].isin(selected_types)]

    if selected_statuses and "Status" in filtered.columns:
        filtered = filtered[filtered["Status"].isin(selected_statuses)]

    if selected_currencies and "Currency" in filtered.columns:
        filtered = filtered[
            filtered["Currency"].isna() | filtered["Currency"].isin(selected_currencies)
        ]

    # Overdue / Current filter
    if selected_overdue_status and len(selected_overdue_status) < 2 and "Days Late" in filtered.columns:
        if "Overdue" in selected_overdue_status and "Current" not in selected_overdue_status:
            filtered = filtered[filtered["Days Late"].fillna(0).apply(lambda x: int(x) if x else 0) > 0]
        elif "Current" in selected_overdue_status and "Overdue" not in selected_overdue_status:
            filtered = filtered[filtered["Days Late"].fillna(0).apply(lambda x: int(x) if x else 0) <= 0]

    return filtered


# ─────────────────────────────────────────────────────────────
# 3b. RENDER FUNCTIONS
# ─────────────────────────────────────────────────────────────

def render_customer_info(meta_list: list):
    """Render customer info bar — supports multiple customers."""
    if not meta_list:
        return
    if len(meta_list) == 1:
        meta = meta_list[0]
        cust_name = meta.get("customer_name", "Unknown Customer")
        cust_id   = meta.get("customer_id", "\u2014")
        contact   = meta.get("contact", "\u2014")
        lpi_rate  = meta.get("lpi_rate")
        avg_late  = meta.get("avg_days_late")
        report_dt = meta.get("report_date")
        st.markdown(f"""
        <div style="background:#FFFFFF; padding:1rem 1.8rem; border-radius:12px; margin-bottom:1.2rem;
                    display:flex; align-items:center; justify-content:space-between; flex-wrap:wrap;
                    box-shadow: 0 2px 10px rgba(0,0,0,.07); border: 1px solid #E0E1EC;">
            <div>
                <span style="font-weight:800; color:#0C0033; font-size:1.15rem;">{cust_name}</span>
                <span class="info-chip">ID: {cust_id}</span>
                <span class="info-chip">{contact}</span>
            </div>
            <div style="display:flex; gap:.8rem; align-items:center; flex-wrap:wrap;">
                {"<span class='info-chip'><b>LPI Rate:</b> " + f"{lpi_rate*100:.2f}%" + "</span>" if lpi_rate else ""}
                {"<span class='info-chip'><b>Avg Days Late:</b> " + str(avg_late) + "</span>" if avg_late else ""}
                {"<span class='info-chip'><b>Report:</b> " + report_dt.strftime('%d %b %Y') + "</span>" if report_dt else ""}
            </div>
        </div>
        """, unsafe_allow_html=True)
    else:
        # Multiple customers — show a card per customer
        for meta in meta_list:
            cust_name = meta.get("customer_name", "Unknown Customer")
            cust_id   = meta.get("customer_id", "\u2014")
            contact   = meta.get("contact", "\u2014")
            lpi_rate  = meta.get("lpi_rate")
            avg_late  = meta.get("avg_days_late")
            report_dt = meta.get("report_date")
            st.markdown(f"""
            <div style="background:#FFFFFF; padding:.8rem 1.4rem; border-radius:10px; margin-bottom:.6rem;
                        display:flex; align-items:center; justify-content:space-between; flex-wrap:wrap;
                        box-shadow: 0 1px 6px rgba(0,0,0,.05); border: 1px solid #E0E1EC;">
                <div>
                    <span style="font-weight:800; color:#0C0033; font-size:1.05rem;">{cust_name}</span>
                    <span class="info-chip">ID: {cust_id}</span>
                    <span class="info-chip">{contact}</span>
                </div>
                <div style="display:flex; gap:.6rem; align-items:center; flex-wrap:wrap;">
                    {"<span class='info-chip'><b>LPI Rate:</b> " + f"{lpi_rate*100:.2f}%" + "</span>" if lpi_rate else ""}
                    {"<span class='info-chip'><b>Avg Days Late:</b> " + str(avg_late) + "</span>" if avg_late else ""}
                    {"<span class='info-chip'><b>Report:</b> " + report_dt.strftime('%d %b %Y') + "</span>" if report_dt else ""}
                </div>
            </div>
            """, unsafe_allow_html=True)


def render_kpi_cards(df: pd.DataFrame, grand: dict, avg_late, sections_dict: dict = None):
    """Render the 6 KPI metric cards + credit highlight bar."""
    total_charges = grand.get("total_charges", 0)
    total_credits = grand.get("total_credits", 0)
    net_balance   = grand.get("net_balance", 0)
    total_overdue = grand.get("total_overdue", net_balance)
    item_count    = grand.get("item_count", 0)

    # Initialize session state for overdue filter
    if "filter_overdue_only" not in st.session_state:
        st.session_state["filter_overdue_only"] = False

    kpi_cols = st.columns(6)
    kpis = [
        ("Total Charges", fmt_currency(total_charges, short=True), ""),
        ("Total Credits", fmt_currency(total_credits, short=True), "positive" if total_credits < 0 else ""),
        ("Net Balance", fmt_currency(net_balance, short=True), "negative" if net_balance > 0 else "positive"),
        None,  # Placeholder — Total Overdue handled separately
        ("Avg Days Late", str(avg_late) if avg_late else "\u2014", ""),
        ("Open Items", str(item_count), ""),
    ]
    for i, col in enumerate(kpi_cols):
        if i == 3:
            # Total Overdue — clickable
            with col:
                st.markdown(metric_card("Total Overdue", fmt_currency(total_overdue, short=True), "negative"),
                            unsafe_allow_html=True)
                is_active = st.session_state.get("filter_overdue_only", False)
                btn_label = "Clear overdue filter" if is_active else "Show overdue only"
                if st.button(btn_label, key="overdue_toggle_btn", use_container_width=True):
                    st.session_state["filter_overdue_only"] = not is_active
                    st.rerun()
        elif kpis[i] is not None:
            label, value, cls = kpis[i]
            col.markdown(metric_card(label, value, cls), unsafe_allow_html=True)

    # --- Credit Available Highlight Card ---
    if total_credits != 0:
        credit_parts = ""
        if sections_dict:
            for sn, sd in sections_dict.items():
                avail = sd.get("totals", {}).get("available credit")
                sec_cr = sum(r["Amount"] for r in sd.get("rows", []) if r.get("Amount", 0) < 0)
                if avail is not None:
                    credit_parts += f'<span class="info-chip" style="border-left:3px solid #2E7D32"><b>{sn}:</b> {fmt_currency(avail, True)}</span>'
                elif sec_cr < 0:
                    credit_parts += f'<span class="info-chip" style="border-left:3px solid #2E7D32"><b>{sn}:</b> {fmt_currency(sec_cr, True)}</span>'
        st.markdown(f"""
        <div style="background:#FFFFFF; padding:.9rem 1.5rem; border-radius:10px; margin:.8rem 0;
                    border-left:5px solid #2E7D32; box-shadow:0 2px 10px rgba(0,0,0,.06);
                    display:flex; align-items:center; justify-content:space-between; flex-wrap:wrap;">
            <div>
                <span style="font-weight:700; color:#2E7D32; font-size:1.1rem;">Credit Available: {fmt_currency(abs(total_credits), True)}</span>
            </div>
            <div style="display:flex; gap:.4rem; flex-wrap:wrap;">{credit_parts}</div>
        </div>
        """, unsafe_allow_html=True)


def render_debt_decomposition(df: pd.DataFrame):
    """Render a debt decomposition card showing total debt broken down by invoice type (section)."""
    if df.empty:
        return
    # Compute charges per section
    charges_by_section = df[df["Amount"] > 0].groupby("Section")["Amount"].sum()
    if charges_by_section.empty:
        return
    total_debt = charges_by_section.sum()
    if total_debt <= 0:
        return

    # Build stacked bar segments and text
    segments_html = ""
    breakdown_chips = ""
    colors = SECTION_COLOURS
    for i, (sec_name, amount) in enumerate(charges_by_section.sort_values(ascending=False).items()):
        pct = (amount / total_debt) * 100 if total_debt > 0 else 0
        color = colors[i % len(colors)]
        segments_html += f'<div style="width:{pct:.1f}%;background:{color};height:100%;display:inline-block" title="{sec_name}: {fmt_currency(amount, True)} ({pct:.0f}%)"></div>'
        breakdown_chips += f'<span style="display:inline-block;margin:.2rem .4rem;font-size:.82rem;color:#1a1a2e"><span style="display:inline-block;width:10px;height:10px;background:{color};border-radius:2px;margin-right:4px"></span><b>{sec_name}:</b> {fmt_currency(amount, True)} ({pct:.0f}%)</span>'

    st.markdown(f"""
    <div style="background:#FFFFFF; padding:1rem 1.5rem; border-radius:10px; margin:.5rem 0 1rem 0;
                border-left:5px solid #10069F; box-shadow:0 2px 10px rgba(0,0,0,.06);">
        <div style="font-weight:700; color:#0C0033; font-size:1rem; margin-bottom:.5rem;">
            Total Debt: {fmt_currency(total_debt, True)}
        </div>
        <div style="background:#E8E8EE; border-radius:4px; height:18px; overflow:hidden; margin-bottom:.6rem; display:flex;">
            {segments_html}
        </div>
        <div style="display:flex; flex-wrap:wrap; gap:0;">{breakdown_chips}</div>
    </div>
    """, unsafe_allow_html=True)


def render_executive_overview(df: pd.DataFrame):
    """Render the 3 executive overview charts (donut, grouped bar, aging)."""
    st.markdown('<div class="section-hdr">Executive Overview</div>', unsafe_allow_html=True)

    overview_cols = st.columns([1, 1, 1])

    # Donut: Breakdown by Section
    if not df.empty:
        sec_abs = df.groupby("Section")["Amount"].apply(lambda x: x.abs().sum()).reset_index()
        sec_abs.columns = ["Section", "Absolute Amount"]
        with overview_cols[0]:
            st.plotly_chart(make_donut(sec_abs["Section"], sec_abs["Absolute Amount"],
                                       "Breakdown by Section"), use_container_width=True)

    # Grouped bar: Charges vs Credits per Section
    if not df.empty:
        charge_credit = df.groupby(["Section", "Entry Type"])["Amount"].sum().reset_index()
        charge_credit["Amount"] = charge_credit["Amount"].abs()
        with overview_cols[1]:
            fig_cc = px.bar(charge_credit, x="Section", y="Amount", color="Entry Type",
                            barmode="group", title="<b>Charges vs Credits by Section</b>",
                            color_discrete_map={"Charge": RR_NAVY, "Credit": RR_GREEN},
                            text_auto=".2s")
            fig_cc.update_layout(height=360, margin=dict(t=55, b=45, l=45, r=20),
                                 title=dict(font=dict(size=15, color="#1a1a2e")),
                                 legend=dict(font=dict(size=12, color="#333")),
                                 **CHART_LAYOUT)
            st.plotly_chart(fig_cc, use_container_width=True)

    # Aging Distribution
    if not df.empty and "Days Late" in df.columns:
        df_aging = df.copy()
        df_aging["Aging Bucket"] = df_aging["Days Late"].apply(aging_bucket)
        aging_counts = df_aging.groupby("Aging Bucket").agg(
            Count=("Amount", "count"),
            Total_Amount=("Amount", "sum")
        ).reindex([b for b in AGING_ORDER if b in df_aging["Aging Bucket"].values]).reset_index()

        with overview_cols[2]:
            fig_age = px.bar(aging_counts, x="Aging Bucket", y="Total_Amount", title="<b>Aging Analysis</b>",
                             text_auto=".2s", color="Aging Bucket",
                             color_discrete_map=AGING_COLORS)
            fig_age.update_layout(height=360, showlegend=False,
                                  margin=dict(t=55, b=45, l=45, r=20),
                                  title=dict(font=dict(size=15, color="#1a1a2e")),
                                  **CHART_LAYOUT)
            st.plotly_chart(fig_age, use_container_width=True)

        # ---- Aging Drill-Down: invoices per bucket ----
        with st.expander("View Invoices by Aging Bucket", expanded=False):
            active_buckets = [b for b in AGING_ORDER if b in df_aging["Aging Bucket"].values]
            if active_buckets:
                age_tabs = st.tabs(active_buckets)
                for age_tab, bucket in zip(age_tabs, active_buckets):
                    with age_tab:
                        bucket_df = df_aging[df_aging["Aging Bucket"] == bucket]
                        st.markdown(f"**{len(bucket_df)} invoices** | Total: **{fmt_currency(bucket_df['Amount'].sum(), True)}**")
                        age_display_cols = [c for c in ["Reference", "Text", "Amount", "Due Date",
                                                        "Days Late", "Section", "Status"]
                                            if c in bucket_df.columns and bucket_df[c].notna().any()]
                        age_disp = bucket_df[age_display_cols].copy()
                        if "Amount" in age_disp.columns:
                            age_disp["Amount"] = age_disp["Amount"].apply(lambda x: fmt_currency(x))
                        if "Due Date" in age_disp.columns:
                            age_disp["Due Date"] = age_disp["Due Date"].apply(
                                lambda x: x.strftime("%d/%m/%Y") if pd.notna(x) and hasattr(x, "strftime") else str(x) if x else "\u2014")
                        st.dataframe(age_disp, use_container_width=True, height=min(350, 35 * len(age_disp) + 60))


def render_bilateral_position(df: pd.DataFrame):
    """Render the bilateral position charts."""
    st.markdown('<div class="section-hdr">Bilateral Position</div>', unsafe_allow_html=True)

    bp_cols = st.columns([1, 1])

    if not df.empty:
        they_owe = df.loc[df["Amount"] > 0, "Amount"].sum()   # Customer owes RR
        we_owe   = df.loc[df["Amount"] < 0, "Amount"].sum()   # RR credits/owes customer

        with bp_cols[0]:
            bp_data = pd.DataFrame({
                "Direction": ["Customer \u2192 RR (Charges)", "RR \u2192 Customer (Credits)"],
                "Amount": [they_owe, abs(we_owe)]
            })
            fig_bp = px.bar(bp_data, x="Direction", y="Amount", title="<b>Bilateral Position</b>",
                            color="Direction",
                            color_discrete_map={
                                "Customer \u2192 RR (Charges)": RR_NAVY,
                                "RR \u2192 Customer (Credits)": RR_GREEN
                            }, text_auto=".3s")
            fig_bp.update_layout(height=360, showlegend=False,
                                 margin=dict(t=55, b=65, l=45, r=20),
                                 title=dict(font=dict(size=15, color="#1a1a2e")),
                                 **CHART_LAYOUT)
            st.plotly_chart(fig_bp, use_container_width=True)

        with bp_cols[1]:
            sec_net = df.groupby("Section")["Amount"].sum().reset_index()
            sec_net.columns = ["Section", "Net Amount"]
            sec_net["Color"] = sec_net["Net Amount"].apply(lambda x: "Owed to RR" if x > 0 else "Credit to Customer")
            fig_sn = px.bar(sec_net, y="Section", x="Net Amount", orientation="h",
                            title="<b>Net Balance by Section</b>", color="Color",
                            color_discrete_map={"Owed to RR": RR_NAVY, "Credit to Customer": RR_GREEN},
                            text_auto=".3s")
            fig_sn.update_layout(height=360, margin=dict(t=55, b=45, l=25, r=25),
                                 title=dict(font=dict(size=15, color="#1a1a2e")),
                                 legend=dict(font=dict(size=12, color="#333")),
                                 **CHART_LAYOUT)
            st.plotly_chart(fig_sn, use_container_width=True)


def render_llp_analysis(df: pd.DataFrame):
    """Render LLP CRC / ULW net value analysis. Matches CRC charges to ULW credits by ESN or part number."""
    if df.empty or "Text" not in df.columns:
        return

    # Look for ESN-like identifiers in Text and Assignment columns
    # Common patterns: ESN followed by digits, "ESN" keyword, part numbers like "72-xxxx"
    esn_pattern = re.compile(
        r'(?:ESN[:\s#-]*(\w{3,20}))'         # "ESN 12345" or "ESN: ABC123"
        r'|(?:(?:LLP|CRC|ULW)[:\s#-]*(\w{3,20}))'  # "LLP 12345"
        r'|(?:\b(\d{2,3}-\d{3,6})\b)',        # Part numbers like "72-1234"
        re.IGNORECASE
    )

    records = []
    for _, row in df.iterrows():
        text_fields = str(row.get("Text", "")) + " " + str(row.get("Assignment", ""))
        matches = esn_pattern.findall(text_fields)
        if matches:
            # Take first non-empty match group
            identifier = None
            for m in matches:
                for g in m:
                    if g:
                        identifier = g.upper()
                        break
                if identifier:
                    break
            if identifier:
                records.append({
                    "Identifier": identifier,
                    "Amount": row["Amount"],
                    "Section": row.get("Section", ""),
                    "Text": row.get("Text", ""),
                    "Reference": row.get("Reference", ""),
                    "Entry Type": row.get("Entry Type", ""),
                })

    if not records:
        return  # No LLP-like data found — silently skip

    llp_df = pd.DataFrame(records)

    # Group by identifier: separate CRC charges from ULW credits
    grouped = llp_df.groupby("Identifier").agg(
        CRC_Amount=("Amount", lambda x: x[x > 0].sum()),
        ULW_Amount=("Amount", lambda x: x[x < 0].sum()),
        Total_Items=("Amount", "count"),
    ).reset_index()
    grouped["Net Value"] = grouped["CRC_Amount"] + grouped["ULW_Amount"]

    # Only show if we have pairs (both charges and credits for at least one identifier)
    has_pairs = ((grouped["CRC_Amount"] > 0) & (grouped["ULW_Amount"] < 0)).any()
    if not has_pairs:
        return

    st.markdown('<div class="section-hdr">LLP CRC / ULW Analysis</div>', unsafe_allow_html=True)

    # Summary card
    total_crc = grouped["CRC_Amount"].sum()
    total_ulw = grouped["ULW_Amount"].sum()
    total_net = grouped["Net Value"].sum()
    sk = st.columns(3)
    sk[0].markdown(metric_card("CRC Charges", fmt_currency(total_crc, True)), unsafe_allow_html=True)
    sk[1].markdown(metric_card("ULW Credits", fmt_currency(total_ulw, True), "positive"), unsafe_allow_html=True)
    sk[2].markdown(metric_card("Net Value", fmt_currency(total_net, True),
                               "negative" if total_net > 0 else "positive"), unsafe_allow_html=True)

    # Detail table
    disp = grouped.copy()
    disp["CRC_Amount"] = disp["CRC_Amount"].apply(lambda x: fmt_currency(x))
    disp["ULW_Amount"] = disp["ULW_Amount"].apply(lambda x: fmt_currency(x))
    disp["Net Value"] = disp["Net Value"].apply(lambda x: fmt_currency(x))
    disp.columns = ["ESN / Part", "CRC Amount", "ULW Amount", "Items", "Net Value"]
    st.dataframe(disp, use_container_width=True, height=min(300, 35 * len(disp) + 60))


def render_section_tabs(df: pd.DataFrame, sections_dict: dict, show_credits: bool):
    """Render the section breakdown tabs with KPIs, charts, tables."""
    st.markdown('<div class="section-hdr">Section Breakdown</div>', unsafe_allow_html=True)

    if not sections_dict:
        return

    tab_names = list(sections_dict.keys())
    tabs = st.tabs(tab_names)

    for tab, sec_name in zip(tabs, tab_names):
        sec = sections_dict[sec_name]
        sec_rows = sec["rows"]
        sec_totals = sec["totals"]

        if not sec_rows:
            with tab:
                st.info(f"No line items found in **{sec_name}**.")
            continue

        sec_df = pd.DataFrame(sec_rows)

        # If global filter is active, filter section rows to match the filtered df
        if not df.empty and "Source File" in df.columns and "Source File" in sec_df.columns:
            # Only show rows that survived global filtering
            sec_section_name = sec_name.split(" \u2014 ")[-1] if " \u2014 " in sec_name else sec_name
            matching = df[df["Section"] == sec_section_name]
            if not matching.empty:
                sec_df = sec_df[sec_df.index.isin(
                    sec_df.merge(matching, on=["Amount", "Section"], how="inner", suffixes=("", "_r")).index
                )]

        with tab:
            # Section KPIs
            sec_total   = sec_totals.get("total", sec_df["Amount"].sum())
            sec_overdue = sec_totals.get("overdue", None)
            sec_credits_avail = sec_totals.get("available credit", None)
            sec_charges = sec_df.loc[sec_df["Amount"] > 0, "Amount"].sum() if not sec_df.empty else 0
            sec_credits = sec_df.loc[sec_df["Amount"] < 0, "Amount"].sum() if not sec_df.empty else 0
            sec_items   = len(sec_df)

            sk = st.columns(5)
            sk[0].markdown(metric_card("Section Total", fmt_currency(sec_total, True)), unsafe_allow_html=True)
            sk[1].markdown(metric_card("Charges", fmt_currency(sec_charges, True)), unsafe_allow_html=True)
            sk[2].markdown(metric_card("Credits", fmt_currency(sec_credits, True), "positive"), unsafe_allow_html=True)
            if sec_overdue is not None:
                sk[3].markdown(metric_card("Overdue", fmt_currency(sec_overdue, True), "negative"), unsafe_allow_html=True)
            else:
                sk[3].markdown(metric_card("Items", str(sec_items)), unsafe_allow_html=True)
            if sec_credits_avail is not None:
                sk[4].markdown(metric_card("Available Credit", fmt_currency(sec_credits_avail, True), "positive"), unsafe_allow_html=True)
            else:
                sk[4].markdown(metric_card("Net", fmt_currency(sec_charges + sec_credits, True)), unsafe_allow_html=True)

            st.markdown("")

            # Two-column: Status distribution + Top items
            chart_cols = st.columns([1, 1])

            with chart_cols[0]:
                if "Status" in sec_df.columns and not sec_df.empty:
                    status_df = sec_df.copy()
                    status_df["Status Clean"] = status_df["Status"].fillna("Unknown").apply(
                        lambda s: s[:40] + "..." if len(str(s)) > 40 else s
                    )
                    status_counts = status_df["Status Clean"].value_counts().reset_index()
                    status_counts.columns = ["Status", "Count"]
                    if len(status_counts) > 0:
                        pie_html = make_chartjs_pie(
                            labels=status_counts["Status"].tolist(),
                            values=status_counts["Count"].tolist(),
                            title="Status Distribution",
                            colors=SECTION_COLOURS[:len(status_counts)],
                            height=350,
                        )
                        components.html(pie_html, height=350)

            with chart_cols[1]:
                avail_cols = [c for c in ["Text", "Amount", "Reference"] if c in sec_df.columns]
                if "Amount" in avail_cols and not sec_df.empty:
                    top_items = sec_df.nlargest(8, "Amount")[avail_cols].copy()
                    if top_items.empty:
                        top_items = sec_df.head(8)[avail_cols].copy()
                    text_col = "Text" if "Text" in top_items.columns else (
                        "Assignment" if "Assignment" in sec_df.columns else None)
                    ref_col = "Reference" if "Reference" in top_items.columns else None
                    if text_col and text_col not in top_items.columns:
                        top_items[text_col] = sec_df.nlargest(8, "Amount")[text_col]
                    if not top_items.empty and text_col and text_col in top_items.columns:
                        top_items["Label"] = top_items.apply(
                            lambda r: (str(r.get(text_col, ""))[:30] + " (" + str(r.get(ref_col, "")) + ")")
                            if ref_col and r.get(ref_col) else str(r.get(text_col, ""))[:40], axis=1)
                        fig_top = px.bar(top_items, y="Label", x="Amount", orientation="h",
                                         title="<b>Top Items by Amount</b>", text_auto=".3s",
                                         color_discrete_sequence=[RR_NAVY])
                        fig_top.update_layout(height=350, margin=dict(t=55, b=25, l=25, r=25),
                                              title=dict(font=dict(size=15, color="#1a1a2e")),
                                              **CHART_LAYOUT)
                        st.plotly_chart(fig_top, use_container_width=True)

            # Detailed data table
            st.markdown("**Detailed Line Items**")

            display_cols = []
            preferred_order = ["Reference", "Document No", "Document Date", "Due Date", "Amount",
                               "Currency", "Text", "Type", "Assignment", "R-R Comments",
                               "Status", "Action Owner", "Days Late", "Customer Comments",
                               "Customer Name", "PO Reference", "LPI Cumulated", "Interest Method",
                               "Entry Type"]
            for c in preferred_order:
                if c in sec_df.columns and sec_df[c].notna().any():
                    display_cols.append(c)

            for c in sec_df.columns:
                if c not in display_cols and c not in ("Section", "Source File") and sec_df[c].notna().any():
                    display_cols.append(c)

            if not display_cols:
                st.info("No columns with data to display.")
                continue

            display_df = sec_df[display_cols].copy()

            if not show_credits:
                if "Entry Type" in display_df.columns:
                    display_df = display_df[display_df["Entry Type"] == "Charge"]

            if "Amount" in display_df.columns:
                display_df["Amount"] = display_df["Amount"].apply(lambda x: fmt_currency(x) if x else "\u2014")
            if "Document Date" in display_df.columns:
                display_df["Document Date"] = display_df["Document Date"].apply(
                    lambda x: x.strftime("%d/%m/%Y") if pd.notna(x) and hasattr(x, "strftime") else str(x) if x else "\u2014")
            if "Due Date" in display_df.columns:
                display_df["Due Date"] = display_df["Due Date"].apply(
                    lambda x: x.strftime("%d/%m/%Y") if pd.notna(x) and hasattr(x, "strftime") else str(x) if x else "\u2014")

            st.dataframe(display_df, use_container_width=True, height=min(400, 35 * len(display_df) + 60))


def render_invoice_register(df: pd.DataFrame):
    """Render the complete invoice register with local filters."""
    st.markdown('<div class="section-hdr">Complete Invoice Register</div>', unsafe_allow_html=True)

    if df.empty:
        st.info("No data to display in the invoice register.")
        return

    # Apply overdue-only filter from clickable KPI
    if st.session_state.get("filter_overdue_only") and "Days Late" in df.columns:
        df = df[df["Days Late"].fillna(0).apply(lambda x: int(x) if x else 0) > 0].copy()
        st.info("Showing **overdue invoices only** (filtered from Total Overdue KPI). Click 'Clear overdue filter' above to show all.")

    filter_cols = st.columns(5)
    with filter_cols[0]:
        sec_options = sorted(df["Section"].unique())
        sec_filter = st.multiselect("Invoice Type (Section)", sec_options, default=sec_options, key="reg_section")
    with filter_cols[1]:
        type_options = sorted(df["Entry Type"].unique())
        type_filter = st.multiselect("Type", type_options, default=type_options, key="reg_type")
    with filter_cols[2]:
        reg_overdue_filter = st.multiselect("Overdue Status", ["Overdue", "Current"],
                                            default=["Overdue", "Current"], key="reg_overdue")
    with filter_cols[3]:
        if "Status" in df.columns:
            statuses = sorted(df["Status"].dropna().unique())
            status_filter = st.multiselect("Status", statuses, default=[], key="reg_status")
        else:
            status_filter = []
    with filter_cols[4]:
        max_amt = float(df["Amount"].abs().max()) if not df.empty else 1e6
        amount_range = st.slider("Amount Range (absolute)", 0.0, max_amt, (0.0, max_amt), key="reg_amount")

    filtered = df.copy()
    if sec_filter:
        filtered = filtered[filtered["Section"].isin(sec_filter)]
    if type_filter:
        filtered = filtered[filtered["Entry Type"].isin(type_filter)]
    if status_filter:
        filtered = filtered[filtered["Status"].isin(status_filter)]
    # Overdue / Current filter
    if reg_overdue_filter and len(reg_overdue_filter) < 2 and "Days Late" in filtered.columns:
        if "Overdue" in reg_overdue_filter and "Current" not in reg_overdue_filter:
            filtered = filtered[filtered["Days Late"].fillna(0).apply(lambda x: int(x) if x else 0) > 0]
        elif "Current" in reg_overdue_filter and "Overdue" not in reg_overdue_filter:
            filtered = filtered[filtered["Days Late"].fillna(0).apply(lambda x: int(x) if x else 0) <= 0]
    filtered = filtered[
        (filtered["Amount"].abs() >= amount_range[0]) & (filtered["Amount"].abs() <= amount_range[1])
    ]

    show_cols = [c for c in ["Section", "Source File", "Reference", "Document No", "Document Date",
                             "Due Date", "Amount", "Currency", "Text", "Type", "Assignment",
                             "R-R Comments", "Status", "Action Owner", "Days Late",
                             "Customer Comments", "Customer Name", "PO Reference",
                             "LPI Cumulated", "Entry Type", "Interest Method"]
                 if c in filtered.columns and filtered[c].notna().any()]

    # Hide Source File column when only one file is uploaded
    if "Source File" in show_cols and filtered["Source File"].nunique() <= 1:
        show_cols.remove("Source File")

    disp = filtered[show_cols].copy()

    if "Amount" in disp.columns:
        disp["Amount"] = disp["Amount"].apply(lambda x: fmt_currency(x))
    for dcol in ["Document Date", "Due Date"]:
        if dcol in disp.columns:
            disp[dcol] = disp[dcol].apply(
                lambda x: x.strftime("%d/%m/%Y") if pd.notna(x) and hasattr(x, "strftime") else str(x) if x else "\u2014")

    st.dataframe(disp, use_container_width=True, height=min(600, 35 * len(disp) + 60))

    sum_cols = st.columns(3)
    sum_cols[0].metric("Filtered Items", len(filtered))
    sum_cols[1].metric("Filtered Total", fmt_currency(filtered["Amount"].sum()))
    if "Days Late" in filtered.columns:
        overdue_mask = filtered["Days Late"].fillna(0).apply(lambda x: int(x) if x else 0) > 0
        overdue_total = filtered.loc[overdue_mask, "Amount"].sum()
    else:
        overdue_total = 0
    sum_cols[2].metric("Filtered Overdue", fmt_currency(overdue_total))


def render_presentation_mode(df, grand, avg_late, meta_list, sections_dict, show_credits):
    """Slide-by-slide presentation mode for Teams call screen-sharing."""
    # Build list of slides
    slide_names = ["Customer Overview", "Key Metrics", "Executive Charts", "Bilateral Position"]
    sec_names = list(sections_dict.keys()) if sections_dict else []
    for sn in sec_names:
        slide_names.append(f"Section: {sn}")
    slide_names.append("Invoice Summary")

    # Session state for current slide
    if "pres_slide" not in st.session_state:
        st.session_state["pres_slide"] = 0
    current = st.session_state["pres_slide"]
    total = len(slide_names)

    # Navigation bar
    nav_cols = st.columns([1, 6, 1])
    with nav_cols[0]:
        if st.button("Previous", disabled=(current == 0), key="pres_prev", use_container_width=True):
            st.session_state["pres_slide"] = max(0, current - 1)
            st.rerun()
    with nav_cols[1]:
        st.markdown(f"""<div style="text-align:center; color:#555; font-size:.95rem; padding:.4rem 0;">
            <b>{slide_names[current]}</b> &nbsp;&mdash;&nbsp; Slide {current + 1} of {total}
        </div>""", unsafe_allow_html=True)
    with nav_cols[2]:
        if st.button("Next", disabled=(current >= total - 1), key="pres_next", use_container_width=True):
            st.session_state["pres_slide"] = min(total - 1, current + 1)
            st.rerun()

    st.markdown("---")

    # Render the current slide
    slide = slide_names[current]

    if slide == "Customer Overview":
        render_customer_info(meta_list)
        render_debt_decomposition(df)

    elif slide == "Key Metrics":
        render_kpi_cards(df, grand, avg_late, sections_dict)

    elif slide == "Executive Charts":
        render_executive_overview(df)

    elif slide == "Bilateral Position":
        render_bilateral_position(df)

    elif slide == "Invoice Summary":
        # Show high-level totals and overdue summary
        st.markdown('<div class="section-hdr">Invoice Summary</div>', unsafe_allow_html=True)
        total_items = len(df) if not df.empty else 0
        total_overdue_items = len(df[df["Days Late"].fillna(0).apply(lambda x: int(x) if x else 0) > 0]) if not df.empty and "Days Late" in df.columns else 0
        sc = st.columns(3)
        sc[0].metric("Total Invoices", total_items)
        sc[1].metric("Overdue Invoices", total_overdue_items)
        sc[2].metric("Net Balance", fmt_currency(grand.get("net_balance", 0), True))

    elif slide.startswith("Section: "):
        sec_name = slide[len("Section: "):]
        if sec_name in sections_dict:
            sec = sections_dict[sec_name]
            sec_rows = sec.get("rows", [])
            if sec_rows:
                sec_df = pd.DataFrame(sec_rows)
                sec_totals = sec.get("totals", {})
                st.markdown(f'<div class="section-hdr">{sec_name}</div>', unsafe_allow_html=True)
                sec_total = sec_totals.get("total", sec_df["Amount"].sum())
                sec_charges = sec_df.loc[sec_df["Amount"] > 0, "Amount"].sum() if not sec_df.empty else 0
                sec_credits = sec_df.loc[sec_df["Amount"] < 0, "Amount"].sum() if not sec_df.empty else 0
                sk = st.columns(4)
                sk[0].markdown(metric_card("Section Total", fmt_currency(sec_total, True)), unsafe_allow_html=True)
                sk[1].markdown(metric_card("Charges", fmt_currency(sec_charges, True)), unsafe_allow_html=True)
                sk[2].markdown(metric_card("Credits", fmt_currency(sec_credits, True), "positive"), unsafe_allow_html=True)
                sk[3].markdown(metric_card("Items", str(len(sec_df))), unsafe_allow_html=True)
            else:
                st.info(f"No line items in **{sec_name}**.")


def render_comparison_mode(parsed_files: dict, selected_sources: list):
    """Render comparison view — side-by-side metrics for 2+ data sources."""
    st.markdown('<div class="section-hdr">Comparison Mode</div>', unsafe_allow_html=True)

    if len(selected_sources) < 2:
        st.warning("Comparison Mode requires at least 2 uploaded files. Please upload more files or switch to Standard view.")
        return

    # Let user pick which 2 sources to compare
    compare_sources = st.multiselect(
        "Select sources to compare", selected_sources,
        default=selected_sources[:2], max_selections=4, key="compare_select"
    )

    if len(compare_sources) < 2:
        st.info("Select at least 2 sources to compare.")
        return

    # Side-by-side columns
    cols = st.columns(len(compare_sources))

    for col, fname in zip(cols, compare_sources):
        data = parsed_files[fname]
        meta = data["metadata"]
        grand = data["grand_totals"]
        df = data["all_items"]

        with col:
            # Customer info
            cust_name = meta.get("customer_name", "Unknown Customer")
            cust_id = meta.get("customer_id", "\u2014")
            report_dt = meta.get("report_date")
            st.markdown(f"""
            <div style="background:#FFFFFF; padding:.8rem 1rem; border-radius:10px; margin-bottom:.8rem;
                        box-shadow: 0 1px 6px rgba(0,0,0,.05); border: 1px solid #E0E1EC; text-align:center;">
                <div style="font-weight:800; color:#0C0033; font-size:1rem;">{cust_name}</div>
                <div style="font-size:.8rem; color:#666;">ID: {cust_id}
                {"&nbsp;&bull;&nbsp;" + report_dt.strftime('%d %b %Y') if report_dt else ""}</div>
                <div style="font-size:.7rem; color:#999; margin-top:.2rem;">{fname}</div>
            </div>
            """, unsafe_allow_html=True)

            # KPI cards — 3 per column
            total_charges = grand.get("total_charges", 0)
            total_credits = grand.get("total_credits", 0)
            net_balance   = grand.get("net_balance", 0)

            st.markdown(metric_card("Total Charges", fmt_currency(total_charges, short=True)), unsafe_allow_html=True)
            st.markdown(metric_card("Total Credits", fmt_currency(total_credits, short=True), "positive" if total_credits < 0 else ""), unsafe_allow_html=True)
            st.markdown(metric_card("Net Balance", fmt_currency(net_balance, short=True), "negative" if net_balance > 0 else "positive"), unsafe_allow_html=True)

            # Summary bar chart per source
            if not df.empty:
                sec_totals = df.groupby("Section")["Amount"].sum().reset_index()
                sec_totals.columns = ["Section", "Amount"]
                fig = px.bar(sec_totals, x="Section", y="Amount", title="<b>By Section</b>",
                             text_auto=".2s", color_discrete_sequence=[RR_NAVY])
                fig.update_layout(height=300, margin=dict(t=45, b=35, l=40, r=15),
                                  title=dict(font=dict(size=13, color="#1a1a2e")),
                                  **CHART_LAYOUT)
                st.plotly_chart(fig, use_container_width=True)

    # Comparison table below
    st.markdown('<div class="section-hdr">Key Metrics Comparison</div>', unsafe_allow_html=True)

    comparison_rows = []
    for fname in compare_sources:
        data = parsed_files[fname]
        g = data["grand_totals"]
        m = data["metadata"]
        comparison_rows.append({
            "Source": fname,
            "Customer": m.get("customer_name", "Unknown"),
            "Total Charges": fmt_currency(g.get("total_charges", 0)),
            "Total Credits": fmt_currency(g.get("total_credits", 0)),
            "Net Balance": fmt_currency(g.get("net_balance", 0)),
            "Total Overdue": fmt_currency(g.get("total_overdue", 0)),
            "Items": g.get("item_count", 0),
            "Avg Days Late": m.get("avg_days_late", "\u2014"),
            "LPI Rate": f"{m['lpi_rate']*100:.2f}%" if m.get("lpi_rate") else "\u2014",
        })

    comp_df = pd.DataFrame(comparison_rows)
    st.dataframe(comp_df, use_container_width=True, hide_index=True)

    # Section-level comparison
    st.markdown("**Section Totals Comparison**")
    all_section_names = set()
    for fname in compare_sources:
        for sn in parsed_files[fname]["sections"].keys():
            all_section_names.add(sn)

    sec_comp_rows = []
    for sn in sorted(all_section_names):
        row = {"Section": sn}
        for fname in compare_sources:
            sections = parsed_files[fname]["sections"]
            if sn in sections:
                stot = sections[sn]["totals"].get("total", None)
                if stot is None:
                    rows = sections[sn]["rows"]
                    stot = sum(r.get("Amount", 0) for r in rows) if rows else 0
                row[fname] = fmt_currency(stot)
            else:
                row[fname] = "\u2014"
        sec_comp_rows.append(row)

    if sec_comp_rows:
        sec_comp_df = pd.DataFrame(sec_comp_rows)
        st.dataframe(sec_comp_df, use_container_width=True, hide_index=True)


# ─────────────────────────────────────────────────────────────
# 4.  SIDEBAR — FILE UPLOAD & SETTINGS
# ─────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("""
    <div style="text-align:center; padding: 1rem 0;">
        <div style="color:#FFF; font-size:1.2rem; font-weight:700; letter-spacing:2px;
                    border:2px solid #C0C0C0; display:inline-block; padding:.4rem 1rem; border-radius:4px;">
            ROLLS-ROYCE
        </div>
        <div style="color:#C0C0C0; font-size:.7rem; margin-top:.3rem; letter-spacing:1.5px;">CIVIL AEROSPACE</div>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("##### Upload Statement of Account")
    uploaded_files = st.file_uploader(
        "Upload .xlsx files", type=["xlsx", "xls"],
        accept_multiple_files=True, label_visibility="collapsed"
    )

    st.markdown("---")
    st.markdown("##### Dashboard Settings")
    dashboard_view = st.selectbox("Dashboard View", ["Standard", "Executive Summary", "Presentation Mode", "Comparison Mode"], index=0)
    currency_symbol = st.selectbox("Currency Display", ["USD", "GBP", "EUR"], index=0)
    show_credits_in_tables = st.checkbox("Show credits in line tables", value=True)

    st.markdown("---")
    st.markdown("##### Share with Customer")
    sidebar_export_clicked = st.button("📄 Export PDF Report", type="primary", use_container_width=True, key="sidebar_pdf_btn")


# ─────────────────────────────────────────────────────────────
# 5.  MAIN DASHBOARD
# ─────────────────────────────────────────────────────────────

# Header
st.markdown("""
<div class="rr-header">
    <div>
        <div style="color:#FFFFFF; margin:0; font-size:1.65rem; font-weight:700; letter-spacing:.5px;">Statement of Account Dashboard</div>
        <div style="color:#D0D0E0; font-size:.92rem; margin-top:.2rem; font-weight:400;">Rolls-Royce Civil Aerospace &mdash; Finance & Receivables</div>
    </div>
    <div style="color:#FFFFFF; font-size:1.15rem; font-weight:700; letter-spacing:3px; text-transform:uppercase; border:2px solid rgba(255,255,255,.6); padding:.4rem 1rem; border-radius:4px;">ROLLS-ROYCE</div>
</div>
""", unsafe_allow_html=True)

if not uploaded_files:
    st.info("Upload one or more Statement of Account Excel files (.xlsx) using the sidebar to get started.")
    st.markdown("""
    <div style="background:#FFF; padding:2rem; border-radius:12px; text-align:center; margin:2rem 0; box-shadow: 0 2px 8px rgba(0,0,0,.06);">
        <div style="font-size:3rem; margin-bottom:1rem;">&#9992;</div>
        <div style="color:#10069F; margin-bottom:.5rem; font-size:1.3rem; font-weight:700;">Welcome to the RR SOA Dashboard</div>
        <p style="color:#666; max-width:500px; margin:0 auto;">
            Upload any Rolls-Royce Statement of Account workbook. The dashboard automatically
            detects sections like <b>TotalCare</b>, <b>CRC Payments</b>, <b>Spare Parts</b>,
            <b>Late Payment Interest</b>, and more &mdash; regardless of layout variations.
            <br><br>
            <b>New:</b> Upload multiple files to compare across customers or periods.
        </p>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# ─────────────────────────────────────────────────────────────
# 5a. PARSE ALL UPLOADED FILES
# ─────────────────────────────────────────────────────────────

parsed_files = {}
parse_errors = []

with st.spinner("Parsing workbook(s)..."):
    for uf in uploaded_files:
        try:
            file_bytes = uf.getvalue()
            parsed_files[uf.name] = _cached_parse(file_bytes, uf.name)
        except Exception as e:
            parse_errors.append((uf.name, str(e)))

if parse_errors:
    for fname, err in parse_errors:
        st.error(f"Failed to parse **{fname}**: {err}")

if not parsed_files:
    st.warning("No files could be parsed successfully.")
    st.stop()

# ─────────────────────────────────────────────────────────────
# 5b. DATA SOURCE SELECTOR (multi-file only)
# ─────────────────────────────────────────────────────────────

all_source_names = list(parsed_files.keys())
selected_sources = all_source_names  # default: all

with st.sidebar:
    if len(all_source_names) > 1:
        st.markdown("---")
        st.markdown("##### Data Sources")
        selected_sources = st.multiselect(
            "Select files to include", all_source_names,
            default=all_source_names, key="data_sources"
        )
        if not selected_sources:
            selected_sources = all_source_names

# ─────────────────────────────────────────────────────────────
# 5c. MERGE DATA FROM SELECTED SOURCES
# ─────────────────────────────────────────────────────────────

df_all, merged_sections, merged_grand, all_metadata = _merge_parsed_files(parsed_files, selected_sources)

# Compute avg_days_late across all metadata
_avg_late_values = [m.get("avg_days_late") for m in all_metadata if m.get("avg_days_late")]
avg_late_combined = int(np.mean(_avg_late_values)) if _avg_late_values else None

# ─────────────────────────────────────────────────────────────
# 5d. GLOBAL FILTERS (sidebar)
# ─────────────────────────────────────────────────────────────

# Defaults for filter variables (in case expander is not opened / df is empty)
g_selected_sections = sorted(df_all["Section"].unique().tolist()) if not df_all.empty else []
g_selected_types = sorted(df_all["Entry Type"].unique().tolist()) if not df_all.empty else []
g_selected_statuses = []
g_selected_currencies = []
g_selected_customers = []
g_selected_overdue_status = ["Overdue", "Current"]

with st.sidebar:
    with st.expander("Global Filters", expanded=True):
        # Customer Name filter
        customer_names = list(set(
            m.get("customer_name", "Unknown") for m in all_metadata
            if m.get("customer_name") and m.get("customer_name") != "Unknown"
        ))
        if customer_names:
            g_selected_customers = st.multiselect(
                "Customer", sorted(customer_names), default=sorted(customer_names), key="gf_customer"
            )

        # Section / Invoice Type filter
        sections_available = sorted(df_all["Section"].unique().tolist()) if not df_all.empty else []
        if sections_available:
            g_selected_sections = st.multiselect(
                "Invoice Type (Section)", sections_available, default=sections_available, key="gf_section"
            )

        # Type filter
        type_options = sorted(df_all["Entry Type"].unique().tolist()) if not df_all.empty else []
        if type_options:
            g_selected_types = st.multiselect(
                "Type", type_options, default=type_options, key="gf_type"
            )

        # Overdue Status filter
        g_selected_overdue_status = st.multiselect(
            "Overdue Status", ["Overdue", "Current"],
            default=["Overdue", "Current"], key="gf_overdue"
        )

        # Status filter
        if not df_all.empty and "Status" in df_all.columns:
            status_options = sorted(df_all["Status"].dropna().unique().tolist())
            if status_options:
                g_selected_statuses = st.multiselect(
                    "Status", status_options, default=[], key="gf_status"
                )

        # Currency filter
        if not df_all.empty and "Currency" in df_all.columns:
            currency_options = sorted(df_all["Currency"].dropna().unique().tolist())
            if currency_options:
                g_selected_currencies = st.multiselect(
                    "Currency", currency_options, default=currency_options, key="gf_currency"
                )

# Apply global filters
df_filtered = _apply_global_filters(
    df_all, g_selected_sections, g_selected_types,
    g_selected_statuses, g_selected_currencies,
    g_selected_customers, parsed_files, selected_sources,
    g_selected_overdue_status
)

# Recompute grand totals from filtered data
if not df_filtered.empty:
    filtered_grand = {
        "total_charges": df_filtered.loc[df_filtered["Amount"] > 0, "Amount"].sum(),
        "total_credits": df_filtered.loc[df_filtered["Amount"] < 0, "Amount"].sum(),
        "net_balance": df_filtered["Amount"].sum(),
        "item_count": len(df_filtered),
        "total_overdue": merged_grand.get("total_overdue", df_filtered["Amount"].sum()),
    }
else:
    filtered_grand = {"total_charges": 0, "total_credits": 0, "net_balance": 0, "item_count": 0, "total_overdue": 0}


# ─────────────────────────────────────────────────────────────
# 6.  RENDER DASHBOARD BASED ON VIEW MODE
# ─────────────────────────────────────────────────────────────

if dashboard_view == "Standard":
    render_customer_info(all_metadata)
    render_kpi_cards(df_filtered, filtered_grand, avg_late_combined, merged_sections)
    render_debt_decomposition(df_filtered)
    render_executive_overview(df_filtered)
    render_bilateral_position(df_filtered)
    render_llp_analysis(df_filtered)
    render_section_tabs(df_filtered, merged_sections, show_credits_in_tables)
    render_invoice_register(df_filtered)

elif dashboard_view == "Executive Summary":
    render_customer_info(all_metadata)
    render_kpi_cards(df_filtered, filtered_grand, avg_late_combined, merged_sections)
    render_debt_decomposition(df_filtered)
    render_executive_overview(df_filtered)
    # Executive Summary: charts only, no tables, no section tabs, no invoice register

elif dashboard_view == "Presentation Mode":
    render_presentation_mode(df_filtered, filtered_grand, avg_late_combined, all_metadata,
                             merged_sections, show_credits_in_tables)

elif dashboard_view == "Comparison Mode":
    if len(selected_sources) < 2:
        st.warning("Comparison Mode requires at least 2 uploaded files. Showing Standard view instead.")
        render_customer_info(all_metadata)
        render_kpi_cards(df_filtered, filtered_grand, avg_late_combined, merged_sections)
        render_debt_decomposition(df_filtered)
        render_executive_overview(df_filtered)
        render_bilateral_position(df_filtered)
        render_llp_analysis(df_filtered)
        render_section_tabs(df_filtered, merged_sections, show_credits_in_tables)
        render_invoice_register(df_filtered)
    else:
        render_comparison_mode(parsed_files, selected_sources)


# ─────────────────────────────────────────────────────────────
# 9.  PDF EXPORT
# ─────────────────────────────────────────────────────────────

# --- PDF Generation Helper (shared by sidebar and bottom buttons) ---
def _generate_pdf():
    _sections_summary = {}
    for _sec_name in merged_sections:
        _sec = merged_sections[_sec_name]
        _sec_df = df_filtered[df_filtered["Section"] == _sec_name] if not df_filtered.empty else pd.DataFrame()
        _sections_summary[_sec_name] = {
            "total": _sec.get("totals", {}).get("total", _sec_df["Amount"].sum() if not _sec_df.empty else 0),
            "charges": _sec_df.loc[_sec_df["Amount"] > 0, "Amount"].sum() if not _sec_df.empty else 0,
            "credits": _sec_df.loc[_sec_df["Amount"] < 0, "Amount"].sum() if not _sec_df.empty else 0,
            "overdue": _sec.get("totals", {}).get("overdue", 0),
            "items": len(_sec_df),
        }
    _pdf_meta = all_metadata[0] if all_metadata else {}
    try:
        _pdf_bytes = generate_pdf_report(
            metadata=_pdf_meta, grand_totals=filtered_grand,
            filtered_df=df_filtered, sections_summary=_sections_summary,
            source_files=selected_sources if len(selected_sources) > 1 else None,
            currency_symbol=currency_symbol,
        )
        st.session_state["pdf_bytes"] = _pdf_bytes
        st.session_state["pdf_ready"] = True
    except Exception as e:
        st.error(f"PDF generation failed: {e}")

# Trigger from sidebar button
if sidebar_export_clicked:
    _generate_pdf()

# Show sidebar download button if ready
with st.sidebar:
    if st.session_state.get("pdf_ready"):
        _cust = all_metadata[0].get("customer_name", "Report") if all_metadata else "Report"
        st.download_button(
            label="⬇ Download PDF",
            data=st.session_state["pdf_bytes"],
            file_name=f"SOA_Report_{_cust.replace(' ', '_')}.pdf",
            mime="application/pdf",
            use_container_width=True,
            key="sidebar_pdf_download",
        )

st.markdown('<div class="section-hdr">Export Report</div>', unsafe_allow_html=True)
_pdf_col1, _pdf_col2 = st.columns([1, 3])
with _pdf_col1:
    if st.button("Generate PDF Report", type="primary", use_container_width=True):
        _generate_pdf()
with _pdf_col2:
    if st.session_state.get("pdf_ready"):
        _cust = all_metadata[0].get("customer_name", "Report") if all_metadata else "Report"
        st.download_button(
            label="Download PDF", data=st.session_state["pdf_bytes"],
            file_name=f"SOA_Report_{_cust.replace(' ', '_')}.pdf",
            mime="application/pdf", use_container_width=False,
        )
        st.success("PDF report generated successfully.")


# ─────────────────────────────────────────────────────────────
# 10.  FOOTER
# ─────────────────────────────────────────────────────────────

st.markdown("---")
st.markdown("""
<div style="text-align:center; color:#999; font-size:.75rem; padding:1rem 0;">
    <b>ROLLS-ROYCE</b> CIVIL AEROSPACE &mdash; Statement of Account Dashboard<br>
    Data sourced from uploaded workbook(s) &bull; For internal use only
</div>
""", unsafe_allow_html=True)
