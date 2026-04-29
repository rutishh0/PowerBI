"""
Build a self-contained HTML visualization benchmark for ME_Employee_Whereabouts.xlsx.

Independent of V6/parser.py. Reads xlsx via openpyxl directly.
Emits a canonical JSON payload + ApexCharts visualisation as a single HTML file.

Expected sheets: 'Mar 2026', 'Apr 2026 ', 'May 2026' (or similar "<Mon> YYYY").
"""
from __future__ import annotations

import calendar
import json
import re
from datetime import datetime, date
from pathlib import Path

import openpyxl

SRC = Path(r"C:\Users\Rutishkrishna\Desktop\RR\RR Powerbi\V6\New info\ME_Employee_Whereabouts.xlsx")
OUT_HTML = Path(r"C:\Users\Rutishkrishna\Desktop\RR\RR Powerbi\V6\TESTEXCEL\ME_Employee_Whereabouts.html")
OUT_JSON = Path(r"C:\Users\Rutishkrishna\Desktop\RR\RR Powerbi\V6\TESTEXCEL\_payloads\ME_Employee_Whereabouts.json")

MONTH_YEAR_RE = re.compile(
    r"^\s*(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*\s*[-/ ]?\s*(\d{4})\s*$",
    re.IGNORECASE,
)
MONTH_TO_NUM = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dec": 12,
}

LEGEND_DEFAULTS = {
    "O": "Office",
    "H": "Home / WFH",
    "WFH": "Work From Home",
    "L": "Leave",
    "B": "Business Trip",
    "S": "Sick",
    "PL": "Personal Leave",
    "CB": "Cross Border",
    "EB": "Easter Break",
    "HOL": "Holiday",
}


# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------
def _clean(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    s = str(v).strip()
    if not s or s in {".", "\xa0"}:
        return None
    return s


def _is_blank(v):
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def parse_month_sheet_name(name: str):
    m = MONTH_YEAR_RE.match(name.strip())
    if not m:
        return None
    token = m.group(1).lower()
    if token.startswith("sept"):
        key = "sept"
    else:
        key = token[:3]
    month_num = MONTH_TO_NUM.get(key)
    if month_num is None:
        return None
    year = int(m.group(2))
    _, days_in_month = calendar.monthrange(year, month_num)
    return {
        "sheet": name,
        "year": year,
        "month": month_num,
        "days_in_month": days_in_month,
        "start_date": f"{year:04d}-{month_num:02d}-01",
        "end_date": f"{year:04d}-{month_num:02d}-{days_in_month:02d}",
    }


def find_header_row(ws, max_scan=20):
    for r in range(1, min(max_scan, ws.max_row) + 1):
        for c in range(1, 10):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            if isinstance(v, str) and "employee number" in v.lower():
                return r
    return -1


def legend_label(code: str) -> str:
    if not code:
        return code
    upper = code.strip().upper()
    if upper in LEGEND_DEFAULTS:
        return LEGEND_DEFAULTS[upper]
    low = code.lower()
    if "eid" in low or "holiday" in low:
        return "Public Holiday"
    if "leave" in low:
        return "Leave"
    if "sick" in low:
        return "Sick"
    if "busines" in low or "trip" in low:
        return "Business Trip"
    if "home" in low or "wfh" in low:
        return "Home / WFH"
    if "office" in low:
        return "Office"
    if "easter" in low:
        return "Easter Break"
    if "cross" in low and "border" in low:
        return "Cross Border"
    return code


# ------------------------------------------------------------------
# Main extraction
# ------------------------------------------------------------------
def build_payload(src: Path) -> dict:
    wb = openpyxl.load_workbook(str(src), data_only=True)

    months_meta = []
    sheets_parsed = []
    sheets_ignored = []
    whereabouts = {}
    employees_by_num = {}
    status_totals_by_month = {}
    daily_office_count_by_month = {}
    by_country = {}
    by_sector = {}
    all_codes = {}

    for sheet_name in wb.sheetnames:
        mmeta = parse_month_sheet_name(sheet_name)
        if mmeta is None:
            sheets_ignored.append(sheet_name)
            continue
        ws = wb[sheet_name]
        hdr = find_header_row(ws)
        if hdr < 0:
            sheets_ignored.append(sheet_name)
            continue

        # Map columns from header row
        no_col = emp_num_col = name_col = sector_col = country_col = notes_col = None
        day_cols = []
        # Scan up to a sane max (38 in sample) — we cap at 50
        scan_end = min(ws.max_column, 60)
        for c in range(1, scan_end + 1):
            v = ws.cell(row=hdr, column=c).value
            if v is None:
                continue
            if isinstance(v, str):
                t = v.strip().lower()
                if t in ("no.", "no", "#"):
                    if no_col is None:
                        no_col = c
                elif "employee number" in t or t in ("emp no", "emp number", "employee no"):
                    emp_num_col = c
                elif "employee name" in t or t == "name":
                    name_col = c
                elif "business sector" in t or t == "sector":
                    sector_col = c
                elif t == "country" or "country" in t:
                    if country_col is None:
                        country_col = c
                elif "note" in t:
                    notes_col = c
            elif isinstance(v, (int, float)):
                n = int(v)
                if 1 <= n <= 31:
                    day_cols.append((c, n))

        if not day_cols and country_col:
            end_c = notes_col or (country_col + mmeta["days_in_month"] + 1)
            for c in range(country_col + 1, min(end_c, country_col + mmeta["days_in_month"] + 1)):
                day_cols.append((c, c - country_col))

        day_cols = [(c, d) for (c, d) in day_cols if d <= mmeta["days_in_month"]]
        day_cols.sort(key=lambda t: t[1])

        if emp_num_col is None and name_col is None:
            sheets_ignored.append(sheet_name)
            continue
        if not day_cols:
            sheets_ignored.append(sheet_name)
            continue

        year = mmeta["year"]
        month = mmeta["month"]

        sheet_records = []
        sheet_status_totals = {}
        sheet_daily_office = {
            f"{year:04d}-{month:02d}-{d:02d}": 0
            for d in range(1, mmeta["days_in_month"] + 1)
        }

        for r in range(hdr + 1, ws.max_row + 1):
            emp_num_raw = ws.cell(row=r, column=emp_num_col).value if emp_num_col else None
            name_raw = ws.cell(row=r, column=name_col).value if name_col else None

            if _is_blank(emp_num_raw) and _is_blank(name_raw):
                continue

            if emp_num_raw is None:
                nm = _clean(name_raw)
                emp_num = f"UNKNOWN-{nm[:30]}" if nm else None
            else:
                emp_num = str(emp_num_raw).strip()

            if not emp_num:
                continue

            name = _clean(name_raw)
            sector = _clean(ws.cell(row=r, column=sector_col).value) if sector_col else None
            country = _clean(ws.cell(row=r, column=country_col).value) if country_col else None

            if emp_num not in employees_by_num:
                employees_by_num[emp_num] = {
                    "employee_number": emp_num,
                    "name": name,
                    "business_sector": sector,
                    "country": country,
                }
                if country:
                    by_country[country] = by_country.get(country, 0) + 1
                if sector:
                    by_sector[sector] = by_sector.get(sector, 0) + 1

            daily_status = {}
            status_counts = {}

            for (col_idx, day_num) in day_cols:
                iso = f"{year:04d}-{month:02d}-{day_num:02d}"
                cell = ws.cell(row=r, column=col_idx).value
                if _is_blank(cell):
                    daily_status[iso] = None
                    status_counts["_blank"] = status_counts.get("_blank", 0) + 1
                    sheet_status_totals["_blank"] = sheet_status_totals.get("_blank", 0) + 1
                else:
                    code = _clean(cell)
                    if not code:
                        daily_status[iso] = None
                        status_counts["_blank"] = status_counts.get("_blank", 0) + 1
                        sheet_status_totals["_blank"] = sheet_status_totals.get("_blank", 0) + 1
                    else:
                        daily_status[iso] = code
                        status_counts[code] = status_counts.get(code, 0) + 1
                        sheet_status_totals[code] = sheet_status_totals.get(code, 0) + 1
                        all_codes[code] = all_codes.get(code, 0) + 1
                        upper = code.strip().upper()
                        if upper == "O" or upper == "OFFICE" or "office" in code.lower():
                            sheet_daily_office[iso] = sheet_daily_office.get(iso, 0) + 1

            sheet_records.append({
                "employee_number": emp_num,
                "name": name,
                "country": country,
                "daily_status": daily_status,
                "status_counts": status_counts,
            })

        whereabouts[sheet_name] = sheet_records
        status_totals_by_month[sheet_name] = sheet_status_totals
        daily_office_count_by_month[sheet_name] = sheet_daily_office
        months_meta.append(mmeta)
        sheets_parsed.append(sheet_name)

    # Legend
    legend = {}
    for code in sorted(all_codes.keys(), key=lambda s: (len(s), s)):
        legend[code] = legend_label(code)
    for k, v in LEGEND_DEFAULTS.items():
        legend.setdefault(k, v)

    employees_list = sorted(employees_by_num.values(), key=lambda e: e.get("employee_number") or "")
    unique_countries = sorted({e.get("country") for e in employees_list if e.get("country")})
    unique_sectors = sorted({e.get("business_sector") for e in employees_list if e.get("business_sector")})

    payload = {
        "file_type": "EMPLOYEE_WHEREABOUTS",
        "metadata": {
            "source_file": src.name,
            "sheets_parsed": sheets_parsed,
            "sheets_ignored": sheets_ignored,
            "months": months_meta,
            "total_employees": len(employees_list),
            "unique_countries": unique_countries,
            "unique_sectors": unique_sectors,
        },
        "employees": employees_list,
        "whereabouts": whereabouts,
        "legend": legend,
        "aggregates": {
            "by_country": by_country,
            "by_sector": by_sector,
            "daily_office_count_by_month": daily_office_count_by_month,
            "status_totals_by_month": status_totals_by_month,
        },
        "errors": [],
    }
    return payload


# ------------------------------------------------------------------
# HTML template
# ------------------------------------------------------------------
HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<title>Middle East Employee Whereabouts - {plan_year_title}</title>
<script src="https://cdn.jsdelivr.net/npm/apexcharts@3.49.0/dist/apexcharts.min.js"></script>
<style>
  :root {{
    --rr-blue: #0b2b54;
    --rr-navy: #061a36;
    --rr-accent: #00a9e0;
    --rr-gold:  #f6b221;
    --office:   #16a34a;
    --home:     #2563eb;
    --leave:    #f59e0b;
    --holiday:  #ef4444;
    --trip:     #8b5cf6;
    --blank:    #e5e7eb;
    --ink:      #0f172a;
    --muted:    #64748b;
    --card-bg:  #ffffff;
    --page-bg:  #f5f7fb;
  }}
  * {{ box-sizing: border-box; }}
  body {{
    margin: 0; padding: 0;
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
    background: var(--page-bg);
    color: var(--ink);
  }}
  header.hero {{
    background: linear-gradient(135deg, var(--rr-navy) 0%, var(--rr-blue) 60%, var(--rr-accent) 140%);
    color: #fff;
    padding: 28px 32px 22px;
    box-shadow: 0 2px 10px rgba(0,0,0,0.08);
  }}
  header.hero h1 {{
    margin: 0 0 6px;
    font-size: 26px;
    letter-spacing: 0.2px;
  }}
  header.hero .subtitle {{
    color: rgba(255,255,255,0.82);
    font-size: 13px;
  }}
  header.hero .badge {{
    display: inline-block;
    padding: 3px 9px;
    background: rgba(255,255,255,0.15);
    border-radius: 999px;
    font-size: 11px;
    margin-right: 6px;
    letter-spacing: 0.3px;
  }}
  main {{ padding: 22px 32px 60px; max-width: 1480px; margin: 0 auto; }}
  section.kpis {{
    display: grid;
    grid-template-columns: repeat(6, minmax(0, 1fr));
    gap: 14px;
    margin-bottom: 22px;
  }}
  @media (max-width: 1100px) {{ section.kpis {{ grid-template-columns: repeat(3, 1fr); }} }}
  @media (max-width: 640px)  {{ section.kpis {{ grid-template-columns: repeat(2, 1fr); }} }}
  .kpi {{
    background: var(--card-bg);
    border-radius: 10px;
    padding: 14px 16px;
    border: 1px solid #e5e7eb;
    box-shadow: 0 1px 2px rgba(0,0,0,0.03);
  }}
  .kpi .label {{ font-size: 11px; text-transform: uppercase; color: var(--muted); letter-spacing: 0.4px; }}
  .kpi .value {{ font-size: 24px; font-weight: 600; color: var(--rr-blue); margin-top: 4px; }}
  .kpi .unit  {{ font-size: 12px; color: var(--muted); }}

  .tabs {{
    display: flex;
    gap: 4px;
    margin-bottom: 16px;
    flex-wrap: wrap;
  }}
  .tab-btn {{
    background: #fff;
    border: 1px solid #e5e7eb;
    color: var(--ink);
    padding: 9px 16px;
    border-radius: 8px;
    cursor: pointer;
    font-size: 13px;
    font-weight: 500;
  }}
  .tab-btn.active {{
    background: var(--rr-blue);
    color: #fff;
    border-color: var(--rr-blue);
  }}

  .card {{
    background: var(--card-bg);
    border-radius: 10px;
    border: 1px solid #e5e7eb;
    padding: 16px 18px;
    margin-bottom: 16px;
    box-shadow: 0 1px 2px rgba(0,0,0,0.03);
  }}
  .card h3 {{ margin: 0 0 10px; font-size: 14px; color: var(--rr-navy); text-transform: uppercase; letter-spacing: 0.3px; }}

  .grid-two {{
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 16px;
  }}
  .grid-three {{
    display: grid;
    grid-template-columns: 2fr 1fr 1fr;
    gap: 16px;
  }}
  @media (max-width: 900px) {{
    .grid-two, .grid-three {{ grid-template-columns: 1fr; }}
  }}

  table.heatmap {{
    border-collapse: separate;
    border-spacing: 1px;
    width: 100%;
    table-layout: fixed;
    font-size: 10px;
  }}
  table.heatmap th, table.heatmap td {{
    padding: 2px 2px;
    text-align: center;
    border-radius: 2px;
    overflow: hidden;
    white-space: nowrap;
  }}
  table.heatmap th {{ font-weight: 600; background: #eef2f7; color: var(--ink); }}
  table.heatmap th.name {{ text-align: left; padding-left: 6px; }}
  table.heatmap td.name {{ text-align: left; padding-left: 6px; background: #f8fafc; font-weight: 500; }}
  table.heatmap td.cell-office   {{ background: var(--office);  color: #fff; }}
  table.heatmap td.cell-home     {{ background: var(--home);    color: #fff; }}
  table.heatmap td.cell-leave    {{ background: var(--leave);   color: #fff; }}
  table.heatmap td.cell-holiday  {{ background: var(--holiday); color: #fff; }}
  table.heatmap td.cell-trip     {{ background: var(--trip);    color: #fff; }}
  table.heatmap td.cell-other    {{ background: #94a3b8; color: #fff; }}
  table.heatmap td.cell-blank    {{ background: var(--blank); }}

  .emp-table-wrap {{ overflow-x: auto; }}
  table.employees {{
    width: 100%; border-collapse: collapse; font-size: 13px;
  }}
  table.employees th, table.employees td {{
    padding: 8px 10px; border-bottom: 1px solid #e5e7eb; text-align: left;
  }}
  table.employees th {{
    background: #eef2f7; font-weight: 600; text-transform: uppercase;
    font-size: 11px; letter-spacing: 0.3px;
    cursor: pointer; user-select: none;
  }}
  table.employees tr:hover td {{ background: #f8fafc; }}

  .filter-bar {{
    display: flex; gap: 10px; margin-bottom: 10px; flex-wrap: wrap;
    align-items: center;
  }}
  .filter-bar input[type=text], .filter-bar select {{
    padding: 7px 9px; font-size: 13px; border: 1px solid #e5e7eb; border-radius: 6px;
    background: #fff; color: var(--ink);
  }}
  .filter-bar label {{ font-size: 12px; color: var(--muted); }}
  .chips {{ display:flex; flex-wrap:wrap; gap:6px; }}
  .chip {{
    padding: 4px 10px; border-radius: 999px; font-size: 11px;
    background: #eef2f7; color: var(--rr-navy); border: 1px solid #dbe2ea;
    cursor: pointer; user-select: none;
  }}
  .chip.active {{ background: var(--rr-blue); color: #fff; border-color: var(--rr-blue); }}

  .legend-grid {{
    display: grid; grid-template-columns: repeat(auto-fit, minmax(170px, 1fr)); gap: 8px;
  }}
  .legend-item {{
    display: flex; align-items: center; gap: 8px;
    padding: 6px 10px; background: #f8fafc; border: 1px solid #e5e7eb; border-radius: 6px;
    font-size: 12px;
  }}
  .legend-swatch {{
    width: 14px; height: 14px; border-radius: 3px; flex-shrink: 0;
  }}
  .legend-item code {{ font-weight: 600; color: var(--rr-navy); font-family: "SF Mono", Menlo, Consolas, monospace; }}

  footer {{ text-align: center; color: var(--muted); font-size: 11px; padding: 14px 0 10px; }}
</style>
</head>
<body>

<header class="hero">
  <h1>Middle East Employee Whereabouts &mdash; {plan_year_title}</h1>
  <div class="subtitle">
    <span class="badge">{source_file}</span>
    <span class="badge">{n_sheets} months</span>
    <span class="badge">{n_employees} employees</span>
    <span class="badge">Civil Aerospace &middot; Rolls-Royce</span>
  </div>
</header>

<main>

  <section class="kpis" id="kpis"></section>

  <div class="card">
    <h3>Month</h3>
    <div class="tabs" id="month-tabs"></div>
  </div>

  <div class="grid-three">
    <div class="card">
      <h3>Daily Office Attendance</h3>
      <div id="chart-daily"></div>
    </div>
    <div class="card">
      <h3>Status Distribution</h3>
      <div id="chart-status"></div>
    </div>
    <div class="card">
      <h3>Country Breakdown</h3>
      <div id="chart-country"></div>
    </div>
  </div>

  <div class="card">
    <h3>Heatmap &mdash; Top 30 Employees &times; Day</h3>
    <div id="heatmap-wrap" style="overflow-x:auto;"></div>
  </div>

  <div class="card">
    <h3>Status Legend</h3>
    <div class="legend-grid" id="legend-grid"></div>
  </div>

  <div class="card">
    <h3>Employee Directory</h3>
    <div class="filter-bar">
      <input type="text" id="emp-search" placeholder="Search name or number..." />
      <label for="sector-select">Sector:</label>
      <select id="sector-select"><option value="">All</option></select>
      <label>Country:</label>
      <div class="chips" id="country-chips"></div>
    </div>
    <div class="emp-table-wrap">
      <table class="employees" id="employees-table">
        <thead>
          <tr>
            <th data-sort="employee_number">Employee #</th>
            <th data-sort="name">Name</th>
            <th data-sort="business_sector">Sector</th>
            <th data-sort="country">Country</th>
          </tr>
        </thead>
        <tbody></tbody>
      </table>
    </div>
  </div>

  <footer>
    Generated {generated_at} &middot; source: <code>{source_file}</code> &middot;
    TESTEXCEL benchmark independent of parser.py
  </footer>

</main>

<script>
const DATA = {data_json};

// =============== Helpers ===============
function statusClass(code) {{
  if (code == null) return 'cell-blank';
  const up = String(code).trim().toUpperCase();
  const lo = String(code).toLowerCase();
  if (up === 'O' || up === 'OFFICE' || lo.includes('office')) return 'cell-office';
  if (up === 'H' || up === 'WFH' || lo.includes('home') || lo.includes('wfh')) return 'cell-home';
  if (up === 'L' || up === 'PL' || lo.includes('leave')) return 'cell-leave';
  if (up === 'B' || lo.includes('trip') || lo.includes('business t')) return 'cell-trip';
  if (up === 'HOL' || lo.includes('holiday') || lo.includes('easter') || lo.includes('eid')) return 'cell-holiday';
  return 'cell-other';
}}
function statusColor(cls) {{
  switch (cls) {{
    case 'cell-office':  return getComputedStyle(document.documentElement).getPropertyValue('--office').trim()  || '#16a34a';
    case 'cell-home':    return getComputedStyle(document.documentElement).getPropertyValue('--home').trim()    || '#2563eb';
    case 'cell-leave':   return getComputedStyle(document.documentElement).getPropertyValue('--leave').trim()   || '#f59e0b';
    case 'cell-holiday': return getComputedStyle(document.documentElement).getPropertyValue('--holiday').trim() || '#ef4444';
    case 'cell-trip':    return getComputedStyle(document.documentElement).getPropertyValue('--trip').trim()    || '#8b5cf6';
    case 'cell-other':   return '#94a3b8';
    default:             return '#e5e7eb';
  }}
}}
function shortCode(code) {{
  if (code == null) return '';
  const s = String(code).trim();
  if (s.length <= 3) return s;
  const up = s.toUpperCase();
  if (up.startsWith('EID')) return 'EID';
  if (up.startsWith('EASTER')) return 'EB';
  if (up.startsWith('HOL')) return 'HOL';
  return s.slice(0, 3);
}}

// =============== KPI Row ===============
function renderKpis() {{
  const months = DATA.metadata.months || [];
  const totalDays = months.reduce((acc, m) => acc + (m.days_in_month || 0), 0);
  // avg office days / employee
  let officeTotal = 0;
  const wb = DATA.whereabouts || {{}};
  const empOfficeDays = {{}};
  Object.keys(wb).forEach(sheet => {{
    (wb[sheet] || []).forEach(rec => {{
      const c = (rec.status_counts || {{}});
      const officeHits = Object.keys(c).filter(k => {{
        const up = k.trim().toUpperCase();
        return up === 'O' || up === 'OFFICE' || k.toLowerCase().includes('office');
      }}).reduce((a, k) => a + c[k], 0);
      officeTotal += officeHits;
      empOfficeDays[rec.employee_number] = (empOfficeDays[rec.employee_number] || 0) + officeHits;
    }});
  }});
  const emps = DATA.metadata.total_employees || 1;
  const avgOffice = (officeTotal / emps).toFixed(1);

  const tiles = [
    {{ label: 'Employees',       value: DATA.metadata.total_employees || 0 }},
    {{ label: 'Countries',       value: (DATA.metadata.unique_countries || []).length }},
    {{ label: 'Sectors',         value: (DATA.metadata.unique_sectors || []).length }},
    {{ label: 'Months Tracked',  value: months.length }},
    {{ label: 'Total Days',      value: totalDays }},
    {{ label: 'Avg Office Days', value: avgOffice, unit: '/ employee' }},
  ];
  const kpiEl = document.getElementById('kpis');
  kpiEl.innerHTML = tiles.map(t => `
    <div class="kpi">
      <div class="label">${{t.label}}</div>
      <div class="value">${{t.value}}</div>
      ${{t.unit ? `<div class="unit">${{t.unit}}</div>` : ''}}
    </div>
  `).join('');
}}

// =============== Tabs ===============
let currentSheet = null;
let charts = {{}};

function renderTabs() {{
  const sheets = DATA.metadata.sheets_parsed || [];
  const wrap = document.getElementById('month-tabs');
  wrap.innerHTML = sheets.map((s, i) =>
    `<button class="tab-btn ${{i===0?'active':''}}" data-sheet="${{s.replace(/"/g,'&quot;')}}">${{s.trim()}}</button>`
  ).join('');
  wrap.querySelectorAll('.tab-btn').forEach(btn => {{
    btn.addEventListener('click', () => {{
      wrap.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
      btn.classList.add('active');
      currentSheet = btn.getAttribute('data-sheet');
      renderMonth(currentSheet);
    }});
  }});
  if (sheets.length) {{
    currentSheet = sheets[0];
    renderMonth(currentSheet);
  }}
}}

// =============== Charts ===============
function destroyChart(key) {{
  if (charts[key]) {{ try {{ charts[key].destroy(); }} catch(e) {{}} delete charts[key]; }}
}}

function renderDailyChart(sheet) {{
  const daily = (DATA.aggregates.daily_office_count_by_month || {{}})[sheet] || {{}};
  const keys = Object.keys(daily).sort();
  const series = [{{ name: 'Office attendees', data: keys.map(k => daily[k]) }}];
  const opts = {{
    chart: {{ type: 'line', height: 260, toolbar: {{ show: false }}, animations: {{ enabled: false }} }},
    colors: ['#16a34a'],
    stroke: {{ curve: 'smooth', width: 3 }},
    markers: {{ size: 4 }},
    xaxis: {{ categories: keys.map(k => k.slice(8)), title: {{ text: 'Day of month' }} }},
    yaxis: {{ title: {{ text: 'Employees in office' }}, min: 0, forceNiceScale: true }},
    dataLabels: {{ enabled: false }},
    grid: {{ borderColor: '#e5e7eb' }},
    tooltip: {{ x: {{ formatter: (_, opts) => keys[opts.dataPointIndex] }} }},
    fill: {{ type: 'gradient', gradient: {{ opacityFrom: 0.35, opacityTo: 0.05 }} }},
  }};
  destroyChart('daily');
  charts.daily = new ApexCharts(document.querySelector('#chart-daily'), opts);
  charts.daily.render();
}}

function renderStatusChart(sheet) {{
  const totals = (DATA.aggregates.status_totals_by_month || {{}})[sheet] || {{}};
  // Group into broad buckets
  const buckets = {{ Office: 0, 'Home/WFH': 0, Leave: 0, 'Business Trip': 0, Holiday: 0, Other: 0, Blank: 0 }};
  Object.keys(totals).forEach(k => {{
    const v = totals[k];
    if (k === '_blank') {{ buckets.Blank += v; return; }}
    const up = k.trim().toUpperCase();
    const lo = k.toLowerCase();
    if (up === 'O' || lo.includes('office')) buckets.Office += v;
    else if (up === 'H' || up === 'WFH' || lo.includes('home') || lo.includes('wfh')) buckets['Home/WFH'] += v;
    else if (up === 'L' || up === 'PL' || lo.includes('leave')) buckets.Leave += v;
    else if (up === 'B' || lo.includes('trip') || lo.includes('business t')) buckets['Business Trip'] += v;
    else if (up === 'HOL' || lo.includes('holiday') || lo.includes('easter') || lo.includes('eid')) buckets.Holiday += v;
    else buckets.Other += v;
  }});
  const labels = [], series = [];
  Object.entries(buckets).forEach(([k, v]) => {{ if (v > 0) {{ labels.push(k); series.push(v); }} }});
  const colors = {{ Office: '#16a34a', 'Home/WFH': '#2563eb', Leave: '#f59e0b', 'Business Trip': '#8b5cf6', Holiday: '#ef4444', Other: '#94a3b8', Blank: '#e5e7eb' }};
  const opts = {{
    chart: {{ type: 'donut', height: 260, animations: {{ enabled: false }} }},
    labels: labels,
    colors: labels.map(l => colors[l] || '#64748b'),
    legend: {{ position: 'bottom', fontSize: '11px' }},
    dataLabels: {{ enabled: true, formatter: v => v.toFixed(0) + '%' }},
    stroke: {{ width: 0 }},
    series: series,
  }};
  destroyChart('status');
  charts.status = new ApexCharts(document.querySelector('#chart-status'), opts);
  charts.status.render();
}}

function renderCountryChart(sheet) {{
  // Employees-per-country for employees actually appearing in this sheet
  const records = (DATA.whereabouts || {{}})[sheet] || [];
  const ctry = {{}};
  records.forEach(r => {{
    const c = r.country || '(Unknown)';
    ctry[c] = (ctry[c] || 0) + 1;
  }});
  const pairs = Object.entries(ctry).sort((a, b) => b[1] - a[1]);
  const opts = {{
    chart: {{ type: 'bar', height: 260, toolbar: {{ show: false }}, animations: {{ enabled: false }} }},
    colors: ['#0b2b54'],
    plotOptions: {{ bar: {{ horizontal: true, barHeight: '70%', borderRadius: 3 }} }},
    dataLabels: {{ enabled: true, style: {{ colors: ['#fff'] }}, offsetX: -6 }},
    xaxis: {{ categories: pairs.map(p => p[0]) }},
    grid: {{ borderColor: '#e5e7eb' }},
    series: [{{ name: 'Employees', data: pairs.map(p => p[1]) }}],
  }};
  destroyChart('country');
  charts.country = new ApexCharts(document.querySelector('#chart-country'), opts);
  charts.country.render();
}}

// =============== Heatmap ===============
function renderHeatmap(sheet) {{
  const records = ((DATA.whereabouts || {{}})[sheet] || []).slice(0, 30);
  const monthMeta = (DATA.metadata.months || []).find(m => m.sheet === sheet);
  if (!monthMeta) return;
  const days = monthMeta.days_in_month;

  let html = '<table class="heatmap">';
  html += '<thead><tr><th class="name" style="width:220px">Employee</th>';
  for (let d = 1; d <= days; d++) html += `<th>${{d}}</th>`;
  html += '</tr></thead><tbody>';
  records.forEach(r => {{
    html += `<tr><td class="name" title="${{r.employee_number}} · ${{r.country || ''}}">${{(r.name || r.employee_number)}}</td>`;
    for (let d = 1; d <= days; d++) {{
      const iso = `${{monthMeta.year}}-${{String(monthMeta.month).padStart(2,'0')}}-${{String(d).padStart(2,'0')}}`;
      const v = r.daily_status[iso];
      const cls = statusClass(v);
      const label = v ? shortCode(v) : '';
      const tip = v ? String(v).replace(/"/g, '&quot;') : 'Blank';
      html += `<td class="${{cls}}" title="${{iso}}: ${{tip}}">${{label}}</td>`;
    }}
    html += '</tr>';
  }});
  html += '</tbody></table>';
  document.getElementById('heatmap-wrap').innerHTML = html;
}}

function renderMonth(sheet) {{
  renderDailyChart(sheet);
  renderStatusChart(sheet);
  renderCountryChart(sheet);
  renderHeatmap(sheet);
}}

// =============== Legend ===============
function renderLegend() {{
  const leg = DATA.legend || {{}};
  const html = Object.entries(leg).map(([code, label]) => {{
    const cls = statusClass(code);
    return `<div class="legend-item">
      <span class="legend-swatch" style="background:${{statusColor(cls)}}"></span>
      <code>${{code}}</code><span>&nbsp;&mdash;&nbsp;${{label}}</span>
    </div>`;
  }}).join('');
  document.getElementById('legend-grid').innerHTML = html;
}}

// =============== Employee table ===============
let sortField = 'employee_number', sortAsc = true;
let activeCountries = new Set();

function renderEmployees() {{
  const employees = DATA.employees || [];
  const q = (document.getElementById('emp-search').value || '').toLowerCase();
  const sector = document.getElementById('sector-select').value;

  let rows = employees.filter(e => {{
    if (sector && (e.business_sector || '') !== sector) return false;
    if (activeCountries.size && !activeCountries.has(e.country)) return false;
    if (!q) return true;
    return (e.name || '').toLowerCase().includes(q) ||
           (e.employee_number || '').toLowerCase().includes(q);
  }});
  rows.sort((a, b) => {{
    const av = (a[sortField] || '').toString().toLowerCase();
    const bv = (b[sortField] || '').toString().toLowerCase();
    if (av < bv) return sortAsc ? -1 : 1;
    if (av > bv) return sortAsc ? 1 : -1;
    return 0;
  }});

  const tbody = document.querySelector('#employees-table tbody');
  tbody.innerHTML = rows.map(e => `
    <tr>
      <td><code>${{e.employee_number || ''}}</code></td>
      <td>${{e.name || ''}}</td>
      <td>${{e.business_sector || ''}}</td>
      <td>${{e.country || ''}}</td>
    </tr>
  `).join('');
}}

function initEmployeeTable() {{
  // Populate sector dropdown
  const sectors = DATA.metadata.unique_sectors || [];
  const sel = document.getElementById('sector-select');
  sectors.forEach(s => {{
    const o = document.createElement('option'); o.value = s; o.textContent = s;
    sel.appendChild(o);
  }});
  sel.addEventListener('change', renderEmployees);
  document.getElementById('emp-search').addEventListener('input', renderEmployees);

  // Country chips
  const ctryWrap = document.getElementById('country-chips');
  const countries = DATA.metadata.unique_countries || [];
  ctryWrap.innerHTML = countries.map(c =>
    `<span class="chip" data-country="${{c.replace(/"/g,'&quot;')}}">${{c}}</span>`
  ).join('');
  ctryWrap.querySelectorAll('.chip').forEach(chip => {{
    chip.addEventListener('click', () => {{
      const c = chip.getAttribute('data-country');
      if (activeCountries.has(c)) {{ activeCountries.delete(c); chip.classList.remove('active'); }}
      else {{ activeCountries.add(c); chip.classList.add('active'); }}
      renderEmployees();
    }});
  }});

  // Sort headers
  document.querySelectorAll('#employees-table th').forEach(th => {{
    th.addEventListener('click', () => {{
      const f = th.getAttribute('data-sort');
      if (sortField === f) sortAsc = !sortAsc; else {{ sortField = f; sortAsc = true; }}
      renderEmployees();
    }});
  }});

  renderEmployees();
}}

// =============== Bootstrap ===============
document.addEventListener('DOMContentLoaded', () => {{
  renderKpis();
  renderTabs();
  renderLegend();
  initEmployeeTable();
}});
</script>

</body>
</html>
"""


def build_html(payload: dict) -> str:
    months = payload["metadata"].get("months", [])
    plan_year_title = (
        f"{months[0]['year']}" if months else "Plan"
    )
    # If multiple years span, show range
    if len({m["year"] for m in months}) > 1:
        ys = sorted({m["year"] for m in months})
        plan_year_title = f"{ys[0]}-{ys[-1]}"

    html = HTML_TEMPLATE.format(
        plan_year_title=plan_year_title,
        source_file=payload["metadata"]["source_file"],
        n_sheets=len(payload["metadata"]["sheets_parsed"]),
        n_employees=payload["metadata"]["total_employees"],
        generated_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),
        data_json=json.dumps(payload, default=str, ensure_ascii=False),
    )
    return html


def main():
    if not SRC.exists():
        raise SystemExit(f"Source not found: {SRC}")

    payload = build_payload(SRC)

    # Write JSON sidecar
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, default=str, ensure_ascii=False)

    # Write HTML
    html = build_html(payload)
    OUT_HTML.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_HTML, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"  JSON size: {OUT_JSON.stat().st_size / 1024:.1f} KB  -> {OUT_JSON}")
    print(f"  HTML size: {OUT_HTML.stat().st_size / 1024:.1f} KB  -> {OUT_HTML}")
    print(f"  Employees: {payload['metadata']['total_employees']}")
    print(f"  Months:    {len(payload['metadata']['months'])}")


if __name__ == "__main__":
    main()
