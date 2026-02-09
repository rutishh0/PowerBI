"""
Rolls-Royce Civil Aerospace — Statement of Account Dashboard
=============================================================
A professional, adaptive dashboard that ingests any RR-style SOA
Excel workbook and visualises CRC payments, TotalCare, Spare Parts,
Late Payment Interest, credits, balances, ageing, and more.

NOTHING is hard-coded to a particular customer or column position.
The parser detects sections, headers, and amounts dynamically.
"""

import io
import re
import math
from datetime import datetime, timedelta
from collections import OrderedDict

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook

# ─────────────────────────────────────────────────────────────
# 0.  PAGE CONFIG & BRANDING
# ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Rolls-Royce Civil Aerospace · SOA Dashboard",
    page_icon="✈",
    layout="wide",
    initial_sidebar_state="expanded",
)

# -- Rolls-Royce Civil Aerospace colour palette --
RR_NAVY   = "#10069F"
RR_DARK   = "#0C0033"
RR_SILVER = "#C0C0C0"
RR_LIGHT  = "#E8E8EE"
RR_WHITE  = "#FFFFFF"
RR_GOLD   = "#B8860B"
RR_RED    = "#D32F2F"
RR_GREEN  = "#2E7D32"
RR_BLUE2  = "#1565C0"
RR_AMBER  = "#F9A825"

SECTION_COLOURS = [RR_NAVY, "#1565C0", "#5E35B1", "#00838F", "#C62828", "#EF6C00", "#2E7D32", "#6A1B9A"]

# Inject CSS for RR look-and-feel
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

/* ---------- global ---------- */
html, body, [class*="css"] { font-family: 'Inter', sans-serif !important; }
.stApp { background: #F0F1F5; }

/* Force dark readable text everywhere in main area */
.stApp p, .stApp span, .stApp label, .stApp div { color: #1a1a2e; }
.stMarkdown p { color: #1a1a2e !important; font-size: 0.95rem; }
.stMarkdown strong { color: #0C0033 !important; }

/* ---------- header bar ---------- */
.rr-header {
    background: linear-gradient(135deg, #10069F 0%, #0C0033 100%);
    padding: 1.8rem 2.2rem;
    border-radius: 14px;
    margin-bottom: 1.4rem;
    display: flex; align-items: center; justify-content: space-between;
    box-shadow: 0 6px 24px rgba(16,6,159,.30);
}
.rr-header h1 {
    color: #FFFFFF !important; margin: 0; font-size: 1.65rem; font-weight: 700; letter-spacing: .5px;
}
.rr-header .subtitle {
    color: #D0D0E0 !important; font-size: .92rem; margin-top: .2rem; font-weight: 400;
}
.rr-logo-text {
    color: #FFFFFF !important; font-size: 1.15rem; font-weight: 700; letter-spacing: 3px; text-transform: uppercase;
    border: 2px solid rgba(255,255,255,.6); padding: .4rem 1rem; border-radius: 4px;
}

/* ---------- metric cards ---------- */
.metric-card {
    background: #FFFFFF;
    border-radius: 12px;
    padding: 1.2rem 1rem;
    text-align: center;
    border-left: 5px solid #10069F;
    box-shadow: 0 2px 12px rgba(0,0,0,.08);
    transition: transform .15s, box-shadow .15s;
}
.metric-card:hover { transform: translateY(-3px); box-shadow: 0 6px 20px rgba(0,0,0,.12); }
.metric-card .label {
    color: #444 !important;
    font-size: .78rem !important;
    font-weight: 700 !important;
    text-transform: uppercase;
    letter-spacing: .9px;
    margin-bottom: .35rem;
}
.metric-card .value {
    color: #0C0033 !important;
    font-size: 1.45rem !important;
    font-weight: 800 !important;
}
.metric-card .value.negative { color: #C62828 !important; }
.metric-card .value.positive { color: #1B5E20 !important; }

/* ---------- section header ---------- */
.section-hdr {
    background: linear-gradient(90deg, #10069F, #1565C0);
    color: #FFFFFF !important;
    padding: .7rem 1.4rem;
    border-radius: 8px;
    font-weight: 700;
    font-size: 1.05rem;
    margin: 1.8rem 0 1rem 0;
    letter-spacing: .5px;
    box-shadow: 0 2px 8px rgba(16,6,159,.18);
}

/* ---------- tables ---------- */
div[data-testid="stDataFrame"] {
    border-radius: 10px;
    overflow: hidden;
    box-shadow: 0 2px 10px rgba(0,0,0,.08);
    font-size: .88rem !important;
}
div[data-testid="stDataFrame"] th {
    background: #10069F !important;
    color: #FFF !important;
    font-weight: 600 !important;
    font-size: .82rem !important;
}
div[data-testid="stDataFrame"] td {
    color: #1a1a2e !important;
    font-size: .84rem !important;
}

/* ---------- sidebar ---------- */
section[data-testid="stSidebar"] { background: #0C0033 !important; }
section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3,
section[data-testid="stSidebar"] h4,
section[data-testid="stSidebar"] h5 { color: #FFFFFF !important; font-weight: 700 !important; }
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] span,
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] div { color: #E0E0F0 !important; }
section[data-testid="stSidebar"] .stSelectbox label,
section[data-testid="stSidebar"] .stMultiSelect label,
section[data-testid="stSidebar"] .stCheckbox label { color: #FFFFFF !important; font-weight: 500 !important; }
section[data-testid="stSidebar"] hr { border-color: rgba(255,255,255,.15) !important; }
section[data-testid="stSidebar"] .stFileUploader label { color: #FFFFFF !important; }
section[data-testid="stSidebar"] .stFileUploader div { color: #D0D0E0 !important; }

/* ---------- tabs ---------- */
.stTabs [data-baseweb="tab-list"] { gap: 6px; }
.stTabs [data-baseweb="tab"] {
    background: #FFFFFF; border-radius: 8px 8px 0 0; padding: .65rem 1.3rem;
    font-weight: 600; color: #333 !important; border: 1px solid #CCC; border-bottom: none;
    font-size: .9rem;
}
.stTabs [aria-selected="true"] {
    background: #10069F !important; color: #FFFFFF !important; border-color: #10069F !important;
}
.stTabs [aria-selected="true"] p,
.stTabs [aria-selected="true"] span,
.stTabs [aria-selected="true"] div {
    color: #FFFFFF !important;
}

/* ---------- selectbox / multiselect in main area ---------- */
.stSelectbox label, .stMultiSelect label, .stSlider label, .stCheckbox label {
    color: #1a1a2e !important;
    font-weight: 600 !important;
    font-size: .88rem !important;
}

/* ---------- metrics (st.metric) ---------- */
div[data-testid="stMetric"] label { color: #555 !important; font-weight: 600 !important; font-size: .82rem !important; }
div[data-testid="stMetric"] div[data-testid="stMetricValue"] { color: #0C0033 !important; font-weight: 700 !important; }

/* ---------- info chip ---------- */
.info-chip {
    display: inline-block;
    background: #E0E1EC;
    padding: .3rem .75rem;
    border-radius: 6px;
    font-size: .82rem;
    color: #1a1a2e !important;
    font-weight: 500;
    margin-right: .5rem;
    margin-bottom: .35rem;
    border: 1px solid #D0D1DC;
}
.info-chip b { color: #0C0033 !important; }

/* ---------- plotly chart text ---------- */
.js-plotly-plot .plotly text { fill: #333 !important; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────
# 1.  SMART EXCEL PARSER
# ─────────────────────────────────────────────────────────────

# Keywords that signal a section header row
SECTION_KEYWORDS = [
    "charges", "credits", "credit", "totalcare", "familycare", "missioncare",
    "spare parts", "late payment", "interest", "customer respon",
    "customer responsibility", "usable", "offset",
]

SUMMARY_KEYWORDS = ["total", "overdue", "available credit", "total overdue", "net balance"]

HEADER_KEYWORDS = [
    "company", "account", "reference", "document", "date", "amount", "curr",
    "text", "assignment", "arrangement", "comments", "status", "action",
    "days", "late", "lpi", "invoice", "type", "interest", "net due",
]


def _is_section_header(row_values: list, col_count: int) -> bool:
    """Return True if this row looks like a section heading."""
    non_empty = [(i, v) for i, v in enumerate(row_values) if v is not None]
    if not non_empty:
        return False
    # Typically section headers only populate col A (and sometimes 1-2 more)
    if len(non_empty) > 3:
        return False
    text = str(non_empty[0][1]).strip().lower()
    # Must not be purely numeric
    try:
        float(text.replace(",", ""))
        return False
    except ValueError:
        pass
    # Exclude summary-row patterns (e.g. "Available Credit:", "Total Overdue")
    text_clean = text.rstrip(":")
    if text_clean in ("total", "overdue", "available credit", "total overdue", "net balance"):
        return False
    # Also exclude if the SECOND cell is a number (looks like a summary: "Total  | 12345.00")
    if len(non_empty) == 2:
        try:
            float(str(non_empty[1][1]).replace(",", "").replace("$", "").strip())
            # Row has a label + number → summary, not section header
            if any(sw in text_clean for sw in ("total", "overdue", "credit", "balance")):
                return False
        except (ValueError, TypeError):
            pass
    return any(kw in text for kw in SECTION_KEYWORDS)


def _is_header_row(row_values: list) -> bool:
    """Return True if row looks like a column header row.

    Strict checks:
      1. At least 4 non-empty cells
      2. Most values must be short text (< 35 chars) — rules out data rows
      3. No large numeric values (> 100) — header cells are labels, not amounts
      4. At least 3 keyword hits among the SHORT text values
    """
    non_empty = [v for v in row_values if v is not None]
    if len(non_empty) < 4:
        return False
    # Reject if any cell is a large number (data rows have amounts)
    for v in non_empty:
        try:
            n = float(str(v).replace(",", "").replace("$", "").strip())
            if abs(n) > 100:
                return False  # Likely a data row
        except (ValueError, TypeError):
            pass
    # Only count keywords in SHORT text cells (column names are concise)
    short_texts = [str(v).strip().lower() for v in non_empty if len(str(v).strip()) < 35]
    if len(short_texts) < 3:
        return False
    hits = sum(1 for t in short_texts for kw in HEADER_KEYWORDS if kw in t)
    return hits >= 3


def _is_summary_row(row_values: list) -> str | None:
    """Return the summary type if this looks like a Total/Overdue row, else None.

    The match must be a STANDALONE label (short text, essentially the keyword itself),
    not a substring of a description like 'TotalCare' or 'Late Payment Invoice'.
    """
    for v in row_values:
        if v is None:
            continue
        t = str(v).strip().lower().rstrip(":")
        # Must be a short cell value (labels like "Total" or "Overdue")
        if len(t) > 25:
            continue
        if t in ("total", "overdue", "available credit", "total overdue",
                 "net balance", "total:", "overdue:"):
            return t.rstrip(":")
        # Also match patterns like "Total  " with trailing spaces already stripped
    return None


def _find_amount_col(header: list) -> int | None:
    """Find the column index that holds amounts."""
    for i, h in enumerate(header):
        if h is None:
            continue
        hl = str(h).lower()
        if "amount" in hl:
            return i
    return None


def _coerce_amount(val) -> float | None:
    """Convert a cell value to a float, handling $, commas, etc."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace("$", "").replace(" ", "")
    if s in ("", "-", "$ -", "$-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _coerce_date(val) -> pd.Timestamp | None:
    """Try to parse a date from various formats."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return pd.Timestamp(val)
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return pd.Timestamp(datetime.strptime(s, fmt))
        except ValueError:
            continue
    return None


def _coerce_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _normalise_header(raw: list) -> list:
    """Clean header names so they are consistent."""
    result = []
    for h in raw:
        if h is None:
            result.append(None)
            continue
        s = str(h).strip()
        # Collapse whitespace / newlines
        s = re.sub(r"\s+", " ", s)
        result.append(s)
    return result


def _map_columns(header: list) -> dict:
    """Map semantic roles to column indices from the header row.
    Returns dict with keys like 'amount', 'date', 'due_date', 'reference', etc."""
    mapping = {}
    hl = [str(h).lower() if h else "" for h in header]
    for i, h in enumerate(hl):
        if not h:
            continue
        if "amount" in h:
            mapping["amount"] = i
        elif h in ("curr", "currency"):
            mapping["currency"] = i
        elif "net due" in h or "due date" in h:
            mapping["due_date"] = i
        elif "document" in h and "date" in h:
            mapping["doc_date"] = i
        elif "document" in h and "no" in h:
            mapping["doc_no"] = i
        elif "invoice date" in h:
            mapping["doc_date"] = i
        elif h == "reference" or "reference" in h:
            mapping["reference"] = i
        elif h == "company" or "company" in h:
            mapping["company"] = i
        elif h == "account" or "account" in h:
            mapping["account"] = i
        elif "text" == h:
            mapping["text"] = i
        elif "assignment" in h or "arrangement" in h:
            mapping["assignment"] = i
        elif "r-r comment" in h or "rr comment" in h:
            mapping["rr_comments"] = i
        elif "action" in h or "reqd" in h:
            mapping["action_owner"] = i
        elif "days" in h and "late" in h:
            mapping["days_late"] = i
        elif "rata" in h:
            mapping["rata_date"] = i          # RATA Date is a date, not days-late int
        elif "comment" in h and "r-r" not in h and "rr" not in h:
            mapping["customer_comments"] = i
        elif "status" in h:
            mapping["status"] = i
        elif "customer" in h and "comment" not in h and "name" not in h and "n" not in h and "respon" not in h:
            mapping["customer_name"] = i
        elif "lpi" in h:
            mapping["lpi_cumulated"] = i
        elif "etr" in h or "po" in h or "pr" in h:
            mapping["po_reference"] = i
        elif "type" in h:
            mapping["type"] = i
        elif "interest" in h or "calc" in h:
            mapping["interest_method"] = i

    # If no explicit date columns, try to find 'date' generically
    if "doc_date" not in mapping:
        for i, h in enumerate(hl):
            if "date" in h and i not in mapping.values():
                mapping["doc_date"] = i
                break
    if "due_date" not in mapping:
        for i, h in enumerate(hl):
            if "due" in h and i not in mapping.values():
                mapping["due_date"] = i
                break
    return mapping


# ---------- Main parse function ----------

def parse_soa_workbook(file) -> dict:
    """Parse a Rolls-Royce Statement of Account workbook.

    Returns a dict:
        metadata  : dict of customer info, LPI rate, avg days late, etc.
        sections  : OrderedDict  section_name -> { header, colmap, rows (list[dict]), totals }
        all_items : pd.DataFrame  flattened across all sections
        grand_totals : dict
    """
    wb = load_workbook(file, data_only=True)
    ws = wb[wb.sheetnames[0]]  # first sheet

    max_col = ws.max_column or 20
    all_rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_col, values_only=True):
        all_rows.append(list(row))

    # ---- PASS 1: Metadata ----
    metadata = {}
    for idx, row in enumerate(all_rows[:15]):
        joined = " ".join(str(v) for v in row if v is not None).lower()
        for v in row:
            if v is None:
                continue
            s = str(v).strip()
            sl = s.lower()
            if "statement of account" in sl:
                metadata["title"] = s
            if "customer" in sl and ("name" in sl or ":" in sl) and "customer_name" not in metadata:
                # next non-None in same row is the value
                vi = row.index(v)
                for nv in row[vi+1:]:
                    if nv is not None:
                        metadata["customer_name"] = str(nv).strip()
                        break
            if ("customer" in sl and ("#" in sl or "n" in sl and ":" in sl)) or "customer n" in sl:
                vi = row.index(v)
                for nv in row[vi+1:]:
                    if nv is not None:
                        metadata["customer_id"] = str(nv).strip()
                        break
            if "contact" in sl:
                vi = row.index(v)
                for nv in row[vi+1:]:
                    if nv is not None:
                        metadata["contact"] = str(nv).strip()
                        break
            if "lpi" in sl or "lp ratio" in sl or "lp rate" in sl:
                for nv in row:
                    if nv is None:
                        continue
                    # Handle percentage strings like "1.5000%" or "3.500%"
                    nv_str = str(nv).strip()
                    if "%" in nv_str:
                        try:
                            metadata["lpi_rate"] = float(nv_str.replace("%", "")) / 100.0
                            break
                        except ValueError:
                            pass
                    amt = _coerce_amount(nv)
                    if amt is not None and 0 < abs(amt) < 1:
                        metadata["lpi_rate"] = amt
                        break
            if "average days late" in sl or "avg days late" in sl or "average days late" in joined:
                for nv in row:
                    if nv is None:
                        continue
                    val = _coerce_int(nv)
                    if val is not None and val > 0:
                        metadata["avg_days_late"] = val
                        break
            if "today" in sl:
                for nv in row:
                    d = _coerce_date(nv)
                    if d is not None:
                        metadata["report_date"] = d
                        break

    # ---- PASS 2: Identify section boundaries and headers ----
    master_header = None
    master_header_idx = None
    sections_info = []  # list of (name, start_row_idx)

    for idx, row in enumerate(all_rows):
        if _is_header_row(row) and master_header is None:
            master_header = _normalise_header(row)
            master_header_idx = idx
            continue
        if _is_section_header(row, max_col):
            name = str([v for v in row if v is not None][0]).strip()
            sections_info.append({"name": name, "start": idx})

    # Assign end boundaries
    for i, sec in enumerate(sections_info):
        if i + 1 < len(sections_info):
            sec["end"] = sections_info[i + 1]["start"]
        else:
            sec["end"] = len(all_rows)

    # ---- PASS 3: Parse each section ----
    sections = OrderedDict()
    all_items_list = []

    for sec in sections_info:
        sec_name = sec["name"]
        start = sec["start"]
        end = sec["end"]

        # Determine the header row for this section
        header = master_header
        header_idx = master_header_idx
        col_map = None

        # Check if section has its own header row (within first 3 rows after start)
        for offset in range(1, 4):
            ri = start + offset
            if ri >= end:
                break
            if _is_header_row(all_rows[ri]):
                header = _normalise_header(all_rows[ri])
                header_idx = ri
                break

        if header:
            col_map = _map_columns(header)
        else:
            col_map = {}

        # Ensure we have an amount column
        amt_idx = col_map.get("amount")
        if amt_idx is None and header:
            amt_idx = _find_amount_col(header)
            if amt_idx is not None:
                col_map["amount"] = amt_idx

        # Parse data rows
        data_rows = []
        totals = {}
        data_start = (header_idx + 1) if header_idx and header_idx >= start else start + 1

        for ri in range(data_start, end):
            row = all_rows[ri]
            # Check for summary row
            summary_type = _is_summary_row(row)
            if summary_type:
                # Find the numeric value next to the keyword
                for v in row:
                    amt = _coerce_amount(v)
                    if amt is not None:
                        totals[summary_type] = amt
                        break
                continue

            # Check if it's a section header sneaking in (skip)
            if _is_section_header(row, max_col):
                continue
            # Also skip rows that look like headers for subsections
            if _is_header_row(row):
                header = _normalise_header(row)
                col_map = _map_columns(header)
                amt_idx = col_map.get("amount")
                if amt_idx is None and header:
                    amt_idx = _find_amount_col(header)
                    if amt_idx is not None:
                        col_map["amount"] = amt_idx
                continue

            # Must have at least an amount to be a data row
            amt_val = None
            if amt_idx is not None and amt_idx < len(row):
                amt_val = _coerce_amount(row[amt_idx])
            # Fallback: scan for a plausible amount
            if amt_val is None:
                for ci, cv in enumerate(row):
                    a = _coerce_amount(cv)
                    if a is not None and abs(a) > 0.01:
                        # Avoid picking up "days late" small int or dates
                        if abs(a) > 100 or (col_map.get("days_late") is not None and ci != col_map.get("days_late")):
                            amt_val = a
                            break

            if amt_val is None:
                continue  # Skip empty/irrelevant rows

            record = {
                "Section": sec_name,
                "Amount": amt_val,
            }

            # Populate from column map
            def _get(key, coerce=str):
                ci = col_map.get(key)
                if ci is None or ci >= len(row):
                    return None
                v = row[ci]
                if v is None:
                    return None
                if coerce == float:
                    return _coerce_amount(v)
                if coerce == "date":
                    return _coerce_date(v)
                if coerce == int:
                    return _coerce_int(v)
                return str(v).strip()

            record["Company"]         = _get("company")
            record["Account"]         = _get("account")
            record["Reference"]       = _get("reference")
            record["Document Date"]   = _get("doc_date", "date")
            record["Due Date"]        = _get("due_date", "date")
            record["Currency"]        = _get("currency")
            record["Text"]            = _get("text")
            record["Assignment"]      = _get("assignment")
            record["R-R Comments"]    = _get("rr_comments")
            record["Action Owner"]    = _get("action_owner")
            record["Days Late"]       = _get("days_late", int)
            record["Customer Comments"] = _get("customer_comments")
            record["Status"]          = _get("status")
            record["PO Reference"]    = _get("po_reference")
            record["LPI Cumulated"]   = _get("lpi_cumulated")
            record["Type"]            = _get("type")
            record["Document No"]     = _get("doc_no")
            record["Interest Method"] = _get("interest_method")
            record["Customer Name"]   = _get("customer_name")

            # Auto-compute Days Late from Due Date when not explicitly available
            if record["Days Late"] is None and record["Due Date"] is not None:
                try:
                    due = record["Due Date"]
                    today = pd.Timestamp.now().normalize()
                    if due < today:
                        record["Days Late"] = (today - due).days
                    else:
                        record["Days Late"] = 0
                except Exception:
                    pass

            # Derive a unified Status field
            if not record.get("Status"):
                for field in ["R-R Comments", "Action Owner", "Customer Comments"]:
                    v = record.get(field, "")
                    if v and any(kw in v.lower() for kw in ["ready for payment", "under approval", "under review",
                                                             "dispute", "ongoing", "et to process", "payment pending",
                                                             "invoice sent", "credit note", "approved",
                                                             "transfer", "invoice approved", "pending for payment"]):
                        record["Status"] = v
                        break
            if not record.get("Status"):
                rrc = record.get("R-R Comments", "")
                if rrc:
                    record["Status"] = rrc

            # Determine if debit or credit
            record["Entry Type"] = "Credit" if amt_val < 0 else "Charge"

            data_rows.append(record)
            all_items_list.append(record)

        sections[sec_name] = {
            "header": header,
            "colmap": col_map,
            "rows": data_rows,
            "totals": totals,
        }

    # Build DataFrame
    df = pd.DataFrame(all_items_list)
    if df.empty:
        df = pd.DataFrame(columns=["Section", "Amount", "Entry Type"])

    # ---- Grand totals ----
    grand = {}
    for sec_name, sec_data in sections.items():
        for k, v in sec_data["totals"].items():
            if "total overdue" in k:
                grand["total_overdue"] = v
            elif "overdue" in k:
                grand.setdefault("section_overdue", {})[sec_name] = v
            elif "available credit" in k:
                grand.setdefault("available_credits", {})[sec_name] = v
            elif "total" in k:
                grand.setdefault("section_totals", {})[sec_name] = v

    if not df.empty:
        grand["total_charges"]  = df.loc[df["Amount"] > 0, "Amount"].sum()
        grand["total_credits"]  = df.loc[df["Amount"] < 0, "Amount"].sum()
        grand["net_balance"]    = df["Amount"].sum()
        grand["item_count"]     = len(df)
        if "total_overdue" not in grand:
            overdue_sum = sum(grand.get("section_overdue", {}).values())
            if overdue_sum:
                grand["total_overdue"] = overdue_sum

    return {
        "metadata": metadata,
        "sections": sections,
        "all_items": df,
        "grand_totals": grand,
    }


# ─────────────────────────────────────────────────────────────
# 2.  HELPER VIS FUNCTIONS
# ─────────────────────────────────────────────────────────────

def fmt_currency(val, short=False):
    """Format a number as USD currency string."""
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return "—"
    neg = val < 0
    av = abs(val)
    if short and av >= 1_000_000:
        s = f"${av/1_000_000:,.2f}M"
    elif short and av >= 1_000:
        s = f"${av/1_000:,.1f}K"
    else:
        s = f"${av:,.2f}"
    return f"-{s}" if neg else s


def metric_card(label, value, color_class=""):
    return f"""
    <div class="metric-card">
        <div class="label">{label}</div>
        <div class="value {color_class}">{value}</div>
    </div>"""


def make_donut(labels, values, title, colors=None):
    if colors is None:
        colors = SECTION_COLOURS[:len(labels)]
    fig = go.Figure(go.Pie(
        labels=labels, values=values, hole=.55,
        marker=dict(colors=colors, line=dict(color="#FFF", width=2)),
        textinfo="label+percent", textposition="outside",
        textfont=dict(size=12, color="#222"),
    ))
    fig.update_layout(
        title=dict(text=f"<b>{title}</b>", font=dict(size=15, color="#1a1a2e")),
        showlegend=False, margin=dict(t=55, b=25, l=25, r=25),
        height=360, paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
    )
    return fig


# Shared chart layout defaults for readable text
CHART_LAYOUT = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    font=dict(size=12, color="#1a1a2e", family="Inter, sans-serif"),
    xaxis=dict(gridcolor="#E0E0E0", tickfont=dict(size=11, color="#333")),
    yaxis=dict(gridcolor="#E0E0E0", tickfont=dict(size=11, color="#333")),
)


def make_bar(df_plot, x, y, title, color=RR_NAVY, horizontal=False, color_col=None, text_auto=True):
    if horizontal:
        fig = px.bar(df_plot, y=x, x=y, orientation="h", title=title, text_auto=".2s" if text_auto else False,
                     color=color_col, color_discrete_sequence=SECTION_COLOURS if color_col else [color])
    else:
        fig = px.bar(df_plot, x=x, y=y, title=title, text_auto=".2s" if text_auto else False,
                     color=color_col, color_discrete_sequence=SECTION_COLOURS if color_col else [color])
    fig.update_layout(
        height=360, margin=dict(t=55, b=45, l=60, r=20),
        showlegend=bool(color_col),
        title=dict(font=dict(size=15, color="#1a1a2e")),
        **CHART_LAYOUT,
    )
    return fig


def aging_bucket(days) -> str:
    if days is None or (isinstance(days, float) and math.isnan(days)):
        return "Unknown"
    d = int(days)
    if d <= 0:
        return "Current"
    elif d <= 30:
        return "1-30 Days"
    elif d <= 60:
        return "31-60 Days"
    elif d <= 90:
        return "61-90 Days"
    elif d <= 180:
        return "91-180 Days"
    else:
        return "180+ Days"


AGING_ORDER = ["Current", "1-30 Days", "31-60 Days", "61-90 Days", "91-180 Days", "180+ Days", "Unknown"]
AGING_COLORS = {"Current": "#2E7D32", "1-30 Days": "#66BB6A", "31-60 Days": "#F9A825",
                "61-90 Days": "#EF6C00", "91-180 Days": "#D32F2F", "180+ Days": "#B71C1C", "Unknown": "#9E9E9E"}


# ─────────────────────────────────────────────────────────────
# 3.  SIDEBAR — FILE UPLOAD
# ─────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("""
    <div style="text-align:center; padding: 1rem 0;">
        <div style="color:#FFF; font-size:1.2rem; font-weight:700; letter-spacing:2px;
                    border:2px solid #C0C0C0; display:inline-block; padding:.4rem 1rem; border-radius:4px;">
            ROLLS-ROYCE
        </div>
        <div style="color:#C0C0C0; font-size:.7rem; margin-top:.3rem; letter-spacing:1.5px;">CIVIL AEROSPACE</div>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("##### Upload Statement of Account")
    uploaded_file = st.file_uploader("Upload .xlsx file", type=["xlsx", "xls"], label_visibility="collapsed")

    st.markdown("---")
    st.markdown("##### Dashboard Settings")
    currency_symbol = st.selectbox("Currency Display", ["USD", "GBP", "EUR"], index=0)
    show_credits_in_tables = st.checkbox("Show credits in line tables", value=True)


# ─────────────────────────────────────────────────────────────
# 4.  MAIN DASHBOARD
# ─────────────────────────────────────────────────────────────

# Header
st.markdown("""
<div class="rr-header">
    <div>
        <div style="color:#FFFFFF; margin:0; font-size:1.65rem; font-weight:700; letter-spacing:.5px;">Statement of Account Dashboard</div>
        <div style="color:#D0D0E0; font-size:.92rem; margin-top:.2rem; font-weight:400;">Rolls-Royce Civil Aerospace &mdash; Finance & Receivables</div>
    </div>
    <div style="color:#FFFFFF; font-size:1.15rem; font-weight:700; letter-spacing:3px; text-transform:uppercase; border:2px solid rgba(255,255,255,.6); padding:.4rem 1rem; border-radius:4px;">ROLLS-ROYCE</div>
</div>
""", unsafe_allow_html=True)

if uploaded_file is None:
    st.info("Upload a Statement of Account Excel file (.xlsx) using the sidebar to get started.")
    st.markdown("""
    <div style="background:#FFF; padding:2rem; border-radius:12px; text-align:center; margin:2rem 0; box-shadow: 0 2px 8px rgba(0,0,0,.06);">
        <div style="font-size:3rem; margin-bottom:1rem;">✈</div>
        <h3 style="color:#10069F; margin-bottom:.5rem;">Welcome to the RR SOA Dashboard</h3>
        <p style="color:#666; max-width:500px; margin:0 auto;">
            Upload any Rolls-Royce Statement of Account workbook. The dashboard automatically
            detects sections like <b>TotalCare</b>, <b>CRC Payments</b>, <b>Spare Parts</b>,
            <b>Late Payment Interest</b>, and more — regardless of layout variations.
        </p>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# ---- Parse the file ----
with st.spinner("Parsing workbook..."):
    data = parse_soa_workbook(uploaded_file)

meta      = data["metadata"]
sections  = data["sections"]
df_all    = data["all_items"]
grand     = data["grand_totals"]

# ---- Customer Info Bar ----
cust_name = meta.get("customer_name", "Unknown Customer")
cust_id   = meta.get("customer_id", "—")
contact   = meta.get("contact", "—")
lpi_rate  = meta.get("lpi_rate")
avg_late  = meta.get("avg_days_late")
report_dt = meta.get("report_date")
title     = meta.get("title", "Statement of Account")

st.markdown(f"""
<div style="background:#FFFFFF; padding:1rem 1.8rem; border-radius:12px; margin-bottom:1.2rem;
            display:flex; align-items:center; justify-content:space-between; flex-wrap:wrap;
            box-shadow: 0 2px 10px rgba(0,0,0,.07); border: 1px solid #E0E1EC;">
    <div>
        <span style="font-weight:800; color:#0C0033; font-size:1.15rem;">{cust_name}</span>
        <span class="info-chip">ID: {cust_id}</span>
        <span class="info-chip">{contact}</span>
    </div>
    <div style="display:flex; gap:.8rem; align-items:center; flex-wrap:wrap;">
        {"<span class='info-chip'><b>LPI Rate:</b> " + f"{lpi_rate*100:.2f}%" + "</span>" if lpi_rate else ""}
        {"<span class='info-chip'><b>Avg Days Late:</b> " + str(avg_late) + "</span>" if avg_late else ""}
        {"<span class='info-chip'><b>Report:</b> " + report_dt.strftime('%d %b %Y') + "</span>" if report_dt else ""}
    </div>
</div>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────
# 5.  KPI CARDS
# ─────────────────────────────────────────────────────────────

total_charges = grand.get("total_charges", 0)
total_credits = grand.get("total_credits", 0)
net_balance   = grand.get("net_balance", 0)
total_overdue = grand.get("total_overdue", net_balance)
item_count    = grand.get("item_count", 0)

kpi_cols = st.columns(6)
kpis = [
    ("Total Charges", fmt_currency(total_charges, short=True), ""),
    ("Total Credits", fmt_currency(total_credits, short=True), "positive" if total_credits < 0 else ""),
    ("Net Balance", fmt_currency(net_balance, short=True), "negative" if net_balance > 0 else "positive"),
    ("Total Overdue", fmt_currency(total_overdue, short=True), "negative"),
    ("Avg Days Late", str(avg_late) if avg_late else "—", ""),
    ("Open Items", str(item_count), ""),
]
for col, (label, value, cls) in zip(kpi_cols, kpis):
    col.markdown(metric_card(label, value, cls), unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────
# 6.  EXECUTIVE OVERVIEW CHARTS
# ─────────────────────────────────────────────────────────────

st.markdown('<div class="section-hdr">Executive Overview</div>', unsafe_allow_html=True)

overview_cols = st.columns([1, 1, 1])

# 6a. Breakdown by Section (donut)
if not df_all.empty:
    sec_abs = df_all.groupby("Section")["Amount"].apply(lambda x: x.abs().sum()).reset_index()
    sec_abs.columns = ["Section", "Absolute Amount"]
    with overview_cols[0]:
        st.plotly_chart(make_donut(sec_abs["Section"], sec_abs["Absolute Amount"],
                                   "Breakdown by Section"), use_container_width=True)

# 6b. Charges vs Credits per Section
if not df_all.empty:
    charge_credit = df_all.groupby(["Section", "Entry Type"])["Amount"].sum().reset_index()
    charge_credit["Amount"] = charge_credit["Amount"].abs()
    with overview_cols[1]:
        fig_cc = px.bar(charge_credit, x="Section", y="Amount", color="Entry Type",
                        barmode="group", title="<b>Charges vs Credits by Section</b>",
                        color_discrete_map={"Charge": RR_NAVY, "Credit": RR_GREEN},
                        text_auto=".2s")
        fig_cc.update_layout(height=360, margin=dict(t=55, b=45, l=45, r=20),
                             title=dict(font=dict(size=15, color="#1a1a2e")),
                             legend=dict(font=dict(size=12, color="#333")),
                             **CHART_LAYOUT)
        st.plotly_chart(fig_cc, use_container_width=True)

# 6c. Aging Distribution
if not df_all.empty and "Days Late" in df_all.columns:
    df_aging = df_all.copy()
    df_aging["Aging Bucket"] = df_aging["Days Late"].apply(aging_bucket)
    aging_counts = df_aging.groupby("Aging Bucket").agg(
        Count=("Amount", "count"),
        Total_Amount=("Amount", "sum")
    ).reindex([b for b in AGING_ORDER if b in df_aging["Aging Bucket"].values]).reset_index()

    with overview_cols[2]:
        fig_age = px.bar(aging_counts, x="Aging Bucket", y="Total_Amount", title="<b>Aging Analysis</b>",
                         text_auto=".2s", color="Aging Bucket",
                         color_discrete_map=AGING_COLORS)
        fig_age.update_layout(height=360, showlegend=False,
                              margin=dict(t=55, b=45, l=45, r=20),
                              title=dict(font=dict(size=15, color="#1a1a2e")),
                              **CHART_LAYOUT)
        st.plotly_chart(fig_age, use_container_width=True)


# ─────────────────────────────────────────────────────────────
# 7.  BILATERAL POSITION: What We Owe Them vs What They Owe Us
# ─────────────────────────────────────────────────────────────

st.markdown('<div class="section-hdr">Bilateral Position</div>', unsafe_allow_html=True)

bp_cols = st.columns([1, 1])

if not df_all.empty:
    they_owe = df_all.loc[df_all["Amount"] > 0, "Amount"].sum()   # Customer owes RR
    we_owe   = df_all.loc[df_all["Amount"] < 0, "Amount"].sum()   # RR credits/owes customer

    with bp_cols[0]:
        bp_data = pd.DataFrame({
            "Direction": [f"Customer → RR (Charges)", f"RR → Customer (Credits)"],
            "Amount": [they_owe, abs(we_owe)]
        })
        fig_bp = px.bar(bp_data, x="Direction", y="Amount", title="<b>Bilateral Position</b>",
                        color="Direction",
                        color_discrete_map={
                            f"Customer → RR (Charges)": RR_NAVY,
                            f"RR → Customer (Credits)": RR_GREEN
                        }, text_auto=".3s")
        fig_bp.update_layout(height=360, showlegend=False,
                             margin=dict(t=55, b=65, l=45, r=20),
                             title=dict(font=dict(size=15, color="#1a1a2e")),
                             **CHART_LAYOUT)
        st.plotly_chart(fig_bp, use_container_width=True)

    with bp_cols[1]:
        # Net per section
        sec_net = df_all.groupby("Section")["Amount"].sum().reset_index()
        sec_net.columns = ["Section", "Net Amount"]
        sec_net["Color"] = sec_net["Net Amount"].apply(lambda x: "Owed to RR" if x > 0 else "Credit to Customer")
        fig_sn = px.bar(sec_net, y="Section", x="Net Amount", orientation="h",
                        title="<b>Net Balance by Section</b>", color="Color",
                        color_discrete_map={"Owed to RR": RR_NAVY, "Credit to Customer": RR_GREEN},
                        text_auto=".3s")
        fig_sn.update_layout(height=360, margin=dict(t=55, b=45, l=25, r=25),
                             title=dict(font=dict(size=15, color="#1a1a2e")),
                             legend=dict(font=dict(size=12, color="#333")),
                             **CHART_LAYOUT)
        st.plotly_chart(fig_sn, use_container_width=True)


# ─────────────────────────────────────────────────────────────
# 8.  SECTION DETAIL TABS
# ─────────────────────────────────────────────────────────────

st.markdown('<div class="section-hdr">Section Breakdown</div>', unsafe_allow_html=True)

if sections:
    tab_names = list(sections.keys())
    tabs = st.tabs(tab_names)

    for tab, sec_name in zip(tabs, tab_names):
        sec = sections[sec_name]
        sec_rows = sec["rows"]
        sec_totals = sec["totals"]

        if not sec_rows:
            with tab:
                st.info(f"No line items found in **{sec_name}**.")
            continue

        sec_df = pd.DataFrame(sec_rows)

        with tab:
            # Section KPIs
            sec_total   = sec_totals.get("total", sec_df["Amount"].sum())
            sec_overdue = sec_totals.get("overdue", None)
            sec_credits_avail = sec_totals.get("available credit", None)
            sec_charges = sec_df.loc[sec_df["Amount"] > 0, "Amount"].sum()
            sec_credits = sec_df.loc[sec_df["Amount"] < 0, "Amount"].sum()
            sec_items   = len(sec_df)

            sk = st.columns(5)
            sk[0].markdown(metric_card("Section Total", fmt_currency(sec_total, True)), unsafe_allow_html=True)
            sk[1].markdown(metric_card("Charges", fmt_currency(sec_charges, True)), unsafe_allow_html=True)
            sk[2].markdown(metric_card("Credits", fmt_currency(sec_credits, True), "positive"), unsafe_allow_html=True)
            if sec_overdue is not None:
                sk[3].markdown(metric_card("Overdue", fmt_currency(sec_overdue, True), "negative"), unsafe_allow_html=True)
            else:
                sk[3].markdown(metric_card("Items", str(sec_items)), unsafe_allow_html=True)
            if sec_credits_avail is not None:
                sk[4].markdown(metric_card("Available Credit", fmt_currency(sec_credits_avail, True), "positive"), unsafe_allow_html=True)
            else:
                sk[4].markdown(metric_card("Net", fmt_currency(sec_charges + sec_credits, True)), unsafe_allow_html=True)

            st.markdown("")

            # Two-column: Status distribution + Top items
            chart_cols = st.columns([1, 1])

            with chart_cols[0]:
                # Status distribution
                if "Status" in sec_df.columns:
                    status_df = sec_df.copy()
                    status_df["Status Clean"] = status_df["Status"].fillna("Unknown").apply(
                        lambda s: s[:40] + "..." if len(str(s)) > 40 else s
                    )
                    status_counts = status_df["Status Clean"].value_counts().reset_index()
                    status_counts.columns = ["Status", "Count"]
                    if len(status_counts) > 0:
                        fig_st = px.pie(status_counts, names="Status", values="Count",
                                        title="<b>Status Distribution</b>",
                                        color_discrete_sequence=SECTION_COLOURS)
                        fig_st.update_layout(height=350, margin=dict(t=55, b=25, l=25, r=25),
                                             paper_bgcolor="rgba(0,0,0,0)",
                                             font=dict(size=12, color="#1a1a2e"),
                                             title=dict(font=dict(size=15, color="#1a1a2e")),
                                             showlegend=True,
                                             legend=dict(font=dict(size=11, color="#333")))
                        fig_st.update_traces(textposition="inside", textinfo="value",
                                             textfont=dict(size=12, color="#FFF"))
                        st.plotly_chart(fig_st, use_container_width=True)

            with chart_cols[1]:
                # Top items by amount
                avail_cols = [c for c in ["Text", "Amount", "Reference"] if c in sec_df.columns]
                if "Amount" in avail_cols:
                    top_items = sec_df.nlargest(8, "Amount")[avail_cols].copy()
                    if top_items.empty:
                        top_items = sec_df.head(8)[avail_cols].copy()
                    text_col = "Text" if "Text" in top_items.columns else (
                        "Assignment" if "Assignment" in sec_df.columns else None)
                    ref_col = "Reference" if "Reference" in top_items.columns else None
                    if text_col and text_col not in top_items.columns:
                        top_items[text_col] = sec_df.nlargest(8, "Amount")[text_col]
                    if not top_items.empty and text_col and text_col in top_items.columns:
                        top_items["Label"] = top_items.apply(
                            lambda r: (str(r.get(text_col, ""))[:30] + " (" + str(r.get(ref_col, "")) + ")")
                            if ref_col and r.get(ref_col) else str(r.get(text_col, ""))[:40], axis=1)
                        fig_top = px.bar(top_items, y="Label", x="Amount", orientation="h",
                                         title="<b>Top Items by Amount</b>", text_auto=".3s",
                                         color_discrete_sequence=[RR_NAVY])
                        fig_top.update_layout(height=350, margin=dict(t=55, b=25, l=25, r=25),
                                              title=dict(font=dict(size=15, color="#1a1a2e")),
                                              **CHART_LAYOUT)
                        st.plotly_chart(fig_top, use_container_width=True)

            # Detailed data table
            st.markdown("**Detailed Line Items**")

            # Prepare display dataframe
            display_cols = []
            preferred_order = ["Reference", "Document No", "Document Date", "Due Date", "Amount",
                               "Currency", "Text", "Type", "Assignment", "R-R Comments",
                               "Status", "Action Owner", "Days Late", "Customer Comments",
                               "Customer Name", "PO Reference", "LPI Cumulated", "Interest Method",
                               "Entry Type"]
            for c in preferred_order:
                if c in sec_df.columns and sec_df[c].notna().any():
                    display_cols.append(c)

            # Add any remaining columns
            for c in sec_df.columns:
                if c not in display_cols and c != "Section" and sec_df[c].notna().any():
                    display_cols.append(c)

            display_df = sec_df[display_cols].copy()

            if not show_credits_in_tables:
                display_df = display_df[display_df.get("Entry Type", "Charge") == "Charge"]

            # Format amount for display
            if "Amount" in display_df.columns:
                display_df["Amount"] = display_df["Amount"].apply(lambda x: fmt_currency(x) if x else "—")
            if "Document Date" in display_df.columns:
                display_df["Document Date"] = display_df["Document Date"].apply(
                    lambda x: x.strftime("%d/%m/%Y") if pd.notna(x) and hasattr(x, "strftime") else str(x) if x else "—")
            if "Due Date" in display_df.columns:
                display_df["Due Date"] = display_df["Due Date"].apply(
                    lambda x: x.strftime("%d/%m/%Y") if pd.notna(x) and hasattr(x, "strftime") else str(x) if x else "—")

            st.dataframe(display_df, use_container_width=True, height=min(400, 35 * len(display_df) + 60))


# ─────────────────────────────────────────────────────────────
# 9.  FULL DATA TABLE (All Sections)
# ─────────────────────────────────────────────────────────────

st.markdown('<div class="section-hdr">Complete Invoice Register</div>', unsafe_allow_html=True)

if not df_all.empty:
    # Filters
    filter_cols = st.columns(4)
    with filter_cols[0]:
        sec_filter = st.multiselect("Section", df_all["Section"].unique(), default=df_all["Section"].unique())
    with filter_cols[1]:
        type_filter = st.multiselect("Type", df_all["Entry Type"].unique(), default=df_all["Entry Type"].unique())
    with filter_cols[2]:
        if "Status" in df_all.columns:
            statuses = df_all["Status"].dropna().unique()
            status_filter = st.multiselect("Status", statuses, default=[])
        else:
            status_filter = []
    with filter_cols[3]:
        amount_range = st.slider("Amount Range (absolute)",
                                  0.0, float(df_all["Amount"].abs().max()) if not df_all.empty else 1e6,
                                  (0.0, float(df_all["Amount"].abs().max()) if not df_all.empty else 1e6))

    filtered = df_all.copy()
    filtered = filtered[filtered["Section"].isin(sec_filter)]
    filtered = filtered[filtered["Entry Type"].isin(type_filter)]
    if status_filter:
        filtered = filtered[filtered["Status"].isin(status_filter)]
    filtered = filtered[(filtered["Amount"].abs() >= amount_range[0]) & (filtered["Amount"].abs() <= amount_range[1])]

    # Show columns that have data
    show_cols = [c for c in ["Section", "Reference", "Document No", "Document Date", "Due Date",
                             "Amount", "Currency", "Text", "Type", "Assignment", "R-R Comments",
                             "Status", "Action Owner", "Days Late", "Customer Comments",
                             "Customer Name", "PO Reference", "LPI Cumulated", "Entry Type",
                             "Interest Method"]
                 if c in filtered.columns and filtered[c].notna().any()]

    disp = filtered[show_cols].copy()

    # Format for display
    if "Amount" in disp.columns:
        disp["Amount"] = disp["Amount"].apply(lambda x: fmt_currency(x))
    for dcol in ["Document Date", "Due Date"]:
        if dcol in disp.columns:
            disp[dcol] = disp[dcol].apply(
                lambda x: x.strftime("%d/%m/%Y") if pd.notna(x) and hasattr(x, "strftime") else str(x) if x else "—")

    st.dataframe(disp, use_container_width=True, height=min(600, 35 * len(disp) + 60))

    # Summary below table
    sum_cols = st.columns(3)
    sum_cols[0].metric("Filtered Items", len(filtered))
    sum_cols[1].metric("Filtered Total", fmt_currency(filtered["Amount"].sum()))
    if "Days Late" in filtered.columns:
        overdue_mask = filtered["Days Late"].fillna(0).apply(lambda x: int(x) if x else 0) > 0
        overdue_total = filtered.loc[overdue_mask, "Amount"].sum()
    else:
        overdue_total = 0
    sum_cols[2].metric("Filtered Overdue", fmt_currency(overdue_total))


# ─────────────────────────────────────────────────────────────
# 10.  FOOTER
# ─────────────────────────────────────────────────────────────

st.markdown("---")
st.markdown("""
<div style="text-align:center; color:#999; font-size:.75rem; padding:1rem 0;">
    <b>ROLLS-ROYCE</b> CIVIL AEROSPACE &mdash; Statement of Account Dashboard<br>
    Data sourced from uploaded workbook &bull; For internal use only
</div>
""", unsafe_allow_html=True)
