"""
parser.py — Universal Rolls-Royce Excel Parser
================================================
Handles all known RR Civil Aerospace file formats:
  • SOA  (Statement of Account)          — multi-section, heuristic layout
  • INVOICE_LIST                          — flat open-items register (EPI style)
  • OPPORTUNITY_TRACKER (MEA Tracker)    — wide opportunity log (L2/L3 sheets)
  • SHOP_VISIT_HISTORY                   — Trent shop visit event history
  • SVRG_MASTER                          — Trent 900 guarantee administration

Design principles
-----------------
  1. NEVER crash — every path ends in a valid dict with an "errors" key.
  2. NEVER hard-code column positions; use dynamic header discovery.
  3. NEVER trust data types — dates can be datetime | str | float | None.
  4. Detect file type by scoring sheet names + cell content.
  5. Return a canonical JSON-serialisable dict for every file type.

Usage
-----
    from parser import parse_file

    # From a file path:
    result = parse_file("/path/to/file.xlsx")

    # From a base64-encoded string (browser upload):
    result = parse_file(base64_string, filename="ETH SOA.xlsx", is_base64=True)
"""

from __future__ import annotations

import base64
import io
import logging
import re
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Union

import pandas as pd

logger = logging.getLogger(__name__)

# ══════════════════════════════════════════════════════════════════════════════
# Primitive helpers
# ══════════════════════════════════════════════════════════════════════════════

_DATE_FMTS = [
    "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y",
    "%Y/%m/%d", "%d.%m.%Y", "%m.%d.%Y",
    "%d/%m/%y", "%m/%d/%y",
]


# Excel error cell strings that should be treated as None
_EXCEL_ERRORS = frozenset({
    "#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!",
    "#GETTING_DATA", "#SPILL!", "#CALC!", "#FIELD!",
})


def _is_excel_error(v: Any) -> bool:
    """Detect Excel error cell strings."""
    if isinstance(v, str):
        s = v.strip().upper()
        return s in _EXCEL_ERRORS
    return False


def _clean(v: Any) -> str:
    """Return stripped string; '' when None/NaN/Excel-error."""
    if v is None:
        return ""
    if isinstance(v, float) and (v != v):          # NaN fast-path
        return ""
    if _is_excel_error(v):
        return ""
    return str(v).strip()


def _is_blank(v: Any) -> bool:
    return (
        v is None
        or (isinstance(v, float) and (v != v))
        or _is_excel_error(v)
        or _clean(v) == ""
    )


def _last_nonempty_col(header_row: Any, hard_cap: int = 500) -> int:
    """Return 1 + index of last non-empty cell in header row.
    Guards against Excel's phantom 16383 / 16384 widths.
    Accepts a pandas Series or a list-like."""
    try:
        n = len(header_row)
    except Exception:
        return hard_cap
    last = 0
    end = min(n, hard_cap)
    for i in range(end):
        try:
            v = header_row.iloc[i] if hasattr(header_row, "iloc") else header_row[i]
        except Exception:
            continue
        if not _is_blank(v):
            last = i + 1
    return last or end


def _is_ws_chartsheet(ws) -> bool:
    """Cheap isinstance check for openpyxl Chartsheet (no max_row attr)."""
    try:
        from openpyxl.chartsheet import Chartsheet  # type: ignore
        return isinstance(ws, Chartsheet)
    except Exception:
        return not hasattr(ws, "iter_rows")


def _to_date(v: Any) -> Optional[str]:
    """Convert any date-like value → ISO 'YYYY-MM-DD' string, or None."""
    if _is_blank(v):
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = _clean(v)
    if not s or s.lower() in ("nat", "none", "nan"):
        return None
    for fmt in _DATE_FMTS:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    try:
        return pd.to_datetime(s, dayfirst=True).strftime("%Y-%m-%d")
    except Exception:
        return s                                     # return as-is if unparseable


def _to_float(v: Any) -> Optional[float]:
    """Convert any currency-like value → float, or None.
    Returns None for Excel error cells and non-numeric placeholder text
    (e.g. "TBD", "tbc", "Confirm with Harry")."""
    if _is_blank(v):
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if v != v:                                   # NaN
            return None
        return float(v)
    s = _clean(v).replace("$", "").replace("£", "").replace("€", "").replace(",", "").replace(" ", "")
    # parentheses → negative  e.g. (1,234.56) → -1234.56
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    if s in ("-", "", "n/a", "#n/a", "#value!", "#ref!", "#div/0!", "#name?"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _is_numeric_text(v: Any) -> bool:
    """True if v is numeric or parses cleanly to float (for mixed-column detection)."""
    return _to_float(v) is not None


def _to_str_ref(v: Any) -> Optional[str]:
    """Normalise reference numbers to strings (strip .0 suffix from floats).
    Always returns a string to preserve leading zeros on invoice references."""
    if _is_blank(v):
        return None
    # Excel dates masquerading as references: preserve ISO form
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, float):
        if v != v:
            return None
        if v == int(v):
            return str(int(v))
        return repr(v)
    if isinstance(v, int):
        return str(v)
    return _clean(v) or None


def _row_values(row: pd.Series) -> List[Any]:
    return list(row)


def _non_blank_vals(row: pd.Series) -> List[str]:
    return [_clean(v) for v in row if not _is_blank(v)]


# ══════════════════════════════════════════════════════════════════════════════
# Excel loader — accepts path, BytesIO, or base64 string
# ══════════════════════════════════════════════════════════════════════════════

def _load_workbook(source: Union[str, bytes, io.BytesIO], filename: str = "") -> Optional[Dict[str, pd.DataFrame]]:
    """
    Load every sheet of an Excel workbook into a {name: DataFrame} dict.
    DataFrames have no header applied (header=None); raw rows preserved.
    Returns None on failure (error logged).
    """
    try:
        if isinstance(source, str) and not source.endswith((".xlsx", ".xls", ".xlsb", ".xlsm")):
            # Assume base64
            raw = base64.b64decode(source)
            buf = io.BytesIO(raw)
        elif isinstance(source, (bytes, bytearray)):
            buf = io.BytesIO(source)
        elif isinstance(source, io.BytesIO):
            buf = source
        else:
            buf = source                             # file path string

        xl = pd.ExcelFile(buf)
        sheets: Dict[str, pd.DataFrame] = {}
        for name in xl.sheet_names:
            try:
                df = xl.parse(name, header=None, dtype=object)
                sheets[name] = df
            except Exception as e:
                logger.warning("Skipped sheet '%s': %s", name, e)
        return sheets if sheets else None
    except Exception as e:
        logger.error("Failed to load workbook: %s", e)
        return None


# ══════════════════════════════════════════════════════════════════════════════
# File-type detection
# ══════════════════════════════════════════════════════════════════════════════

_TYPE_SIGNALS: Dict[str, List[str]] = {
    "SOA": [
        "statement of account", "lpi rate", "lpi rate >",
        "customer name:", "totalcare charges", "customer responsible charges",
        "spare parts charges", "late payment interest", "credits usable",
        "total overdue",
    ],
    "INVOICE_LIST": [
        "reference key 3", "amount in doc. curr.", "net due date", "document date",
        "document currency",
    ],
    "OPPORTUNITY_TRACKER": [
        "opp log sheet", "type of opportunity", "external probability",
        "internal complexity", "evaluation level", "term benefit", "away day date",
        "ict estimates", "commercial optimisation", "profit opportunities tracker",
        "account management evaluations", "cash reciepts", "in year profit",
        "existing deal", "new deal", "resource prioritization",
        "contracting actuals", "financial criteria",
    ],
    "SHOP_VISIT": [
        "event item part number", "event item serial number", "action code",
        "rework level", "service event number", "shopvisit_type", "shopvisit_location",
    ],
    "SVRG_MASTER": [
        "trent 900 guarantee", "trent 900 guarantee administration",
        "claims summary", "event entry", "hptb", "svrg", "esvrg",
        "enhanced guarantees",
    ],
    "GLOBAL_HOPPER": [
        "commercial optimisation opportunity report", "global commercial optimisation hopper",
        "crp term benefit", "restructure type", "opportunity maturity",
        "signature ap", "engine value stream", "top level evs",
        "vp/account manager", "onerous/non onerous", "project plan requirements",
        "expected year of signature",
    ],
    "EMPLOYEE_WHEREABOUTS": [
        "employee number", "employee name", "business sector",
        "personal leave inside", "personal leave outside",
        "approved cross border", "approved business t",
        "easter break",
    ],
}

# Hard boosts keyed on lower-case sheet names
_SHEET_NAME_BOOSTS: Dict[str, str] = {
    "soa": "SOA", "soa summary": "SOA", "soa 26.1": "SOA",
    "l2": "OPPORTUNITY_TRACKER", "l3": "OPPORTUNITY_TRACKER",
    "mea log": "OPPORTUNITY_TRACKER", "opps and threats": "OPPORTUNITY_TRACKER",
    "date input": "OPPORTUNITY_TRACKER", "count": "OPPORTUNITY_TRACKER",
    "sum": "OPPORTUNITY_TRACKER", "timeline": "OPPORTUNITY_TRACKER",
    "input": "OPPORTUNITY_TRACKER", "cover": "OPPORTUNITY_TRACKER",
    "menu": "SVRG_MASTER", "claims summary": "SVRG_MASTER", "event entry": "SVRG_MASTER",
    "glossary_2": "SHOP_VISIT",
    "global log": "GLOBAL_HOPPER", "detail_report": "GLOBAL_HOPPER",
    "exec_report": "GLOBAL_HOPPER", "data validations": "GLOBAL_HOPPER",
}

_SHEET_PREFIX_BOOSTS: Dict[str, str] = {
    "report page": "SHOP_VISIT",
    "soa ": "SOA",
}


_COMMERCIAL_PLAN_FILENAME_RE = re.compile(
    r"(?i)\b(\d{4}[_\s]plan|annual\s+plan|commercial\s+plan)\b"
)


# EMPLOYEE_WHEREABOUTS — Middle East regional monthly attendance tracker
_MONTH_YEAR_SHEET_RE = re.compile(
    r"(?i)^\s*(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*\s*[-/ ]?\s*(\d{4})\s*$"
)
_MONTH_NAME_TO_NUM = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dec": 12,
}


def _whereabouts_signals(all_sheets: Dict[str, pd.DataFrame], filename: str = "") -> int:
    """
    Score EMPLOYEE_WHEREABOUTS signals:
      * sheet name matches month-year pattern (e.g. 'Apr 2026') — PRIMARY
      * a sheet has >=3 of the 4 header cells (Employee Number, Employee Name,
        Business Sector, Country) within its first 12 rows — STRONG
      * filename contains 'whereabouts' — SUPPORTING
    Returns an integer score (2+ => definitive EMPLOYEE_WHEREABOUTS).
    """
    hits = 0

    # Month-year sheet name hit
    for sn in all_sheets:
        if _MONTH_YEAR_SHEET_RE.match(sn.strip()):
            hits += 1
            break  # count only once — one such sheet is enough

    # Header-keyword hit: need >=3 of the 4 keywords in same workbook
    required = {"employee number", "employee name", "business sector", "country"}
    found_keywords: set = set()
    for sn, df in all_sheets.items():
        try:
            sample = df.head(12)
            for row in sample.values:
                for v in row:
                    if _is_blank(v):
                        continue
                    t = _clean(v).lower()
                    for kw in required:
                        if kw == t or (kw in t and len(t) < len(kw) + 12):
                            found_keywords.add(kw)
        except Exception:
            continue
    if len(found_keywords) >= 3:
        hits += 1

    # Filename hit
    if filename and "whereabouts" in filename.lower():
        hits += 1

    return hits


def _commercial_plan_signals(all_sheets: Dict[str, pd.DataFrame], filename: str = "") -> int:
    """
    Count COMMERCIAL_PLAN detection hits:
      * sheet name matches 1YP or 5YP (any combination)
      * a sheet contains 'Week Beginning' in its first 10 rows
      * any sheet name matches the annual-plan regex
      * filename matches the annual-plan regex
    Returns an integer score (2+ => definitive COMMERCIAL_PLAN).
    """
    hits = 0
    for sn in all_sheets:
        sn_l = sn.lower().strip()
        if sn_l == "1yp" or sn_l == "5yp" or sn_l.startswith("1yp") or sn_l.startswith("5yp"):
            hits += 1
        if _COMMERCIAL_PLAN_FILENAME_RE.search(sn):
            hits += 1
    for sn, df in all_sheets.items():
        try:
            sample = df.head(10)
            for row in sample.values:
                for v in row:
                    if not _is_blank(v) and "week beginning" in _clean(v).lower():
                        hits += 1
                        break
                else:
                    continue
                break
        except Exception:
            continue
    if filename and _COMMERCIAL_PLAN_FILENAME_RE.search(filename):
        hits += 1
    return hits


def detect_file_type(all_sheets: Dict[str, pd.DataFrame], filename: str = "") -> str:
    # COMMERCIAL_PLAN takes priority when two or more of its signals hit —
    # its sheet names ("1YP", "5YP SPE SALES") are distinctive enough that
    # no other file type can trigger them.
    if _commercial_plan_signals(all_sheets, filename) >= 2:
        return "COMMERCIAL_PLAN"

    # EMPLOYEE_WHEREABOUTS — checked before generic content-signal scoring
    # because its month-year sheet names are very distinctive.
    if _whereabouts_signals(all_sheets, filename) >= 2:
        return "EMPLOYEE_WHEREABOUTS"

    scores: Dict[str, int] = {k: 0 for k in _TYPE_SIGNALS}

    for sheet_name, df in all_sheets.items():
        sn_lower = sheet_name.lower().strip()

        # Sheet-name hard boosts
        if sn_lower in _SHEET_NAME_BOOSTS:
            scores[_SHEET_NAME_BOOSTS[sn_lower]] += 8
        for prefix, ftype in _SHEET_PREFIX_BOOSTS.items():
            if sn_lower.startswith(prefix):
                scores[ftype] += 8

        # Content signal scan (only first 25 rows to keep it fast)
        try:
            sample = df.head(25)
            text_blob = " ".join(
                _clean(v).lower()
                for row in sample.values
                for v in row
                if not _is_blank(v)
            )
            for ftype, signals in _TYPE_SIGNALS.items():
                for sig in signals:
                    if sig in text_blob:
                        scores[ftype] += 1
        except Exception:
            pass

    # Disambiguate: GLOBAL_HOPPER takes priority if "global log" sheet exists
    # (shared sheets like COVER, COUNT, SUM would otherwise boost OPPORTUNITY_TRACKER)
    has_global_log = any("global log" in sn.lower() for sn in all_sheets)
    if has_global_log and scores.get("GLOBAL_HOPPER", 0) > 0:
        return "GLOBAL_HOPPER"

    best = max(scores, key=scores.get)
    return best if scores[best] > 0 else "UNKNOWN"


# ══════════════════════════════════════════════════════════════════════════════
# SOA Parser
# ══════════════════════════════════════════════════════════════════════════════

_SOA_SECTION_KW = [
    "credits usable",
    "totalcare charges",
    "totalcare",
    "familycare charges",           # Rutish: FamilyCare banner → maps to TotalCare group
    "familycare",
    "customer responsible charges",
    "customer responsibility charges",  # Rutish spelling variant
    "customer responsibility",
    "spare parts charges",
    "late payment interest",
    "other charges",
    "other credits",
]

_SOA_SUMMARY_KW = ["total", "overdue", "available credit"]

_SOA_COL_ALIASES: Dict[str, List[str]] = {
    "company_code":        ["company code", "co code", "company"],
    "account":             ["account"],
    "reference":           ["reference", "ref no", "doc no", "document no", "invoice no"],
    "doc_date":            ["document date", "doc date", "posting date", "inv date"],
    "due_date":            ["net due date", "due date", "payment date", "net due"],
    "amount":              ["amount in doc", "amount", "balance", "value"],
    "currency":            ["curr", "currency"],
    "text":                ["text", "description", "narrative", "detail"],
    "assignment":          ["assignment", "assign", "reference key 1", "ref key 1"],
    "rr_comments":         ["r-r comments", "rr comments", "rr note", "comments"],
    "action_owner":        ["action owner", "action", "awaiting approval", "owner"],
    "days_late":           ["days late", "days overdue", "overdue days"],
    "customer_comments":   ["eth comments", "customer comments", "airline comments",
                             "customer note"],
    "po_reference":        ["eth po reference", "eth po", "po reference", "po ref",
                             "purchase order", "po number"],
    "lpi_cumulated":       ["lpi cumulated", "lpi cum", "cumulated lpi"],
}

_SOA_POSITIONAL: Dict[str, int] = {
    "company_code": 0, "account": 1, "reference": 2, "doc_date": 3,
    "due_date": 4, "amount": 5, "currency": 6, "text": 7,
    "assignment": 8, "rr_comments": 9, "action_owner": 10,
    "days_late": 11, "customer_comments": 12, "po_reference": 13,
    "lpi_cumulated": 14,
}


def _map_soa_columns(header_row: pd.Series) -> Dict[str, int]:
    """
    Dynamically map field names to column indices.

    Uses a "first-claim-per-column" strategy: once a column is claimed by a
    field, no other field can steal it.  Fields are processed in dict-insertion
    order (most-specific first), so e.g. 'amount' claims column 5 before
    'currency' can incorrectly grab it via the substring "curr" in
    "Amount in doc. curr.".
    """
    claimed: Dict[int, str] = {}   # col_idx → field that owns it
    mapping: Dict[str, int] = {}

    for field, aliases in _SOA_COL_ALIASES.items():
        for i, val in enumerate(header_row):
            if _is_blank(val) or i in claimed or field in mapping:
                continue
            vl = _clean(val).lower()
            for alias in aliases:
                if alias in vl:
                    mapping[field] = i
                    claimed[i] = field
                    break

    # Fill any missing fields with positional defaults
    for field, idx in _SOA_POSITIONAL.items():
        if field not in mapping:
            mapping[field] = idx
    return mapping


def _get_col(row: pd.Series, col_map: Dict[str, int], field: str) -> Any:
    idx = col_map.get(field)
    if idx is None or idx >= len(row):
        return None
    return row.iloc[idx]


def _soa_is_section_header(row: pd.Series, col_map: Dict[str, int]) -> Optional[str]:
    """Return section label if any early cell is a known section keyword.
    Scans cols 0-3 because Rutish uses merged banners A:D and the value
    lands in col 0 after openpyxl unmerging; ETH SOA uses col 0 plain."""
    for j in range(min(4, len(row))):
        v = row.iloc[j]
        if _is_blank(v) or isinstance(v, (int, float)):
            continue
        s = _clean(v)
        if len(s) > 120:
            continue  # almost certainly free text, not a banner
        s_lower = s.lower()
        for kw in _SOA_SECTION_KW:
            if kw in s_lower:
                return s
    return None


def _soa_is_summary_row(row: pd.Series, col_map: Dict[str, int]) -> Optional[tuple]:
    """
    Detect summary rows that have 'Total', 'Overdue', 'Available Credit' in
    the due_date column (or anywhere in the row).
    Returns (keyword, amount) or None.
    """
    due_val = _get_col(row, col_map, "due_date")
    if not _is_blank(due_val) and isinstance(due_val, str):
        keyword = _clean(due_val).lower()
        for kw in _SOA_SUMMARY_KW:
            if kw in keyword:
                amt = _to_float(_get_col(row, col_map, "amount"))
                return (keyword, amt)
    # Also scan all cells for "total overdue"
    for v in row:
        if not _is_blank(v) and isinstance(v, str):
            vl = _clean(v).lower()
            if "total overdue" in vl:
                # Find adjacent number
                for vv in row:
                    f = _to_float(vv)
                    if f is not None:
                        return ("total overdue", f)
    return None


def _soa_is_data_row(row: pd.Series, col_map: Dict[str, int]) -> bool:
    """Data rows have a numeric amount AND at least one of (reference, doc_no,
    text, account, company_code) non-empty. This relaxes from the prior rule
    that demanded numeric company_code, which silently dropped ~11 rows per
    Ethiopian SOA where company_code is a 4-letter alphabetic code (e.g. 'RAFL').
    """
    amt_raw = _get_col(row, col_map, "amount")
    amt = _to_float(amt_raw)
    if amt is None:
        return False
    # Reject obvious subtotal rows where due_date cell literally says "Total:"
    due_raw = _get_col(row, col_map, "due_date")
    if isinstance(due_raw, str) and "total" in due_raw.lower():
        return False
    # Require at least one identifying field
    ref = _get_col(row, col_map, "reference")
    text = _get_col(row, col_map, "text")
    acct = _get_col(row, col_map, "account")
    comp = _get_col(row, col_map, "company_code")
    if any(not _is_blank(x) for x in (ref, text, acct, comp)):
        return True
    return False


def _soa_extract_item(row: pd.Series, col_map: Dict[str, int]) -> Dict:
    return {
        "company_code":      _to_str_ref(_get_col(row, col_map, "company_code")),
        "account":           _to_str_ref(_get_col(row, col_map, "account")),
        "reference":         _to_str_ref(_get_col(row, col_map, "reference")),
        "doc_date":          _to_date(_get_col(row, col_map, "doc_date")),
        "due_date":          _to_date(_get_col(row, col_map, "due_date")),
        "amount":            _to_float(_get_col(row, col_map, "amount")),
        "currency":          _clean(_get_col(row, col_map, "currency")) or "USD",
        "text":              _clean(_get_col(row, col_map, "text")) or None,
        "assignment":        _clean(_get_col(row, col_map, "assignment")) or None,
        "rr_comments":       _clean(_get_col(row, col_map, "rr_comments")) or None,
        "action_owner":      _clean(_get_col(row, col_map, "action_owner")) or None,
        "days_late":         _to_float(_get_col(row, col_map, "days_late")),
        "customer_comments": _clean(_get_col(row, col_map, "customer_comments")) or None,
        "po_reference":      _clean(_get_col(row, col_map, "po_reference")) or None,
        "lpi_cumulated":     _to_float(_get_col(row, col_map, "lpi_cumulated")),
    }


def _parse_soa(all_sheets: Dict[str, pd.DataFrame], filename: str) -> Dict:
    errors: List[str] = []

    # ── Pick primary data sheet ──────────────────────────────────────────────
    primary_df: Optional[pd.DataFrame] = None
    primary_sheet = ""

    preferred_order = [
        # prefer sheets with 'SoA' but not 'Summary'
        name for name in all_sheets
        if "soa" in name.lower() and "summary" not in name.lower()
    ]
    if not preferred_order:
        # fallback: largest sheet (most non-null cells)
        preferred_order = sorted(
            all_sheets.keys(),
            key=lambda n: all_sheets[n].notna().sum().sum(),
            reverse=True,
        )

    for name in preferred_order:
        df = all_sheets[name]
        text_blob = " ".join(
            _clean(v).lower() for row in df.head(20).values for v in row if not _is_blank(v)
        )
        if any(kw in text_blob for kw in ["customer name", "lpi rate", "totalcare", "amount in doc"]):
            primary_df = df
            primary_sheet = name
            break

    if primary_df is None:
        primary_sheet = preferred_order[0] if preferred_order else list(all_sheets.keys())[0]
        primary_df = all_sheets[primary_sheet]

    rows = [primary_df.iloc[i] for i in range(len(primary_df))]

    # ── Metadata scan (first 15 rows) ────────────────────────────────────────
    metadata: Dict[str, Any] = {
        "title": None,
        "customer_name": None,
        "customer_number": None,
        "contact_email": None,
        "lpi_rate": None,
        "report_date": None,
        "avg_days_late": None,
        "source_file": filename,
        "source_sheet": primary_sheet,
    }

    def _look_right(start_j: int, row: pd.Series) -> Any:
        for k in range(start_j + 1, min(start_j + 8, len(row))):
            v = row.iloc[k]
            if not _is_blank(v):
                return v
        return None

    for row in rows[:15]:
        for j, val in enumerate(row):
            if _is_blank(val):
                continue
            s = _clean(val)
            sl = s.lower()

            if "statement of account" in sl:
                if metadata["title"] is None:
                    metadata["title"] = s
                # Rutish: title row "Rutish Airways Statement of Account with Rolls-Royce"
                # carries customer name in the leading phrase
                if metadata["customer_name"] is None:
                    m = re.match(r"^(.+?)\s+statement of account", s, re.IGNORECASE)
                    if m:
                        metadata["customer_name"] = m.group(1).strip()
            elif ("customer name" in sl or re.match(r"^\s*customer\s*:", sl)) \
                    and metadata["customer_name"] is None:
                v = _look_right(j, row)
                if v is not None:
                    metadata["customer_name"] = _clean(v)
            elif (re.search(r"customer\s+n[oº°#]", sl) or re.match(r"^\s*customer\s*#", sl)) \
                    and "name" not in sl and metadata["customer_number"] is None:
                v = _look_right(j, row)
                if v is not None:
                    metadata["customer_number"] = _clean(v)
            elif ("contact email" in sl or re.match(r"^\s*contact\s*:", sl)) \
                    and metadata["contact_email"] is None:
                v = _look_right(j, row)
                if v is not None:
                    metadata["contact_email"] = _clean(v)
            elif "lpi rate" in sl or re.match(r"^\s*lp\s+ratio", sl):
                for k in range(j + 1, min(j + 8, len(row))):
                    raw = row.iloc[k]
                    if _is_blank(raw):
                        continue
                    # Handle percentage strings like "1.5000%"
                    rs = _clean(raw)
                    if rs.endswith("%"):
                        f = _to_float(rs[:-1])
                        if f is not None:
                            metadata["lpi_rate"] = f / 100.0
                            break
                    f = _to_float(raw)
                    if f is not None:
                        metadata["lpi_rate"] = f
                        break
            elif "today" in sl or re.match(r"^\s*date\s*:", sl) or "report date" in sl:
                for k in range(j + 1, min(j + 8, len(row))):
                    d = _to_date(row.iloc[k])
                    if d:
                        metadata["report_date"] = d
                        break
            elif "average days late" in sl or "avg days late" in sl:
                for k in range(j + 1, min(j + 8, len(row))):
                    f = _to_float(row.iloc[k])
                    if f is not None:
                        metadata["avg_days_late"] = round(f, 2)
                        break

    # ── Discover header row ──────────────────────────────────────────────────
    col_map: Dict[str, int] = dict(_SOA_POSITIONAL)   # start with positional
    header_row_idx: int = 6                            # safe default

    for i, row in enumerate(rows[:20]):
        match_count = sum(
            1 for v in row
            if not _is_blank(v) and any(
                kw in _clean(v).lower()
                for kw in ["reference", "amount", "document", "company code",
                           "days late", "assignment", "net due"]
            )
        )
        if match_count >= 3:
            header_row_idx = i
            col_map = _map_soa_columns(row)
            break

    # ── Section + line-item parsing ──────────────────────────────────────────
    sections: List[Dict] = []
    current_section: Optional[Dict] = None
    grand_totals: Dict[str, Optional[float]] = {
        "total_overdue": None,
        "total_credits": None,
        "net_balance": None,
    }

    def _flush_section():
        nonlocal current_section
        if current_section is not None:
            sections.append(current_section)
            current_section = None

    # Detect if a row looks like a NEW header row (for per-section remap).
    # Rutish SOA has different column layouts per section (FamilyCare has
    # "Net due date", Spare Parts has "Due Date", etc).
    def _is_new_header_row(row: pd.Series) -> bool:
        match_count = sum(
            1 for v in row
            if not _is_blank(v) and any(
                kw in _clean(v).lower()
                for kw in ["reference", "amount", "document", "company",
                           "net due", "due date", "invoice date"]
            )
        )
        return match_count >= 3

    i = header_row_idx + 1
    while i < len(rows):
        row = rows[i]
        nb = _non_blank_vals(row)
        if not nb:
            i += 1
            continue

        # Per-section header remap
        if _is_new_header_row(row):
            col_map = _map_soa_columns(row)
            i += 1
            continue

        # Summary row?
        summary = _soa_is_summary_row(row, col_map)
        if summary:
            kw, amt = summary
            if "total overdue" in kw:
                grand_totals["total_overdue"] = amt
            elif current_section:
                if "total" in kw and "overdue" not in kw:
                    current_section["total"] = amt
                elif "overdue" in kw:
                    current_section["overdue"] = amt
                elif "available credit" in kw:
                    current_section["available_credit"] = amt
            i += 1
            continue

        # Section header?
        sec_name = _soa_is_section_header(row, col_map)
        if sec_name:
            _flush_section()
            current_section = {
                "name": sec_name.rstrip(),
                "section_type": _classify_section(sec_name),
                "items": [],
                "total": None,
                "overdue": None,
                "available_credit": None,
            }
            i += 1
            continue

        # Data row?
        if _soa_is_data_row(row, col_map):
            if current_section is None:
                current_section = {
                    "name": "General",
                    "section_type": "charges",
                    "items": [],
                    "total": None,
                    "overdue": None,
                    "available_credit": None,
                }
            item = _soa_extract_item(row, col_map)
            if item["amount"] is not None:
                current_section["items"].append(item)
            else:
                logger.warning("SOA: row %d dropped — amount could not be parsed", i + 1)
        i += 1

    _flush_section()

    # ── Grand totals ─────────────────────────────────────────────────────────
    if grand_totals["total_overdue"] is None:
        grand_totals["total_overdue"] = round(sum(
            (s["overdue"] or 0) for s in sections if (s["overdue"] or 0) > 0
        ), 2) or None

    all_amounts = [
        it["amount"] for s in sections for it in s["items"] if it["amount"] is not None
    ]
    credits = [a for a in all_amounts if a < 0]
    grand_totals["total_credits"] = round(sum(credits), 2) if credits else 0.0
    grand_totals["net_balance"] = round(sum(all_amounts), 2) if all_amounts else 0.0

    # ── Summary sheet (SoA Summary) ──────────────────────────────────────────
    summary_sheet: Dict[str, float] = {}
    for name, df in all_sheets.items():
        if "summary" in name.lower():
            for _, row in df.iterrows():
                nb_cells = [(j, v) for j, v in enumerate(row) if not _is_blank(v)]
                if len(nb_cells) >= 2:
                    label = _clean(nb_cells[0][1])
                    val = _to_float(nb_cells[-1][1])
                    if label and val is not None:
                        summary_sheet[label] = val
            break

    # ── Aging buckets (7 canonical buckets, days_late OR due_date derived) ──
    aging: Dict[str, float] = {
        "current": 0.0, "1_30_days": 0.0, "31_60_days": 0.0,
        "61_90_days": 0.0, "91_180_days": 0.0, "over_180_days": 0.0,
        "unknown": 0.0,
    }
    # Derive days_late from due_date when the column is missing
    report_dt_iso = metadata.get("report_date")
    report_dt: Optional[datetime] = None
    if report_dt_iso:
        try:
            report_dt = datetime.strptime(report_dt_iso, "%Y-%m-%d")
        except Exception:
            report_dt = None
    if report_dt is None:
        report_dt = datetime.today()

    for sec in sections:
        for it in sec["items"]:
            amt = it.get("amount") or 0.0
            if amt <= 0:
                continue
            d = it.get("days_late")
            if d is None and it.get("due_date"):
                try:
                    due = datetime.strptime(it["due_date"], "%Y-%m-%d")
                    d = (report_dt - due).days
                except Exception:
                    d = None
            if d is None:
                aging["unknown"] += amt
            elif d <= 0:
                aging["current"] += amt
            elif d <= 30:
                aging["1_30_days"] += amt
            elif d <= 60:
                aging["31_60_days"] += amt
            elif d <= 90:
                aging["61_90_days"] += amt
            elif d <= 180:
                aging["91_180_days"] += amt
            else:
                aging["over_180_days"] += amt
    aging = {k: round(v, 2) for k, v in aging.items()}

    # Human-readable breakdown matching visualizer AGING_ORDER constant
    aging_breakdown = {
        "Current": aging["current"],
        "1-30 Days": aging["1_30_days"],
        "31-60 Days": aging["31_60_days"],
        "61-90 Days": aging["61_90_days"],
        "91-180 Days": aging["91_180_days"],
        "180+ Days": aging["over_180_days"],
        "Unknown": aging["unknown"],
    }

    # ── Auxiliary sheets (Offset, Paymen/Payment, 2022 Cash, 2022 Credit) ──
    aux: Dict[str, Any] = {}
    for sheet_name, df in all_sheets.items():
        sl = sheet_name.lower().strip()
        # Skip the primary data sheet and the summary sheet already processed
        if sheet_name == primary_sheet:
            continue
        if "summary" in sl:
            continue
        if sheet_name in aux:
            continue
        # Detect header in first 10 rows
        try:
            hdr_i = 0
            for i in range(min(10, len(df))):
                nb = _non_blank_vals(df.iloc[i])
                if len(nb) >= 2 and any(kw in " ".join(nb).lower()
                                        for kw in ["date", "reference", "amount", "currency",
                                                   "transaction", "method"]):
                    hdr_i = i
                    break
            hdr = df.iloc[hdr_i] if hdr_i < len(df) else None
            if hdr is None:
                continue
            eff_width = _last_nonempty_col(hdr)
            headers = [_clean(hdr.iloc[j]) if j < len(hdr) else "" for j in range(eff_width)]
            # Drop leading/trailing empties
            aux_rows: List[Dict] = []
            aux_totals: Dict[str, float] = {}
            for i in range(hdr_i + 1, len(df)):
                r = df.iloc[i]
                if not _non_blank_vals(r):
                    continue
                rec = {}
                for j in range(min(eff_width, len(r))):
                    v = r.iloc[j]
                    if _is_blank(v):
                        continue
                    h = headers[j] or f"col_{j}"
                    if isinstance(v, (datetime, date)):
                        rec[h] = v.strftime("%Y-%m-%d")
                    elif isinstance(v, float):
                        rec[h] = None if v != v else v
                    else:
                        rec[h] = _clean(v) or None
                if rec:
                    aux_rows.append(rec)
            # Summary: numeric column sums
            for j in range(eff_width):
                h = headers[j] or f"col_{j}"
                nums = []
                for i in range(hdr_i + 1, len(df)):
                    r = df.iloc[i]
                    if j < len(r):
                        f = _to_float(r.iloc[j])
                        if f is not None:
                            nums.append(f)
                if nums:
                    aux_totals[h] = round(sum(nums), 2)
            aux[sheet_name] = {
                "header_row": hdr_i + 1,   # 1-based for humans
                "headers": headers,
                "row_count": len(aux_rows),
                "rows": aux_rows,
                "summary": aux_totals,
            }
        except Exception as e:
            logger.warning("SOA: failed to parse auxiliary sheet '%s': %s", sheet_name, e)

    return {
        "file_type": "SOA",
        "metadata": metadata,
        "sections": sections,
        "grand_totals": grand_totals,
        "summary_sheet": summary_sheet,
        "aging_buckets": aging,             # legacy key (snake_case, 7 buckets)
        "aging_breakdown": aging_breakdown,  # new canonical key (matches AGING_ORDER)
        "auxiliary_sheets": aux,
        "all_sheets": list(all_sheets.keys()),
        "errors": errors,
    }


def _classify_section(name: str) -> str:
    nl = name.lower()
    if "credit" in nl:
        return "credits"
    if "totalcare" in nl or "familycare" in nl:   # FamilyCare → TotalCare group (Rutish)
        return "totalcare"
    if "spare" in nl or "parts" in nl:
        return "spare_parts"
    if "late payment" in nl or "lpi" in nl or "interest" in nl:
        return "lpi"
    if "customer responsibility" in nl or "customer responsible" in nl or "crc" in nl:
        return "crc"
    return "charges"


# ══════════════════════════════════════════════════════════════════════════════
# Invoice List Parser (EPI style)
# ══════════════════════════════════════════════════════════════════════════════

_INV_COL_ALIASES: Dict[str, List[str]] = {
    "reference":      ["reference", "doc no", "invoice no", "document no"],
    "doc_date":       ["document date", "doc date", "posting date"],
    "due_date":       ["net due date", "due date", "payment date"],
    "currency":       ["document currency", "currency", "curr"],
    "amount":         ["amount in doc", "amount", "value", "balance"],
    "reference_key3": ["reference key 3", "ref key 3", "ref 3"],
    "text":           ["text", "description", "narrative"],
    "assignment":     ["assignment", "assign", "ref key 1"],
}


def _find_header_row(df: pd.DataFrame, required_kws: List[str], max_scan: int = 20) -> int:
    """Return 0-indexed row index of the row that best matches required keywords."""
    best_idx, best_score = 0, 0
    for i in range(min(max_scan, len(df))):
        row = df.iloc[i]
        score = sum(
            1 for v in row
            if not _is_blank(v) and any(kw in _clean(v).lower() for kw in required_kws)
        )
        if score > best_score:
            best_score, best_idx = score, i
    return best_idx


def _map_generic_columns(header_row: pd.Series, aliases: Dict[str, List[str]]) -> Dict[str, int]:
    """
    Same first-claim-per-column strategy as _map_soa_columns:
    once column i is owned by a field, no other field can take it.
    """
    claimed: Dict[int, str] = {}
    mapping: Dict[str, int] = {}
    for field, kws in aliases.items():
        for i, val in enumerate(header_row):
            if _is_blank(val) or i in claimed or field in mapping:
                continue
            vl = _clean(val).lower()
            for kw in kws:
                if kw in vl:
                    mapping[field] = i
                    claimed[i] = field
                    break
    return mapping


def _parse_invoice_list(all_sheets: Dict[str, pd.DataFrame], filename: str) -> Dict:
    errors: List[str] = []
    # Use the first / largest sheet
    sheet_name = sorted(all_sheets.keys(), key=lambda n: all_sheets[n].notna().sum().sum(), reverse=True)[0]
    df = all_sheets[sheet_name]

    hdr_idx = _find_header_row(df, ["reference", "amount", "document", "net due", "currency"])
    col_map = _map_generic_columns(df.iloc[hdr_idx], _INV_COL_ALIASES)

    items: List[Dict] = []
    subtotals: List[Dict] = []
    for i in range(hdr_idx + 1, len(df)):
        row = df.iloc[i]
        if not _non_blank_vals(row):
            continue
        ref      = _to_str_ref(row.iloc[col_map.get("reference", 0)])
        doc_date = _to_date(row.iloc[col_map["doc_date"]] if "doc_date" in col_map else None)
        due_date = _to_date(row.iloc[col_map["due_date"]] if "due_date" in col_map else None)
        text     = _clean(row.iloc[col_map["text"]] if "text" in col_map else None) or None
        assign   = _clean(row.iloc[col_map["assignment"]] if "assignment" in col_map else None) or None
        amt      = _to_float(row.iloc[col_map.get("amount", 4)])

        if amt is None and ref is None:
            continue

        # Pure-total/subtotal rows: amount-only, all identifying fields are blank.
        # These are Excel formatting rows (colored running totals) — skip them.
        is_pure_total = (
            ref is None
            and doc_date is None
            and due_date is None
            and text is None
            and assign is None
        )
        if is_pure_total:
            if amt is not None:
                subtotals.append({"row_index": i + 1, "amount": amt})
            continue

        items.append({
            "reference":      ref,
            "doc_date":       doc_date,
            "due_date":       due_date,
            "currency":       _clean(row.iloc[col_map["currency"]] if "currency" in col_map else None) or "USD",
            "amount":         amt,
            "reference_key3": _clean(row.iloc[col_map["reference_key3"]] if "reference_key3" in col_map else None) or None,
            "text":           text,
            "assignment":     assign,
        })

    pos_amounts = [it["amount"] for it in items if it["amount"] and it["amount"] > 0]
    neg_amounts = [it["amount"] for it in items if it["amount"] and it["amount"] < 0]

    # Aging breakdown (7 canonical buckets) derived from net_due_date vs today
    today = datetime.today()
    aging_breakdown = {
        "Current": 0.0, "1-30 Days": 0.0, "31-60 Days": 0.0, "61-90 Days": 0.0,
        "91-180 Days": 0.0, "180+ Days": 0.0, "Unknown": 0.0,
    }
    for it in items:
        amt = it.get("amount") or 0.0
        if amt <= 0:
            continue
        due_iso = it.get("due_date")
        if not due_iso:
            aging_breakdown["Unknown"] += amt
            continue
        try:
            due = datetime.strptime(due_iso, "%Y-%m-%d")
            d = (today - due).days
        except Exception:
            aging_breakdown["Unknown"] += amt
            continue
        if d <= 0:        aging_breakdown["Current"] += amt
        elif d <= 30:     aging_breakdown["1-30 Days"] += amt
        elif d <= 60:     aging_breakdown["31-60 Days"] += amt
        elif d <= 90:     aging_breakdown["61-90 Days"] += amt
        elif d <= 180:    aging_breakdown["91-180 Days"] += amt
        else:             aging_breakdown["180+ Days"] += amt
    aging_breakdown = {k: round(v, 2) for k, v in aging_breakdown.items()}

    return {
        "file_type": "INVOICE_LIST",
        "metadata": {
            "source_file": filename,
            "source_sheet": sheet_name,
            "total_items": len(items),
            "currencies": list({it["currency"] for it in items if it["currency"]}),
        },
        "items": items,
        "totals": {
            "total_amount":   round(sum(it["amount"] for it in items if it["amount"]), 2),
            "total_positive": round(sum(pos_amounts), 2),
            "total_negative": round(sum(neg_amounts), 2),
            "item_count":     len(items),
        },
        "aging_breakdown": aging_breakdown,
        # Subtotal/running-total rows found in the file (excluded from items).
        # These are the coloured formatting rows (e.g. orange, gold, green totals).
        "sheet_subtotals": subtotals,
        "all_sheets": list(all_sheets.keys()),
        "errors": errors,
    }


# ══════════════════════════════════════════════════════════════════════════════
# Opportunity Tracker Parser (MEA Commercial Optimisation style)
# ══════════════════════════════════════════════════════════════════════════════
#
# Handles the full MEA Profit Opportunities Tracker workbook:
#   • MEA LOG  — Account Management level estimates (hopper)
#   • L2       — ICT-evaluated estimates
#   • L3       — Contract-level estimates
#   • Summary  — Project-level strategic/financial overview
#   • Opps and Threats — Pre-existing opportunities & 5YP improvement
#   • Timeline / Date Input — Project milestone timelines
#   • COUNT    — Analytics/aggregation by customer
#   • INPUT    — Reference/config data (dropdowns, weights, statuses)
#   • SUM      — Calculator/filter sheet
#   • COVER    — Report cover metadata
#   • Sheet1   — Lookup/reference data
# ══════════════════════════════════════════════════════════════════════════════

_OPP_HDR_SIGNALS = [
    "#", "project", "programme", "customer", "region", "asks",
    "type of opportunity", "priority", "status", "term benefit",
    "external probability", "internal complexity",
]

# Core columns present in MEA LOG, L2, and L3 sheets (cols 2-21)
# "benefit_2026"/"benefit_2027" use exact-match via trailing space/pipe
# to avoid catching header cells that merely contain the substring "2026".
_OPP_CORE_COLS: Dict[str, List[str]] = {
    "number":               ["#"],
    "project":              ["project"],
    "programme":            ["programme", "program"],
    "customer":             ["customer"],
    "region":               ["region"],
    "asks":                 ["asks"],
    "opportunity_type":     ["type of opportunity"],
    "levers":               ["levers"],
    "priority":             ["priority"],
    "spe_related":          ["spe related"],
    "num_spe":              ["no,of spe", "no. of spe", "no of spe"],
    "crp_pct":              ["crp%", "crp %"],
    "ext_probability":      ["external probability"],
    "int_complexity":       ["internal complexity"],
    "status":               ["status"],
    "evaluation_level":     ["evaluation level"],
    "term_benefit":         ["term benefit"],
    "benefit_2026":         ["|2026|"],  # exact-match sentinel (handled specially below)
    "benefit_2027":         ["|2027|"],
    "sum_26_27":            ["sum of 26/27", "sum 26/27", "26/27"],
}


def _map_opp_columns(header_row: pd.Series, aliases: Dict[str, List[str]]) -> Dict[str, int]:
    """Opp-log column mapper that exactly-matches bare year headers like '2026'
    (to avoid snagging a sibling '=SUM(...)' cell or "Cash Receipts 2026-2030"
    style merged-banner)."""
    claimed: Dict[int, str] = {}
    mapping: Dict[str, int] = {}
    # Pass 1: exact-match numeric-year fields
    for field, kws in aliases.items():
        for kw in kws:
            if not (kw.startswith("|") and kw.endswith("|")):
                continue
            year = kw.strip("|")
            for i, val in enumerate(header_row):
                if _is_blank(val) or i in claimed or field in mapping:
                    continue
                s = _clean(val).lower()
                if s == year:
                    mapping[field] = i
                    claimed[i] = field
                    break
    # Pass 2: regular substring matching for non-sentinel aliases
    for field, kws in aliases.items():
        if field in mapping:
            continue
        for i, val in enumerate(header_row):
            if _is_blank(val) or i in claimed:
                continue
            vl = _clean(val).lower()
            for kw in kws:
                if kw.startswith("|") and kw.endswith("|"):
                    continue
                if kw in vl:
                    mapping[field] = i
                    claimed[i] = field
                    break
            if field in mapping:
                break
    return mapping

# Supporting financial info columns (cols ~51-58)
_OPP_SUPPORT_FIN_COLS: Dict[str, List[str]] = {
    "deal_benefits":        ["deal benefits"],
    "expected_deal_costs":  ["expected deal costs"],
    "inyear_profit_impact": ["in-year profit impact", "in year profit impact"],
    "fyp_profit_improvement": ["5yp profit improvement", "5yp improvement"],
    "term_profit_improvement": ["term profit impr"],
    "total_crp_term_revenue":  ["total crp term rev"],
    "total_crp_term_margin":   ["total crp term margin"],
    "crp_margin_pct":       ["crp margin %"],
}

# "To Go" columns (cols ~60-62)
_OPP_TOGO_COLS: Dict[str, List[str]] = {
    "togo_term_revenue":    ["term revenue"],
    "togo_term_cost":       ["term cost"],
    "togo_term_profit":     ["term profit"],
}

# Resource Prioritisation columns (cols ~66-71)
_OPP_RESOURCE_COLS: Dict[str, List[str]] = {
    "res_account_mgmt":       ["account management"],
    "res_contract_mgmt":      ["contract management"],
    "res_service_business":   ["service business"],
    "res_business_evaluation":["business eval"],
    "res_sales_contracting":  ["sales & contracting", "sales and contracting"],
    "res_customer_ops":       ["customer operations"],
}

# ── Financial column layout in L2/L3/MEA LOG (cols 25-49) ──────────────────
# The wide financial area has 4 groups, each with years 2025-2030:
#   cols 25-30: Existing Deal — Cash Receipts
#   cols 31-36: Existing Deal — In Year Profit
#   cols 38-43: New Deal — Cash Receipts
#   cols 44-49: New Deal — In Year Profit
_FINANCIAL_GROUPS = [
    ("existing_deal_cash",   25, 30),   # Existing Deal Cash Receipts
    ("existing_deal_profit", 31, 36),   # Existing Deal In Year Profit
    ("new_deal_cash",        38, 43),   # New Deal Cash Receipts
    ("new_deal_profit",      44, 49),   # New Deal In Year Profit
]
_FINANCIAL_YEARS = [2025, 2026, 2027, 2028, 2029, 2030]


def _classify_opp_sheet(sheet_name: str) -> str:
    """Map sheet name → estimation level label."""
    sn = sheet_name.lower().strip()
    if sn == "l2":
        return "ICT"
    elif sn == "l3":
        return "Contract"
    elif "mea log" in sn or "opp log" in sn:
        return "Hopper"
    return "Unknown"


def _extract_opp_sums_row(df: pd.DataFrame) -> Dict[str, Optional[float]]:
    """Extract the SUMS row (row 11 typically) from an opp log sheet."""
    sums: Dict[str, Optional[float]] = {}
    for i in range(min(15, len(df))):
        row = df.iloc[i]
        for j, val in enumerate(row):
            if not _is_blank(val) and "SUMS" in _clean(val).upper():
                # Found the sums row — extract known positions
                sums["term_benefit_sum"] = _to_float(row.iloc[18]) if 18 < len(row) else None
                sums["sum_2026"] = _to_float(row.iloc[19]) if 19 < len(row) else None
                sums["sum_2027"] = _to_float(row.iloc[20]) if 20 < len(row) else None
                sums["sum_26_27"] = _to_float(row.iloc[21]) if 21 < len(row) else None
                return sums
    return sums


def _extract_financial_breakdown(row: pd.Series) -> Dict[str, Dict[str, Optional[float]]]:
    """
    Extract the 4-group financial breakdown from a wide opp row.
    Returns {group_name: {yr_2025: val, yr_2026: val, ...}}.
    """
    fin: Dict[str, Dict[str, Optional[float]]] = {}
    for group_name, start_col, end_col in _FINANCIAL_GROUPS:
        year_vals: Dict[str, Optional[float]] = {}
        for offset, year in enumerate(_FINANCIAL_YEARS):
            col_idx = start_col + offset
            if col_idx <= end_col and col_idx < len(row):
                year_vals[f"yr_{year}"] = _to_float(row.iloc[col_idx])
            else:
                year_vals[f"yr_{year}"] = None
        fin[group_name] = year_vals
    return fin


def _extract_supporting_financials(row: pd.Series, col_map: Dict[str, int]) -> Dict[str, Optional[float]]:
    """Extract supporting financial info columns (Deal Benefits, etc.)."""
    return {
        field: _to_float(_get_generic(row, col_map, field))
        for field in _OPP_SUPPORT_FIN_COLS
    }


def _extract_togo(row: pd.Series, col_map: Dict[str, int]) -> Dict[str, Optional[float]]:
    """Extract 'To Go' columns."""
    return {
        field: _to_float(_get_generic(row, col_map, field))
        for field in _OPP_TOGO_COLS
    }


def _extract_resource_priority(row: pd.Series, col_map: Dict[str, int]) -> Dict[str, Any]:
    """Extract resource prioritisation columns."""
    return {
        field: _to_float(_get_generic(row, col_map, field))
        for field in _OPP_RESOURCE_COLS
    }


def _parse_opp_log_sheet(df: pd.DataFrame, sheet_name: str) -> Dict:
    """
    Parse a single opportunity log sheet (MEA LOG, L2, or L3).
    These all share the same 83-column layout.

    Critical: the real header row is row 14 (1-based) / index 13. Row 3 is a
    merged "OPP LOG SHEET" banner that tricks naive header detection. We
    require at least 5 core-column signals so the title banner cannot win.
    """
    estimation_level = _classify_opp_sheet(sheet_name)
    sums = _extract_opp_sums_row(df)

    # Find header row with strong signal requirement
    hdr_idx = None
    best_score = 0
    required = ["project", "customer", "region", "programme", "term benefit",
                "priority", "status", "asks", "type of opportunity",
                "external probability", "internal complexity"]
    for i in range(min(30, len(df))):
        row = df.iloc[i]
        score = sum(
            1 for v in row
            if not _is_blank(v) and any(kw in _clean(v).lower() for kw in required)
        )
        if score > best_score:
            best_score = score
            hdr_idx = i
    # Fallback to index 13 if no strong match found (documented row 14 1-based)
    if hdr_idx is None or best_score < 5:
        hdr_idx = 13 if len(df) > 14 else (hdr_idx or 0)
    hdr_row = df.iloc[hdr_idx]

    # Map core columns — use exact-match for year headers
    col_map = _map_opp_columns(hdr_row, _OPP_CORE_COLS)

    # Map supporting financial, to-go, and resource columns
    support_map = _map_generic_columns(hdr_row, _OPP_SUPPORT_FIN_COLS)
    togo_map = _map_generic_columns(hdr_row, _OPP_TOGO_COLS)
    resource_map = _map_generic_columns(hdr_row, _OPP_RESOURCE_COLS)

    records: List[Dict] = []
    for i in range(hdr_idx + 1, len(df)):
        row = df.iloc[i]
        if not _non_blank_vals(row):
            continue
        num_val = _to_float(_get_generic(row, col_map, "number"))
        if num_val is None:
            continue

        # The spreadsheet pre-populates row numbers 1-500 but most are
        # empty template rows.  Require at least one key data field to
        # be populated (project, customer, asks, or status).
        project_v  = _clean(_get_generic(row, col_map, "project")) or None
        customer_v = _clean(_get_generic(row, col_map, "customer")) or None
        asks_v     = _clean(_get_generic(row, col_map, "asks")) or None
        status_v   = _clean(_get_generic(row, col_map, "status")) or None
        if not any([project_v, customer_v, asks_v, status_v]):
            continue

        rec: Dict[str, Any] = {
            "number":           int(num_val),
            "project":          project_v,
            "programme":        _clean(_get_generic(row, col_map, "programme")) or None,
            "customer":         customer_v,
            "region":           _clean(_get_generic(row, col_map, "region")) or None,
            "asks":             asks_v,
            "opportunity_type": _clean(_get_generic(row, col_map, "opportunity_type")) or None,
            "levers":           _clean(_get_generic(row, col_map, "levers")) or None,
            "priority":         _to_float(_get_generic(row, col_map, "priority")),
            "spe_related":      _clean(_get_generic(row, col_map, "spe_related")) or None,
            "num_spe":          _to_float(_get_generic(row, col_map, "num_spe")),
            "crp_pct":          _to_float(_get_generic(row, col_map, "crp_pct")),
            "ext_probability":  _clean(_get_generic(row, col_map, "ext_probability")) or None,
            "int_complexity":   _clean(_get_generic(row, col_map, "int_complexity")) or None,
            "status":           status_v,
            "evaluation_level": _clean(_get_generic(row, col_map, "evaluation_level")) or None,
            "term_benefit":     _to_float(_get_generic(row, col_map, "term_benefit")),
            "benefit_2026":     _to_float(_get_generic(row, col_map, "benefit_2026")),
            "benefit_2027":     _to_float(_get_generic(row, col_map, "benefit_2027")),
            "benefit_2026_note": (_clean(_get_generic(row, col_map, "benefit_2026"))
                                  if _to_float(_get_generic(row, col_map, "benefit_2026")) is None
                                  and not _is_blank(_get_generic(row, col_map, "benefit_2026"))
                                  else None),
            "benefit_2027_note": (_clean(_get_generic(row, col_map, "benefit_2027"))
                                  if _to_float(_get_generic(row, col_map, "benefit_2027")) is None
                                  and not _is_blank(_get_generic(row, col_map, "benefit_2027"))
                                  else None),
            "sum_26_27":        _to_float(_get_generic(row, col_map, "sum_26_27")),
            # Structured financial breakdown
            "financials":       _extract_financial_breakdown(row),
            # Supporting financial info
            "supporting_financials": _extract_supporting_financials(row, support_map),
            # To Go
            "to_go":            _extract_togo(row, togo_map),
            # Resource prioritisation
            "resource_priority": _extract_resource_priority(row, resource_map),
        }
        records.append(rec)

    return {
        "estimation_level": estimation_level,
        "sheet_name": sheet_name,
        "sums": sums,
        "records": records,
    }


def _parse_opp_summary_sheet(df: pd.DataFrame) -> Dict:
    """
    Parse the Summary sheet — project-level strategic/financial overview.

    Layout:
      Row 1: section labels ("Strategic", "Financials")
      Row 2: headers (Project, Customer, Programme, KAM Pack Complete, ...)
      Row 3: sub-headers for 5YP years (2025-2029)
      Row 4+: data rows with "Launched" / "Pending" group labels in col 1
    """
    projects: List[Dict] = []
    current_group = None

    # Find header row (row with "Project", "Customer", "Programme")
    hdr_idx = _find_header_row(df, ["project", "customer", "programme", "crp", "onerous"], max_scan=5)
    hdr_row = df.iloc[hdr_idx]

    _SUMMARY_COLS: Dict[str, List[str]] = {
        "project":          ["project"],
        "customer":         ["customer"],
        "programme":        ["programme", "program"],
        "kam_pack":         ["kam pack"],
        "crp_deep_dive":    ["crp deep dive"],
        "risk_review":      ["risk review"],
        "contract_dcr":     ["contract dcr"],
        "ov_review":        ["o&v review"],
        "current_crp_margin":  ["current crp margin"],
        "onerous_provision":   ["onerous provision"],
        "current_crp_pct":     ["current crp %"],
        "onerous_2024":        ["2024 onerous"],
        "onerous_2025":        ["2025 onerous"],
        "overall_pack_improvement": ["overall pack improvement"],
    }
    col_map = _map_generic_columns(hdr_row, _SUMMARY_COLS)

    # 5YP year sub-header row (usually hdr_idx + 1)
    fyp_start_col = 15   # columns 15-19 = 2025..2029
    fyp_years = [2025, 2026, 2027, 2028, 2029]

    for i in range(hdr_idx + 1, len(df)):
        row = df.iloc[i]
        nb = _non_blank_vals(row)
        if not nb:
            continue

        # Check for group label in col 1
        v1 = _clean(row.iloc[1]) if 1 < len(row) else ""
        if v1.lower() in ("launched", "pending"):
            current_group = v1
            # May also have data in same row
            proj = _clean(_get_generic(row, col_map, "project"))
            if not proj:
                continue

        proj = _clean(_get_generic(row, col_map, "project"))
        cust = _clean(_get_generic(row, col_map, "customer"))
        prog = _clean(_get_generic(row, col_map, "programme"))

        # Skip sub-header rows (years row)
        if not proj and not cust and not prog:
            # Could be a continuation row with only programme
            if prog:
                pass
            else:
                continue

        # 5YP improvement by year
        fyp: Dict[str, Optional[float]] = {}
        for offset, year in enumerate(fyp_years):
            col_idx = fyp_start_col + offset
            if col_idx < len(row):
                fyp[f"yr_{year}"] = _to_float(row.iloc[col_idx])

        rec = {
            "group":             current_group,
            "project":           proj or None,
            "customer":          cust or None,
            "programme":         prog or None,
            "kam_pack_complete":      _to_date(_get_generic(row, col_map, "kam_pack")),
            "crp_deep_dive_complete": _to_date(_get_generic(row, col_map, "crp_deep_dive")),
            "risk_review_complete":   _to_date(_get_generic(row, col_map, "risk_review")),
            "contract_dcr_complete":  _to_date(_get_generic(row, col_map, "contract_dcr")),
            "ov_review_complete":     _to_date(_get_generic(row, col_map, "ov_review")),
            "current_crp_margin":     _to_float(_get_generic(row, col_map, "current_crp_margin")),
            "onerous_provision":      _to_float(_get_generic(row, col_map, "onerous_provision")),
            "current_crp_pct":        _to_float(_get_generic(row, col_map, "current_crp_pct")),
            "onerous_release_2024":   _to_float(_get_generic(row, col_map, "onerous_2024")),
            "onerous_release_2025":   _to_float(_get_generic(row, col_map, "onerous_2025")),
            "fyp_improvement":        fyp,
            "overall_pack_improvement": _to_float(_get_generic(row, col_map, "overall_pack_improvement")),
        }
        projects.append(rec)

    return {"projects": projects}


def _parse_opp_opps_and_threats(df: pd.DataFrame) -> Dict:
    """
    Parse the 'Opps and Threats' sheet — pre-existing opportunities with
    profit releases, 5YP improvement, owner, status, and comments.
    """
    _OT_COLS: Dict[str, List[str]] = {
        "project":              ["project"],
        "programme":            ["programme", "program"],
        "customer":             ["customer"],
        "opportunity":          ["opportunity"],
        "profit_release_2024":  ["2024 profit release"],
        "profit_release_2025":  ["2025 profit release"],
        "overall_pack_improvement": ["overall pack improvement"],
        "window_for_forecast":  ["window for forecast"],
        "due_date":             ["due date"],
        "owner":                ["owner"],
        "status":               ["status"],
        "comments":             ["comments"],
    }

    # Row 0 has totals, Row 1 has headers, Row 2 has "5YP Improvement" sub-header
    hdr_idx = _find_header_row(df, ["project", "customer", "opportunity", "owner", "due date"], max_scan=5)
    hdr_row = df.iloc[hdr_idx]
    col_map = _map_generic_columns(hdr_row, _OT_COLS)

    # 5YP year columns (2025-2029 in cols 7-11)
    fyp_year_cols: Dict[str, int] = {}
    for i, val in enumerate(hdr_row):
        if _is_blank(val):
            continue
        vl = _clean(val).lower()
        m = re.match(r"^(20\d{2})(\.0)?$", vl)
        if m:
            year = int(m.group(1))
            fyp_year_cols[f"fyp_{year}"] = i

    # Extract totals from row 0
    totals_row = df.iloc[0] if len(df) > 0 else None
    totals: Dict[str, Optional[float]] = {}
    if totals_row is not None:
        for i, val in enumerate(totals_row):
            f = _to_float(val)
            if f is not None:
                totals[f"col_{i}"] = f

    records: List[Dict] = []
    for i in range(hdr_idx + 1, len(df)):
        row = df.iloc[i]
        if not _non_blank_vals(row):
            continue

        proj = _clean(_get_generic(row, col_map, "project")) or None
        prog = _clean(_get_generic(row, col_map, "programme")) or None
        cust = _clean(_get_generic(row, col_map, "customer")) or None
        opp  = _clean(_get_generic(row, col_map, "opportunity")) or None

        # Skip rows that have nothing meaningful
        if not any([proj, prog, cust, opp]):
            continue

        # 5YP improvement by year
        fyp: Dict[str, Optional[float]] = {}
        for key, col_idx in fyp_year_cols.items():
            if col_idx < len(row):
                fyp[key] = _to_float(row.iloc[col_idx])

        rec = {
            "project":               proj,
            "programme":             prog,
            "customer":              cust,
            "opportunity":           opp,
            "profit_release_2024":   _to_float(_get_generic(row, col_map, "profit_release_2024")),
            "profit_release_2025":   _to_float(_get_generic(row, col_map, "profit_release_2025")),
            "fyp_improvement":       fyp,
            "overall_pack_improvement": _to_float(_get_generic(row, col_map, "overall_pack_improvement")),
            "window_for_forecast":   _clean(_get_generic(row, col_map, "window_for_forecast")) or None,
            "due_date":              _to_date(_get_generic(row, col_map, "due_date")),
            "owner":                 _clean(_get_generic(row, col_map, "owner")) or None,
            "status":                _clean(_get_generic(row, col_map, "status")) or None,
            "comments":              _clean(_get_generic(row, col_map, "comments")) or None,
        }
        records.append(rec)

    return {"totals": totals, "items": records}


def _parse_opp_timeline(all_sheets: Dict[str, pd.DataFrame]) -> Dict:
    """
    Parse Timeline + Date Input sheets into a unified milestone map.

    Date Input sheet (preferred): clean table with columns:
      Project | Customer | Idea Generation | Approval to Launch | Strategy Approval |
      BE Generated | Approval | Negotiation Strategy | Proposal Submitted | Proposal Signed

    Timeline sheet: Gantt-style with week columns — milestones placed in cells.
    """
    milestones: List[Dict] = []

    # ── Date Input (structured, preferred) ───────────────────────────────
    for name in ("Date Input", "date input"):
        if name not in all_sheets:
            continue
        df = all_sheets[name]
        hdr_idx = _find_header_row(df, ["project", "customer", "idea generation", "approval"], max_scan=5)
        hdr_row = df.iloc[hdr_idx]

        _DI_COLS: Dict[str, List[str]] = {
            "project":              ["project"],
            "customer":             ["customer"],
            "idea_generation":      ["idea generation"],
            "approval_to_launch":   ["approval to launch"],
            "strategy_approval":    ["strategy approval"],
            "be_generated":         ["be generated"],
            "approval":             ["approval"],
            "negotiation_strategy": ["negotiation strategy"],
            "proposal_submitted":   ["proposal submitted"],
            "proposal_signed":      ["proposal signed"],
        }
        col_map = _map_generic_columns(hdr_row, _DI_COLS)

        _MILESTONE_ORDER = [
            "idea_generation", "approval_to_launch", "strategy_approval",
            "be_generated", "approval", "negotiation_strategy",
            "proposal_submitted", "proposal_signed",
        ]

        for i in range(hdr_idx + 1, len(df)):
            row = df.iloc[i]
            if not _non_blank_vals(row):
                continue
            proj = _clean(_get_generic(row, col_map, "project")) or None
            cust = _clean(_get_generic(row, col_map, "customer")) or None
            if not proj and not cust:
                continue

            ms_dates: Dict[str, Optional[str]] = {}
            for ms_key in _MILESTONE_ORDER:
                ms_dates[ms_key] = _to_date(_get_generic(row, col_map, ms_key))

            # Determine current phase based on latest non-null milestone
            current_phase = None
            for ms_key in reversed(_MILESTONE_ORDER):
                if ms_dates.get(ms_key):
                    current_phase = ms_key
                    break

            milestones.append({
                "project":       proj,
                "customer":      cust,
                "milestones":    ms_dates,
                "current_phase": current_phase,
                "source":        "Date Input",
            })
        break   # only parse once

    # ── Timeline sheet (Gantt-style, supplementary) ──────────────────────
    for name in all_sheets:
        if name.lower().strip() != "timeline":
            continue
        df = all_sheets[name]
        # Timeline row 13 has Project | Customer | week-date headers
        # Data rows have project/customer in cols 2-3 and milestone labels in week columns
        hdr_idx = _find_header_row(df, ["project", "customer"], max_scan=15)
        if hdr_idx < 2:
            hdr_idx = 13   # known position

        for i in range(hdr_idx + 1, min(hdr_idx + 50, len(df))):
            row = df.iloc[i]
            if not _non_blank_vals(row):
                continue
            proj = _clean(row.iloc[2]) if 2 < len(row) else None
            cust = _clean(row.iloc[3]) if 3 < len(row) else None
            if not proj:
                continue

            # Collect milestone labels from week columns
            timeline_milestones: List[Dict] = []
            for j in range(4, len(row)):
                val = _clean(row.iloc[j])
                if val and val.lower() not in ("false", "true", "0", "nan"):
                    timeline_milestones.append({
                        "column_index": j,
                        "milestone_label": val,
                    })

            if timeline_milestones:
                milestones.append({
                    "project":            proj,
                    "customer":           cust,
                    "timeline_milestones": timeline_milestones,
                    "source":             "Timeline",
                })
        break

    return {"milestones": milestones}


def _parse_opp_count_sheet(df: pd.DataFrame) -> Dict:
    """
    Parse the COUNT sheet — analytics/aggregation breakdowns by customer.

    Layout:
      Row 7: "PROJECTS" label
      Row 8: section headers (PROGRAM, PROBABILITY, COMPLEXITY, STATUS, YEARS)
      Row 9: column headers
      Row 10+: data rows by customer
    """
    _COUNT_COLS: Dict[str, List[str]] = {
        "customer":           ["project"],    # labelled "PROJECT" but contains customer names
        "count_of_asks":      ["count of asks"],
        "spe_related":        ["of which spe", "spe related"],
        "num_spe":            ["no.of spe", "no of spe"],
        # Programme breakdown
        "prog_xwb84":         ["trent xwb-84"],
        "prog_xwb97":         ["trent xwb-97"],
        "prog_t1000":         ["trent 1000"],
        "prog_t500":          ["trent 500"],
        "prog_t700":          ["trent 700"],
        "prog_rb211":         ["rb211"],
        "prog_t7000":         ["trent 7000"],
        # Probability
        "count_high_prob":    ["count of high prob"],
        "count_med_prob":     ["count of med prob"],
        "count_low_prob":     ["count of low prob"],
        # Complexity
        "count_high_comp":    ["count of high comp"],
        "count_med_comp":     ["count of med comp"],
        "count_low_comp":     ["count of low comp"],
        # Status
        "count_completed":    ["count of completed"],
        "count_contracting":  ["count of contracting"],
        "count_negotiations": ["count of negotiations"],
        "count_ict":          ["count of ict"],
        "count_hopper":       ["count of hopper"],
        # Financial sums
        "sum_2026":           ["2026"],
        "sum_2027":           ["2027"],
        "sum_26_27":          ["sum of 26/27"],
        "sum_term_benefit":   ["sum of term benefit"],
    }

    hdr_idx = _find_header_row(df, ["project", "count of asks", "trent", "count of high"], max_scan=15)
    col_map = _map_generic_columns(df.iloc[hdr_idx], _COUNT_COLS)

    customers: List[Dict] = []
    for i in range(hdr_idx + 1, len(df)):
        row = df.iloc[i]
        if not _non_blank_vals(row):
            continue
        cust = _clean(_get_generic(row, col_map, "customer"))
        if not cust:
            continue

        rec: Dict[str, Any] = {"customer": cust}
        for field in _COUNT_COLS:
            if field == "customer":
                continue
            rec[field] = _to_float(_get_generic(row, col_map, field))
        customers.append(rec)

    # Also extract sorted rankings if present (cols 29+)
    sorted_rankings: Dict[str, List[Dict]] = {}
    hdr_row = df.iloc[hdr_idx]
    # Look for "SORTED CUSTOMER" / "SORTED TERM IMPACT" etc. in header
    sort_col_pairs: List[tuple] = []
    for j, val in enumerate(hdr_row):
        if not _is_blank(val) and "sorted" in _clean(val).lower():
            label = _clean(val)
            sort_col_pairs.append((j, label))

    # Parse sorted columns in pairs (customer_col, value_col)
    for idx in range(0, len(sort_col_pairs) - 1, 2):
        cust_col_idx = sort_col_pairs[idx][0]
        val_col_idx = sort_col_pairs[idx + 1][0]
        label = sort_col_pairs[idx + 1][1]

        ranking: List[Dict] = []
        for i in range(hdr_idx + 1, len(df)):
            row = df.iloc[i]
            c = _clean(row.iloc[cust_col_idx]) if cust_col_idx < len(row) else None
            v = _to_float(row.iloc[val_col_idx]) if val_col_idx < len(row) else None
            if c:
                ranking.append({"customer": c, "value": v})
        if ranking:
            sorted_rankings[label] = ranking

    return {"customers": customers, "sorted_rankings": sorted_rankings}


def _parse_opp_input_sheet(df: pd.DataFrame) -> Dict:
    """
    Parse the INPUT sheet — reference/config data for dropdowns and weights.

    Contains:
      - Probability scores (High=3, Med=2, Low=1)
      - Complexity scores (High=1, Med=1, Low=1)
      - Weights (PROB=1, COMPLEX=1)
      - Years list (2026-2031)
      - Status values
      - Customer/Project/Programme lookup lists
    """
    config: Dict[str, Any] = {
        "probability_scores": {},
        "complexity_scores": {},
        "weights": {},
        "years": [],
        "statuses": [],
        "customers": [],
        "projects": [],
        "programmes": [],
    }

    # Scan rows for known config sections
    for i in range(min(50, len(df))):
        row = df.iloc[i]
        for j, val in enumerate(row):
            if _is_blank(val):
                continue
            s = _clean(val)
            sl = s.lower()

            # Probability table: col j=label, col j+1=score
            if sl in ("high", "med", "low") and j + 1 < len(row):
                score = _to_float(row.iloc[j + 1])
                if score is not None:
                    # Determine if this is probability or complexity by checking header above
                    context_label = ""
                    for ci in range(max(0, i - 3), i):
                        for cj, cv in enumerate(df.iloc[ci]):
                            if not _is_blank(cv):
                                cvl = _clean(cv).lower()
                                if "prob" in cvl:
                                    context_label = "probability"
                                elif "complex" in cvl:
                                    context_label = "complexity"
                    if context_label == "probability" and sl not in config["probability_scores"]:
                        config["probability_scores"][s] = score
                    elif context_label == "complexity" and sl not in config["complexity_scores"]:
                        config["complexity_scores"][s] = score

            # Status values
            if sl in ("hopper", "ict", "negotiations", "contracting", "completed"):
                if s not in config["statuses"]:
                    config["statuses"].append(s)

            # Years
            m = re.match(r"^(20\d{2})(\.0)?$", sl)
            if m:
                year = int(m.group(1))
                if year not in config["years"]:
                    config["years"].append(year)

    # Extract lookup lists from the lower part of INPUT (rows 16+)
    # Typically: col 5 = CUSTOMER, col 6 = PROJECT, col 8 = PROGRAMME
    for i in range(16, min(50, len(df))):
        row = df.iloc[i]
        for j, val in enumerate(row):
            if _is_blank(val):
                continue
            s = _clean(val)
            if not s or s == "-":
                continue
            # Heuristic: col 5 → customer, col 6 → project, col 8 → programme
            # Skip header labels
            if s.upper() in ("CUSTOMER", "PROJECT", "PROGRAMME", "PROEJCT", "STATUS"):
                continue
            if j == 5 and s not in config["customers"]:
                config["customers"].append(s)
            elif j == 6 and s not in config["projects"]:
                config["projects"].append(s)
            elif j == 8 and s not in config["programmes"]:
                config["programmes"].append(s)

    config["years"].sort()
    return config


def _parse_opp_cover(df: pd.DataFrame) -> Dict:
    """Extract metadata from the COVER sheet."""
    cover: Dict[str, Optional[str]] = {"title": None, "subtitle": None}
    for i in range(min(15, len(df))):
        row = df.iloc[i]
        for val in row:
            if _is_blank(val):
                continue
            s = _clean(val)
            if "commercial optimisation" in s.lower() or "opportunity report" in s.lower():
                if cover["title"] is None:
                    cover["title"] = s
                else:
                    cover["subtitle"] = s
    return cover


def _parse_opp_sum_sheet(df: pd.DataFrame) -> Dict:
    """
    Parse the SUM calculator sheet — filter/aggregation tool.
    Returns the filter criteria and computed sums.
    """
    result: Dict[str, Any] = {"filters": {}, "computed_sums": {}}

    hdr_idx = _find_header_row(df, ["probability", "complexity", "status", "sum", "term"], max_scan=12)
    if hdr_idx >= len(df) - 1:
        return result

    hdr_row = df.iloc[hdr_idx]
    _SUM_COLS: Dict[str, List[str]] = {
        "probability": ["probability"],
        "complexity":  ["complexity"],
        "status":      ["status"],
        "sum_term":    ["term impact", "sum"],
        "sum_2026":    ["2026"],
        "sum_2027":    ["2027"],
    }
    col_map = _map_generic_columns(hdr_row, _SUM_COLS)

    # First data row = filter criteria
    if hdr_idx + 1 < len(df):
        frow = df.iloc[hdr_idx + 1]
        result["filters"] = {
            "probability": _clean(_get_generic(frow, col_map, "probability")) or None,
            "complexity":  _clean(_get_generic(frow, col_map, "complexity")) or None,
            "status":      _clean(_get_generic(frow, col_map, "status")) or None,
        }
        result["computed_sums"] = {
            "term_impact": _to_float(_get_generic(frow, col_map, "sum_term")),
            "sum_2026":    _to_float(_get_generic(frow, col_map, "sum_2026")),
            "sum_2027":    _to_float(_get_generic(frow, col_map, "sum_2027")),
        }

    # Check for totals row (usually last non-blank row)
    for i in range(len(df) - 1, hdr_idx + 1, -1):
        row = df.iloc[i]
        if _non_blank_vals(row):
            result["totals"] = {
                "term_impact": _to_float(_get_generic(row, col_map, "sum_term")),
                "sum_2026":    _to_float(_get_generic(row, col_map, "sum_2026")),
                "sum_2027":    _to_float(_get_generic(row, col_map, "sum_2027")),
            }
            break

    return result


def _parse_opp_reference_data(df: pd.DataFrame) -> Dict:
    """
    Parse Sheet1 — lookup/reference data containing valid values for
    Programme, Operator, Project, TCA Agreement types, SPE/Services types,
    Opportunity types, Lever types, etc.
    """
    ref: Dict[str, List[str]] = {
        "programmes": [],
        "operators": [],
        "projects": [],
        "tca_agreement_types": [],
        "spe_services_types": [],
        "opportunity_types": [],
        "lever_types": [],
    }

    # Row 0 has headers: Programme | Operator | Project | TCA_AGREEMENT | SPE__SERVICES...
    # Row 1+ has values
    if len(df) < 2:
        return ref

    hdr_row = df.iloc[0]
    col_mapping: Dict[str, int] = {}
    for j, val in enumerate(hdr_row):
        if _is_blank(val):
            continue
        s = _clean(val).lower()
        if "programme" in s:
            col_mapping["programmes"] = j
        elif "operator" in s:
            col_mapping["operators"] = j
        elif "project" in s:
            col_mapping["projects"] = j
        elif "tca" in s:
            col_mapping["tca_agreement_types"] = j
        elif "spe" in s:
            col_mapping["spe_services_types"] = j

    # Also scan for opportunity type / lever columns (cols 7-13 area)
    for j, val in enumerate(hdr_row):
        if _is_blank(val):
            continue
        s = _clean(val)
        # These are deeper columns with asks, levers, etc.

    for i in range(1, len(df)):
        row = df.iloc[i]
        for field, col_idx in col_mapping.items():
            if col_idx < len(row):
                v = _clean(row.iloc[col_idx])
                if v and v not in ref[field]:
                    ref[field].append(v)

        # Also extract opportunity type/lever from known positions
        # Col 7 = broader category, col 8 = specific asks, col 9/10 = levers
        for col_idx, field_name in [(7, "opportunity_types"), (9, "lever_types"), (13, "lever_types")]:
            if col_idx < len(row):
                v = _clean(row.iloc[col_idx])
                if v and v not in ref.get(field_name, []):
                    ref.setdefault(field_name, []).append(v)

    return ref


def _parse_opportunity_tracker(all_sheets: Dict[str, pd.DataFrame], filename: str) -> Dict:
    """
    Master parser for the MEA Profit Opportunities Tracker workbook.
    Parses all sheets and returns a comprehensive, structured result.
    """
    errors: List[str] = []
    metadata: Dict[str, Any] = {
        "source_file": filename,
        "away_day_date": None,
        "exchange_rate": None,
        "report_title": None,
    }
    opportunities_by_sheet: Dict[str, Dict] = {}

    # ── Identify opportunity log sheets (MEA LOG, L2, L3) ────────────────
    opp_log_sheets = [
        name for name in all_sheets
        if any(kw in name.lower() for kw in ["l2", "l3", "mea log", "opp log"])
    ]
    if not opp_log_sheets:
        # Fallback: look for sheets with opportunity header signals
        for name, df in all_sheets.items():
            text_blob = " ".join(
                _clean(v).lower() for row in df.head(20).values for v in row if not _is_blank(v)
            )
            if "opp log sheet" in text_blob or "type of opportunity" in text_blob:
                opp_log_sheets.append(name)

    # ── Extract global metadata from first opp log sheet ─────────────────
    for sheet_name in opp_log_sheets[:1]:
        df = all_sheets[sheet_name]
        for i in range(min(15, len(df))):
            row = df.iloc[i]
            for j, val in enumerate(row):
                if _is_blank(val):
                    continue
                sl = _clean(val).lower()
                if "away day date" in sl:
                    for k in range(j + 1, min(j + 5, len(row))):
                        d = _to_date(row.iloc[k])
                        if d:
                            metadata["away_day_date"] = d
                            break

    # ── Parse each opportunity log sheet ─────────────────────────────────
    for sheet_name in opp_log_sheets:
        try:
            df = all_sheets[sheet_name]
            parsed_sheet = _parse_opp_log_sheet(df, sheet_name)
            if parsed_sheet["records"]:
                opportunities_by_sheet[sheet_name] = parsed_sheet
        except Exception as e:
            errors.append(f"Failed to parse opp sheet '{sheet_name}': {e}")

    # ── Parse Summary sheet ──────────────────────────────────────────────
    summary_data: Dict = {}
    for name in all_sheets:
        if name.lower().strip() == "summary":
            try:
                summary_data = _parse_opp_summary_sheet(all_sheets[name])
            except Exception as e:
                errors.append(f"Failed to parse Summary sheet: {e}")
            break

    # ── Parse Opps and Threats sheet ─────────────────────────────────────
    opps_threats_data: Dict = {}
    for name in all_sheets:
        if "opps" in name.lower() and "threat" in name.lower():
            try:
                opps_threats_data = _parse_opp_opps_and_threats(all_sheets[name])
            except Exception as e:
                errors.append(f"Failed to parse Opps and Threats sheet: {e}")
            break

    # ── Parse Timeline + Date Input ──────────────────────────────────────
    timeline_data: Dict = {}
    try:
        timeline_data = _parse_opp_timeline(all_sheets)
    except Exception as e:
        errors.append(f"Failed to parse Timeline/Date Input: {e}")

    # ── Parse COUNT sheet ────────────────────────────────────────────────
    count_data: Dict = {}
    for name in all_sheets:
        if name.lower().strip() == "count":
            try:
                count_data = _parse_opp_count_sheet(all_sheets[name])
            except Exception as e:
                errors.append(f"Failed to parse COUNT sheet: {e}")
            break

    # ── Parse INPUT sheet ────────────────────────────────────────────────
    input_config: Dict = {}
    for name in all_sheets:
        if name.lower().strip() == "input":
            try:
                input_config = _parse_opp_input_sheet(all_sheets[name])
            except Exception as e:
                errors.append(f"Failed to parse INPUT sheet: {e}")
            break

    # ── Parse SUM sheet ──────────────────────────────────────────────────
    sum_data: Dict = {}
    for name in all_sheets:
        if name.lower().strip() == "sum":
            try:
                sum_data = _parse_opp_sum_sheet(all_sheets[name])
            except Exception as e:
                errors.append(f"Failed to parse SUM sheet: {e}")
            break

    # ── Parse COVER sheet ────────────────────────────────────────────────
    cover_data: Dict = {}
    for name in all_sheets:
        if name.lower().strip() == "cover":
            try:
                cover_data = _parse_opp_cover(all_sheets[name])
                if cover_data.get("title"):
                    metadata["report_title"] = cover_data["title"]
            except Exception as e:
                errors.append(f"Failed to parse COVER sheet: {e}")
            break

    # ── Parse Sheet1 (reference data) ────────────────────────────────────
    reference_data: Dict = {}
    for name in all_sheets:
        if name.lower().strip() == "sheet1":
            try:
                reference_data = _parse_opp_reference_data(all_sheets[name])
            except Exception as e:
                errors.append(f"Failed to parse reference data (Sheet1): {e}")
            break

    # ── Build global summary statistics ──────────────────────────────────
    # Flatten all records across all estimation levels
    all_records: List[Dict] = []
    for sheet_data in opportunities_by_sheet.values():
        all_records.extend(sheet_data.get("records", []))

    statuses: Dict[str, int] = {}
    for r in all_records:
        s = r.get("status") or "Unknown"
        statuses[s] = statuses.get(s, 0) + 1

    programmes: Dict[str, int] = {}
    for r in all_records:
        p = r.get("programme") or "Unknown"
        programmes[p] = programmes.get(p, 0) + 1

    customers: Dict[str, int] = {}
    for r in all_records:
        c = r.get("customer") or "Unknown"
        customers[c] = customers.get(c, 0) + 1

    opp_types: Dict[str, int] = {}
    for r in all_records:
        t = r.get("opportunity_type") or "Unknown"
        opp_types[t] = opp_types.get(t, 0) + 1

    # Per-estimation-level sums
    estimation_sums: Dict[str, Dict] = {}
    for sheet_name, sheet_data in opportunities_by_sheet.items():
        level = sheet_data.get("estimation_level", "Unknown")
        recs = sheet_data.get("records", [])
        estimation_sums[level] = {
            "sheet_name": sheet_name,
            "count": len(recs),
            "total_term_benefit": round(sum(r.get("term_benefit") or 0 for r in recs), 2),
            "total_2026": round(sum(r.get("benefit_2026") or 0 for r in recs), 2),
            "total_2027": round(sum(r.get("benefit_2027") or 0 for r in recs), 2),
            "total_sum_26_27": round(sum(r.get("sum_26_27") or 0 for r in recs), 2),
            "sums_from_sheet": sheet_data.get("sums", {}),
        }

    # ── Build backward-compatible "opportunities" dict ───────────────────
    # Keep the old format (sheet_name → [records]) for backward compat
    opportunities_flat: Dict[str, List[Dict]] = {}
    for sheet_name, sheet_data in opportunities_by_sheet.items():
        opportunities_flat[sheet_name] = sheet_data.get("records", [])

    return {
        "file_type": "OPPORTUNITY_TRACKER",
        "metadata": {
            **metadata,
            "sheets_parsed": list(opportunities_by_sheet.keys()),
            "all_sheets": list(all_sheets.keys()),
            "estimation_levels": {
                sn: sd.get("estimation_level") for sn, sd in opportunities_by_sheet.items()
            },
        },
        # Backward-compatible flat opportunities dict
        "opportunities": opportunities_flat,
        # New: structured per-sheet data with estimation levels
        "opportunities_by_level": {
            sheet_data.get("estimation_level", "Unknown"): {
                "sheet_name": sn,
                "records": sheet_data.get("records", []),
                "sums": sheet_data.get("sums", {}),
            }
            for sn, sheet_data in opportunities_by_sheet.items()
        },
        "summary": {
            "total_opportunities": len(all_records),
            "by_status": statuses,
            "by_programme": programmes,
            "by_customer": customers,
            "by_opportunity_type": opp_types,
            "total_term_benefit": round(
                sum(r.get("term_benefit") or 0 for r in all_records), 2
            ),
            "estimation_level_sums": estimation_sums,
        },
        # New: additional parsed sheets
        "project_summary": summary_data,
        "opps_and_threats": opps_threats_data,
        "timeline": timeline_data,
        "customer_analytics": count_data,
        "calculator": sum_data,
        "cover": cover_data,
        "input_config": input_config,
        "reference_data": reference_data,
        "errors": errors,
    }


def _get_generic(row: pd.Series, col_map: Dict[str, int], field: str) -> Any:
    idx = col_map.get(field)
    if idx is None or idx >= len(row):
        return None
    return row.iloc[idx]


# ══════════════════════════════════════════════════════════════════════════════
# Shop Visit History Parser
# ══════════════════════════════════════════════════════════════════════════════

_SV_COL_ALIASES: Dict[str, List[str]] = {
    "part_number":    ["event item part number", "part number"],
    "serial_number":  ["event item serial number", "serial number", "esn"],
    "event_datetime": ["event date time", "event date", "date"],
    "operator":       ["operator"],
    "parent_serial":  ["parent serial number"],
    "registration":   ["parent item registration", "registration"],
    "action_code":    ["action code"],
    "rework_level":   ["rework level"],
    "service_event":  ["service event number"],
    "hsn":            ["hsn"],
    "csn":            ["csn"],
    "hssv":           ["hssv"],
    "cssv":           ["cssv"],
    "sv_type":        ["shopvisit_type", "shop visit type"],
    "sv_location":    ["shopvisit_location", "shop visit location"],
}


def _parse_shop_visit(all_sheets: Dict[str, pd.DataFrame], filename: str) -> Dict:
    errors: List[str] = []

    # Main data sheet
    data_sheet = next(
        (n for n in all_sheets if "report" in n.lower() or "page" in n.lower()),
        list(all_sheets.keys())[0],
    )
    df = all_sheets[data_sheet]

    hdr_idx = _find_header_row(df, list(_SV_COL_ALIASES.keys()) + ["action code", "csn", "hsn"])
    col_map = _map_generic_columns(df.iloc[hdr_idx], _SV_COL_ALIASES)

    current_status_rows: List[Dict] = []
    shop_visit_rows: List[Dict] = []
    maintenance_rows: List[Dict] = []

    for i in range(hdr_idx + 1, len(df)):
        row = df.iloc[i]
        if not _non_blank_vals(row):
            continue

        part_num = _clean(_get_generic(row, col_map, "part_number")) or None
        serial   = _to_str_ref(_get_generic(row, col_map, "serial_number"))
        action   = _clean(_get_generic(row, col_map, "action_code")).lower()
        rework   = _clean(_get_generic(row, col_map, "rework_level")).lower()

        if not serial:
            continue

        rec = {
            "part_number":    part_num,
            "serial_number":  serial,
            "event_datetime": _to_date(_get_generic(row, col_map, "event_datetime")),
            "operator":       _clean(_get_generic(row, col_map, "operator")) or None,
            "parent_serial":  _to_str_ref(_get_generic(row, col_map, "parent_serial")),
            "registration":   _clean(_get_generic(row, col_map, "registration")) or None,
            "action_code":    _clean(_get_generic(row, col_map, "action_code")) or None,
            "rework_level":   _clean(_get_generic(row, col_map, "rework_level")) or None,
            "service_event":  _to_str_ref(_get_generic(row, col_map, "service_event")),
            "hsn":            _to_float(_get_generic(row, col_map, "hsn")),
            "csn":            _to_float(_get_generic(row, col_map, "csn")),
            "hssv":           _to_float(_get_generic(row, col_map, "hssv")),
            "cssv":           _to_float(_get_generic(row, col_map, "cssv")),
            "sv_type":        _clean(_get_generic(row, col_map, "sv_type")) or None,
            "sv_location":    _clean(_get_generic(row, col_map, "sv_location")) or None,
        }

        if "current status" in action or "current status" in rework:
            current_status_rows.append(rec)
        elif "shop visit" in rework:
            shop_visit_rows.append(rec)
        else:
            maintenance_rows.append(rec)

    # Unique serials and operators
    all_serials = {r["serial_number"] for r in shop_visit_rows + maintenance_rows if r["serial_number"]}
    all_operators = {r["operator"] for r in shop_visit_rows if r["operator"] and r["operator"] != "NO OPERATOR"}

    # Detect engine model from part numbers
    part_nums = {r["part_number"] for r in (shop_visit_rows or current_status_rows) if r["part_number"]}
    engine_models = list({pn.split(" ")[0] for pn in part_nums if pn})

    # Derived fields: events by year, events by engine, top operators
    events_by_year: Dict[str, int] = {}
    events_by_engine: Dict[str, int] = {}
    for r in shop_visit_rows + maintenance_rows:
        dt = r.get("event_datetime")
        if dt and isinstance(dt, str) and len(dt) >= 4:
            y = dt[:4]
            events_by_year[y] = events_by_year.get(y, 0) + 1
        esn = r.get("serial_number")
        if esn:
            events_by_engine[esn] = events_by_engine.get(esn, 0) + 1

    op_counts = _count_field(shop_visit_rows, "operator")
    top_operators = sorted(op_counts.items(), key=lambda kv: kv[1], reverse=True)[:15]

    # Engine ages (in years) from first observed event per ESN
    engine_first_seen: Dict[str, str] = {}
    engine_last_seen: Dict[str, str] = {}
    for r in shop_visit_rows + maintenance_rows + current_status_rows:
        esn = r.get("serial_number")
        dt = r.get("event_datetime")
        if not esn or not dt:
            continue
        if esn not in engine_first_seen or dt < engine_first_seen[esn]:
            engine_first_seen[esn] = dt
        if esn not in engine_last_seen or dt > engine_last_seen[esn]:
            engine_last_seen[esn] = dt
    engine_ages = []
    today_iso = datetime.today().strftime("%Y-%m-%d")
    for esn, first_dt in engine_first_seen.items():
        try:
            first = datetime.strptime(first_dt, "%Y-%m-%d")
            age_days = (datetime.today() - first).days
            engine_ages.append({
                "serial_number": esn,
                "first_seen": first_dt,
                "last_seen": engine_last_seen.get(esn),
                "age_days": age_days,
                "age_years": round(age_days / 365.25, 2),
            })
        except Exception:
            continue

    return {
        "file_type": "SHOP_VISIT_HISTORY",
        "metadata": {
            "source_file": filename,
            "source_sheet": data_sheet,
            "engine_models": engine_models,
            "total_engines": len(all_serials),
            "operators": sorted(all_operators),
        },
        "shop_visits": shop_visit_rows,
        # Maintenance actions: non-current-status, non-shop-visit events
        # (historically empty when source only has current-status + shop-visit)
        "maintenance_actions": maintenance_rows,
        "current_status": current_status_rows,
        "statistics": {
            "total_shop_visits":    len(shop_visit_rows),
            "total_maintenance":    len(maintenance_rows),
            "total_engines_tracked": len(all_serials),
            "sv_types": _count_field(shop_visit_rows, "sv_type"),
            "sv_locations": _count_field(shop_visit_rows, "sv_location"),
        },
        # New derived analytics
        "events_by_year": events_by_year,
        "events_by_engine_serial": events_by_engine,
        "top_operators": [{"operator": o, "event_count": c} for o, c in top_operators],
        "engine_ages": engine_ages,
        "all_sheets": list(all_sheets.keys()),
        "errors": errors,
    }


def _count_field(records: List[Dict], field: str) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    for r in records:
        v = r.get(field) or "Unknown"
        counts[v] = counts.get(v, 0) + 1
    return counts


# ══════════════════════════════════════════════════════════════════════════════
# SVRG Master Parser
# ══════════════════════════════════════════════════════════════════════════════

_SVRG_CLAIMS_ALIASES: Dict[str, List[str]] = {
    "date":             ["date"],
    "year":             ["year"],
    "credit_ref":       ["credit note reference", "reference"],
    "guarantee":        ["guarantee"],
    "credit_value":     ["credit note value", "value"],
    "cumulative_value": ["cumulative claim value", "cumulative"],
}

_SVRG_EVENT_ALIASES: Dict[str, List[str]] = {
    "event_type":       ["event type", "#", "type"],
    "date":             ["date"],
    "engine_serial":    ["engine serial no", "esn", "serial"],
    "aircraft":         ["a/c no", "aircraft", "a/c"],
    "tsn_tsr":          ["tsn or tsr", "tsn", "hours"],
    "csn_csr":          ["csn or csr", "csn", "cycles"],
    "description":      ["cause of event", "description", "disruption"],
    "qualification":    ["qualified/non-qualified", "qualification", "emirates input"],
    "justification":    ["justification", "emirates - justification"],
    "rr_input":         ["rr input", "rr qualification"],
    "rr_justification": ["rr - justification", "rr justification"],
    "guarantee_coverage":["guarantee coverage", "coverage"],
    "comments":         ["further comments", "comments"],
}


def _clamp_date(v: Any) -> Optional[str]:
    """Convert date-like value but reject Excel 1900-epoch artifacts (<1970)
    and far-future ghosts (>2075). Returns ISO yyyy-mm-dd or None.

    Upper bound is 2075 (not 2050) because SVRG MASTER plan/forecast rows
    legitimately contain 2050s–2070s dates for engine lifetime projections.
    """
    if _is_blank(v):
        return None
    if isinstance(v, (datetime, date)):
        y = v.year
        if y < 1970 or y > 2075:
            return None
        return v.strftime("%Y-%m-%d")
    iso = _to_date(v)
    if not iso:
        return None
    try:
        y = int(iso[:4])
        if y < 1970 or y > 2075:
            return None
    except Exception:
        return None
    return iso


def _svrg_extract_header_row(df: pd.DataFrame, keywords: List[str], max_scan: int = 20) -> int:
    """Find the header row index (0-based) whose row matches the most keywords.
    Returns -1 if no row scores >= 2."""
    best_idx, best_score = -1, 1
    for i in range(min(max_scan, len(df))):
        row = df.iloc[i]
        score = sum(
            1 for v in row
            if not _is_blank(v) and any(kw in _clean(v).lower() for kw in keywords)
        )
        if score > best_score:
            best_score, best_idx = score, i
    return best_idx


def _svrg_parse_tabular_sheet(
    df: pd.DataFrame,
    sheet_name: str,
    header_keywords: List[str],
    max_header_scan: int = 20,
    skip_rows_after_header: int = 0,
    max_cols: int = 100,
) -> Dict[str, Any]:
    """Generic SVRG tabular sheet reader. Returns {headers, items, row_count}.
    Clips column width, treats error cells as None, emits warnings on malformed
    rows."""
    hdr_idx = _svrg_extract_header_row(df, header_keywords, max_header_scan)
    if hdr_idx < 0:
        logger.warning("SVRG: sheet '%s' — header row not detected", sheet_name)
        return {"header_row": None, "headers": [], "items": [], "row_count": 0}
    hdr_row = df.iloc[hdr_idx]
    eff_width = min(_last_nonempty_col(hdr_row, hard_cap=max_cols), max_cols)
    headers = [_clean(hdr_row.iloc[j]) if j < len(hdr_row) else "" for j in range(eff_width)]
    # Ensure unique header names
    seen: Dict[str, int] = {}
    unique_headers: List[str] = []
    for h in headers:
        h2 = h or f"col_{len(unique_headers)}"
        if h2 in seen:
            seen[h2] += 1
            unique_headers.append(f"{h2}_{seen[h2]}")
        else:
            seen[h2] = 0
            unique_headers.append(h2)

    items: List[Dict] = []
    start = hdr_idx + 1 + skip_rows_after_header
    skipped_blank = 0
    for i in range(start, len(df)):
        row = df.iloc[i]
        row_vals = [row.iloc[j] if j < len(row) else None for j in range(eff_width)]
        if not any(not _is_blank(v) for v in row_vals):
            skipped_blank += 1
            continue
        rec: Dict[str, Any] = {}
        for j, h in enumerate(unique_headers):
            v = row_vals[j]
            if _is_blank(v):
                continue
            if isinstance(v, (datetime, date)):
                clamped = _clamp_date(v)
                if clamped:
                    rec[h] = clamped
                else:
                    logger.warning("SVRG '%s' row %d: date out of range (%s)", sheet_name, i + 1, v)
            elif isinstance(v, float):
                rec[h] = None if v != v else v
            else:
                s = _clean(v)
                rec[h] = s or None
        if rec:
            items.append(rec)
    return {
        "header_row": hdr_idx + 1,
        "headers": unique_headers,
        "items": items,
        "row_count": len(items),
        "skipped_blank_rows": skipped_blank,
    }


def _svrg_sample_rows(
    df: pd.DataFrame,
    sheet_name: str,
    max_items: int = 500,
    header_scan: int = 20,
    max_cols: int = 50,
) -> Dict[str, Any]:
    """Sample up to max_items rows from a wide/bulk SVRG sheet.

    Unlike _svrg_parse_tabular_sheet, this is tolerant of weak headers and
    never returns zero rows from a non-empty sheet. It also never walks
    the full 12k-row body — it stops once max_items is reached, so cost is
    bounded regardless of sheet size. Preserves raw header labels.

    Returns: {items, row_count_sampled, row_count_total_scanned,
              sample_capped (bool), headers, header_row, capped_at}
    """
    if df is None or df.empty:
        return {
            "items": [], "row_count_sampled": 0, "row_count_total_scanned": 0,
            "sample_capped": False, "headers": [], "header_row": None,
            "capped_at": max_items,
        }

    # Find the best header row. Reject title-banner rows (only 1-2 cells
    # populated) and pick the first row with a meaningful number of labels.
    # We pick the row with the HIGHEST non-blank count within the scan window
    # (ties broken by earliest index), requiring at least 4 cells.
    hdr_idx = 0
    best_score = 0
    for i in range(min(header_scan, len(df))):
        row = df.iloc[i]
        non_blank = sum(1 for v in row if not _is_blank(v))
        if non_blank >= 4 and non_blank > best_score:
            best_score = non_blank
            hdr_idx = i
    # Fallback: if nothing met the >=4 bar, use the first row with >=2 cells
    if best_score == 0:
        for i in range(min(header_scan, len(df))):
            row = df.iloc[i]
            if sum(1 for v in row if not _is_blank(v)) >= 2:
                hdr_idx = i
                break

    hdr_row = df.iloc[hdr_idx]
    eff_width = min(_last_nonempty_col(hdr_row, hard_cap=max_cols), max_cols)
    if eff_width <= 0:
        eff_width = min(max_cols, df.shape[1])
    raw_headers = [_clean(hdr_row.iloc[j]) if j < len(hdr_row) else "" for j in range(eff_width)]
    # Dedupe
    seen: Dict[str, int] = {}
    headers: List[str] = []
    for h in raw_headers:
        h2 = h or f"col_{len(headers)}"
        if h2 in seen:
            seen[h2] += 1
            headers.append(f"{h2}_{seen[h2]}")
        else:
            seen[h2] = 0
            headers.append(h2)

    items: List[Dict] = []
    total_scanned = 0
    for i in range(hdr_idx + 1, len(df)):
        total_scanned += 1
        row = df.iloc[i]
        row_vals = [row.iloc[j] if j < len(row) else None for j in range(eff_width)]
        if not any(not _is_blank(v) for v in row_vals):
            continue
        rec: Dict[str, Any] = {}
        for j, h in enumerate(headers):
            v = row_vals[j]
            if _is_blank(v):
                continue
            if isinstance(v, (datetime, date)):
                clamped = _clamp_date(v)
                if clamped:
                    rec[h] = clamped
            elif isinstance(v, float):
                rec[h] = None if v != v else v
            else:
                s = _clean(v)
                rec[h] = s or None
        if rec:
            items.append(rec)
        if len(items) >= max_items:
            break

    return {
        "items": items,
        "row_count_sampled": len(items),
        "row_count_total_scanned": total_scanned,
        "sample_capped": len(items) >= max_items,
        "headers": headers,
        "header_row": hdr_idx + 1,
        "capped_at": max_items,
    }


# Canonical-field aliases for SVRG shop visits (QUALIFIED SVs) — matches the
# raw column labels observed in VERSION 2 Enhanced SVRG MASTER FILE.
_SVRG_SV_FIELD_MAP: Dict[str, List[str]] = {
    "asset_id":       ["asset#", "asset id", "asset serial number", "asset"],
    "engine_serial":  ["esn", "engine serial", "engine"],   # "Engine" col holds ESN
    "sv_date":        ["date of engine removal", "removal date", "sv date",
                       "shop visit date", "date"],
    "sv_type":        ["shop visit type", "qualified hptb shop visit",
                       "cause for shop visit", "sv type", "type"],
    "sv_location":    ["shop visit location", "location"],
    "operator":       ["operator"],
    "csn":            ["csn"],
    "year_of_sv":     ["year of sv", "year"],
    "rsv_num":        ["rsv#", "rsv"],
}

# Canonical-field aliases for SVRG engines (QUALIFIED ENGINES)
_SVRG_ENGINE_FIELD_MAP: Dict[str, List[str]] = {
    "engine_serial":  ["esn", "engine serial"],
    "asset_id":       ["asset#", "asset"],
    "engine_family":  ["engine family", "rating when new", "rating"],
    "engine_status":  ["engine status when delivered", "engine status", "status"],
    "delivery_date":  ["date of first delivery of engine to emirates",
                       "date of first delivery", "delivery date", "first delivery"],
    "qualified_date": ["date engine becomes a qualified engine",
                       "date engine becomes qualified"],
    "improved_hptb":  ["improved hptb from new?", "improved hptb from new"],
    "final_fix":      ["final fix from new?", "final fix from new"],
}


def _svrg_normalize_items(
    items: List[Dict[str, Any]],
    field_map: Dict[str, List[str]],
) -> List[Dict[str, Any]]:
    """Project each raw row into canonical keys, keeping ALL raw fields under
    a nested 'raw' sub-dict for zero data loss. Matching is case-insensitive
    and picks the first alias that resolves to a non-empty value."""
    out: List[Dict[str, Any]] = []
    for rec in items:
        # Lowercase key index, preserve original keys
        lc_index = {k.lower(): k for k in rec.keys() if isinstance(k, str)}
        normalized: Dict[str, Any] = {}
        for canon, aliases in field_map.items():
            found = None
            for alias in aliases:
                al = alias.lower()
                # Exact match first
                if al in lc_index:
                    v = rec.get(lc_index[al])
                    if v not in (None, ""):
                        found = v
                        break
                # Fallback: substring match
                for lck, origk in lc_index.items():
                    if al in lck:
                        v = rec.get(origk)
                        if v not in (None, ""):
                            found = v
                            break
                if found is not None:
                    break
            normalized[canon] = found
        normalized["raw"] = dict(rec)
        out.append(normalized)
    return out


def _svrg_summary_only(df: pd.DataFrame, sheet_name: str, max_cols: int = 50) -> Dict[str, Any]:
    """For bulk sheets where we want aggregates only — no row dump."""
    eff_width = min(_last_nonempty_col(df.iloc[0] if len(df) else pd.Series(),
                                        hard_cap=max_cols), max_cols)
    row_count = 0
    col_non_blank_counts = [0] * eff_width
    col_numeric_counts = [0] * eff_width
    col_sums: List[float] = [0.0] * eff_width
    col_min: List[Optional[float]] = [None] * eff_width
    col_max: List[Optional[float]] = [None] * eff_width
    min_date: Optional[str] = None
    max_date: Optional[str] = None
    for i in range(len(df)):
        row = df.iloc[i]
        row_has = False
        for j in range(eff_width):
            if j >= len(row):
                continue
            v = row.iloc[j]
            if _is_blank(v):
                continue
            row_has = True
            col_non_blank_counts[j] += 1
            f = _to_float(v)
            if f is not None:
                col_numeric_counts[j] += 1
                col_sums[j] += f
                col_min[j] = f if col_min[j] is None else min(col_min[j], f)
                col_max[j] = f if col_max[j] is None else max(col_max[j], f)
            if isinstance(v, (datetime, date)):
                iso = _clamp_date(v)
                if iso:
                    if min_date is None or iso < min_date:
                        min_date = iso
                    if max_date is None or iso > max_date:
                        max_date = iso
        if row_has:
            row_count += 1
    # Build per-column summary
    cols_summary = {}
    for j in range(eff_width):
        if col_non_blank_counts[j] == 0:
            continue
        cols_summary[f"col_{j}"] = {
            "non_blank": col_non_blank_counts[j],
            "numeric_count": col_numeric_counts[j],
            "sum": round(col_sums[j], 4) if col_numeric_counts[j] else None,
            "min": col_min[j],
            "max": col_max[j],
        }
    return {
        "row_count": row_count,
        "columns": cols_summary,
        "date_range": {"min": min_date, "max": max_date} if (min_date or max_date) else None,
    }


def _parse_svrg_master(all_sheets: Dict[str, pd.DataFrame], filename: str) -> Dict:
    """SVRG MASTER (Trent 900 Guarantee Administration) full-fidelity parser.

    Previously extracted <1% of data (~50 rows out of 14,500). Rewritten to
    cover every non-chart sheet:
      • QUALIFIED ENGINES (167 rows) — engine master list
      • QUALIFIED SVs (1145 rows, header row 6, skip row 7 label)
      • QUALIFIED EFH (394×184) — flight hours, per-engine totals
      • HOURS&CYCLES INPUT (12143 rows) — aggregate summary only
      • CLAIMS SUMMARY — header detection relaxed (6-row preamble)
      • EVENT ENTRY (59 rows, was only 16)
      • SVRG+ESVRG (79 rows) + cross-cutting summaries
    """
    errors: List[str] = []
    metadata: Dict[str, Any] = {"source_file": filename, "customer": None, "engine_model": None}

    # ── Metadata from MENU ──
    if "MENU" in all_sheets:
        menu_df = all_sheets["MENU"]
        for i in range(min(12, len(menu_df))):
            row = menu_df.iloc[i]
            for val in row:
                if _is_blank(val):
                    continue
                s = _clean(val)
                sl = s.lower()
                if metadata["customer"] is None and any(kw in sl for kw in
                        ["emirates", "singapore", "airline", "airways", "aviation"]):
                    metadata["customer"] = s
                if metadata["engine_model"] is None and ("trent" in sl or "t900" in sl):
                    metadata["engine_model"] = s

    # ── QUALIFIED ENGINES ──
    engines_out: Dict[str, Any] = {"items": [], "row_count": 0}
    if "QUALIFIED ENGINES" in all_sheets:
        df = all_sheets["QUALIFIED ENGINES"]
        # Audit: header row 6 (1-based) / index 5; real data from row 7
        # Row 1 has a title banner; row 6 has the real labels
        parsed = _svrg_parse_tabular_sheet(
            df, "QUALIFIED ENGINES",
            header_keywords=["engine status", "rating when new", "esn",
                             "date of first delivery", "asset#"],
            max_header_scan=10, max_cols=45,
        )
        # Normalize engine items to canonical keys; preserve raw under rec["raw"]
        engines_out["items"] = _svrg_normalize_items(parsed["items"], _SVRG_ENGINE_FIELD_MAP)
        engines_out["row_count"] = parsed["row_count"]
        engines_out["headers"] = parsed["headers"]
        engines_out["header_row"] = parsed["header_row"]

    # ── QUALIFIED SVs ──
    svs_out: Dict[str, Any] = {"items": [], "row_count": 0}
    if "QUALIFIED SVs" in all_sheets:
        df = all_sheets["QUALIFIED SVs"]
        # Audit: header row 6, data from row 8 (skip row 7 "Actual shop visits" label)
        parsed = _svrg_parse_tabular_sheet(
            df, "QUALIFIED SVs",
            header_keywords=["asset#", "engine", "date of engine removal",
                             "cause for shop visit", "qualified", "hptb driven"],
            max_header_scan=10,
            skip_rows_after_header=1,   # skip the "Actual shop visits" label row
            max_cols=16,
        )
        # Normalize shop visit items to canonical keys; preserve raw under rec["raw"]
        svs_out["items"] = _svrg_normalize_items(parsed["items"], _SVRG_SV_FIELD_MAP)
        svs_out["row_count"] = parsed["row_count"]
        svs_out["headers"] = parsed["headers"]
        svs_out["header_row"] = parsed["header_row"]

    # ── QUALIFIED EFH (very wide: 184 cols) — aggregate + capped row sample ──
    efh_out: Dict[str, Any] = {"row_count": 0, "summary": {}, "items": []}
    if "QUALIFIED EFH" in all_sheets:
        df = all_sheets["QUALIFIED EFH"]
        # Aggregate to avoid 72k-cell dump, BUT also emit a capped item sample
        # so the visualizer has rows to render (it expects .items).
        efh_out["summary"] = _svrg_summary_only(df, "QUALIFIED EFH", max_cols=200)
        efh_out["row_count"] = efh_out["summary"]["row_count"]
        sample = _svrg_sample_rows(df, "QUALIFIED EFH", max_items=500,
                                    header_scan=10, max_cols=50)
        efh_out["items"] = sample["items"]
        efh_out["sample_capped"] = sample["sample_capped"]
        efh_out["sample_capped_at"] = sample["capped_at"]
        efh_out["sample_header_row"] = sample["header_row"]
        efh_out["sample_headers"] = sample["headers"]

    # ── HOURS&CYCLES INPUT (12143 rows) — summary + capped row sample ──
    hc_out: Dict[str, Any] = {"row_count": 0, "summary": {}, "items": []}
    if "HOURS&CYCLES INPUT" in all_sheets:
        df = all_sheets["HOURS&CYCLES INPUT"]
        hc_out["summary"] = _svrg_summary_only(df, "HOURS&CYCLES INPUT", max_cols=10)
        hc_out["row_count"] = hc_out["summary"]["row_count"]
        sample = _svrg_sample_rows(df, "HOURS&CYCLES INPUT", max_items=500,
                                    header_scan=10, max_cols=10)
        hc_out["items"] = sample["items"]
        hc_out["sample_capped"] = sample["sample_capped"]
        hc_out["sample_capped_at"] = sample["capped_at"]
        hc_out["sample_header_row"] = sample["header_row"]
        hc_out["sample_headers"] = sample["headers"]

    # ── CLAIMS SUMMARY ──
    claims: List[Dict] = []
    claims_parsed = {"items": [], "row_count": 0, "headers": []}
    if "CLAIMS SUMMARY" in all_sheets:
        df = all_sheets["CLAIMS SUMMARY"]
        claims_parsed = _svrg_parse_tabular_sheet(
            df, "CLAIMS SUMMARY",
            header_keywords=["date", "credit note", "guarantee",
                             "cumulative", "reference", "value"],
            max_header_scan=20, max_cols=10,
        )
        # Project to canonical claim schema too
        for rec in claims_parsed["items"]:
            # Try to find canonical fields regardless of exact header text
            def pick(*keys):
                for k in keys:
                    for h, v in rec.items():
                        hl = h.lower()
                        if any(kk in hl for kk in keys):
                            return v
                return None

            date_v = None
            year_v = None
            credit_ref = None
            guarantee = None
            value_v = None
            cumulative = None
            for h, v in rec.items():
                hl = h.lower()
                if date_v is None and "date" in hl:
                    date_v = v if isinstance(v, str) and re.match(r"\d{4}-\d{2}-\d{2}", v) else _clamp_date(v)
                if year_v is None and "year" in hl:
                    year_v = _to_float(v)
                if credit_ref is None and ("reference" in hl or "credit note ref" in hl):
                    credit_ref = _to_str_ref(v)
                if guarantee is None and "guarantee" in hl:
                    guarantee = _clean(v) or None
                if value_v is None and ("value" in hl and "cumulative" not in hl):
                    value_v = _to_float(v)
                if cumulative is None and "cumulative" in hl:
                    cumulative = _to_float(v)
            if date_v or (value_v is not None and value_v != 0):
                claims.append({
                    "date": date_v,
                    "year": year_v,
                    "credit_ref": credit_ref,
                    "guarantee": guarantee,
                    "credit_value": value_v,
                    "cumulative_value": cumulative,
                })

    # ── EVENT ENTRY ──
    events: List[Dict] = []
    event_parsed = {"items": [], "row_count": 0, "headers": []}
    if "EVENT ENTRY" in all_sheets:
        df = all_sheets["EVENT ENTRY"]
        event_parsed = _svrg_parse_tabular_sheet(
            df, "EVENT ENTRY",
            header_keywords=["date", "engine serial", "esn", "a/c",
                             "cause", "qualified", "coverage", "tsn", "csn"],
            max_header_scan=20, max_cols=20,
        )
        # Project to canonical event schema
        for rec in event_parsed["items"]:
            date_v = None
            serial_v = None
            desc_v = None
            aircraft = None
            tsn = None
            csn = None
            qualification = None
            justification = None
            rr_input = None
            rr_just = None
            coverage = None
            comments = None
            event_type = None
            for h, v in rec.items():
                hl = h.lower()
                if date_v is None and "date" in hl:
                    date_v = v if isinstance(v, str) and re.match(r"\d{4}-\d{2}-\d{2}", v) else _clamp_date(v)
                if serial_v is None and ("serial" in hl or "esn" in hl):
                    serial_v = _to_str_ref(v)
                if desc_v is None and ("cause" in hl or "description" in hl or "disruption" in hl):
                    desc_v = _clean(v) or None
                if aircraft is None and ("a/c" in hl or "aircraft" in hl):
                    aircraft = _clean(v) or None
                if tsn is None and ("tsn" in hl or "hours" in hl):
                    tsn = _to_float(v)
                if csn is None and ("csn" in hl or "cycles" in hl):
                    csn = _to_float(v)
                if qualification is None and ("qualified" in hl or "emirates input" in hl):
                    qualification = _clean(v) or None
                if justification is None and "justification" in hl and "rr" not in hl:
                    justification = _clean(v) or None
                if rr_input is None and "rr input" in hl:
                    rr_input = _clean(v) or None
                if rr_just is None and "rr" in hl and "justification" in hl:
                    rr_just = _clean(v) or None
                if coverage is None and "coverage" in hl:
                    coverage = _clean(v) or None
                if comments is None and "comment" in hl:
                    comments = _clean(v) or None
                if event_type is None and ("event type" in hl or hl.strip() == "#"):
                    event_type = _clean(v) or None
            if not any([date_v, serial_v, desc_v, aircraft]):
                continue
            events.append({
                "event_type": event_type,
                "date": date_v,
                "engine_serial": serial_v,
                "aircraft": aircraft,
                "tsn_tsr": tsn,
                "csn_csr": csn,
                "description": desc_v,
                "qualification": qualification,
                "justification": justification,
                "rr_input": rr_input,
                "rr_justification": rr_just,
                "guarantee_coverage": coverage,
                "comments": comments,
            })

    # ── SVRG+ESVRG (summary-level metrics) ──
    svrg_summary: Dict[str, Any] = {"row_count": 0}
    if "SVRG+ESVRG" in all_sheets:
        df = all_sheets["SVRG+ESVRG"]
        svrg_summary = _svrg_summary_only(df, "SVRG+ESVRG", max_cols=30)

    # ── Cross-cutting summaries: one row-count block per remaining sheet ──
    secondary_sheets = [
        "RATE BASED SUMMARY", "DI SUMMARY", "HPTB&VANE SUMMARY", "ELMB SUMMARY",
        "EMISSIONS SUMMARY", "WEIGHT SUMMARY", "OIL  ", "OIL SUMMARY",
        "FBR&UNCONT FAIL", "TGT DETERIORATION", "ASSUMPTIONS", "EHPTB MEASURE",
        "2024 Expected SV's", "EFH AND REV DEPS",
    ]
    extra_summaries: Dict[str, Dict] = {}
    for sn in secondary_sheets:
        if sn in all_sheets:
            ws = all_sheets[sn]
            try:
                # Treat these as simple tabular sheets with 1-row header scan
                parsed = _svrg_parse_tabular_sheet(
                    ws, sn,
                    header_keywords=["rate", "year", "summary", "value", "engine",
                                     "date", "asset", "month", "emission", "oil", "hptb"],
                    max_header_scan=12, max_cols=40,
                )
                extra_summaries[sn] = {
                    "row_count": parsed["row_count"],
                    "headers": parsed["headers"][:20],
                    "items": parsed["items"],
                }
            except Exception as e:
                logger.warning("SVRG: failed secondary sheet '%s': %s", sn, e)
                extra_summaries[sn] = {"row_count": 0, "error": str(e)}

    # ── Available-sheets overview (for debugging) ──
    available_sheets: Dict[str, Dict] = {}
    for sname, df in all_sheets.items():
        if sname in ("MENU", "EVENT ENTRY", "CLAIMS SUMMARY", "Chart1", "Chart2",
                     "DESCRIPTIONS", "Sheet3"):
            continue
        try:
            row_count = int(df.notna().any(axis=1).sum())
        except Exception:
            row_count = 0
        available_sheets[sname] = {
            "row_count": row_count,
            "col_count": int(df.shape[1]),
        }

    return {
        "file_type": "SVRG_MASTER",
        "metadata": {
            **metadata,
            "all_sheets": list(all_sheets.keys()),
        },
        # Legacy keys (visualizer compat)
        "claims_summary": {
            "claims": claims,
            "total_claims": len(claims),
            "total_credit_value": round(sum(c["credit_value"] or 0 for c in claims), 2),
        },
        "event_entries": {
            "events": events,
            "total_events": len(events),
            "qualifications": _count_field(events, "qualification"),
            "guarantee_types": _count_field(events, "guarantee_coverage"),
        },
        # New structured blocks
        "engines": engines_out,
        "shop_visits": svs_out,
        "flight_hours": efh_out,
        "hours_cycles": hc_out,
        "claims": {
            "items": claims_parsed.get("items", []),
            "row_count": claims_parsed.get("row_count", 0),
            "headers": claims_parsed.get("headers", []),
        },
        "events": {
            "items": event_parsed.get("items", []),
            "row_count": event_parsed.get("row_count", 0),
            "headers": event_parsed.get("headers", []),
        },
        "svrg_summary": svrg_summary,
        "secondary_summaries": extra_summaries,
        "available_sheets": available_sheets,
        "errors": errors,
    }


# ══════════════════════════════════════════════════════════════════════════════
# Global Hopper Parser
# ══════════════════════════════════════════════════════════════════════════════

_HOPPER_COL_ALIASES: Dict[str, List[str]] = {
    "region":               ["region"],
    "customer":             ["customer"],
    "engine_value_stream":  ["engine value stream", "evs"],
    "top_level_evs":        ["top level evs", "top level"],
    "vp_owner":             ["vp/account manager", "vp owner", "account manager"],
    "restructure_type":     ["restructure type"],
    "maturity":             ["opportunity maturity", "maturity"],
    "onerous_type":         ["onerous/non onerous", "onerous"],
    "initiative":           ["initative", "initiative"],  # note: typo in actual file
    "project_plan_req":     ["project plan requirements", "project plan"],
    "status":               ["status"],
    "expected_year":        ["expected year of signature", "expected year"],
    "signature_ap":         ["signature ap"],
    "crp_term_benefit":     ["crp term benefit"],
    "profit_2026":          ["profit 2026"],
    "profit_2027":          ["profit 2027"],
    "profit_2028":          ["profit 2028"],
    "profit_2029":          ["profit 2029"],
    "profit_2030":          ["profit 2030"],
}


_HOPPER_VP_ALIASES = {
    "Dan Hector": "Daniel Hector",
    "Nick Chadwick": "Nicholas Chadwick",
}


def _normalize_vp_owner(raw: Optional[str]) -> Optional[str]:
    """Collapse near-duplicate owner names: first-initial + last-name canonical.
    Handles slash-joined duos (`A / B`) by picking the first name."""
    if not raw:
        return None
    s = raw.strip()
    # Handle slash-separated duos: take the primary (first) name
    primary = s.split("/")[0].strip()
    primary = primary.rstrip()
    # Apply explicit alias map first
    if primary in _HOPPER_VP_ALIASES:
        return _HOPPER_VP_ALIASES[primary]
    # Remove trailing spaces (46+ rows have them per spec)
    return primary


def _parse_global_hopper(all_sheets: Dict[str, pd.DataFrame], filename: str) -> Dict:
    """Parse a Global Commercial Optimisation Hopper workbook.

    STRICT per spec (V6/_hopper_global_log_spec.md): parses ONLY the sheet
    literally named 'GLOBAL LOG'. DETAIL_REPORT / EXEC_REPORT are ignored
    and NOT emitted in output — previously they were emitted as empty lists
    (confusing consumers). Downstream MUST read 'opportunities' only.
    """
    errors: List[str] = []
    opportunities: List[Dict] = []
    reference_data: Dict[str, List[str]] = {}
    cover_title = ""

    # ── COVER (title only, not parsed as data) ──
    for sn, df in all_sheets.items():
        if "cover" in sn.lower():
            for i in range(min(10, len(df))):
                row = df.iloc[i]
                for v in row:
                    s = _clean(v)
                    if "commercial optimisation" in s.lower():
                        cover_title = s
                        break
                if cover_title:
                    break
            break

    # ── GLOBAL LOG (core data — STRICT match only) ──
    global_log_df = None
    global_log_sheet = None
    # Exact match first
    for sn, df in all_sheets.items():
        if sn == "GLOBAL LOG":
            global_log_df = df
            global_log_sheet = sn
            break
    # Fallback: stripped/upper match with warning
    if global_log_df is None:
        for sn, df in all_sheets.items():
            if sn.strip().upper() == "GLOBAL LOG":
                global_log_df = df
                global_log_sheet = sn
                logger.warning("GLOBAL_HOPPER: sheet name '%s' matched via fallback "
                               "(expected literal 'GLOBAL LOG')", sn)
                break

    if global_log_df is not None and len(global_log_df) > 2:
        # Spec: header row = 5 (1-based) / index 4; validate col C == "Region"
        hdr_idx = 4
        hdr_row = global_log_df.iloc[hdr_idx] if hdr_idx < len(global_log_df) else None
        # Validate header; if not valid, search rows 0..10
        region_at_c = (
            hdr_row is not None
            and len(hdr_row) > 2
            and _clean(hdr_row.iloc[2]).lower() == "region"
        )
        if not region_at_c:
            for i in range(min(15, len(global_log_df))):
                row = global_log_df.iloc[i]
                if len(row) > 2 and _clean(row.iloc[2]).lower() == "region":
                    hdr_idx = i
                    hdr_row = row
                    break

        # Clip header to last non-empty column — avoids phantom widths
        eff_width = _last_nonempty_col(hdr_row, hard_cap=100)
        hdr_clip = pd.Series(
            [hdr_row.iloc[j] if j < len(hdr_row) else None for j in range(eff_width)]
        )
        col_map = _map_generic_columns(hdr_clip, _HOPPER_COL_ALIASES)

        # Iterate full row range; blanks are skipped but do NOT stop iteration
        # (orphan Uganda row at index ~128 after 15-row gap)
        skipped_blank = 0
        for i in range(hdr_idx + 1, len(global_log_df)):
            row = global_log_df.iloc[i]
            # Restrict to effective width for row-blank check
            row_effective = [row.iloc[j] if j < len(row) else None for j in range(eff_width)]
            if not any(not _is_blank(v) for v in row_effective):
                skipped_blank += 1
                continue

            def _g(field):
                idx = col_map.get(field)
                if idx is None or idx >= len(row):
                    return None
                return row.iloc[idx]

            raw_customer = _clean(_g("customer"))
            region = _clean(_g("region"))
            if not raw_customer and not region:
                logger.warning(
                    "GLOBAL_HOPPER: row %d has no Region or Customer — skipped",
                    i + 1,
                )
                continue

            # Mixed-type money handling: numeric extraction + note surfacing
            def _num_and_note(field: str):
                v = _g(field)
                if _is_blank(v):
                    return None, None
                f = _to_float(v)
                if f is not None:
                    return f, None
                return None, _clean(v) or None

            crp_num, crp_note = _num_and_note("crp_term_benefit")
            p26, p26_note = _num_and_note("profit_2026")
            p27, p27_note = _num_and_note("profit_2027")
            p28, p28_note = _num_and_note("profit_2028")
            p29, p29_note = _num_and_note("profit_2029")
            p30, p30_note = _num_and_note("profit_2030")

            raw_vp = _clean(_g("vp_owner")) or None

            rec = {
                "region": region or None,
                "customer": raw_customer.strip() if raw_customer else None,
                "raw_customer": raw_customer or None,
                "engine_value_stream": _clean(_g("engine_value_stream")) or None,
                "top_level_evs": _clean(_g("top_level_evs")) or None,
                "vp_owner": raw_vp,
                "vp_owner_normalized": _normalize_vp_owner(raw_vp),
                "restructure_type": _clean(_g("restructure_type")) or None,
                "maturity": _clean(_g("maturity")) or None,
                "onerous_type": _clean(_g("onerous_type")) or None,
                "initiative": _clean(_g("initiative")) or None,
                "project_plan_req": _clean(_g("project_plan_req")) or None,
                "status": _clean(_g("status")) or None,
                "expected_year": _to_float(_g("expected_year")),
                "signature_ap": _clean(_g("signature_ap")) or None,
                "crp_term_benefit": crp_num,
                "profit_2026": p26,
                "profit_2027": p27,
                "profit_2028": p28,
                "profit_2029": p29,
                "profit_2030": p30,
            }

            # Attach notes only when non-numeric placeholders exist
            notes = {}
            if crp_note:  notes["crp_term_benefit_note"] = crp_note
            if p26_note:  notes["profit_2026_note"] = p26_note
            if p27_note:  notes["profit_2027_note"] = p27_note
            if p28_note:  notes["profit_2028_note"] = p28_note
            if p29_note:  notes["profit_2029_note"] = p29_note
            if p30_note:  notes["profit_2030_note"] = p30_note
            if notes:
                rec["notes"] = notes

            # Convert expected_year to int if valid
            if rec["expected_year"] is not None:
                try:
                    rec["expected_year"] = int(rec["expected_year"])
                except (ValueError, TypeError):
                    rec["expected_year"] = None

            opportunities.append(rec)

        logger.info("GLOBAL_HOPPER: extracted %d rows, skipped %d blank rows from '%s'",
                    len(opportunities), skipped_blank, global_log_sheet)
    else:
        errors.append("GLOBAL LOG sheet not found or empty.")

    # ── Data Validations (reference data — read-only, does not crash if absent) ──
    for sn, df in all_sheets.items():
        if sn.strip().lower() == "data validations":
            if len(df) > 1:
                headers = [_clean(v) for v in df.iloc[0]]
                for j, h in enumerate(headers):
                    if h:
                        vals = []
                        for i in range(1, len(df)):
                            if j < len(df.iloc[i]) and not _is_blank(df.iloc[i].iloc[j]):
                                vals.append(_clean(df.iloc[i].iloc[j]))
                        if vals:
                            reference_data[h] = vals
            break

    # ── Build summary statistics ──
    summary: Dict[str, Any] = {"total_opportunities": len(opportunities)}

    def _count_by(field):
        counts: Dict[str, int] = {}
        for r in opportunities:
            v = str(r.get(field) or "Unknown")
            counts[v] = counts.get(v, 0) + 1
        return counts

    def _sum_by(field, value_field="crp_term_benefit"):
        sums: Dict[str, float] = {}
        for r in opportunities:
            k = str(r.get(field) or "Unknown")
            v = r.get(value_field) or 0
            sums[k] = sums.get(k, 0) + v
        return sums

    summary["by_region"] = _count_by("region")
    summary["by_region_value"] = _sum_by("region")
    summary["by_status"] = _count_by("status")
    summary["by_status_value"] = _sum_by("status")
    summary["by_restructure_type"] = _count_by("restructure_type")
    summary["by_restructure_type_value"] = _sum_by("restructure_type")
    summary["by_maturity"] = _count_by("maturity")
    summary["by_maturity_value"] = _sum_by("maturity")
    summary["by_evs"] = _count_by("engine_value_stream")
    summary["by_evs_value"] = _sum_by("engine_value_stream")
    summary["by_customer"] = _count_by("customer")
    summary["by_customer_value"] = _sum_by("customer")
    summary["by_top_level_evs"] = _count_by("top_level_evs")
    summary["by_onerous"] = _count_by("onerous_type")
    summary["by_expected_year"] = _count_by("expected_year")
    summary["by_vp_owner"] = _count_by("vp_owner")

    # Pipeline stages in order
    pipeline_order = [
        "Initial idea", "ICT formed", "Strategy Approved",
        "Financial Modelling Started", "Financial Modelling Complete",
        "Financials Approved", "Negotiations Started", "Negotiations Concluded",
        "Contracting Started", "Contracting Concluded",
    ]
    pipeline_stages = []
    for stage in pipeline_order:
        count = summary["by_status"].get(stage, 0)
        value = summary["by_status_value"].get(stage, 0)
        if count > 0 or stage in summary["by_status"]:
            pipeline_stages.append({"stage": stage, "count": count, "value": round(value, 2)})
    summary["pipeline_stages"] = pipeline_stages

    # Financial totals
    total_crp = sum(r.get("crp_term_benefit") or 0 for r in opportunities)
    total_2026 = sum(r.get("profit_2026") or 0 for r in opportunities)
    total_2027 = sum(r.get("profit_2027") or 0 for r in opportunities)
    total_2028 = sum(r.get("profit_2028") or 0 for r in opportunities)
    total_2029 = sum(r.get("profit_2029") or 0 for r in opportunities)
    total_2030 = sum(r.get("profit_2030") or 0 for r in opportunities)

    summary["total_crp_term_benefit"] = round(total_crp, 2)
    summary["total_profit_2026"] = round(total_2026, 2)
    summary["total_profit_2027"] = round(total_2027, 2)
    summary["total_profit_2028"] = round(total_2028, 2)
    summary["total_profit_2029"] = round(total_2029, 2)
    summary["total_profit_2030"] = round(total_2030, 2)

    # Top customers by CRP term benefit
    customer_totals = _sum_by("customer")
    sorted_customers = sorted(customer_totals.items(), key=lambda x: x[1], reverse=True)
    summary["top_customers"] = [{"customer": c, "crp_term_benefit": round(v, 2)} for c, v in sorted_customers[:20]]

    # Unique values for filter options
    summary["unique_regions"] = sorted(set(r.get("region") for r in opportunities if r.get("region")))
    summary["unique_evs"] = sorted(set(r.get("engine_value_stream") for r in opportunities if r.get("engine_value_stream")))
    summary["unique_statuses"] = sorted(set(r.get("status") for r in opportunities if r.get("status")))
    summary["unique_restructure_types"] = sorted(set(r.get("restructure_type") for r in opportunities if r.get("restructure_type")))
    summary["unique_maturities"] = sorted(set(r.get("maturity") for r in opportunities if r.get("maturity")))
    summary["unique_customers"] = sorted(set(r.get("customer") for r in opportunities if r.get("customer")))

    return {
        "file_type": "GLOBAL_HOPPER",
        "metadata": {
            "source_file": filename,
            "title": cover_title or "Commercial Optimisation Opportunity Report",
            "currency": "GBP",
            "total_opportunities": len(opportunities),
            "regions": summary.get("unique_regions", []),
            "all_sheets": list(all_sheets.keys()),
            # Honest: only GLOBAL LOG is parsed — no more lying
            "sheets_parsed": ["GLOBAL LOG"] if global_log_sheet else [],
            # Everything else (DETAIL_REPORT, EXEC_REPORT, COVER, Data Validations,
            # COUNT, SUM, etc.) is intentionally ignored per spec.
            "ignored_sheets": [sn for sn in all_sheets if sn != global_log_sheet],
        },
        "opportunities": opportunities,
        "summary": summary,
        # NOTE: detail_report / exec_report keys removed — we do NOT parse those
        # sheets. Downstream consumers must read "opportunities" only.
        "reference_data": reference_data,
        "errors": errors,
    }


# ══════════════════════════════════════════════════════════════════════════════
# COMMERCIAL_PLAN Parser (Account Management annual plan workbooks)
# ══════════════════════════════════════════════════════════════════════════════
#
# Target shape
# ------------
# Three sheets, each with a distinct layout:
#
#   1YP                 — One Year Plan (category status + weekly timeline)
#   5YP SPE SALES       — flat five-year SPE sales register (forward-filled
#                         Customer column)
#   SPE SALES PER YEAR  — pivoted multi-year Customer/Engines summary with
#                         "Grand Total" rows per year band
#
# The parser produces:
#
#   one_year_plan.items              list of issue rows with category and weekly status
#   five_year_spe_sales.items        list of flat SPE sale rows (Customer is forward-filled)
#   five_year_spe_sales.totals       derived rollups: by_year / by_engine / by_customer
#   annual_summary.by_year           structured pivot of "SPE SALES PER YEAR"
#
# NOTE: written to be crash-proof — every extractor returns a partial result
# and logs a warning rather than raising on malformed rows.


_PLAN_YEAR_RE = re.compile(r"(\d{4})[_\s]+PLAN", re.IGNORECASE)
_PLAN_YEAR_BANNER_MIN = 2020
_PLAN_YEAR_BANNER_MAX = 2040
_PLAN_CATEGORY_ORDER = ["COMMERCIAL", "AM", "SALES", "CUSTOMER OPS"]
_PLAN_STATUS_RE = re.compile(r"^\s*L[1-4]\b", re.IGNORECASE)


def _plan_year_from_filename(filename: str) -> Optional[int]:
    if not filename:
        return None
    m = _PLAN_YEAR_RE.search(filename)
    if m:
        try:
            y = int(m.group(1))
            if _PLAN_YEAR_BANNER_MIN <= y <= _PLAN_YEAR_BANNER_MAX:
                return y
        except Exception:
            return None
    return None


def _plan_to_iso_date(v: Any) -> Optional[str]:
    """Normalise a weekly-date cell to ISO YYYY-MM-DD; return None when blank."""
    if _is_blank(v):
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = _clean(v)
    if not s:
        return None
    # Accept "2024-10-07" or parseable date strings
    for fmt in _DATE_FMTS:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    try:
        return pd.to_datetime(s, dayfirst=True).strftime("%Y-%m-%d")
    except Exception:
        return None


def _plan_find_header_row(df: pd.DataFrame, max_scan: int = 15) -> int:
    """Find the 1YP header row (looks for 'Blue Chip' + 'Customer' + 'Owner')."""
    for i in range(min(max_scan, len(df))):
        row = df.iloc[i]
        vals = {_clean(v).lower() for v in row if not _is_blank(v)}
        # Check for anchor keywords
        hits = sum(
            1 for kw in ("blue chip", "customer", "issue", "description", "owner")
            if any(kw == v or v.startswith(kw) for v in vals)
        )
        if hits >= 3:
            return i
    return -1


def _plan_find_category_row(df: pd.DataFrame, max_scan: int = 10) -> int:
    """
    Find the row with the category banner (COMMERCIAL / AM / SALES / CUSTOMER OPS).
    These labels live in adjacent columns somewhere in the first ~5 rows.
    """
    needles = {"commercial", "am", "sales", "customer ops"}
    for i in range(min(max_scan, len(df))):
        row = df.iloc[i]
        vals = [_clean(v).lower() for v in row]
        hits = sum(1 for v in vals if v in needles)
        if hits >= 3:
            return i
    return -1


def _plan_extract_category_columns(df: pd.DataFrame, cat_row_idx: int) -> Dict[str, int]:
    """Return {CATEGORY: column_index} for the 4 category banner columns."""
    out: Dict[str, int] = {}
    if cat_row_idx < 0 or cat_row_idx >= len(df):
        return out
    row = df.iloc[cat_row_idx]
    canon = {
        "commercial": "COMMERCIAL",
        "am": "AM",
        "sales": "SALES",
        "customer ops": "CUSTOMER OPS",
    }
    for j in range(len(row)):
        v = _clean(row.iloc[j]).lower()
        if v in canon:
            out[canon[v]] = j
    return out


def _plan_cell(row: pd.Series, idx: int) -> Any:
    """Safe positional access on a row."""
    if idx < 0 or idx >= len(row):
        return None
    v = row.iloc[idx]
    if _is_blank(v):
        return None
    return v


def _plan_cell_text(row: pd.Series, idx: int) -> Optional[str]:
    v = _plan_cell(row, idx)
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = _clean(v)
    return s or None


def _plan_parse_1yp(df: pd.DataFrame) -> Dict[str, Any]:
    """Extract the One Year Plan sheet."""
    out: Dict[str, Any] = {
        "week_columns": [],
        "category_columns": [],
        "items": [],
    }
    if df is None or df.empty:
        return out

    hdr_idx = _plan_find_header_row(df)
    if hdr_idx < 0:
        logger.warning("1YP: could not locate header row (Blue Chip/Customer/Owner).")
        return out

    hdr_row = df.iloc[hdr_idx]

    # Map the six leading label columns (Blue Chip | Customer | Issue | Description | Owner | Latest Update)
    label_col: Dict[str, int] = {}
    _label_keys = {
        "blue chip": "blue_chip",
        "customer": "customer",
        "issue": "issue",
        "description": "description",
        "owner": "owner",
        "latest update": "latest_update",
    }
    for j in range(len(hdr_row)):
        key = _clean(hdr_row.iloc[j]).lower()
        if key in _label_keys and _label_keys[key] not in label_col:
            label_col[_label_keys[key]] = j

    # Category banner row (usually row 0) and columns
    cat_row_idx = _plan_find_category_row(df)
    cat_col_map = _plan_extract_category_columns(df, cat_row_idx) if cat_row_idx >= 0 else {}
    out["category_columns"] = [c for c in _PLAN_CATEGORY_ORDER if c in cat_col_map]

    # Week-date columns: any column in the header row whose value parses as a date,
    # plus also scan the row immediately below the header (some workbooks keep the
    # "Week N" label on the header and the actual date one row down).
    week_cols: List[tuple] = []  # (col_index, iso_date)
    seen_cols: set = set()

    def _scan_row_for_dates(row_idx: int) -> None:
        if row_idx < 0 or row_idx >= len(df):
            return
        r = df.iloc[row_idx]
        for j in range(len(r)):
            if j in seen_cols:
                continue
            iso = _plan_to_iso_date(r.iloc[j])
            if iso and iso.startswith(("19", "20")):
                week_cols.append((j, iso))
                seen_cols.add(j)

    _scan_row_for_dates(hdr_idx)
    _scan_row_for_dates(hdr_idx + 1)

    # Sort week columns by column index to preserve left-to-right order
    week_cols.sort(key=lambda x: x[0])
    out["week_columns"] = [iso for _, iso in week_cols]

    # Determine where data rows start: row immediately after header (skip the
    # second-date row if present).
    data_start = hdr_idx + 1
    if data_start < len(df):
        # If the first row after header has ONLY dates/blanks, skip it
        probe = df.iloc[data_start]
        only_dates = True
        any_val = False
        for j in range(len(probe)):
            v = probe.iloc[j]
            if _is_blank(v):
                continue
            any_val = True
            if _plan_to_iso_date(v) is None:
                only_dates = False
                break
        if any_val and only_dates:
            data_start += 1

    # Build items
    items: List[Dict[str, Any]] = []
    category_col_pairs = [(cat, cat_col_map[cat]) for cat in out["category_columns"]]

    for i in range(data_start, len(df)):
        row = df.iloc[i]
        # Skip entirely blank rows
        if not _non_blank_vals(row):
            continue

        blue_chip = _plan_cell_text(row, label_col.get("blue_chip", -1))
        customer = _plan_cell_text(row, label_col.get("customer", -1))
        issue = _plan_cell_text(row, label_col.get("issue", -1))
        description = _plan_cell_text(row, label_col.get("description", -1))
        owner = _plan_cell_text(row, label_col.get("owner", -1))
        latest_update = _plan_cell_text(row, label_col.get("latest_update", -1))

        # A row is a real data row if it has ANY of these meaningful fields
        meaningful = any([blue_chip, customer, issue, description, owner, latest_update])
        has_status = False

        # Category status
        cat_status: Dict[str, Optional[str]] = {c: None for c in out["category_columns"]}
        for cat, col in category_col_pairs:
            v = _plan_cell_text(row, col)
            if v:
                cat_status[cat] = v
                has_status = True

        # Weekly status
        weekly_status: Dict[str, Optional[str]] = {iso: None for _, iso in week_cols}
        for j, iso in week_cols:
            v = _plan_cell_text(row, j)
            if v:
                weekly_status[iso] = v
                has_status = True

        if not (meaningful or has_status):
            continue

        # Derive status_summary
        all_cell_vals = [v for v in weekly_status.values() if v]
        if not owner or not owner.strip():
            status_summary = "Unassigned"
        elif all_cell_vals and all(_clean(v).lower() in ("complete", "completed", "done") for v in all_cell_vals):
            status_summary = "Completed"
        elif any(_PLAN_STATUS_RE.match(_clean(v or "")) for v in list(cat_status.values()) + list(weekly_status.values()) if v):
            status_summary = "Active"
        elif has_status or meaningful:
            status_summary = "Active"
        else:
            status_summary = "Unassigned"

        # Raw passthrough: keep every non-blank cell so nothing is lost downstream
        raw: Dict[str, Any] = {}
        for j in range(len(row)):
            v = row.iloc[j]
            if _is_blank(v):
                continue
            if isinstance(v, (datetime, date)):
                raw[f"col_{j}"] = v.strftime("%Y-%m-%d")
            elif isinstance(v, float) and v == int(v):
                raw[f"col_{j}"] = int(v)
            else:
                raw[f"col_{j}"] = v if not isinstance(v, (bytes, bytearray)) else None

        items.append({
            "row_index": int(i),
            "blue_chip": blue_chip,
            "customer": customer,
            "issue": issue,
            "description": description,
            "owner": owner,
            "latest_update": latest_update,
            "category_status": cat_status,
            "weekly_status": weekly_status,
            "status_summary": status_summary,
            "raw": raw,
        })

    out["items"] = items
    return out


def _plan_parse_5yp(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Extract the 5YP SPE SALES tabular sheet.
    Header is on row 0: Customer | Engine Type | Year | Quarter | Amount | Comments.
    Customer is forward-filled across blank cells.
    """
    out: Dict[str, Any] = {
        "items": [],
        "totals": {
            "by_year": {},
            "by_engine": {},
            "by_customer": {},
            "total_opportunities": 0,
            "total_amount": 0.0,
        },
    }
    if df is None or df.empty:
        return out

    # Find header
    hdr_idx = -1
    want = {"customer", "engine type", "year", "quarter", "amount"}
    for i in range(min(5, len(df))):
        row = df.iloc[i]
        vals = {_clean(v).lower() for v in row if not _is_blank(v)}
        if len(want & vals) >= 4:
            hdr_idx = i
            break
    if hdr_idx < 0:
        logger.warning("5YP SPE SALES: could not locate header row.")
        return out

    hdr_row = df.iloc[hdr_idx]
    col_idx: Dict[str, int] = {}
    aliases = {
        "customer": "customer",
        "engine type": "engine_type",
        "engine": "engine_type",
        "year": "year",
        "quarter": "quarter",
        "amount": "amount",
        "comments": "comments",
        "comment": "comments",
    }
    for j in range(len(hdr_row)):
        k = _clean(hdr_row.iloc[j]).lower()
        if k in aliases and aliases[k] not in col_idx:
            col_idx[aliases[k]] = j

    items: List[Dict[str, Any]] = []
    last_customer: Optional[str] = None

    totals_year: Dict[str, int] = {}
    totals_engine: Dict[str, int] = {}
    totals_customer: Dict[str, int] = {}
    total_amount = 0.0

    for i in range(hdr_idx + 1, len(df)):
        row = df.iloc[i]
        if not _non_blank_vals(row):
            continue

        # Forward-fill customer
        raw_customer = _plan_cell_text(row, col_idx.get("customer", -1))
        if raw_customer:
            last_customer = raw_customer
        customer = raw_customer or last_customer

        engine_type = _plan_cell_text(row, col_idx.get("engine_type", -1))
        year_cell = _plan_cell(row, col_idx.get("year", -1))
        year: Optional[int] = None
        if isinstance(year_cell, (int, float)) and not isinstance(year_cell, bool):
            try:
                yi = int(year_cell)
                if _PLAN_YEAR_BANNER_MIN <= yi <= _PLAN_YEAR_BANNER_MAX:
                    year = yi
            except Exception:
                year = None
        elif isinstance(year_cell, str):
            try:
                yi = int(_clean(year_cell))
                if _PLAN_YEAR_BANNER_MIN <= yi <= _PLAN_YEAR_BANNER_MAX:
                    year = yi
            except Exception:
                year = None

        quarter = _plan_cell_text(row, col_idx.get("quarter", -1))
        amount = _to_float(_plan_cell(row, col_idx.get("amount", -1)))
        comments = _plan_cell_text(row, col_idx.get("comments", -1))

        # A row is a real sale if it has engine_type OR year OR amount
        if not any([engine_type, year, amount, quarter]):
            logger.warning("5YP SPE SALES: skipping row %d (no engine/year/amount).", i)
            continue

        items.append({
            "row_index": int(i),
            "customer": customer,
            "engine_type": engine_type,
            "year": year,
            "quarter": quarter,
            "amount": amount,
            "comments": comments,
        })

        # Rollups — each opportunity counts as 1 unit
        if year is not None:
            totals_year[str(year)] = totals_year.get(str(year), 0) + 1
        if engine_type:
            totals_engine[engine_type] = totals_engine.get(engine_type, 0) + 1
        if customer:
            totals_customer[customer] = totals_customer.get(customer, 0) + 1
        if isinstance(amount, (int, float)):
            total_amount += float(amount)

    out["items"] = items
    out["totals"] = {
        "by_year": dict(sorted(totals_year.items())),
        "by_engine": dict(sorted(totals_engine.items(), key=lambda kv: (-kv[1], kv[0]))),
        "by_customer": dict(sorted(totals_customer.items(), key=lambda kv: (-kv[1], kv[0]))),
        "total_opportunities": len(items),
        "total_amount": round(total_amount, 4),
    }
    return out


def _plan_parse_spe_sales_per_year(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Extract the SPE SALES PER YEAR pivot sheet.
    Layout: year banners sit on row 0 (or row 1). For each banner the next
    2-3 columns form a Customer | Engines | (blank) block. Inside the block
    rows alternate: customer-name row, engine-type row (both with a numeric
    count in the Engines column), ending with a "Grand Total" row.
    """
    out: Dict[str, Any] = {"by_year": {}}
    if df is None or df.empty:
        return out

    # Find the year banner row — scan first 3 rows for integer cells in range
    banner_row_idx = -1
    banners: List[tuple] = []  # (col_index, year)
    for i in range(min(3, len(df))):
        row = df.iloc[i]
        candidates: List[tuple] = []
        for j in range(len(row)):
            v = row.iloc[j]
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                try:
                    yi = int(v)
                    if _PLAN_YEAR_BANNER_MIN <= yi <= _PLAN_YEAR_BANNER_MAX:
                        candidates.append((j, yi))
                except Exception:
                    continue
            elif isinstance(v, str):
                try:
                    yi = int(_clean(v))
                    if _PLAN_YEAR_BANNER_MIN <= yi <= _PLAN_YEAR_BANNER_MAX:
                        candidates.append((j, yi))
                except Exception:
                    continue
        if len(candidates) >= 2:
            banner_row_idx = i
            banners = candidates
            break
    if banner_row_idx < 0:
        logger.warning("SPE SALES PER YEAR: could not locate year banner row.")
        return out

    # Sort banners by column and compute the column span each owns
    banners.sort(key=lambda x: x[0])
    col_bounds: List[tuple] = []  # (customer_col, engines_col, year)
    for idx, (col, yr) in enumerate(banners):
        # Customer column = banner col; Engines column = banner col + 1
        customer_col = col
        engines_col = col + 1
        col_bounds.append((customer_col, engines_col, yr))

    # The header row (with "Customer"/"Engines" labels) is 1-2 rows below banners
    # Data starts after that.
    data_start = banner_row_idx + 1
    # Skip a divider row of '.' characters if present
    if data_start < len(df):
        probe = df.iloc[data_start]
        if any(_clean(v) == "." for v in probe if not _is_blank(v)):
            data_start += 1
    # Skip the "Customer / Engines" header row if present
    if data_start < len(df):
        probe = df.iloc[data_start]
        vals_l = {_clean(v).lower() for v in probe if not _is_blank(v)}
        if "customer" in vals_l or "engines" in vals_l:
            data_start += 1

    by_year: Dict[str, Dict[str, Any]] = {}

    for customer_col, engines_col, yr in col_bounds:
        year_block: Dict[str, Any] = {
            "year": yr,
            "customers": [],
            "grand_total": 0,
        }
        current_customer: Optional[Dict[str, Any]] = None

        for i in range(data_start, len(df)):
            row = df.iloc[i]
            name_cell = _plan_cell_text(row, customer_col)
            count_cell = _plan_cell(row, engines_col)

            # Stop when we hit the Grand Total for this year
            if name_cell and "grand total" in name_cell.lower():
                try:
                    year_block["grand_total"] = int(_to_float(count_cell) or 0)
                except Exception:
                    year_block["grand_total"] = 0
                break

            # Skip fully blank slice
            if not name_cell and _is_blank(count_cell):
                continue

            count: Optional[int] = None
            fv = _to_float(count_cell)
            if fv is not None:
                try:
                    count = int(fv)
                except Exception:
                    count = None

            if name_cell:
                # Heuristic: if the label looks like an engine type (ALL CAPS / known engines),
                # it's an engine row for the CURRENT customer. Otherwise it's a new customer.
                # Engine type patterns seen in the data: XWB-84, XWB-97, TRENT7000,
                # TRENT1000, Trent 7000, etc.
                is_engine = False
                up = name_cell.upper()
                if re.match(r"^(XWB|TRENT)[\s\-]?\d+", up):
                    is_engine = True
                elif up.startswith("TRENT "):
                    is_engine = True

                if is_engine and current_customer is not None:
                    current_customer["engines"].append({
                        "type": name_cell,
                        "count": count if count is not None else 0,
                    })
                    current_customer["total"] = current_customer.get("total", 0) + (count or 0)
                else:
                    # New customer entry
                    current_customer = {
                        "name": name_cell,
                        "engines": [],
                        "total": count if count is not None else 0,
                    }
                    year_block["customers"].append(current_customer)

        # If Grand Total wasn't explicitly found, compute from customers
        if not year_block["grand_total"]:
            year_block["grand_total"] = sum(
                c.get("total", 0) for c in year_block["customers"]
            )

        by_year[str(yr)] = year_block

    out["by_year"] = by_year
    return out


def _parse_commercial_plan(all_sheets: Dict[str, pd.DataFrame], filename: str) -> Dict:
    """Parse an Account Management annual plan workbook (e.g. 2026_PLAN.xlsx)."""
    errors: List[str] = []
    parsed_sheet_names: List[str] = []
    ignored_sheet_names: List[str] = []

    # ── 1YP ──────────────────────────────────────────────────────────────────
    one_year_plan: Dict[str, Any] = {
        "week_columns": [],
        "category_columns": [],
        "items": [],
    }
    for name, df in all_sheets.items():
        if name.strip().lower().startswith("1yp"):
            try:
                one_year_plan = _plan_parse_1yp(df)
                parsed_sheet_names.append(name)
            except Exception as e:
                logger.exception("Failed to parse 1YP sheet '%s'", name)
                errors.append(f"1YP parse error ({name}): {e}")
            break
    else:
        logger.warning("COMMERCIAL_PLAN: no 1YP sheet found.")

    # ── 5YP SPE SALES ────────────────────────────────────────────────────────
    five_year: Dict[str, Any] = {
        "items": [],
        "totals": {
            "by_year": {},
            "by_engine": {},
            "by_customer": {},
            "total_opportunities": 0,
            "total_amount": 0.0,
        },
    }
    for name, df in all_sheets.items():
        nl = name.strip().lower()
        if nl.startswith("5yp") and "spe" in nl and "per year" not in nl:
            try:
                five_year = _plan_parse_5yp(df)
                parsed_sheet_names.append(name)
            except Exception as e:
                logger.exception("Failed to parse 5YP SPE SALES sheet '%s'", name)
                errors.append(f"5YP SPE SALES parse error ({name}): {e}")
            break

    # Fallback: any 5YP sheet if the specific one not found
    if not five_year["items"]:
        for name, df in all_sheets.items():
            nl = name.strip().lower()
            if nl.startswith("5yp") and name not in parsed_sheet_names:
                try:
                    five_year = _plan_parse_5yp(df)
                    parsed_sheet_names.append(name)
                except Exception as e:
                    logger.exception("Failed to parse fallback 5YP sheet '%s'", name)
                    errors.append(f"5YP fallback parse error ({name}): {e}")
                break

    # ── SPE SALES PER YEAR ───────────────────────────────────────────────────
    annual_summary: Dict[str, Any] = {"by_year": {}}
    for name, df in all_sheets.items():
        nl = name.strip().lower()
        if "per year" in nl or (nl.startswith("spe") and "year" in nl):
            try:
                annual_summary = _plan_parse_spe_sales_per_year(df)
                parsed_sheet_names.append(name)
            except Exception as e:
                logger.exception("Failed to parse SPE SALES PER YEAR sheet '%s'", name)
                errors.append(f"SPE SALES PER YEAR parse error ({name}): {e}")
            break

    # Anything else goes into ignored_sheets
    for name in all_sheets:
        if name not in parsed_sheet_names:
            ignored_sheet_names.append(name)

    # Plan-year resolution: filename → earliest year in 5YP → current year
    plan_year = _plan_year_from_filename(filename)
    if plan_year is None:
        years_seen: List[int] = []
        for item in five_year.get("items", []):
            y = item.get("year")
            if isinstance(y, int):
                years_seen.append(y)
        if years_seen:
            plan_year = min(years_seen)
        else:
            plan_year = datetime.utcnow().year

    return {
        "file_type": "COMMERCIAL_PLAN",
        "metadata": {
            "source_file": filename,
            "sheets_parsed": parsed_sheet_names,
            "plan_year": plan_year,
            "sheets_ignored": ignored_sheet_names,
        },
        "one_year_plan": one_year_plan,
        "five_year_spe_sales": five_year,
        "annual_summary": annual_summary,
        "errors": errors,
    }


# ══════════════════════════════════════════════════════════════════════════════
# EMPLOYEE_WHEREABOUTS Parser
# ══════════════════════════════════════════════════════════════════════════════

_WHEREABOUTS_LEGEND_DEFAULTS: Dict[str, str] = {
    "O": "Office",
    "H": "Home / WFH",
    "WFH": "Work From Home",
    "L": "Leave",
    "B": "Business Trip",
    "S": "Sick",
    "PL": "Personal Leave",
    "CB": "Cross Border",
    "EB": "Easter Break",
    "HOL": "Holiday",
}


def _whereabouts_parse_month(sheet_name: str) -> Optional[Dict[str, Any]]:
    """Parse 'Apr 2026' → {sheet, year, month, days_in_month, start_date, end_date}."""
    m = _MONTH_YEAR_SHEET_RE.match(sheet_name.strip())
    if not m:
        return None
    month_token = m.group(1).lower()[:4] if m.group(1).lower().startswith("sept") else m.group(1).lower()[:3]
    month_num = _MONTH_NAME_TO_NUM.get(month_token)
    if month_num is None:
        return None
    try:
        year = int(m.group(2))
    except Exception:
        return None
    # days_in_month via calendar
    import calendar
    _, days_in_month = calendar.monthrange(year, month_num)
    start_date = f"{year:04d}-{month_num:02d}-01"
    end_date = f"{year:04d}-{month_num:02d}-{days_in_month:02d}"
    return {
        "sheet": sheet_name,
        "year": year,
        "month": month_num,
        "days_in_month": days_in_month,
        "start_date": start_date,
        "end_date": end_date,
    }


def _whereabouts_find_header_row(df: pd.DataFrame, max_scan: int = 20) -> int:
    """
    Locate the row that contains 'Employee Number' in an early column.
    Returns the row index (0-based) or -1 if not found.
    """
    scan_rows = min(max_scan, len(df))
    for i in range(scan_rows):
        try:
            row = df.iloc[i]
        except Exception:
            continue
        for j in range(min(10, len(row))):
            v = row.iloc[j]
            if _is_blank(v):
                continue
            t = _clean(v).lower()
            if "employee number" in t or t == "employee no" or t == "emp no":
                return i
    return -1


def _whereabouts_map_columns(
    header_row: pd.Series,
    days_in_month: int,
) -> Dict[str, Any]:
    """
    Map header cells to column indexes. Locates:
      no_col, emp_num_col, name_col, sector_col, country_col
      day_cols: list of (col_idx, day_number)   (1..days_in_month)
      notes_col (optional)
    """
    col_map: Dict[str, Any] = {
        "no": None, "emp_num": None, "name": None,
        "sector": None, "country": None, "notes": None,
        "day_cols": [],
    }
    for j in range(len(header_row)):
        v = header_row.iloc[j]
        if _is_blank(v):
            continue
        t = _clean(v).lower()
        if t == "no." or t == "no" or t == "#":
            if col_map["no"] is None:
                col_map["no"] = j
        elif "employee number" in t or t in ("emp no", "emp number", "employee no"):
            col_map["emp_num"] = j
        elif "employee name" in t or t == "name":
            col_map["name"] = j
        elif "business sector" in t or t == "sector":
            col_map["sector"] = j
        elif t == "country" or "country" in t:
            if col_map["country"] is None:
                col_map["country"] = j
        elif "note" in t:
            col_map["notes"] = j
        else:
            # Day-number header: integer 1..31
            raw = header_row.iloc[j]
            try:
                if isinstance(raw, (int, float)) and not (isinstance(raw, float) and raw != raw):
                    n = int(raw)
                    if 1 <= n <= 31:
                        col_map["day_cols"].append((j, n))
                elif isinstance(raw, str) and raw.strip().isdigit():
                    n = int(raw.strip())
                    if 1 <= n <= 31:
                        col_map["day_cols"].append((j, n))
            except Exception:
                pass

    # If day-number header scan yielded nothing, fall back to position-based:
    # everything right of country through to notes-or-end.
    if not col_map["day_cols"] and col_map["country"] is not None:
        start = col_map["country"] + 1
        end = col_map["notes"] if col_map["notes"] is not None else len(header_row)
        for j in range(start, min(end, start + days_in_month)):
            col_map["day_cols"].append((j, j - start + 1))

    # Trim day_cols that exceed days_in_month (defensive)
    col_map["day_cols"] = [(c, d) for (c, d) in col_map["day_cols"] if d <= days_in_month]
    # Sort by day number ascending
    col_map["day_cols"].sort(key=lambda t: t[1])
    return col_map


def _whereabouts_legend_label(code: str) -> str:
    """Return a friendly label for a status code, falling back to the code itself."""
    if not code:
        return code
    upper = code.strip().upper()
    if upper in _WHEREABOUTS_LEGEND_DEFAULTS:
        return _WHEREABOUTS_LEGEND_DEFAULTS[upper]
    # Heuristic partial matches for free-text values
    low = code.lower()
    if "eid" in low or "holiday" in low:
        return "Public Holiday"
    if "leave" in low:
        return "Leave"
    if "sick" in low:
        return "Sick"
    if "busines" in low or "trip" in low:
        return "Business Trip"
    if "home" in low or "wfh" in low:
        return "Home / WFH"
    if "office" in low:
        return "Office"
    if "easter" in low:
        return "Easter Break"
    if "cross" in low and "border" in low:
        return "Cross Border"
    # Unknown — use the code itself as its own label
    return code


def _parse_employee_whereabouts(all_sheets: Dict[str, pd.DataFrame], filename: str) -> Dict:
    """Parse a Middle East Employee Whereabouts workbook."""
    errors: List[str] = []

    months_meta: List[Dict[str, Any]] = []
    sheets_parsed: List[str] = []
    sheets_ignored: List[str] = []
    whereabouts: Dict[str, List[Dict[str, Any]]] = {}
    employees_by_num: Dict[str, Dict[str, Any]] = {}
    status_totals_by_month: Dict[str, Dict[str, int]] = {}
    daily_office_count_by_month: Dict[str, Dict[str, int]] = {}
    by_country: Dict[str, int] = {}
    by_sector: Dict[str, int] = {}
    all_status_codes: Dict[str, int] = {}  # for legend inference

    for sheet_name, df in all_sheets.items():
        month_meta = _whereabouts_parse_month(sheet_name)
        if not month_meta:
            sheets_ignored.append(sheet_name)
            continue

        hdr_idx = _whereabouts_find_header_row(df, max_scan=20)
        if hdr_idx < 0:
            logger.warning(
                "EMPLOYEE_WHEREABOUTS: sheet '%s' has no 'Employee Number' header; skipping",
                sheet_name,
            )
            errors.append(f"{sheet_name}: header row not found")
            sheets_ignored.append(sheet_name)
            continue

        header_row = df.iloc[hdr_idx]
        col_map = _whereabouts_map_columns(header_row, month_meta["days_in_month"])

        if col_map["emp_num"] is None and col_map["name"] is None:
            logger.warning(
                "EMPLOYEE_WHEREABOUTS: sheet '%s' missing employee columns; skipping",
                sheet_name,
            )
            errors.append(f"{sheet_name}: missing employee column")
            sheets_ignored.append(sheet_name)
            continue

        if not col_map["day_cols"]:
            logger.warning(
                "EMPLOYEE_WHEREABOUTS: sheet '%s' has no day columns; skipping",
                sheet_name,
            )
            errors.append(f"{sheet_name}: no day columns detected")
            sheets_ignored.append(sheet_name)
            continue

        year = month_meta["year"]
        month = month_meta["month"]

        sheet_records: List[Dict[str, Any]] = []
        sheet_status_totals: Dict[str, int] = {}
        sheet_daily_office: Dict[str, int] = {
            f"{year:04d}-{month:02d}-{d:02d}": 0
            for d in range(1, month_meta["days_in_month"] + 1)
        }

        max_col = len(df.columns)

        for ri in range(hdr_idx + 1, len(df)):
            try:
                row = df.iloc[ri]
            except Exception:
                continue

            emp_num_raw = row.iloc[col_map["emp_num"]] if col_map["emp_num"] is not None and col_map["emp_num"] < max_col else None
            name_raw = row.iloc[col_map["name"]] if col_map["name"] is not None and col_map["name"] < max_col else None

            if _is_blank(emp_num_raw) and _is_blank(name_raw):
                continue  # blank row

            emp_num = _to_str_ref(emp_num_raw)
            if emp_num is None:
                # If emp_num is blank but name exists, synthesise from name
                nm = _clean(name_raw)
                if nm:
                    emp_num = f"UNKNOWN-{nm[:30]}"
                else:
                    logger.warning("EMPLOYEE_WHEREABOUTS: row %d in '%s' skipped (no emp number or name)", ri, sheet_name)
                    continue

            name = _clean(name_raw) or None
            sector = _clean(row.iloc[col_map["sector"]]) if col_map["sector"] is not None and col_map["sector"] < max_col else None
            sector = sector or None
            country = _clean(row.iloc[col_map["country"]]) if col_map["country"] is not None and col_map["country"] < max_col else None
            country = country or None

            # Register employee (first occurrence wins)
            if emp_num not in employees_by_num:
                employees_by_num[emp_num] = {
                    "employee_number": emp_num,
                    "name": name,
                    "business_sector": sector,
                    "country": country,
                }
                if country:
                    by_country[country] = by_country.get(country, 0) + 1
                if sector:
                    by_sector[sector] = by_sector.get(sector, 0) + 1

            daily_status: Dict[str, Optional[str]] = {}
            status_counts: Dict[str, int] = {}

            for (col_idx, day_num) in col_map["day_cols"]:
                iso = f"{year:04d}-{month:02d}-{day_num:02d}"
                if col_idx >= max_col:
                    daily_status[iso] = None
                    status_counts["_blank"] = status_counts.get("_blank", 0) + 1
                    sheet_status_totals["_blank"] = sheet_status_totals.get("_blank", 0) + 1
                    continue
                cell = row.iloc[col_idx]
                if _is_blank(cell):
                    daily_status[iso] = None
                    status_counts["_blank"] = status_counts.get("_blank", 0) + 1
                    sheet_status_totals["_blank"] = sheet_status_totals.get("_blank", 0) + 1
                else:
                    code = _clean(cell)
                    if not code:
                        daily_status[iso] = None
                        status_counts["_blank"] = status_counts.get("_blank", 0) + 1
                        sheet_status_totals["_blank"] = sheet_status_totals.get("_blank", 0) + 1
                    else:
                        daily_status[iso] = code
                        status_counts[code] = status_counts.get(code, 0) + 1
                        sheet_status_totals[code] = sheet_status_totals.get(code, 0) + 1
                        all_status_codes[code] = all_status_codes.get(code, 0) + 1
                        # Office count (treat code 'O' or string containing 'office')
                        upper = code.strip().upper()
                        if upper == "O" or upper == "OFFICE" or "office" in code.lower():
                            sheet_daily_office[iso] = sheet_daily_office.get(iso, 0) + 1

            sheet_records.append({
                "employee_number": emp_num,
                "name": name,
                "country": country,
                "daily_status": daily_status,
                "status_counts": status_counts,
            })

        whereabouts[sheet_name] = sheet_records
        status_totals_by_month[sheet_name] = sheet_status_totals
        daily_office_count_by_month[sheet_name] = sheet_daily_office
        months_meta.append(month_meta)
        sheets_parsed.append(sheet_name)

    # Build legend from observed codes
    legend: Dict[str, str] = {}
    # Prefer the canonical short codes first (deterministic order)
    for code in sorted(all_status_codes.keys(), key=lambda s: (len(s), s)):
        legend[code] = _whereabouts_legend_label(code)
    # Ensure defaults present even if not observed (useful for UI legend)
    for k, v in _WHEREABOUTS_LEGEND_DEFAULTS.items():
        legend.setdefault(k, v)

    employees_list = list(employees_by_num.values())
    employees_list.sort(key=lambda e: (e.get("employee_number") or ""))

    unique_countries = sorted({e.get("country") for e in employees_list if e.get("country")})
    unique_sectors = sorted({e.get("business_sector") for e in employees_list if e.get("business_sector")})

    return {
        "file_type": "EMPLOYEE_WHEREABOUTS",
        "metadata": {
            "source_file": filename,
            "sheets_parsed": sheets_parsed,
            "sheets_ignored": sheets_ignored,
            "months": months_meta,
            "total_employees": len(employees_list),
            "unique_countries": unique_countries,
            "unique_sectors": unique_sectors,
        },
        "employees": employees_list,
        "whereabouts": whereabouts,
        "legend": legend,
        "aggregates": {
            "by_country": by_country,
            "by_sector": by_sector,
            "daily_office_count_by_month": daily_office_count_by_month,
            "status_totals_by_month": status_totals_by_month,
        },
        "errors": errors,
    }


# ══════════════════════════════════════════════════════════════════════════════
# Unknown / Fallback Parser
# ══════════════════════════════════════════════════════════════════════════════

def _parse_unknown(all_sheets: Dict[str, pd.DataFrame], filename: str) -> Dict:
    """Best-effort generic extraction for unrecognised files."""
    sheets_data: Dict[str, Any] = {}
    for name, df in all_sheets.items():
        # Try to find a header row
        hdr_idx = 0
        for i in range(min(10, len(df))):
            row = df.iloc[i]
            non_blank = _non_blank_vals(row)
            if len(non_blank) >= 3:
                hdr_idx = i
                break
        headers = [_clean(v) or f"col_{j}" for j, v in enumerate(df.iloc[hdr_idx])]
        rows = []
        for i in range(hdr_idx + 1, len(df)):
            row = df.iloc[i]
            if not _non_blank_vals(row):
                continue
            rec = {}
            for j, h in enumerate(headers):
                if j < len(row) and not _is_blank(row.iloc[j]):
                    v = row.iloc[j]
                    # Serialise
                    if isinstance(v, (datetime, date)):
                        rec[h] = v.strftime("%Y-%m-%d")
                    elif isinstance(v, float) and v == int(v):
                        rec[h] = int(v)
                    else:
                        rec[h] = v if not _is_blank(v) else None
            if rec:
                rows.append(rec)
        sheets_data[name] = {"headers": headers, "rows": rows, "row_count": len(rows)}

    return {
        "file_type": "UNKNOWN",
        "metadata": {"source_file": filename},
        "sheets": sheets_data,
        "errors": ["File type could not be determined; generic extraction applied."],
    }


# ══════════════════════════════════════════════════════════════════════════════
# Public API
# ══════════════════════════════════════════════════════════════════════════════

def parse_file(
    source: Union[str, bytes, io.BytesIO],
    filename: str = "",
    is_base64: bool = False,
) -> Dict:
    """
    Parse any supported RR Excel file and return a canonical dict.

    Parameters
    ----------
    source   : file path (str), raw bytes, BytesIO, or base64-encoded string
    filename : original file name (used for type hints and metadata)
    is_base64: True when ``source`` is a base64-encoded string

    Returns
    -------
    dict with keys: file_type, metadata, [type-specific keys], errors
    The dict is always JSON-serialisable.
    """
    if is_base64 and isinstance(source, str):
        try:
            source = base64.b64decode(source)
        except Exception as e:
            return {"file_type": "ERROR", "errors": [f"base64 decode failed: {e}"]}

    all_sheets = _load_workbook(source, filename)
    if not all_sheets:
        return {
            "file_type": "ERROR",
            "metadata": {"source_file": filename},
            "errors": ["Could not load workbook — file may be corrupt or unsupported."],
        }

    file_type = detect_file_type(all_sheets, filename)

    try:
        if file_type == "SOA":
            return _parse_soa(all_sheets, filename)
        elif file_type == "INVOICE_LIST":
            return _parse_invoice_list(all_sheets, filename)
        elif file_type == "OPPORTUNITY_TRACKER":
            return _parse_opportunity_tracker(all_sheets, filename)
        elif file_type == "GLOBAL_HOPPER":
            return _parse_global_hopper(all_sheets, filename)
        elif file_type == "SHOP_VISIT":
            return _parse_shop_visit(all_sheets, filename)
        elif file_type == "SVRG_MASTER":
            return _parse_svrg_master(all_sheets, filename)
        elif file_type == "COMMERCIAL_PLAN":
            return _parse_commercial_plan(all_sheets, filename)
        elif file_type == "EMPLOYEE_WHEREABOUTS":
            return _parse_employee_whereabouts(all_sheets, filename)
        else:
            return _parse_unknown(all_sheets, filename)
    except Exception as e:
        logger.exception("Parser crashed on file '%s'", filename)
        # Attempt generic fallback so backend never gets an empty response
        try:
            fallback = _parse_unknown(all_sheets, filename)
            fallback["errors"].append(f"Primary parser crashed: {e}. Fell back to generic extraction.")
            fallback["file_type"] = f"{file_type}_FALLBACK"
            return fallback
        except Exception as e2:
            return {
                "file_type": "ERROR",
                "metadata": {"source_file": filename},
                "errors": [f"Primary parser crashed: {e}", f"Fallback also failed: {e2}"],
            }


# ══════════════════════════════════════════════════════════════════════════════
# Multi-file session layer — cross-references + combined views
# ══════════════════════════════════════════════════════════════════════════════

# Regex patterns for extracting engine serial numbers from free-text fields
_ESN_IN_TEXT = re.compile(
    r'\bESN\s*(\d{4,6})\b'            # "ESN 10499"
    r'|\b(9[0-9]{4})\b',              # 5-digit 9xxxx Trent 900 serials
    re.IGNORECASE,
)


def _extract_esns_from_text(text: str) -> List[str]:
    """Pull engine serial numbers from a free-text cell."""
    if not text:
        return []
    found: List[str] = []
    for m in _ESN_IN_TEXT.finditer(text):
        val = m.group(1) or m.group(2)
        if val:
            found.append(val.strip())
    return found


def _extract_file_keys(result: Dict, filename: str) -> List[Dict]:
    """
    Walk a parsed-file result and yield every 'linkable' key with enough
    context to reconstruct where it came from.

    Key types emitted
    -----------------
    invoice_ref   — document / invoice reference numbers
    assignment    — assignment / PO cross-reference values
    account       — customer account numbers
    customer      — customer name strings
    esn           — engine serial numbers (extracted from text or explicit fields)
    """
    keys: List[Dict] = []
    ft = result.get("file_type", "UNKNOWN")

    def _k(key_type: str, value: Any, **ctx) -> None:
        """Append a key entry only when value is meaningful."""
        if value is None:
            return
        v = _clean(str(value))
        if not v or v.lower() in ("none", "nan", "unknown"):
            return
        keys.append({"key_type": key_type, "value": v, "file": filename,
                     "file_type": ft, **ctx})

    # ── SOA ──────────────────────────────────────────────────────────────────
    if ft == "SOA":
        customer = result.get("metadata", {}).get("customer_name")
        _k("customer", customer)

        for sec in result.get("sections", []):
            sname = sec["name"]
            for item in sec.get("items", []):
                _k("invoice_ref", item.get("reference"),
                   section=sname, amount=item.get("amount"),
                   days_late=item.get("days_late"), due_date=item.get("due_date"),
                   text=item.get("text"))
                _k("assignment", item.get("assignment"),
                   section=sname, amount=item.get("amount"))
                _k("account", item.get("account"),
                   section=sname)
                # ESNs buried in text / assignment
                for esn in _extract_esns_from_text(item.get("text") or ""):
                    _k("esn", esn, section=sname,
                       amount=item.get("amount"), ref=item.get("reference"))
                for esn in _extract_esns_from_text(item.get("rr_comments") or ""):
                    _k("esn", esn, section=sname, ref=item.get("reference"))

    # ── INVOICE_LIST ─────────────────────────────────────────────────────────
    elif ft == "INVOICE_LIST":
        for item in result.get("items", []):
            _k("invoice_ref", item.get("reference"),
               amount=item.get("amount"), due_date=item.get("due_date"),
               text=item.get("text"))
            _k("assignment", item.get("assignment"),
               amount=item.get("amount"))
            for esn in _extract_esns_from_text(item.get("text") or ""):
                _k("esn", esn, ref=item.get("reference"),
                   amount=item.get("amount"))

    # ── SHOP_VISIT_HISTORY ───────────────────────────────────────────────────
    elif ft == "SHOP_VISIT_HISTORY":
        for event in (result.get("shop_visits", [])
                      + result.get("maintenance_actions", [])
                      + result.get("current_status", [])):
            _k("esn", event.get("serial_number"),
               event_datetime=event.get("event_datetime"),
               operator=event.get("operator"),
               sv_type=event.get("sv_type"),
               sv_location=event.get("sv_location"))

    # ── SVRG_MASTER ──────────────────────────────────────────────────────────
    elif ft == "SVRG_MASTER":
        _k("customer", result.get("metadata", {}).get("customer"))
        for event in result.get("event_entries", {}).get("events", []):
            _k("esn", event.get("engine_serial"),
               date=event.get("date"),
               description=event.get("description"),
               qualification=event.get("qualification"))

    # ── OPPORTUNITY_TRACKER ──────────────────────────────────────────────────
    elif ft == "OPPORTUNITY_TRACKER":
        for sheet, opps in result.get("opportunities", {}).items():
            for opp in opps:
                _k("customer", opp.get("customer"),
                   section=sheet, project=opp.get("project"),
                   programme=opp.get("programme"), status=opp.get("status"))
                _k("programme", opp.get("programme"),
                   section=sheet, project=opp.get("project"),
                   customer=opp.get("customer"))
                _k("project", opp.get("project"),
                   section=sheet, customer=opp.get("customer"),
                   programme=opp.get("programme"), status=opp.get("status"),
                   term_benefit=opp.get("term_benefit"))

        # Opps and Threats cross-references
        for item in result.get("opps_and_threats", {}).get("items", []):
            _k("customer", item.get("customer"),
               project=item.get("project"), programme=item.get("programme"),
               opportunity=item.get("opportunity"), owner=item.get("owner"))
            _k("project", item.get("project"),
               customer=item.get("customer"), programme=item.get("programme"))

        # Project summary cross-references
        for proj in result.get("project_summary", {}).get("projects", []):
            _k("customer", proj.get("customer"),
               project=proj.get("project"), programme=proj.get("programme"),
               crp_margin=proj.get("current_crp_margin"))
            _k("project", proj.get("project"),
               customer=proj.get("customer"), programme=proj.get("programme"))

        # Timeline cross-references
        for ms in result.get("timeline", {}).get("milestones", []):
            _k("project", ms.get("project"),
               customer=ms.get("customer"), current_phase=ms.get("current_phase"))

    return keys


def _build_cross_references(all_results: Dict[str, Dict]) -> Dict:
    """
    Build a cross-reference index from all parsed files in a session.

    Returns
    -------
    {
        "cross_refs": {
            "invoice_ref": {
                "<ref_value>": [
                    {file, file_type, section, amount, days_late, ...},
                    {file, file_type, section, amount, ...},
                ]
            },
            "assignment": { ... },
            "account":    { ... },
            "customer":   { ... },
            "esn":        { ... },
        },
        "stats": {
            "total_keys_extracted": N,
            "cross_file_matches":   M,
            "matches_by_type":      { invoice_ref: x, ... },
        }
    }

    Only entries that appear in 2 or more *distinct* files are included in
    cross_refs (single-file occurrences are noise, not cross-references).
    """
    from collections import defaultdict

    # Group raw keys by (key_type → value → [occurrences])
    index: Dict[str, Dict[str, List[Dict]]] = defaultdict(lambda: defaultdict(list))
    total_extracted = 0

    for filename, result in all_results.items():
        if result.get("file_type") == "ERROR":
            continue
        keys = _extract_file_keys(result, filename)
        total_extracted += len(keys)
        for k in keys:
            kt, val = k["key_type"], k["value"]
            entry = {kk: vv for kk, vv in k.items() if kk not in ("key_type", "value")}
            index[kt][val].append(entry)

    cross_refs: Dict[str, Dict] = {}
    matches_by_type: Dict[str, int] = {}

    for key_type, value_map in index.items():
        multi: Dict[str, List] = {}
        for value, occurrences in value_map.items():
            distinct_files = {occ["file"] for occ in occurrences}
            if len(distinct_files) >= 2:
                multi[value] = occurrences
        if multi:
            cross_refs[key_type] = multi
            matches_by_type[key_type] = len(multi)

    total_matches = sum(matches_by_type.values())

    return {
        "cross_refs": cross_refs,
        "stats": {
            "total_keys_extracted": total_extracted,
            "cross_file_matches":   total_matches,
            "matches_by_type":      matches_by_type,
        },
    }


def _build_combined_open_items(all_results: Dict[str, Dict]) -> List[Dict]:
    """
    Merge all overdue / open items from SOA and INVOICE_LIST files into a
    single flat list, sorted by days_late descending.

    Each item carries _source_file, _source_section, and _file_type so the
    AI/backend always knows which file a record came from.
    """
    combined: List[Dict] = []

    for fname, result in all_results.items():
        ft = result.get("file_type")

        if ft == "SOA":
            for sec in result.get("sections", []):
                for item in sec.get("items", []):
                    if item.get("amount") is not None:
                        combined.append({
                            **item,
                            "_source_file":    fname,
                            "_source_section": sec["name"],
                            "_file_type":      ft,
                        })

        elif ft == "INVOICE_LIST":
            for item in result.get("items", []):
                if item.get("amount") is not None:
                    combined.append({
                        **item,
                        "_source_file":    fname,
                        "_source_section": None,
                        "_file_type":      ft,
                    })

    # Sort: overdue items first (highest days_late), then by amount descending
    combined.sort(
        key=lambda x: (-(x.get("days_late") or 0), -(abs(x.get("amount") or 0)))
    )
    return combined


def parse_session(
    files: List[Dict],
    is_base64: bool = True,
) -> Dict:
    """
    Parse multiple uploaded Excel files and build a unified session object.

    This is the primary entry point when multiple files are uploaded together.
    It:
      1. Parses every file with the appropriate per-type parser.
      2. Builds a cross-reference index linking matching keys across files
         (invoice refs, assignments, account numbers, ESNs, customer names).
      3. Produces a combined open-items list merging SOA + INVOICE_LIST data.
      4. Returns one dict ready to be stored in flask.session.

    Parameters
    ----------
    files     : list of {"name": "...", "data": "<base64 or path>"}
    is_base64 : True when data fields are base64-encoded (browser upload)

    Returns
    -------
    {
        "files": {
            "ETH SOA.xlsx": { <full parse result> },
            "EPI 16.02.xlsx": { <full parse result> },
            ...
        },
        "cross_references": {
            "cross_refs": {
                "invoice_ref": { "1820146074": [<SOA entry>, <EPI entry>], ... },
                "assignment":  { "DEG 9054":   [...], ... },
                "account":     { "1009374":    [...], ... },
                "esn":         { "91020":      [...], ... },
                "customer":    { "Emirates":   [...], ... },
            },
            "stats": {
                "total_keys_extracted": N,
                "cross_file_matches":   M,
                "matches_by_type":      { invoice_ref: x, ... },
            }
        },
        "combined_open_items": [ ... ],   # all SOA+INVOICE items, sorted
        "session_summary": {
            "files_loaded":       N,
            "file_types_present": [...],
            "cross_file_matches": M,
            "session_errors":     [...],
        }
    }
    """
    parsed: Dict[str, Dict] = {}
    session_errors: List[str] = []

    for f in files:
        name = f.get("name", "unknown.xlsx")
        data = f.get("data", "")
        try:
            result = parse_file(data, filename=name, is_base64=is_base64)
            result["original_filename"] = name
            parsed[name] = result
        except Exception as e:
            logger.exception("Session parse failed for '%s'", name)
            session_errors.append(f"Failed to parse '{name}': {e}")
            parsed[name] = {
                "file_type":          "ERROR",
                "metadata":           {"source_file": name},
                "errors":             [str(e)],
                "original_filename":  name,
            }

    xref             = _build_cross_references(parsed)
    combined_items   = _build_combined_open_items(parsed)
    file_types       = list({r.get("file_type", "UNKNOWN") for r in parsed.values()})

    return {
        "files":               parsed,
        "cross_references":    xref,
        "combined_open_items": combined_items,
        "session_summary": {
            "files_loaded":       len(parsed),
            "file_types_present": file_types,
            "cross_file_matches": xref["stats"]["cross_file_matches"],
            "session_errors":     session_errors,
        },
    }


# ══════════════════════════════════════════════════════════════════════════════
# Convenience: parse a base64 payload as sent by the frontend
# ══════════════════════════════════════════════════════════════════════════════

def parse_upload_payload(payload: Dict) -> Dict:
    """
    Parse the JSON payload sent by app.js to /api/upload.

    Expected format:
        { "files": [ {"name": "...", "data": "<base64>"}, ... ] }

    Returns a full session dict (files + cross_references + combined_open_items).
    The return type changed from List[Dict] → Dict to support multi-file linking.

    server.py migration note
    ------------------------
    Old:  results = parse_upload_payload(payload)   # list
          session["files_data"] = results

    New:  session_data = parse_upload_payload(payload)
          session["files_data"] = session_data["files"]   # individual results
          session["session"]    = session_data             # full session
    """
    files = payload.get("files", [])
    return parse_session(files, is_base64=True)


# ══════════════════════════════════════════════════════════════════════════════
# LEGACY SOA PARSER (migrated verbatim from the old parser.py)
# ------------------------------------------------------------------------------
# These symbols are consumed by server.py:
#   parse_soa_workbook, serialize_parsed_data, aging_bucket, fmt_currency,
#   AGING_ORDER, AGING_COLORS
# Supporting helpers/constants are kept private (leading underscore) to avoid
# collision with the universal parser above.
# ══════════════════════════════════════════════════════════════════════════════

import math as _legacy_math
from collections import OrderedDict as _LegacyOrderedDict
from datetime import timedelta as _legacy_timedelta  # noqa: F401
import json as _legacy_json  # noqa: F401

import numpy as _legacy_np
from openpyxl import load_workbook as _legacy_load_workbook


# ─────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────

RR_NAVY   = "#10069F"
RR_DARK   = "#0C0033"
RR_SILVER = "#C0C0C0"
RR_LIGHT  = "#E8E8EE"
RR_WHITE  = "#FFFFFF"
RR_GOLD   = "#B8860B"
RR_RED    = "#D32F2F"
RR_GREEN  = "#2E7D32"
RR_BLUE2  = "#1565C0"
RR_AMBER  = "#F9A825"

SECTION_COLOURS = [RR_NAVY, "#1565C0", "#5E35B1", "#00838F", "#C62828", "#EF6C00", "#2E7D32", "#6A1B9A"]

AGING_ORDER = ["Current", "1-30 Days", "31-60 Days", "61-90 Days", "91-180 Days", "180+ Days", "Unknown"]
AGING_COLORS = {
    "Current": "#2E7D32", "1-30 Days": "#66BB6A", "31-60 Days": "#F9A825",
    "61-90 Days": "#EF6C00", "91-180 Days": "#D32F2F", "180+ Days": "#B71C1C", "Unknown": "#9E9E9E",
}

# Keywords that signal a section header row
SECTION_KEYWORDS = [
    "charges", "credits", "credit", "totalcare", "familycare", "missioncare",
    "spare parts", "late payment", "interest", "customer respon",
    "customer responsibility", "usable", "offset",
]

SUMMARY_KEYWORDS = ["total", "overdue", "available credit", "total overdue", "net balance"]

HEADER_KEYWORDS = [
    "company", "account", "reference", "document", "date", "amount", "curr",
    "text", "assignment", "arrangement", "comments", "status", "action",
    "days", "late", "lpi", "invoice", "type", "interest", "net due",
]


# ─────────────────────────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────────────────────────

def _is_section_header(row_values: list, col_count: int) -> bool:
    """Return True if this row looks like a section heading."""
    non_empty = [(i, v) for i, v in enumerate(row_values) if v is not None]
    if not non_empty:
        return False
    if len(non_empty) > 3:
        return False
    text = str(non_empty[0][1]).strip().lower()
    try:
        float(text.replace(",", ""))
        return False
    except ValueError:
        pass
    text_clean = text.rstrip(":")
    if text_clean in ("total", "overdue", "available credit", "total overdue", "net balance"):
        return False
    if len(non_empty) == 2:
        try:
            float(str(non_empty[1][1]).replace(",", "").replace("$", "").strip())
            if any(sw in text_clean for sw in ("total", "overdue", "credit", "balance")):
                return False
        except (ValueError, TypeError):
            pass
    return any(kw in text for kw in SECTION_KEYWORDS)


def _is_header_row(row_values: list) -> bool:
    """Return True if row looks like a column header row."""
    non_empty = [v for v in row_values if v is not None]
    if len(non_empty) < 4:
        return False
    for v in non_empty:
        try:
            n = float(str(v).replace(",", "").replace("$", "").strip())
            if abs(n) > 100:
                return False
        except (ValueError, TypeError):
            pass
    short_texts = [str(v).strip().lower() for v in non_empty if len(str(v).strip()) < 35]
    if len(short_texts) < 3:
        return False
    hits = sum(1 for t in short_texts for kw in HEADER_KEYWORDS if kw in t)
    return hits >= 3


def _is_summary_row(row_values: list):
    """Return the summary type if this looks like a Total/Overdue row, else None."""
    for v in row_values:
        if v is None:
            continue
        t = str(v).strip().lower().rstrip(":")
        if len(t) > 25:
            continue
        if t in ("total", "overdue", "available credit", "total overdue",
                 "net balance", "total:", "overdue:"):
            return t.rstrip(":")
    return None


def _find_amount_col(header: list):
    """Find the column index that holds amounts."""
    for i, h in enumerate(header):
        if h is None:
            continue
        hl = str(h).lower()
        if "amount" in hl:
            return i
    return None


def _coerce_amount(val):
    """Convert a cell value to a float, handling $, commas, etc."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace("$", "").replace(" ", "")
    if s in ("", "-", "$ -", "$-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _coerce_date(val):
    """Try to parse a date from various formats. Returns datetime or None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, pd.Timestamp):
        return val.to_pydatetime()
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _coerce_int(val):
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _normalise_header(raw: list) -> list:
    """Clean header names so they are consistent."""
    result = []
    for h in raw:
        if h is None:
            result.append(None)
            continue
        s = str(h).strip()
        s = re.sub(r"\s+", " ", s)
        result.append(s)
    return result


def _map_columns(header: list) -> dict:
    """Map semantic roles to column indices from the header row."""
    mapping = {}
    hl = [str(h).lower() if h else "" for h in header]
    for i, h in enumerate(hl):
        if not h:
            continue
        if "amount" in h:
            mapping["amount"] = i
        elif h in ("curr", "currency"):
            mapping["currency"] = i
        elif "net due" in h or "due date" in h:
            mapping["due_date"] = i
        elif "document" in h and "date" in h:
            mapping["doc_date"] = i
        elif "document" in h and "no" in h:
            mapping["doc_no"] = i
        elif "invoice date" in h:
            mapping["doc_date"] = i
        elif h == "reference" or "reference" in h:
            mapping["reference"] = i
        elif h == "company" or "company" in h:
            mapping["company"] = i
        elif h == "account" or "account" in h:
            mapping["account"] = i
        elif "text" == h:
            mapping["text"] = i
        elif "assignment" in h or "arrangement" in h:
            mapping["assignment"] = i
        elif "r-r comment" in h or "rr comment" in h:
            mapping["rr_comments"] = i
        elif "action" in h or "reqd" in h:
            mapping["action_owner"] = i
        elif "days" in h and "late" in h:
            mapping["days_late"] = i
        elif "rata" in h:
            mapping["rata_date"] = i
        elif "comment" in h and "r-r" not in h and "rr" not in h:
            mapping["customer_comments"] = i
        elif "status" in h:
            mapping["status"] = i
        elif "customer" in h and "comment" not in h and "name" not in h and "n" not in h and "respon" not in h:
            mapping["customer_name"] = i
        elif "lpi" in h:
            mapping["lpi_cumulated"] = i
        elif "etr" in h or "po" in h or "pr" in h:
            mapping["po_reference"] = i
        elif "type" in h:
            mapping["type"] = i
        elif "interest" in h or "calc" in h:
            mapping["interest_method"] = i

    if "doc_date" not in mapping:
        for i, h in enumerate(hl):
            if "date" in h and i not in mapping.values():
                mapping["doc_date"] = i
                break
    if "due_date" not in mapping:
        for i, h in enumerate(hl):
            if "due" in h and i not in mapping.values():
                mapping["due_date"] = i
                break
    return mapping


# ─────────────────────────────────────────────────────────────
# AGING HELPERS
# ─────────────────────────────────────────────────────────────

def aging_bucket(days) -> str:
    """Classify days late into aging buckets."""
    if days is None or (isinstance(days, float) and _legacy_math.isnan(days)):
        return "Unknown"
    d = int(days)
    if d <= 0:
        return "Current"
    elif d <= 30:
        return "1-30 Days"
    elif d <= 60:
        return "31-60 Days"
    elif d <= 90:
        return "61-90 Days"
    elif d <= 180:
        return "91-180 Days"
    else:
        return "180+ Days"


def fmt_currency(val, short=False):
    """Format a number as USD currency string."""
    if val is None or (isinstance(val, float) and _legacy_math.isnan(val)):
        return "—"
    neg = val < 0
    av = abs(val)
    if short and av >= 1_000_000:
        s = f"${av/1_000_000:,.2f}M"
    elif short and av >= 1_000:
        s = f"${av/1_000:,.1f}K"
    else:
        s = f"${av:,.2f}"
    return f"-{s}" if neg else s


# ─────────────────────────────────────────────────────────────
# MAIN PARSE FUNCTION
# ─────────────────────────────────────────────────────────────

def parse_soa_workbook(file) -> dict:
    """Parse a Rolls-Royce Statement of Account workbook.

    Returns a dict:
        metadata   : dict of customer info, LPI rate, avg days late, etc.
        sections   : OrderedDict  section_name -> { header, colmap, rows (list[dict]), totals }
        all_items  : list[dict]  flattened across all sections (JSON-serializable)
        grand_totals : dict
    """
    wb = _legacy_load_workbook(file, data_only=True)

    # Global accumulators across all sheets
    all_metadata = {}
    all_sections = _LegacyOrderedDict()
    all_items_list = []

    # Iterate over all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_col = ws.max_column or 20
        all_rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_col, values_only=True):
            all_rows.append(list(row))

        if not all_rows:
            continue

        # ---- PASS 1: Metadata ----
        # We try to fill metadata from every sheet, favoring the first non-empty values we find.
        # If all_metadata is already populated, we might just update missing fields.

        local_metadata = {}
        for idx, row in enumerate(all_rows[:15]):
            joined = " ".join(str(v) for v in row if v is not None).lower()
            for v in row:
                if v is None:
                    continue
                s = str(v).strip()
                sl = s.lower()
                if "statement of account" in sl:
                    local_metadata["title"] = s
                if "customer" in sl and ("name" in sl or ":" in sl) and "customer_name" not in local_metadata:
                    vi = row.index(v)
                    for nv in row[vi+1:]:
                        if nv is not None:
                            local_metadata["customer_name"] = str(nv).strip()
                            break
                if ("customer" in sl and ("#" in sl or "n" in sl and ":" in sl)) or "customer n" in sl:
                    vi = row.index(v)
                    for nv in row[vi+1:]:
                        if nv is not None:
                            local_metadata["customer_id"] = str(nv).strip()
                            break
                if "contact" in sl:
                    vi = row.index(v)
                    for nv in row[vi+1:]:
                        if nv is not None:
                            local_metadata["contact"] = str(nv).strip()
                            break
                if "lpi" in sl or "lp ratio" in sl or "lp rate" in sl:
                    for nv in row:
                        if nv is None:
                            continue
                        nv_str = str(nv).strip()
                        if "%" in nv_str:
                            try:
                                local_metadata["lpi_rate"] = float(nv_str.replace("%", "")) / 100.0
                                break
                            except ValueError:
                                pass
                        amt = _coerce_amount(nv)
                        if amt is not None and 0 < abs(amt) < 1:
                            local_metadata["lpi_rate"] = amt
                            break
                if "average days late" in sl or "avg days late" in sl or "average days late" in joined:
                    for nv in row:
                        if nv is None:
                            continue
                        val = _coerce_int(nv)
                        if val is not None and val > 0:
                            local_metadata["avg_days_late"] = val
                            break
                if "today" in sl:
                    for nv in row:
                        d = _coerce_date(nv)
                        if d is not None:
                            local_metadata["report_date"] = d
                            break

        # Merge local metadata into global if not present
        for k, v in local_metadata.items():
            if k not in all_metadata or all_metadata[k] is None:
                all_metadata[k] = v

        # ---- PASS 2: Identify section boundaries and headers ----
        master_header = None
        master_header_idx = None
        sections_info = []

        for idx, row in enumerate(all_rows):
            if _is_header_row(row) and master_header is None:
                master_header = _normalise_header(row)
                master_header_idx = idx
                continue
            if _is_section_header(row, max_col):
                name = str([v for v in row if v is not None][0]).strip()
                sections_info.append({"name": name, "start": idx})

        for i, sec in enumerate(sections_info):
            if i + 1 < len(sections_info):
                sec["end"] = sections_info[i + 1]["start"]
            else:
                sec["end"] = len(all_rows)

        # ---- PASS 3: Parse each section ----
        for sec in sections_info:
            sec_name = sec["name"]
            start = sec["start"]
            end = sec["end"]

            header = master_header
            header_idx = master_header_idx
            col_map = None

            for offset in range(1, 4):
                ri = start + offset
                if ri >= end:
                    break
                if _is_header_row(all_rows[ri]):
                    header = _normalise_header(all_rows[ri])
                    header_idx = ri
                    break

            if header:
                col_map = _map_columns(header)
            else:
                col_map = {}

            amt_idx = col_map.get("amount")
            if amt_idx is None and header:
                amt_idx = _find_amount_col(header)
                if amt_idx is not None:
                    col_map["amount"] = amt_idx

            data_rows = []
            totals = {}
            data_start = (header_idx + 1) if header_idx and header_idx >= start else start + 1

            for ri in range(data_start, end):
                row = all_rows[ri]
                summary_type = _is_summary_row(row)
                if summary_type:
                    for v in row:
                        amt = _coerce_amount(v)
                        if amt is not None:
                            totals[summary_type] = amt
                            break
                    continue

                if _is_section_header(row, max_col):
                    continue
                if _is_header_row(row):
                    header = _normalise_header(row)
                    col_map = _map_columns(header)
                    amt_idx = col_map.get("amount")
                    if amt_idx is None and header:
                        amt_idx = _find_amount_col(header)
                        if amt_idx is not None:
                            col_map["amount"] = amt_idx
                    continue

                amt_val = None
                if amt_idx is not None and amt_idx < len(row):
                    amt_val = _coerce_amount(row[amt_idx])
                if amt_val is None:
                    for ci, cv in enumerate(row):
                        a = _coerce_amount(cv)
                        if a is not None and abs(a) > 0.01:
                            if abs(a) > 100 or (col_map.get("days_late") is not None and ci != col_map.get("days_late")):
                                amt_val = a
                                break

                if amt_val is None:
                    continue

                record = {
                    "Section": sec_name,
                    "Amount": amt_val,
                }

                def _get(key, coerce=str):
                    ci = col_map.get(key)
                    if ci is None or ci >= len(row):
                        return None
                    v = row[ci]
                    if v is None:
                        return None
                    if coerce == float:
                        return _coerce_amount(v)
                    if coerce == "date":
                        return _coerce_date(v)
                    if coerce == int:
                        return _coerce_int(v)
                    return str(v).strip()

                record["Company"]           = _get("company")
                record["Account"]           = _get("account")
                record["Reference"]         = _get("reference")
                record["Document Date"]     = _get("doc_date", "date")
                record["Due Date"]          = _get("due_date", "date")
                record["Currency"]          = _get("currency")
                record["Text"]              = _get("text")
                record["Assignment"]        = _get("assignment")
                record["R-R Comments"]      = _get("rr_comments")
                record["Action Owner"]      = _get("action_owner")
                record["Days Late"]         = _get("days_late", int)
                record["Customer Comments"] = _get("customer_comments")
                record["Status"]            = _get("status")
                record["PO Reference"]      = _get("po_reference")
                record["LPI Cumulated"]     = _get("lpi_cumulated")
                record["Type"]              = _get("type")
                record["Document No"]       = _get("doc_no")
                record["Interest Method"]   = _get("interest_method")
                record["Customer Name"]     = _get("customer_name")

                # Auto-compute Days Late from Due Date
                if record["Days Late"] is None and record["Due Date"] is not None:
                    try:
                        due = record["Due Date"]
                        # Use report date if available, else today
                        anchor_date = all_metadata.get("report_date") or datetime.now()
                        anchor_date = anchor_date.replace(hour=0, minute=0, second=0, microsecond=0)

                        if due < anchor_date:
                            record["Days Late"] = (anchor_date - due).days
                        else:
                            record["Days Late"] = 0
                    except Exception:
                        pass

                # Derive a unified Status field
                if not record.get("Status"):
                    for field in ["R-R Comments", "Action Owner", "Customer Comments"]:
                        v = record.get(field, "")
                        if v and any(kw in v.lower() for kw in [
                            "ready for payment", "under approval", "under review",
                            "dispute", "ongoing", "et to process", "payment pending",
                            "invoice sent", "credit note", "approved",
                            "transfer", "invoice approved", "pending for payment",
                        ]):
                            record["Status"] = v
                            break
                if not record.get("Status"):
                    rrc = record.get("R-R Comments", "")
                    if rrc:
                        record["Status"] = rrc

                record["Entry Type"] = "Credit" if amt_val < 0 else "Charge"

                data_rows.append(record)
                all_items_list.append(record)

            # --- Merge logic used to combine sections across sheets ---
            if sec_name not in all_sections:
                all_sections[sec_name] = {
                    "header": header,
                    "colmap": col_map,
                    "rows": data_rows,
                    "totals": totals,
                }
            else:
                # Append rows
                all_sections[sec_name]["rows"].extend(data_rows)
                # Aggregate totals (sum numeric values)
                existing_totals = all_sections[sec_name]["totals"]
                for k, v in totals.items():
                    if k in existing_totals:
                        existing_totals[k] += v
                    else:
                        existing_totals[k] = v

    # ---- Grand totals (Recalculate from aggregated items) ----
    df = pd.DataFrame(all_items_list)
    if df.empty:
        df = pd.DataFrame(columns=["Section", "Amount", "Entry Type"])

    grand = {}
    for sec_name, sec_data in all_sections.items():
        for k, v in sec_data["totals"].items():
            if "total overdue" in k:
                grand["total_overdue"] = grand.get("total_overdue", 0) + v
            elif "overdue" in k:
                grand.setdefault("section_overdue", {})[sec_name] = grand.setdefault("section_overdue", {}).get(sec_name, 0) + v
            elif "available credit" in k:
                grand.setdefault("available_credits", {})[sec_name] = grand.setdefault("available_credits", {}).get(sec_name, 0) + v
            elif "total" in k:
                grand.setdefault("section_totals", {})[sec_name] = grand.setdefault("section_totals", {}).get(sec_name, 0) + v

    if not df.empty:
        grand["total_charges"] = float(df.loc[df["Amount"] > 0, "Amount"].sum())
        grand["total_credits"] = float(df.loc[df["Amount"] < 0, "Amount"].sum())
        grand["net_balance"]   = float(df["Amount"].sum())
        grand["item_count"]    = int(len(df))
        if "total_overdue" not in grand:
            overdue_sum = sum(grand.get("section_overdue", {}).values())
            if overdue_sum:
                grand["total_overdue"] = overdue_sum
            else:
                # Fallback: sum of positive amounts if no explicit overdue field found?
                # Or just use total charges?
                # Existing logic used section_overdue. If not present, maybe use net_balance or 0.
                pass

    return {
        "metadata": all_metadata,
        "sections": all_sections,
        "all_items": all_items_list,
        "grand_totals": grand,
    }


# ─────────────────────────────────────────────────────────────
# JSON SERIALIZATION HELPERS
# ─────────────────────────────────────────────────────────────

def _serialize_value(val):
    """Convert a single value to JSON-serializable form."""
    if val is None:
        return None
    if isinstance(val, (datetime, pd.Timestamp)):
        return val.isoformat()
    if isinstance(val, float):
        if _legacy_math.isnan(val) or _legacy_math.isinf(val):
            return None
        return val
    if isinstance(val, (_legacy_np.integer,)):
        return int(val)
    if isinstance(val, (_legacy_np.floating,)):
        f = float(val)
        return None if _legacy_math.isnan(f) else f
    if isinstance(val, (_legacy_np.bool_,)):
        return bool(val)
    return val


def serialize_parsed_data(parsed: dict) -> dict:
    """Convert parsed workbook data to fully JSON-serializable dict.

    Handles: datetime -> ISO string, NaN -> None, numpy types -> Python types,
    OrderedDict -> dict, DataFrame -> list of dicts.
    """
    result = {}

    # Metadata
    meta = {}
    for k, v in parsed["metadata"].items():
        meta[k] = _serialize_value(v)
    result["metadata"] = meta

    # Sections
    sections = {}
    for sec_name, sec_data in parsed["sections"].items():
        rows = []
        for row in sec_data["rows"]:
            rows.append({k: _serialize_value(v) for k, v in row.items()})
        sections[sec_name] = {
            "header": sec_data.get("header"),
            "rows": rows,
            "totals": {k: _serialize_value(v) for k, v in sec_data.get("totals", {}).items()},
        }
    result["sections"] = sections

    # All items (already list of dicts from our modified parser)
    items = parsed.get("all_items", [])
    if isinstance(items, pd.DataFrame):
        items = items.to_dict("records")
    result["all_items"] = [
        {k: _serialize_value(v) for k, v in row.items()} for row in items
    ]

    # Grand totals
    grand = {}
    for k, v in parsed["grand_totals"].items():
        if isinstance(v, dict):
            grand[k] = {sk: _serialize_value(sv) for sk, sv in v.items()}
        else:
            grand[k] = _serialize_value(v)
    result["grand_totals"] = grand

    return result
