"""
Build a self-contained HTML visualization for Rutish_Airways_Statement_of_Account.xlsx (SOA).

Ground-truth benchmark independent of V6/parser.py. Reads the xlsx directly with openpyxl.
Emits V6/TESTEXCEL/Rutish_Airways_SOA.html embedding all data as JSON, rendered with
ApexCharts 3.49.0 via CDN.

File characteristics (from audit):
  Sheets:       SOA.26.1.26 | Offset (empty) | Payment (15 rows)
  Main sheet:   104 x 15, 4 sections demarcated by merged banner rows
                  A9:D9   -> FamilyCare Charges          (canonical: TotalCare)
                  A15:D15 -> Customer Responsibility     (canonical: CRC)
                  A56:D56 -> Spare Parts Charges         (canonical: Spare Parts)
                  A87:D87 -> Late Payment Interest       (canonical: LPI)
  Dates are strings in dd/mm/yyyy format.
  Section schemas differ - each has its own header row directly below the banner.
  Metadata in rows 1-8: customer name/number, contact email, LP ratio, avg days late.
"""

from __future__ import annotations

import json
from datetime import datetime, date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SRC_PATH = Path(
    r"C:\Users\Rutishkrishna\Desktop\RR\RR Powerbi\V6\New info\Rutish_Airways_Statement_of_Account.xlsx"
)
OUT_PATH = Path(
    r"C:\Users\Rutishkrishna\Desktop\RR\RR Powerbi\V6\TESTEXCEL\Rutish_Airways_SOA.html"
)

# Today's date for aging computation
TODAY = datetime(2026, 4, 22)

AGING_BUCKETS = [
    "Not due",
    "0-30",
    "31-60",
    "61-90",
    "91-180",
    "181-365",
    "365+",
]

# Canonical section mapping. Keys are substrings that identify banner text.
SECTION_CANONICAL = [
    ("familycare",               "TotalCare"),
    ("customer responsibility",  "CRC"),
    ("spare parts",              "Spare Parts"),
    ("late payment interest",    "LPI"),
]

# Each section has a distinct physical schema. For a given canonical name,
# we know WHICH cell holds what semantic value.
# Tuple fields: (due_date_col_idx, amount_col_idx, currency_col_idx,
#                status_col_idx_or_none, text_col_idx_or_none,
#                reference_col_idx, doc_no_col_idx_or_none)
SECTION_SCHEMA = {
    # 1-based column indices. SOA sheet layout:
    #   A=Company, B=Account, C=Reference, D=Doc No, E=Net due date,
    #   F=Amount in due curr, G=Curr, H=Text, I=Arrangement, J=R-R Comments,
    #   K=Action Reqd, L=RATA Date, M=GTR Comments, N=ETR PR, O=LPI Consolidated
    "TotalCare":    {"due": 5, "amount": 6,  "curr": 7,  "status": None, "text": 8,  "ref": 3, "doc": 4, "account": 2},
    "CRC":          {"due": 5, "amount": 6,  "curr": 7,  "status": None, "text": 8,  "ref": 3, "doc": 4, "account": 2},
    # Spare Parts header row 57:
    #   A=Company, B=Account (SP-code), C=Reference (AR-code), D=Due Date,
    #   E=Amount, F=Curr, G=Text, H=Type, I=Arrangement, J=Customer,
    #   K=Status, L=Comments
    "Spare Parts":  {"due": 4, "amount": 5,  "curr": 6,  "status": 11,   "text": 7,  "ref": 3, "doc": None, "account": 2, "customer": 10, "type": 8},
    # LPI header row 88:
    #   A=Company, B=Account (LP-code), C=Reference (dd/mm/yyyy charge date),
    #   D=Invoice Date, E=Amount, F=Curr, G=Type, H=Interest Calc Method, I=Comments
    "LPI":          {"due": 4, "amount": 5,  "curr": 6,  "status": None, "text": 9,  "ref": 2, "doc": None, "account": 2, "type": 7, "calc": 8},
}


def aging_bucket(days_late: int | None) -> str:
    if days_late is None:
        return "Unknown"
    if days_late <= 0:
        return "Not due"
    if days_late <= 30:
        return "0-30"
    if days_late <= 60:
        return "31-60"
    if days_late <= 90:
        return "61-90"
    if days_late <= 180:
        return "91-180"
    if days_late <= 365:
        return "181-365"
    return "365+"


def parse_ddmmyyyy(v: Any) -> datetime | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime.combine(v, datetime.min.time())
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def to_iso(v: Any) -> str | None:
    dt = parse_ddmmyyyy(v)
    if dt is None:
        if v is None or v == "":
            return None
        return str(v)
    return dt.strftime("%Y-%m-%d")


def as_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def infer_status(canonical: str, row: dict[str, Any]) -> str:
    """Infer a status for charting/filtering when source has no explicit status col."""
    explicit = row.get("source_status")
    if explicit:
        return str(explicit)
    amount = row.get("amount") or 0
    days_late = row.get("days_late")
    text = (row.get("text") or "").lower()
    arr  = (row.get("arrangement") or "").lower()
    if canonical == "LPI":
        return "LPI Accrued"
    if isinstance(amount, (int, float)) and amount < 0:
        return "Credit"
    if "credit" in text or "credit" in arr:
        return "Credit"
    if "disput" in text or "disput" in arr:
        return "Disputed"
    if days_late is not None and days_late > 0:
        return "Overdue"
    return "Outstanding"


def find_section_banners(ws) -> list[tuple[int, str, str]]:
    """Scan col A rows 1..max for merged banner rows. Returns list of (row, raw_label, canonical)."""
    merged_ranges = list(ws.merged_cells.ranges)
    banners: list[tuple[int, str, str]] = []
    for rng in merged_ranges:
        if rng.min_col != 1:
            continue
        # Only banner-like merges (4+ columns wide), skip the top title A1:F1
        if rng.min_row <= 1:
            continue
        if rng.max_col - rng.min_col < 3:
            continue
        raw = ws.cell(row=rng.min_row, column=1).value
        if not raw:
            continue
        raw_s = str(raw).strip()
        low = raw_s.lower()
        canonical = None
        for key, can in SECTION_CANONICAL:
            if key in low:
                canonical = can
                break
        if canonical:
            banners.append((rng.min_row, raw_s, canonical))
    banners.sort(key=lambda x: x[0])
    return banners


def extract_metadata(ws) -> dict[str, Any]:
    """Parse rows 1-8 for customer/LPI/contact metadata."""
    meta: dict[str, Any] = {
        "title": None,
        "customer_name": None,
        "customer_number": None,
        "contact_email": None,
        "lp_ratio": None,
        "average_days_late": None,
        "report_date": None,
    }
    meta["title"] = as_str(ws.cell(row=1, column=1).value) or None

    for r in range(2, 9):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v).strip()
            low = s.lower()
            if "customer:" in low or low == "customer":
                nxt = ws.cell(row=r, column=c + 1).value
                if nxt: meta["customer_name"] = str(nxt).strip()
            elif "customer #" in low or "customer#" in low or low.startswith("customer #"):
                nxt = ws.cell(row=r, column=c + 1).value
                if nxt: meta["customer_number"] = str(nxt).strip()
            elif "contact" in low and ":" in low:
                nxt = ws.cell(row=r, column=c + 1).value
                if nxt: meta["contact_email"] = str(nxt).strip()
            elif "lp ratio" in low or "lpi rate" in low or "lp rate" in low:
                nxt = ws.cell(row=r, column=c + 1).value
                if nxt: meta["lp_ratio"] = str(nxt).strip()
            elif "average days late" in low:
                nxt = ws.cell(row=r, column=c + 1).value
                if nxt is not None:
                    meta["average_days_late"] = nxt
            elif "report date" in low or "as of" in low or "statement date" in low:
                nxt = ws.cell(row=r, column=c + 1).value
                if nxt: meta["report_date"] = to_iso(nxt)
    return meta


def is_total_row(row_vals: list[Any]) -> bool:
    """A 'Total:' / 'Overdue:' / 'Available Credit:' row is a summary footer."""
    for v in row_vals[:6]:
        if v is None:
            continue
        s = str(v).strip().lower()
        if s in {"total:", "overdue:", "available credit:", "total overdue:"} or s.startswith("total:"):
            return True
    return False


def parse_section_rows(
    ws,
    canonical: str,
    start_row: int,
    end_row: int,
    schema: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """
    Parse rows [start_row..end_row] for a section. start_row is the FIRST data row
    (i.e. banner_row + 2, skipping the per-section header).
    Returns (items, source_totals) where source_totals are 'Total:/Overdue:/Available Credit:' rows.
    """
    items: list[dict[str, Any]] = []
    source_totals: list[dict[str, Any]] = []
    for r in range(start_row, end_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or v == "" for v in row_vals):
            continue
        if is_total_row(row_vals):
            # capture the label + value (amount is in the col that holds "Total:" + 1)
            label = None
            val = None
            for ci, v in enumerate(row_vals[:6], start=1):
                if v is None:
                    continue
                s = str(v).strip()
                low = s.lower().rstrip(":")
                if low in {"total", "overdue", "available credit", "total overdue"}:
                    label = s
                    # value is in col E (amount col) - grab schema["amount"] - 1 index
                    amt_idx = schema["amount"] - 1
                    if amt_idx < len(row_vals):
                        val = row_vals[amt_idx]
                    break
            source_totals.append({"row": r, "label": label, "value": val})
            continue

        amount = row_vals[schema["amount"] - 1]
        if not isinstance(amount, (int, float)):
            # Skip non-numeric amount rows (defensive)
            continue

        due_raw = row_vals[schema["due"] - 1]
        due_dt = parse_ddmmyyyy(due_raw)
        due_iso = due_dt.strftime("%Y-%m-%d") if due_dt else (str(due_raw) if due_raw else None)
        days_late = int((TODAY - due_dt).days) if due_dt else None

        item = {
            "row": r,
            "section": canonical,
            "account":       as_str(row_vals[schema["account"] - 1]) if schema.get("account") else "",
            "reference":     as_str(row_vals[schema["ref"] - 1]) if schema.get("ref") else "",
            "document_no":   as_str(row_vals[schema["doc"] - 1]) if schema.get("doc") else "",
            "due_date_raw":  as_str(due_raw) if due_raw else "",
            "due_date":      due_iso,
            "amount":        float(amount),
            "currency":      as_str(row_vals[schema["curr"] - 1]) if schema.get("curr") else "",
            "text":          as_str(row_vals[schema["text"] - 1]) if schema.get("text") else "",
            "days_late":     days_late,
            "aging_bucket":  aging_bucket(days_late),
        }
        # Section-specific extras (1-based col indices aligned with row_vals 0-based)
        if canonical in ("TotalCare", "CRC"):
            item["arrangement"]   = as_str(ws.cell(row=r, column=9).value)   # I
            item["rr_comments"]   = as_str(ws.cell(row=r, column=10).value)  # J
            item["action"]        = as_str(ws.cell(row=r, column=11).value)  # K
            item["rata_date"]     = to_iso(ws.cell(row=r, column=12).value)  # L
            item["gtr_comments"]  = as_str(ws.cell(row=r, column=13).value)  # M
            item["etr_pr"]        = as_str(ws.cell(row=r, column=14).value)  # N
            lpi_val = ws.cell(row=r, column=15).value                         # O
            item["lpi_consolidated"] = float(lpi_val) if isinstance(lpi_val, (int, float)) else None
        elif canonical == "Spare Parts":
            item["type"]          = as_str(row_vals[schema["type"] - 1])     # H
            item["arrangement"]   = as_str(row_vals[9 - 1])                   # I
            item["customer"]      = as_str(row_vals[schema["customer"] - 1]) # J
            item["source_status"] = as_str(row_vals[schema["status"] - 1])   # K
            item["comments"]      = as_str(row_vals[12 - 1])                  # L
        elif canonical == "LPI":
            # Col C labeled "Reference" holds an LPI-charge date (string dd/mm/yyyy)
            item["reference"]        = as_str(row_vals[2 - 1])                # B (LP-code)
            item["lpi_account_code"] = as_str(row_vals[2 - 1])                # B
            item["lpi_charge_date"]  = to_iso(row_vals[3 - 1])                # C
            item["type"]             = as_str(row_vals[schema["type"] - 1])   # G
            item["calc_method"]      = as_str(row_vals[schema["calc"] - 1])   # H
            item["comments"]         = as_str(row_vals[schema["text"] - 1])   # I

        item["status"] = infer_status(canonical, item)
        items.append(item)
    return items, source_totals


def parse_payment_sheet(wb) -> dict[str, Any]:
    ws = wb["Payment"]
    header = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column + 1)]
    items = []
    for r in range(4, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or v == "" for v in row_vals):
            continue
        items.append({
            "row": r,
            "date_raw":  as_str(row_vals[0]),
            "date":      to_iso(row_vals[0]),
            "reference": as_str(row_vals[1]),
            "amount":    float(row_vals[2]) if isinstance(row_vals[2], (int, float)) else None,
            "currency":  as_str(row_vals[3]),
            "method":    as_str(row_vals[4]),
            "status":    as_str(row_vals[5]),
        })
    return {"header": [as_str(h) for h in header], "items": items}


def extract() -> dict[str, Any]:
    wb = load_workbook(SRC_PATH, data_only=True)
    ws = wb["SOA.26.1.26"]

    metadata = extract_metadata(ws)
    banners = find_section_banners(ws)

    # Compute per-section row ranges:
    #   FamilyCare's header sits at r8 (ABOVE banner r9). Other sections have a
    #   repeated header row IMMEDIATELY BELOW the banner. Detect dynamically by
    #   checking whether row (banner+1) contains a header-keyword ("Company" in col A).
    # data_end = (next banner row - 1) or ws.max_row
    sections: list[dict[str, Any]] = []
    all_items: list[dict[str, Any]] = []
    for i, (br, raw, canonical) in enumerate(banners):
        data_end = banners[i + 1][0] - 1 if i + 1 < len(banners) else ws.max_row
        # Probe the row right below the banner
        below_a = ws.cell(row=br + 1, column=1).value
        below_a_s = str(below_a).strip().lower() if below_a is not None else ""
        if below_a_s in {"company", "account"}:
            # header immediately below banner
            header_row = br + 1
            data_start = br + 2
        else:
            # Header sits above banner (FamilyCare case) - reuse nearest prior header row
            header_row = br - 1 if ws.cell(row=br - 1, column=1).value else br - 2
            if header_row < 1:
                header_row = br + 1
            data_start = br + 1

        schema = SECTION_SCHEMA[canonical]
        # Capture the per-section header row verbatim
        section_header = [
            as_str(ws.cell(row=header_row, column=c).value)
            for c in range(1, ws.max_column + 1)
        ]
        items, source_totals = parse_section_rows(ws, canonical, data_start, data_end, schema)
        all_items.extend(items)
        total_amount = sum(it["amount"] for it in items)
        sections.append({
            "canonical": canonical,
            "raw_label": raw,
            "banner_row": br,
            "data_start": data_start,
            "data_end": data_end,
            "header": section_header,
            "item_count": len(items),
            "total_amount": total_amount,
            "source_totals": source_totals,
            "items": items,
        })

    # Parse Payment sheet
    payment = parse_payment_sheet(wb)

    # Aggregations across all items
    total_items     = len(all_items)
    total_amount    = sum(i["amount"] for i in all_items)
    total_debits    = sum(i["amount"] for i in all_items if i["amount"] > 0)
    total_credits   = sum(i["amount"] for i in all_items if i["amount"] < 0)
    net_balance     = total_debits + total_credits  # credits are negative
    overdue_items   = [i for i in all_items if i["status"] == "Overdue"]
    total_overdue   = sum(i["amount"] for i in overdue_items)
    avg_days_late   = (
        sum(i["days_late"] for i in overdue_items if i["days_late"] is not None) / len(overdue_items)
        if overdue_items else 0
    )

    # Aging distribution (counts + amounts) - canonical order + "Unknown"
    aging_counts = {b: 0 for b in AGING_BUCKETS}
    aging_amounts = {b: 0.0 for b in AGING_BUCKETS}
    aging_counts["Unknown"] = 0
    aging_amounts["Unknown"] = 0.0
    for i in all_items:
        b = i["aging_bucket"]
        aging_counts[b] = aging_counts.get(b, 0) + 1
        aging_amounts[b] = aging_amounts.get(b, 0.0) + i["amount"]

    # Status distribution (counts + amounts)
    status_counts: dict[str, int] = {}
    status_amounts: dict[str, float] = {}
    for i in all_items:
        st = i["status"]
        status_counts[st] = status_counts.get(st, 0) + 1
        status_amounts[st] = status_amounts.get(st, 0.0) + i["amount"]

    # Section totals for chart (in canonical order)
    section_totals = [
        {"section": s["canonical"], "count": s["item_count"], "amount": s["total_amount"]}
        for s in sections
    ]

    # Top 10 overdue by amount magnitude
    top10 = sorted(overdue_items, key=lambda x: abs(x["amount"]), reverse=True)[:10]

    return {
        "meta": {
            "source_path": str(SRC_PATH),
            "sheet": "SOA.26.1.26",
            "all_sheets": wb.sheetnames,
            "title": metadata["title"],
            "customer_name": metadata["customer_name"],
            "customer_number": metadata["customer_number"],
            "contact_email": metadata["contact_email"],
            "lp_ratio": metadata["lp_ratio"],
            "average_days_late_src": metadata["average_days_late"],
            "report_date": metadata["report_date"],
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "today_anchor": TODAY.strftime("%Y-%m-%d"),
            "row_count": total_items,
            "note": "Synthetic test data - fictional Rutish Airways. Ground-truth benchmark for parser validation.",
        },
        "kpis": {
            "total_items":        total_items,
            "total_amount":       total_amount,
            "total_debits":       total_debits,
            "total_credits":      total_credits,
            "net_balance":        net_balance,
            "total_overdue":      total_overdue,
            "avg_days_late":      avg_days_late,
            "overdue_count":      len(overdue_items),
            "section_count":      len(sections),
        },
        "aging_counts":    aging_counts,
        "aging_amounts":   aging_amounts,
        "status_counts":   status_counts,
        "status_amounts":  status_amounts,
        "section_totals":  section_totals,
        "top10_overdue":   top10,
        "sections":        sections,
        "auxiliary_sheets": {
            "Offset":  {"header": [], "items": [], "note": "empty sheet"},
            "Payment": payment,
        },
        "all_items": all_items,
    }


HTML_TEMPLATE = r"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8" />
<meta name="viewport" content="width=device-width, initial-scale=1" />
<title>Rutish Airways SOA — Benchmark</title>
<script src="https://cdn.jsdelivr.net/npm/apexcharts@3.49.0/dist/apexcharts.min.js"></script>
<style>
  :root {
    --bg: #0b1220;
    --panel: #111a2e;
    --panel-2: #162241;
    --border: #24324f;
    --text: #e6ecf7;
    --muted: #8b99b8;
    --accent: #5aa2ff;
    --ok: #22c55e;
    --warn: #f59e0b;
    --bad: #ef4444;
    --purple: #a78bfa;
  }
  * { box-sizing: border-box; }
  html, body { margin: 0; padding: 0; background: var(--bg); color: var(--text);
               font: 14px/1.45 "Segoe UI", Roboto, system-ui, -apple-system, Arial, sans-serif; }
  a { color: var(--accent); }
  header { padding: 22px 28px; border-bottom: 1px solid var(--border);
           background: linear-gradient(180deg, #0f1a35, #0b1220); }
  header h1 { margin: 0 0 6px; font-size: 22px; font-weight: 600; letter-spacing: .2px; }
  header .sub { color: var(--muted); font-size: 12.5px; }
  header .badge {
    display: inline-block; padding: 2px 8px; border-radius: 999px;
    background: #3a2a14; color: #f5c26b; font-size: 11px; font-weight: 600;
    margin-left: 10px;
  }
  main { padding: 22px 28px 60px; max-width: 1400px; margin: 0 auto; }
  .kpis { display: grid; grid-template-columns: repeat(6, minmax(0,1fr)); gap: 14px; margin: 18px 0 22px; }
  .kpi { background: var(--panel); border: 1px solid var(--border); border-radius: 10px; padding: 14px 16px; }
  .kpi .label { color: var(--muted); font-size: 11.5px; text-transform: uppercase; letter-spacing: .8px; }
  .kpi .val { font-size: 20px; font-weight: 600; margin-top: 6px; }
  .kpi .val.small { font-size: 16px; }
  .kpi.ok .val { color: var(--ok); }
  .kpi.warn .val { color: var(--warn); }
  .kpi.bad .val { color: var(--bad); }
  .charts { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 18px; }
  .charts-3 { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 16px; margin-bottom: 18px; }
  .card { background: var(--panel); border: 1px solid var(--border); border-radius: 10px; padding: 14px 16px; }
  .card h2 { margin: 0 0 12px; font-size: 14px; font-weight: 600; color: #cfd8ec; letter-spacing: .3px; }
  .card h2 .hint { color: var(--muted); font-weight: 400; font-size: 12px; margin-left: 8px; }
  .filters { display: flex; flex-wrap: wrap; gap: 10px; margin: 4px 0 14px; align-items: center; }
  .filters input, .filters select {
    background: var(--panel-2); color: var(--text); border: 1px solid var(--border);
    border-radius: 8px; padding: 8px 10px; font-size: 13px; min-width: 140px;
  }
  .filters input[type=search] { min-width: 260px; }
  .filters .count { color: var(--muted); margin-left: auto; font-size: 12.5px; }
  table { width: 100%; border-collapse: collapse; font-size: 12.5px; }
  th, td { padding: 8px 10px; text-align: left; border-bottom: 1px solid var(--border); }
  thead th { background: var(--panel-2); color: #cfd8ec; font-weight: 600; cursor: pointer; user-select: none; position: sticky; top: 0; }
  thead th:hover { background: #1c2a4e; }
  tbody tr:hover { background: #16213e; }
  td.num { text-align: right; font-variant-numeric: tabular-nums; }
  td.ref { font-family: ui-monospace, Consolas, monospace; font-size: 11.5px; }
  .pill { display: inline-block; padding: 2px 8px; border-radius: 999px; font-size: 11px; font-weight: 600; }
  .pill.Paid        { background: #0f3b22; color: #5ee89a; }
  .pill.Outstanding { background: #0f2a4a; color: #7ab5ff; }
  .pill.Overdue     { background: #3a1414; color: #ff8b8b; }
  .pill.Disputed    { background: #3a2a14; color: #f5c26b; }
  .pill.Credit      { background: #28124a; color: #c7a3ff; }
  .pill.LPI         { background: #1f2937; color: #facc15; }
  .pill.Unknown     { background: #222; color: #aaa; }
  .pill.Shipped     { background: #0f2a4a; color: #7ab5ff; }
  .pill.Delivered   { background: #0f3b22; color: #5ee89a; }
  .pill.Pending     { background: #3a2a14; color: #f5c26b; }
  .pill.Processing  { background: #28124a; color: #c7a3ff; }
  .pill.Completed   { background: #0f3b22; color: #5ee89a; }
  .pill[class*="In"] { background: #28124a; color: #c7a3ff; }
  .pill.LPIAccrued  { background: #1f2937; color: #facc15; }
  .section-wrap { margin-bottom: 14px; }
  details.sec {
    background: var(--panel); border: 1px solid var(--border); border-radius: 10px;
    margin-bottom: 12px; overflow: hidden;
  }
  details.sec > summary {
    list-style: none; cursor: pointer; padding: 14px 16px;
    display: flex; align-items: center; gap: 14px;
  }
  details.sec > summary::-webkit-details-marker { display: none; }
  details.sec > summary .caret { color: var(--muted); transition: transform .15s ease; }
  details.sec[open] > summary .caret { transform: rotate(90deg); }
  details.sec > summary .title { font-weight: 600; font-size: 15px; }
  details.sec > summary .raw { color: var(--muted); font-size: 12px; font-family: ui-monospace, Consolas, monospace; }
  details.sec > summary .rowcount {
    background: var(--panel-2); border: 1px solid var(--border);
    border-radius: 999px; padding: 2px 10px; font-size: 12px; color: #cfd8ec;
    margin-left: auto;
  }
  details.sec > summary .secamt { font-size: 14px; font-weight: 600; color: #cfd8ec; margin-left: 10px; }
  details.sec > .secbody { padding: 0 16px 16px; }
  .table-wrap { max-height: 560px; overflow: auto; border: 1px solid var(--border); border-radius: 10px; }
  footer { padding: 20px 28px; color: var(--muted); font-size: 12px; border-top: 1px solid var(--border); }
  .desc-trunc { max-width: 340px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; display: inline-block; }
  .meta-grid { display: grid; grid-template-columns: repeat(4, 1fr); gap: 10px; margin-top: 10px; }
  .meta-item { background: var(--panel-2); border: 1px solid var(--border);
               border-radius: 8px; padding: 8px 10px; font-size: 12.5px; }
  .meta-item .mlbl { color: var(--muted); font-size: 10.5px; text-transform: uppercase; letter-spacing: .6px; }
  .meta-item .mval { margin-top: 2px; font-weight: 500; word-break: break-word; }
  @media (max-width: 1100px) {
    .kpis { grid-template-columns: repeat(3, 1fr); }
    .charts, .charts-3 { grid-template-columns: 1fr; }
    .meta-grid { grid-template-columns: repeat(2, 1fr); }
  }
</style>
</head>
<body>
<header>
  <h1>Rutish Airways — Statement of Account <span class="badge">SYNTHETIC TEST DATA</span></h1>
  <div class="sub" id="headerSub"></div>
  <div class="meta-grid" id="metaGrid"></div>
</header>
<main>
  <section class="kpis" id="kpis"></section>

  <section class="charts-3">
    <div class="card">
      <h2>Aging distribution <span class="hint">count per bucket</span></h2>
      <div id="chartAging"></div>
    </div>
    <div class="card">
      <h2>Section totals <span class="hint">amount by section</span></h2>
      <div id="chartSections"></div>
    </div>
    <div class="card">
      <h2>Balance overview <span class="hint">debits vs credits</span></h2>
      <div id="chartBalance"></div>
    </div>
  </section>

  <section class="charts">
    <div class="card">
      <h2>Top 10 overdue invoices <span class="hint">by amount magnitude</span></h2>
      <div id="chartTop10"></div>
    </div>
    <div class="card">
      <h2>Payment status <span class="hint">from Payment sheet</span></h2>
      <div id="chartPayStatus"></div>
    </div>
  </section>

  <section class="card">
    <h2>Global filters <span class="hint">apply across all section tables</span></h2>
    <div class="filters">
      <input id="q" type="search" placeholder="Search reference, text, comments..." />
      <select id="fSection"><option value="">All sections</option></select>
      <select id="fStatus"><option value="">All statuses</option></select>
      <select id="fBucket"><option value="">All aging buckets</option></select>
      <span class="count" id="globalCount"></span>
    </div>
  </section>

  <div id="sectionTables" class="section-wrap"></div>

  <details class="sec" open>
    <summary>
      <span class="caret">▶</span>
      <span class="title">Payment sheet</span>
      <span class="raw">sheet: Payment · all-string dates (dd/mm/yyyy)</span>
      <span class="rowcount" id="payRowCount"></span>
    </summary>
    <div class="secbody">
      <div class="table-wrap">
        <table id="payTbl">
          <thead>
            <tr>
              <th data-k="date">Date</th>
              <th data-k="reference">Reference</th>
              <th data-k="amount" class="num">Amount</th>
              <th data-k="currency">Cur</th>
              <th data-k="method">Method</th>
              <th data-k="status">Status</th>
            </tr>
          </thead>
          <tbody id="payBody"></tbody>
        </table>
      </div>
    </div>
  </details>
</main>
<footer>
  Generated <span id="gen-ts"></span> ·
  Source <span id="f-src"></span> ·
  Sheets <span id="f-sheets"></span> ·
  Today anchor <span id="f-today"></span>
  <br/>
  <em>This is synthetic test data (fictional Rutish Airways). Used as a ground-truth benchmark for the production parser.</em>
</footer>

<script id="soa-data" type="application/json">__DATA_JSON__</script>
<script>
(function(){
  const D = JSON.parse(document.getElementById('soa-data').textContent);
  const fmtMoney = (v, ccy) => {
    const n = Number(v || 0);
    const sign = n < 0 ? '-' : '';
    return sign + (ccy ? ccy + ' ' : '$') + Math.abs(n).toLocaleString(undefined, {maximumFractionDigits: 2});
  };
  const fmtMoneyPlain = (v) => {
    const n = Number(v || 0);
    const sign = n < 0 ? '-' : '';
    return sign + '$' + Math.abs(n).toLocaleString(undefined, {maximumFractionDigits: 2});
  };
  const fmtInt = (v) => Number(v || 0).toLocaleString();

  // --- Header / Meta ---
  const subParts = [];
  if (D.meta.customer_name)   subParts.push('Customer <b>' + D.meta.customer_name + '</b>');
  if (D.meta.customer_number) subParts.push('#' + D.meta.customer_number);
  if (D.meta.lp_ratio)        subParts.push('LP ratio <b>' + D.meta.lp_ratio + '</b>');
  subParts.push(D.meta.row_count + ' line items across ' + D.kpis.section_count + ' sections');
  document.getElementById('headerSub').innerHTML = subParts.join(' · ');

  const metaItems = [
    ['Customer',        D.meta.customer_name || '—'],
    ['Customer #',      D.meta.customer_number || '—'],
    ['Contact',         D.meta.contact_email || '—'],
    ['LP ratio',        D.meta.lp_ratio || '—'],
    ['Avg days late (source)', D.meta.average_days_late_src != null ? D.meta.average_days_late_src : '—'],
    ['Report date',     D.meta.report_date || '—'],
    ['Sheet',           D.meta.sheet],
    ['Generated',       D.meta.generated_at],
  ];
  document.getElementById('metaGrid').innerHTML = metaItems.map(([l, v]) =>
    '<div class="meta-item"><div class="mlbl">' + l + '</div><div class="mval">' + v + '</div></div>'
  ).join('');

  // --- KPIs ---
  const kpis = [
    {label:'Total overdue',   val: fmtMoneyPlain(D.kpis.total_overdue), cls:'bad'},
    {label:'Avg days late',   val: Math.round(D.kpis.avg_days_late) + 'd'},
    {label:'LP ratio',        val: D.meta.lp_ratio || '—'},
    {label:'Total credits',   val: fmtMoneyPlain(D.kpis.total_credits), cls:'warn small'},
    {label:'Net balance',     val: fmtMoneyPlain(D.kpis.net_balance), cls:'small'},
    {label:'Total line items',val: fmtInt(D.kpis.total_items)},
  ];
  document.getElementById('kpis').innerHTML = kpis.map(k =>
    '<div class="kpi ' + (k.cls||'').split(' ')[0] + '">' +
    '<div class="label">' + k.label + '</div>' +
    '<div class="val ' + ((k.cls||'').includes('small') ? 'small' : '') + '">' + k.val + '</div>' +
    '</div>'
  ).join('');

  // --- Chart 1: Aging horizontal bar ---
  const agingOrder = ["Not due","0-30","31-60","61-90","91-180","181-365","365+","Unknown"];
  const agingData = agingOrder.map(b => ({b, n: D.aging_counts[b] || 0}));
  new ApexCharts(document.getElementById('chartAging'), {
    chart: {type:'bar', height: 300, toolbar:{show:false}, background:'transparent', foreColor:'#cfd8ec'},
    theme: {mode:'dark'},
    series: [{name:'Items', data: agingData.map(d => d.n)}],
    plotOptions: {bar: {horizontal: true, borderRadius: 3, distributed: true}},
    colors: ['#4ade80','#60a5fa','#fbbf24','#f97316','#ef4444','#b91c1c','#7f1d1d','#475569'],
    xaxis: {categories: agingData.map(d => d.b)},
    dataLabels: {enabled: true, style:{colors:['#fff']}},
    legend: {show: false},
    grid: {borderColor: '#24324f'},
  }).render();

  // --- Chart 2: Section totals horizontal bar ---
  const secLabels = D.section_totals.map(s => s.section);
  const secAmts   = D.section_totals.map(s => s.amount);
  new ApexCharts(document.getElementById('chartSections'), {
    chart: {type:'bar', height: 300, toolbar:{show:false}, background:'transparent', foreColor:'#cfd8ec'},
    theme: {mode:'dark'},
    series: [{name:'Amount', data: secAmts}],
    plotOptions: {bar: {horizontal: true, borderRadius: 3, distributed: true}},
    colors: ['#60a5fa','#f59e0b','#22c55e','#a78bfa'],
    xaxis: {categories: secLabels, labels: {formatter: (v) => fmtMoneyPlain(v)}},
    dataLabels: {enabled: true, formatter: (v) => fmtMoneyPlain(v), style: {colors:['#fff'], fontSize:'11px'}},
    legend: {show: false},
    grid: {borderColor: '#24324f'},
    tooltip: {y: {formatter: (v) => fmtMoneyPlain(v)}},
  }).render();

  // --- Chart 3: Balance overview ---
  new ApexCharts(document.getElementById('chartBalance'), {
    chart: {type:'bar', height: 300, toolbar:{show:false}, background:'transparent', foreColor:'#cfd8ec'},
    theme: {mode:'dark'},
    series: [{name:'Amount', data:[D.kpis.total_debits, D.kpis.total_credits, D.kpis.net_balance]}],
    plotOptions: {bar: {distributed:true, borderRadius: 4, columnWidth:'55%'}},
    colors: ['#60a5fa','#a78bfa','#22c55e'],
    xaxis: {categories: ['Debits','Credits','Net balance'],
            labels:{formatter:(v)=>fmtMoneyPlain(v)}},
    dataLabels: {enabled: true, formatter: (v) => fmtMoneyPlain(v), style: {colors:['#fff'], fontSize:'11px'}},
    legend: {show:false},
    grid: {borderColor: '#24324f'},
    tooltip: {y: {formatter: (v) => fmtMoneyPlain(v)}},
  }).render();

  // --- Chart 4: Top 10 overdue invoices ---
  new ApexCharts(document.getElementById('chartTop10'), {
    chart: {type:'bar', height: 380, toolbar:{show:false}, background:'transparent', foreColor:'#cfd8ec'},
    theme: {mode:'dark'},
    series: [{name:'Overdue amount', data: D.top10_overdue.map(i => i.amount)}],
    plotOptions: {bar: {horizontal: true, borderRadius: 2, distributed: false}},
    colors: ['#ef4444'],
    xaxis: {
      categories: D.top10_overdue.map(i => (i.reference || i.document_no || '(r' + i.row + ')') + ' · ' + i.section),
      labels: {formatter: (v) => fmtMoneyPlain(v)},
    },
    dataLabels: {enabled: true, formatter: (v) => fmtMoneyPlain(v), style:{colors:['#fff'], fontSize:'11px'}},
    legend: {show:false},
    grid: {borderColor: '#24324f'},
    tooltip: {y: {formatter: (v) => fmtMoneyPlain(v)}},
  }).render();

  // --- Chart 5: Payment status donut ---
  const payItems = D.auxiliary_sheets.Payment.items;
  const payStatusCounts = {};
  payItems.forEach(p => { payStatusCounts[p.status] = (payStatusCounts[p.status]||0) + 1; });
  const payLabels = Object.keys(payStatusCounts);
  const payValues = payLabels.map(l => payStatusCounts[l]);
  const payColorMap = {Completed:'#22c55e', Processing:'#a78bfa', Pending:'#f59e0b'};
  new ApexCharts(document.getElementById('chartPayStatus'), {
    chart: {type:'donut', height: 380, background:'transparent', foreColor:'#cfd8ec'},
    theme: {mode:'dark'},
    series: payValues,
    labels: payLabels,
    colors: payLabels.map(l => payColorMap[l] || '#64748b'),
    legend: {position: 'bottom'},
    dataLabels: {enabled: true},
    plotOptions: {pie: {donut: {size:'60%'}}},
  }).render();

  // --- Filter bar wiring ---
  const qEl = document.getElementById('q');
  const fSection = document.getElementById('fSection');
  const fStatus  = document.getElementById('fStatus');
  const fBucket  = document.getElementById('fBucket');
  const globalCountEl = document.getElementById('globalCount');

  const uniq = (arr) => Array.from(new Set(arr)).filter(x => x !== undefined && x !== null && x !== '');
  uniq(D.all_items.map(i => i.section)).sort().forEach(s => {
    const o = document.createElement('option'); o.value=s; o.textContent=s; fSection.appendChild(o);
  });
  uniq(D.all_items.map(i => i.status)).sort().forEach(s => {
    const o = document.createElement('option'); o.value=s; o.textContent=s; fStatus.appendChild(o);
  });
  ["Not due","0-30","31-60","61-90","91-180","181-365","365+","Unknown"].forEach(b => {
    const o = document.createElement('option'); o.value=b; o.textContent=b; fBucket.appendChild(o);
  });

  function truncate(s, n) { s = s || ''; return s.length > n ? s.slice(0, n-1) + '…' : s; }

  // --- Per-section collapsible tables ---
  const sectionHost = document.getElementById('sectionTables');
  const sectionState = {}; // canonical -> {sortKey, sortDir, node}

  function headerFor(section) {
    if (section === "TotalCare" || section === "CRC") {
      return [
        ["reference","Reference"], ["document_no","Doc #"], ["due_date","Due Date"],
        ["amount","Amount","num"], ["currency","Cur"], ["text","Text"],
        ["arrangement","Arrangement"], ["rr_comments","R-R Comments"],
        ["days_late","Days Late","num"], ["aging_bucket","Aging"], ["status","Status"],
      ];
    }
    if (section === "Spare Parts") {
      return [
        ["reference","SP Ref"], ["account","AR-Account"], ["due_date","Due Date"],
        ["amount","Amount","num"], ["currency","Cur"], ["type","Type"],
        ["customer","Customer"], ["source_status","Src Status"],
        ["days_late","Days Late","num"], ["aging_bucket","Aging"], ["comments","Comments"],
      ];
    }
    if (section === "LPI") {
      return [
        ["reference","LP Code"], ["lpi_charge_date","LPI Date"], ["due_date","Invoice Date"],
        ["amount","Amount","num"], ["currency","Cur"], ["calc_method","Calc Method"],
        ["comments","Comments"], ["status","Status"],
      ];
    }
    return [];
  }

  function renderSectionTable(sec) {
    const st = sectionState[sec.canonical];
    const cols = headerFor(sec.canonical);
    const q = qEl.value.trim().toLowerCase();
    const fs = fSection.value, fst = fStatus.value, fb = fBucket.value;

    let rows = sec.items.filter(i => {
      if (fs && i.section !== fs) return false;
      if (fst && i.status !== fst) return false;
      if (fb && i.aging_bucket !== fb) return false;
      if (q) {
        const hay = [i.reference, i.document_no, i.text, i.comments, i.arrangement,
                     i.customer, i.source_status, i.calc_method]
          .filter(Boolean).join(' ').toLowerCase();
        if (!hay.includes(q)) return false;
      }
      return true;
    });

    rows.sort((a,b) => {
      const va = a[st.sortKey], vb = b[st.sortKey];
      if (va == null && vb == null) return 0;
      if (va == null) return 1;
      if (vb == null) return -1;
      if (typeof va === 'number' && typeof vb === 'number') return (va - vb) * st.sortDir;
      return String(va).localeCompare(String(vb)) * st.sortDir;
    });

    st.node.querySelector('.rowcount').textContent = rows.length + ' of ' + sec.items.length + ' rows';
    const thead = '<thead><tr>' + cols.map(c =>
      '<th data-k="' + c[0] + '" ' + (c[2] ? 'class="' + c[2] + '"' : '') + '>' + c[1] + '</th>'
    ).join('') + '</tr></thead>';
    const tbody = '<tbody>' + rows.map(i => '<tr>' + cols.map(c => {
      const k = c[0]; let v = i[k];
      const cls = c[2] || '';
      if (k === 'amount') return '<td class="num">' + fmtMoneyPlain(v) + '</td>';
      if (k === 'status') return '<td><span class="pill ' + (v||'Unknown').replace(/\s+/g,'') + '">' + (v||'') + '</span></td>';
      if (k === 'source_status') return '<td><span class="pill ' + (v||'Unknown').replace(/\s+/g,'') + '">' + (v||'') + '</span></td>';
      if (k === 'text' || k === 'comments' || k === 'rr_comments') {
        const s = String(v || '');
        return '<td><span class="desc-trunc" title="' + s.replace(/"/g,'&quot;') + '">' + truncate(s, 48) + '</span></td>';
      }
      if (k === 'reference' || k === 'document_no' || k === 'account') {
        return '<td class="ref">' + (v || '') + '</td>';
      }
      if (v == null) return '<td' + (cls ? ' class="'+cls+'"' : '') + '></td>';
      return '<td' + (cls ? ' class="'+cls+'"' : '') + '>' + v + '</td>';
    }).join('') + '</tr>').join('') + '</tbody>';
    st.node.querySelector('.secTableHost').innerHTML = '<div class="table-wrap"><table>' + thead + tbody + '</table></div>';

    // sort listeners
    st.node.querySelectorAll('thead th').forEach(th => {
      th.addEventListener('click', () => {
        const k = th.dataset.k; if (!k) return;
        if (st.sortKey === k) st.sortDir = -st.sortDir;
        else { st.sortKey = k; st.sortDir = (k === 'amount' || k === 'days_late') ? -1 : 1; }
        renderSectionTable(sec);
      });
    });
  }

  D.sections.forEach(sec => {
    const detail = document.createElement('details');
    detail.className = 'sec';
    detail.open = true;
    detail.innerHTML =
      '<summary>' +
        '<span class="caret">▶</span>' +
        '<span class="title">' + sec.canonical + '</span>' +
        '<span class="raw">source banner: "' + sec.raw_label + '" · rows ' + sec.data_start + '–' + sec.data_end + '</span>' +
        '<span class="secamt">' + fmtMoneyPlain(sec.total_amount) + '</span>' +
        '<span class="rowcount"></span>' +
      '</summary>' +
      '<div class="secbody">' +
        '<div class="secTableHost"></div>' +
      '</div>';
    sectionHost.appendChild(detail);
    sectionState[sec.canonical] = {
      sortKey: 'amount', sortDir: -1, node: detail,
    };
    renderSectionTable(sec);
  });

  function renderAllSections() {
    let total = 0, shown = 0;
    D.sections.forEach(sec => {
      total += sec.items.length;
      const st = sectionState[sec.canonical];
      renderSectionTable(sec);
      shown += +st.node.querySelector('.rowcount').textContent.split(' ')[0];
    });
    globalCountEl.textContent = shown + ' of ' + total + ' SOA line items';
  }

  [qEl, fSection, fStatus, fBucket].forEach(el => el.addEventListener('input', renderAllSections));
  renderAllSections();

  // --- Payment sheet table ---
  const payBody = document.getElementById('payBody');
  const payRowCount = document.getElementById('payRowCount');
  let paySortKey = 'date', paySortDir = -1;
  function renderPay() {
    const rows = [...payItems].sort((a,b) => {
      const va = a[paySortKey], vb = b[paySortKey];
      if (va == null && vb == null) return 0;
      if (va == null) return 1;
      if (vb == null) return -1;
      if (typeof va === 'number' && typeof vb === 'number') return (va - vb) * paySortDir;
      return String(va).localeCompare(String(vb)) * paySortDir;
    });
    payRowCount.textContent = rows.length + ' rows';
    payBody.innerHTML = rows.map(i => {
      const pillCls = (i.status || 'Unknown').replace(/\s+/g,'');
      return '<tr>' +
        '<td>' + (i.date_raw || '') + '</td>' +
        '<td class="ref">' + (i.reference || '') + '</td>' +
        '<td class="num">' + fmtMoneyPlain(i.amount) + '</td>' +
        '<td>' + (i.currency || '') + '</td>' +
        '<td>' + (i.method || '') + '</td>' +
        '<td><span class="pill ' + pillCls + '">' + (i.status || '') + '</span></td>' +
      '</tr>';
    }).join('');
  }
  document.querySelectorAll('#payTbl thead th').forEach(th => {
    th.addEventListener('click', () => {
      const k = th.dataset.k; if (!k) return;
      if (paySortKey === k) paySortDir = -paySortDir;
      else { paySortKey = k; paySortDir = (k === 'amount') ? -1 : 1; }
      renderPay();
    });
  });
  renderPay();

  // --- Footer ---
  document.getElementById('gen-ts').textContent = D.meta.generated_at;
  document.getElementById('f-src').textContent = D.meta.source_path;
  document.getElementById('f-sheets').textContent = D.meta.all_sheets.join(', ');
  document.getElementById('f-today').textContent = D.meta.today_anchor;
})();
</script>
</body>
</html>
"""


def build() -> None:
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    data = extract()

    payload = json.dumps(data, default=str, ensure_ascii=False).replace("</", "<\\/")
    html = HTML_TEMPLATE.replace("__DATA_JSON__", payload)
    OUT_PATH.write_text(html, encoding="utf-8")

    # Verification report
    meta = data["meta"]
    kpis = data["kpis"]
    print(f"[OK] Source                : {SRC_PATH}")
    print(f"[OK] Sheets                : {meta['all_sheets']}")
    print(f"[OK] Customer name         : {meta['customer_name']!r}")
    print(f"[OK] Customer number       : {meta['customer_number']!r}")
    print(f"[OK] Contact               : {meta['contact_email']!r}")
    print(f"[OK] LP ratio              : {meta['lp_ratio']!r}")
    print(f"[OK] Avg days late (src)   : {meta['average_days_late_src']!r}")
    print()
    print(f"[OK] Sections detected     : {len(data['sections'])}")
    for s in data["sections"]:
        print(f"       {s['canonical']:<12} items={s['item_count']:>3}  "
              f"total={s['total_amount']:>14,.2f}  raw='{s['raw_label']}'")
    print()
    print(f"[OK] Grand total amount    : {kpis['total_amount']:>14,.2f}")
    print(f"[OK] Total debits          : {kpis['total_debits']:>14,.2f}")
    print(f"[OK] Total credits         : {kpis['total_credits']:>14,.2f}")
    print(f"[OK] Net balance           : {kpis['net_balance']:>14,.2f}")
    print(f"[OK] Total overdue (items) : {kpis['total_overdue']:>14,.2f}")
    print(f"[OK] Overdue count         : {kpis['overdue_count']}")
    print(f"[OK] Avg days late (calc)  : {kpis['avg_days_late']:.1f}")
    print()
    print(f"[OK] Aging counts          : {data['aging_counts']}")
    print(f"[OK] Status counts         : {data['status_counts']}")
    print()
    pay = data["auxiliary_sheets"]["Payment"]
    print(f"[OK] Payment sheet rows    : {len(pay['items'])}")
    print(f"[OK] Total line items      : {kpis['total_items']}")
    print()
    print(f"[OK] Output                : {OUT_PATH}")
    print(f"[OK] Size                  : {OUT_PATH.stat().st_size:,} bytes")


if __name__ == "__main__":
    build()
