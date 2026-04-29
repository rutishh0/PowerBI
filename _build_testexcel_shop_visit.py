"""
Builder: Trent 900 Shop Visit History HTML visualization.

Reads the SHOP_VISIT Excel file via openpyxl (independent of V6/parser.py),
classifies each row as Current Status vs Shop Visit, aggregates data for
charts, and emits a self-contained HTML page with embedded JSON.

Output: V6/TESTEXCEL/Trent_900_Shop_Visit_History.html
"""
from __future__ import annotations

import glob
import json
import os
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
NEW_INFO_DIR = HERE / "New info"
OUT_DIR = HERE / "TESTEXCEL"
OUT_HTML = OUT_DIR / "Trent_900_Shop_Visit_History.html"

SOURCE_GLOB = "SV008RV08_Trent 900 Shop Visit History Report*.xlsx"


def find_source() -> Path:
    matches = sorted(NEW_INFO_DIR.glob(SOURCE_GLOB))
    if not matches:
        raise FileNotFoundError(f"No file matching {SOURCE_GLOB} in {NEW_INFO_DIR}")
    # Prefer the latest mtime
    return max(matches, key=lambda p: p.stat().st_mtime)


def to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def iso(dt):
    if isinstance(dt, datetime):
        return dt.isoformat()
    return None


def classify(action_code, sv_type):
    """Classify a row as Current Status vs Shop Visit."""
    if action_code == "Current Status":
        return "current_status"
    # If Shop Visit Type is empty/None, treat as current status per audit
    if sv_type is None or (isinstance(sv_type, str) and not sv_type.strip()):
        return "current_status"
    return "shop_visit"


def engine_family(part_number: str | None) -> str:
    if not part_number:
        return "Unknown"
    # Part number patterns like "TRENT900 [TRENT970-84]"
    # Extract the thing inside brackets
    if "[" in part_number and "]" in part_number:
        inside = part_number.split("[", 1)[1].rstrip("]").strip()
        return inside or "Unknown"
    return part_number


def build():
    src = find_source()
    print(f"[build] Source: {src}")

    wb = load_workbook(src, data_only=True, read_only=True)
    main_sheet_name = wb.sheetnames[0]
    ws = wb[main_sheet_name]

    # Row 2 is the header row.
    header = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
    print(f"[build] Header: {header}")

    # Map header index
    def col(name):
        return header.index(name)

    ix_part = col("Event Item Part Number")
    ix_serial = col("Event Item Serial Number")
    ix_date = col("Event Date Time")
    ix_operator = col("Operator")
    ix_parent_serial = col("Parent Serial Number")
    ix_parent_reg = col("Parent Item Registration")
    ix_action = col("Action Code")
    ix_rework = col("Rework Level")
    ix_sen = col("Service Event Number")
    ix_hsn = col("HSN")
    ix_csn = col("CSN")
    ix_hssv = col("HSSV")
    ix_cssv = col("CSSV")
    ix_svtype = col("ShopVisit_Type")
    ix_svloc = col("ShopVisit_Location")

    shop_visits = []
    current_status = []
    all_dates: list[datetime] = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        if row is None or all(v is None for v in row):
            continue

        part = row[ix_part]
        serial = row[ix_serial]
        dt = row[ix_date]
        operator = row[ix_operator]
        parent_serial = row[ix_parent_serial]
        parent_reg = row[ix_parent_reg]
        action = row[ix_action]
        rework = row[ix_rework]
        sen = row[ix_sen]
        hsn = to_float(row[ix_hsn])
        csn = to_float(row[ix_csn])
        hssv = to_float(row[ix_hssv])
        cssv = to_float(row[ix_cssv])
        svtype = row[ix_svtype]
        svloc = row[ix_svloc]

        # Normalize blanks
        if isinstance(svtype, str) and not svtype.strip():
            svtype = None
        if isinstance(svloc, str) and not svloc.strip():
            svloc = None

        rec = {
            "part_number": str(part) if part is not None else None,
            "engine_family": engine_family(part),
            "serial_number": str(serial) if serial is not None else None,
            "datetime": iso(dt),
            "year": dt.year if isinstance(dt, datetime) else None,
            "operator": operator,
            "parent_serial": str(parent_serial) if parent_serial is not None else None,
            "parent_reg": parent_reg,
            "action_code": action,
            "rework_level": rework,
            "service_event_number": str(sen).strip() if sen is not None else None,
            "hsn": hsn,
            "csn": csn,
            "hssv": hssv,
            "cssv": cssv,
            "shop_visit_type": svtype,
            "shop_visit_location": svloc,
        }

        kind = classify(action, svtype)
        if kind == "current_status":
            current_status.append(rec)
        else:
            shop_visits.append(rec)

        if isinstance(dt, datetime):
            all_dates.append(dt)

    # ---- Aggregates for charts ----
    # Chart 1: per year
    year_counts = Counter()
    for r in shop_visits:
        if r["year"] is not None:
            year_counts[r["year"]] += 1
    per_year = [{"year": y, "count": c} for y, c in sorted(year_counts.items())]

    # Chart 2: per operator (top 10)
    op_counts = Counter(r["operator"] or "(blank)" for r in shop_visits)
    by_operator = [
        {"operator": k, "count": v}
        for k, v in op_counts.most_common(10)
    ]

    # Chart 3: per location
    loc_counts = Counter(r["shop_visit_location"] or "(blank)" for r in shop_visits)
    by_location = [
        {"location": k, "count": v}
        for k, v in sorted(loc_counts.items(), key=lambda kv: -kv[1])
    ]

    # Chart 4: per type
    type_counts = Counter(r["shop_visit_type"] or "(blank)" for r in shop_visits)
    by_type = [
        {"type": k, "count": v}
        for k, v in sorted(type_counts.items(), key=lambda kv: -kv[1])
    ]

    # Chart 5: engine lifeline (scatter).
    # Limit to top-N engines by visit count so the chart stays readable but
    # covers the bulk of events.
    engine_visit_counts = Counter(r["serial_number"] for r in shop_visits if r["serial_number"])
    top_engines = [sn for sn, _ in engine_visit_counts.most_common(60)]
    # Keep ordering deterministic by sorting those serials numerically where possible
    def _sort_key(sn):
        try:
            return (0, int(sn))
        except Exception:
            return (1, sn)
    top_engines_sorted = sorted(top_engines, key=_sort_key)

    lifeline_points = []
    for r in shop_visits:
        if r["serial_number"] in top_engines and r["datetime"]:
            lifeline_points.append({
                "serial": r["serial_number"],
                "datetime": r["datetime"],
                "hsn": r["hsn"],
                "hssv": r["hssv"],
                "type": r["shop_visit_type"],
                "location": r["shop_visit_location"],
                "operator": r["operator"],
            })

    # Chart 6: HSN histogram (bucket width 5000 hours)
    hsn_values = [r["hsn"] for r in shop_visits if r["hsn"] is not None and r["hsn"] > 0]
    bucket = 5000
    max_hsn = max(hsn_values) if hsn_values else 0
    buckets = []
    if hsn_values:
        top = int(((max_hsn // bucket) + 1) * bucket)
        edges = list(range(0, top + 1, bucket))
        counts = [0] * (len(edges) - 1)
        for v in hsn_values:
            idx = min(int(v // bucket), len(counts) - 1)
            counts[idx] += 1
        for i in range(len(counts)):
            buckets.append({
                "label": f"{edges[i]:,}-{edges[i+1]:,}",
                "lo": edges[i],
                "hi": edges[i+1],
                "count": counts[i],
            })

    # Chart 7: CSN vs HSN scatter
    csn_hsn_pairs = [
        {"hsn": r["hsn"], "csn": r["csn"], "serial": r["serial_number"]}
        for r in shop_visits
        if r["hsn"] is not None and r["csn"] is not None and r["hsn"] > 0 and r["csn"] > 0
    ]

    # ---- KPIs ----
    total_shop_visits = len(shop_visits)
    total_current_status = len(current_status)
    total_rows = total_shop_visits + total_current_status
    unique_engines = len({r["serial_number"] for r in shop_visits if r["serial_number"]})
    unique_operators = len({r["operator"] for r in shop_visits if r["operator"]})
    avg_hsn = (sum(r["hsn"] for r in shop_visits if r["hsn"] is not None) /
               sum(1 for r in shop_visits if r["hsn"] is not None)) if shop_visits else 0
    earliest = min((r["datetime"] for r in shop_visits if r["datetime"]), default=None)
    latest = max((r["datetime"] for r in shop_visits if r["datetime"]), default=None)

    # Engine family breakdown
    family_counts = Counter(r["engine_family"] for r in shop_visits if r["engine_family"])
    families = [{"family": k, "count": v} for k, v in family_counts.most_common()]

    # Filter dropdown values
    operators_list = sorted({r["operator"] for r in shop_visits if r["operator"]})
    locations_list = sorted({r["shop_visit_location"] for r in shop_visits if r["shop_visit_location"]})
    types_list = sorted({r["shop_visit_type"] for r in shop_visits if r["shop_visit_type"]})
    years_list = sorted({r["year"] for r in shop_visits if r["year"] is not None})

    payload = {
        "meta": {
            "source_file": src.name,
            "sheet": main_sheet_name,
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "total_rows": total_rows,
            "shop_visit_rows": total_shop_visits,
            "current_status_rows": total_current_status,
            "date_range": {
                "earliest": earliest,
                "latest": latest,
            },
        },
        "kpis": {
            "total_shop_visits": total_shop_visits,
            "unique_engines": unique_engines,
            "unique_operators": unique_operators,
            "avg_hsn_at_visit": round(avg_hsn, 1) if avg_hsn else 0,
            "current_status_rows": total_current_status,
            "earliest": earliest,
            "latest": latest,
        },
        "charts": {
            "per_year": per_year,
            "by_operator": by_operator,
            "by_location": by_location,
            "by_type": by_type,
            "lifeline": {
                "serials": top_engines_sorted,
                "points": lifeline_points,
            },
            "hsn_histogram": buckets,
            "csn_vs_hsn": csn_hsn_pairs,
            "engine_families": families,
        },
        "filters": {
            "operators": operators_list,
            "locations": locations_list,
            "types": types_list,
            "years": years_list,
        },
        "shop_visits": shop_visits,
        "current_status": current_status,
    }

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    html = render_html(payload)
    OUT_HTML.write_text(html, encoding="utf-8")

    print(f"[build] shop_visits={total_shop_visits}  current_status={total_current_status}  total={total_rows}")
    print(f"[build] Unique engines (in shop visits): {unique_engines}")
    print(f"[build] Unique operators: {unique_operators}")
    print(f"[build] Date range: {earliest} -> {latest}")
    print(f"[build] Top operators: {op_counts.most_common(5)}")
    print(f"[build] Engine families: {family_counts.most_common()}")
    print(f"[build] Wrote: {OUT_HTML}")


def render_html(payload: dict) -> str:
    # JSON default for datetime-as-ISO (we already convert via iso())
    data_json = json.dumps(payload, ensure_ascii=False, default=str)
    template = _HTML_TEMPLATE
    return template.replace("__DATA_JSON__", data_json)


_HTML_TEMPLATE = r"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Trent 900 Shop Visit History</title>
<meta name="viewport" content="width=device-width, initial-scale=1">
<script src="https://cdn.jsdelivr.net/npm/apexcharts@3.49.0/dist/apexcharts.min.js"></script>
<style>
:root{
  --bg:#0b1220;
  --panel:#121a2b;
  --panel-2:#18223a;
  --ink:#e6ebf5;
  --muted:#8b99b5;
  --accent:#4aa3ff;
  --accent-2:#6ee7b7;
  --warn:#fbbf24;
  --bad:#f87171;
  --border:#1f2a44;
}
*{box-sizing:border-box}
html,body{margin:0;padding:0;background:var(--bg);color:var(--ink);font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Helvetica,Arial,sans-serif;font-size:14px;line-height:1.45}
a{color:var(--accent)}
header{padding:22px 28px 14px;border-bottom:1px solid var(--border)}
h1{margin:0 0 4px;font-size:22px;letter-spacing:.3px}
.sub{color:var(--muted);font-size:13px}
main{padding:18px 28px 40px;max-width:1600px;margin:0 auto}
.kpis{display:grid;grid-template-columns:repeat(6,minmax(0,1fr));gap:12px;margin:14px 0 22px}
.kpi{background:var(--panel);border:1px solid var(--border);border-radius:10px;padding:12px 14px}
.kpi .k{color:var(--muted);font-size:11px;text-transform:uppercase;letter-spacing:.6px}
.kpi .v{font-size:20px;font-weight:600;margin-top:4px}
.kpi .v.small{font-size:14px;font-weight:500}
.grid{display:grid;grid-template-columns:repeat(12,minmax(0,1fr));gap:14px}
.card{background:var(--panel);border:1px solid var(--border);border-radius:10px;padding:12px}
.card h3{margin:0 0 10px;font-size:13px;color:var(--muted);text-transform:uppercase;letter-spacing:.5px}
.col-6{grid-column:span 6}
.col-4{grid-column:span 4}
.col-8{grid-column:span 8}
.col-12{grid-column:span 12}
.tabs{display:flex;gap:8px;margin:18px 0 8px}
.tabs button{background:var(--panel);color:var(--ink);border:1px solid var(--border);border-radius:999px;padding:6px 14px;font-size:12px;cursor:pointer}
.tabs button.active{background:var(--accent);border-color:var(--accent);color:#0b1220}
.filters{display:grid;grid-template-columns:repeat(5,minmax(0,1fr));gap:8px;margin:10px 0 14px}
.filters label{display:flex;flex-direction:column;gap:4px;font-size:11px;color:var(--muted)}
.filters input,.filters select{background:var(--panel-2);color:var(--ink);border:1px solid var(--border);border-radius:6px;padding:6px 8px;font-size:13px}
.year-range{display:flex;gap:6px;align-items:center}
.year-range input{flex:1}
table{width:100%;border-collapse:collapse;font-size:12.5px}
th,td{padding:6px 8px;border-bottom:1px solid var(--border);white-space:nowrap;text-align:left}
th{position:sticky;top:0;background:var(--panel-2);cursor:pointer;user-select:none;font-weight:600;font-size:11px;text-transform:uppercase;letter-spacing:.4px;color:var(--muted)}
th.sorted::after{content:" \25BE";color:var(--accent)}
th.sorted.asc::after{content:" \25B4";color:var(--accent)}
tbody tr:hover{background:#17203a}
td.num{text-align:right;font-variant-numeric:tabular-nums}
.table-wrap{max-height:520px;overflow:auto;border:1px solid var(--border);border-radius:8px}
.count-bar{display:flex;justify-content:space-between;align-items:center;color:var(--muted);font-size:12px;margin-bottom:6px}
footer{margin-top:30px;padding:14px 28px;border-top:1px solid var(--border);color:var(--muted);font-size:12px}
.panel-hidden{display:none}
.pill{display:inline-block;background:var(--panel-2);border:1px solid var(--border);border-radius:999px;padding:2px 8px;font-size:11px;color:var(--muted);margin-left:6px}
</style>
</head>
<body>
<header>
  <h1>Trent 900 Shop Visit History</h1>
  <div class="sub" id="subtitle"></div>
</header>
<main>
  <section class="kpis" id="kpis"></section>

  <section class="grid">
    <div class="card col-6"><h3>Shop visits per year</h3><div id="chart-year"></div></div>
    <div class="card col-6"><h3>Shop visits by type</h3><div id="chart-type"></div></div>
    <div class="card col-6"><h3>Shop visits by operator (top 10)</h3><div id="chart-operator"></div></div>
    <div class="card col-6"><h3>Shop visits by location</h3><div id="chart-location"></div></div>
    <div class="card col-12"><h3>Engine lifeline &mdash; visits over time (top engines, bubble size = HSSV)</h3><div id="chart-lifeline"></div></div>
    <div class="card col-6"><h3>HSN distribution at visit</h3><div id="chart-hsn"></div></div>
    <div class="card col-6"><h3>CSN vs HSN at visit</h3><div id="chart-scatter"></div></div>
  </section>

  <section style="margin-top:22px">
    <div class="tabs">
      <button id="tab-sv" class="active" type="button">Shop Visits <span class="pill" id="pill-sv"></span></button>
      <button id="tab-cs" type="button">Current Status (latest per engine) <span class="pill" id="pill-cs"></span></button>
    </div>

    <div id="panel-sv">
      <div class="filters">
        <label>Year range
          <div class="year-range">
            <input id="f-year-min" type="number" />
            <span>to</span>
            <input id="f-year-max" type="number" />
          </div>
        </label>
        <label>Operator
          <select id="f-operator"><option value="">All</option></select>
        </label>
        <label>Location
          <select id="f-location"><option value="">All</option></select>
        </label>
        <label>Visit type
          <select id="f-type"><option value="">All</option></select>
        </label>
        <label>Serial search
          <input id="f-serial" placeholder="e.g. 91020" />
        </label>
      </div>
      <div class="count-bar">
        <span id="sv-count"></span>
        <span id="sv-note" class="sub"></span>
      </div>
      <div class="table-wrap">
        <table id="sv-table">
          <thead><tr>
            <th data-k="serial_number">Serial</th>
            <th data-k="part_number">Part Number</th>
            <th data-k="action_code">Action</th>
            <th data-k="datetime">Date</th>
            <th data-k="operator">Operator</th>
            <th data-k="hsn">HSN</th>
            <th data-k="csn">CSN</th>
            <th data-k="hssv">HSSV</th>
            <th data-k="cssv">CSSV</th>
            <th data-k="shop_visit_type">Type</th>
            <th data-k="shop_visit_location">Location</th>
          </tr></thead>
          <tbody></tbody>
        </table>
      </div>
    </div>

    <div id="panel-cs" class="panel-hidden">
      <div class="count-bar">
        <span id="cs-count"></span>
        <span class="sub">These are current-status snapshot rows (not actual shop visits).</span>
      </div>
      <div class="table-wrap">
        <table id="cs-table">
          <thead><tr>
            <th data-k="serial_number">Serial</th>
            <th data-k="part_number">Part Number</th>
            <th data-k="action_code">Action</th>
            <th data-k="datetime">Snapshot Date</th>
            <th data-k="operator">Operator</th>
            <th data-k="hsn">HSN</th>
            <th data-k="csn">CSN</th>
            <th data-k="hssv">HSSV</th>
            <th data-k="cssv">CSSV</th>
            <th data-k="shop_visit_type">Type</th>
            <th data-k="shop_visit_location">Location</th>
          </tr></thead>
          <tbody></tbody>
        </table>
      </div>
    </div>
  </section>
</main>

<footer>
  <div id="footer-text"></div>
</footer>

<script id="payload" type="application/json">__DATA_JSON__</script>
<script>
(function(){
  const DATA = JSON.parse(document.getElementById('payload').textContent);
  const fmtInt = n => (n===null||n===undefined||Number.isNaN(n)) ? '' : Math.round(n).toLocaleString();
  const fmtNum = n => (n===null||n===undefined||Number.isNaN(n)) ? '' : Number(n).toLocaleString(undefined,{maximumFractionDigits:1});
  const fmtDate = s => {
    if(!s) return '';
    const d = new Date(s);
    if(Number.isNaN(d.getTime())) return s;
    return d.toISOString().slice(0,10);
  };
  const shortDate = s => {
    if(!s) return '';
    return String(s).slice(0,10);
  };

  // Subtitle + footer
  const m = DATA.meta;
  document.getElementById('subtitle').innerHTML =
    `Source: <strong>${m.source_file}</strong> &middot; Sheet: <strong>${m.sheet}</strong> &middot; `+
    `Total rows: <strong>${m.total_rows.toLocaleString()}</strong> `+
    `(<strong>${m.shop_visit_rows.toLocaleString()}</strong> shop visits, <strong>${m.current_status_rows.toLocaleString()}</strong> current-status snapshots) &middot; `+
    `Range: <strong>${fmtDate(m.date_range.earliest)}</strong> &rarr; <strong>${fmtDate(m.date_range.latest)}</strong>`;

  document.getElementById('footer-text').innerHTML =
    `Generated ${m.generated_at} &middot; Ground-truth benchmark &middot; `+
    `${m.current_status_rows.toLocaleString()} "Current Status" rows excluded from event totals.`;

  // KPIs
  const kp = DATA.kpis;
  const kpiEl = document.getElementById('kpis');
  const kpis = [
    ['Total shop visits', kp.total_shop_visits.toLocaleString()],
    ['Unique engines', kp.unique_engines.toLocaleString()],
    ['Unique operators', kp.unique_operators.toLocaleString()],
    ['Avg HSN at visit', fmtNum(kp.avg_hsn_at_visit) + ' hrs'],
    ['Current status rows', kp.current_status_rows.toLocaleString()],
    ['Date range', `${fmtDate(kp.earliest)} → ${fmtDate(kp.latest)}`, true],
  ];
  kpis.forEach(([k,v,small]) => {
    const d = document.createElement('div');
    d.className='kpi';
    d.innerHTML = `<div class="k">${k}</div><div class="v${small?' small':''}">${v}</div>`;
    kpiEl.appendChild(d);
  });

  // ---- Charts ----
  const baseOpts = {
    chart: { foreColor: '#c2cdde', toolbar: { show:false }, fontFamily:'inherit' },
    grid: { borderColor:'#1f2a44', strokeDashArray:3 },
    theme: { mode:'dark' },
  };

  // Chart 1 - per year
  new ApexCharts(document.getElementById('chart-year'), {
    ...baseOpts,
    chart: { ...baseOpts.chart, type:'bar', height:260 },
    series: [{ name:'Visits', data: DATA.charts.per_year.map(d=>d.count) }],
    xaxis: { categories: DATA.charts.per_year.map(d=>d.year) },
    colors: ['#4aa3ff'],
    dataLabels: { enabled:false },
    plotOptions: { bar: { borderRadius:3, columnWidth:'60%' } },
  }).render();

  // Chart 4 - donut by type
  new ApexCharts(document.getElementById('chart-type'), {
    ...baseOpts,
    chart: { ...baseOpts.chart, type:'donut', height:260 },
    series: DATA.charts.by_type.map(d=>d.count),
    labels: DATA.charts.by_type.map(d=>d.type),
    legend: { position:'right', fontSize:'12px' },
    dataLabels: { enabled:true, formatter: (val, opts) => opts.w.config.series[opts.seriesIndex] },
    colors: ['#4aa3ff','#6ee7b7','#fbbf24','#f87171','#a78bfa','#f472b6','#60a5fa','#34d399'],
  }).render();

  // Chart 2 - horizontal bar by operator
  new ApexCharts(document.getElementById('chart-operator'), {
    ...baseOpts,
    chart: { ...baseOpts.chart, type:'bar', height:320 },
    series: [{ name:'Visits', data: DATA.charts.by_operator.map(d=>d.count) }],
    xaxis: { categories: DATA.charts.by_operator.map(d=>d.operator) },
    plotOptions: { bar: { horizontal:true, borderRadius:3 } },
    colors: ['#6ee7b7'],
    dataLabels: { enabled:true, style:{colors:['#0b1220']}, formatter:v=>v.toLocaleString() },
  }).render();

  // Chart 3 - horizontal bar by location
  new ApexCharts(document.getElementById('chart-location'), {
    ...baseOpts,
    chart: { ...baseOpts.chart, type:'bar', height:320 },
    series: [{ name:'Visits', data: DATA.charts.by_location.map(d=>d.count) }],
    xaxis: { categories: DATA.charts.by_location.map(d=>d.location) },
    plotOptions: { bar: { horizontal:true, borderRadius:3 } },
    colors: ['#fbbf24'],
    dataLabels: { enabled:true, style:{colors:['#0b1220']}, formatter:v=>v.toLocaleString() },
  }).render();

  // Chart 5 - Engine lifeline scatter (bubble)
  const serials = DATA.charts.lifeline.serials;
  const lifelineBySerial = new Map(serials.map(s=>[s,[]]));
  DATA.charts.lifeline.points.forEach(p=>{
    if(!lifelineBySerial.has(p.serial)) return;
    const t = new Date(p.datetime).getTime();
    const z = p.hssv && p.hssv > 0 ? Math.min(40, 4 + Math.sqrt(p.hssv)/6) : 6;
    lifelineBySerial.get(p.serial).push({ x: t, y: p.serial, z: z, meta: p });
  });
  const bubbleSeries = [{
    name: 'Shop visits',
    data: Array.from(lifelineBySerial.values()).flat(),
  }];
  new ApexCharts(document.getElementById('chart-lifeline'), {
    ...baseOpts,
    chart: { ...baseOpts.chart, type:'bubble', height:560, zoom:{enabled:true,type:'xy'} },
    series: bubbleSeries,
    xaxis: { type:'datetime' },
    yaxis: { type:'category', categories: serials, reversed:false, labels:{style:{fontSize:'11px'}} },
    colors: ['#4aa3ff'],
    dataLabels: { enabled:false },
    tooltip: {
      custom: ({seriesIndex, dataPointIndex, w}) => {
        const p = w.config.series[seriesIndex].data[dataPointIndex];
        const m = p.meta || {};
        return `<div style="padding:8px 10px;background:#121a2b;border:1px solid #1f2a44;border-radius:6px">
          <div><strong>Serial ${m.serial||''}</strong></div>
          <div>${shortDate(m.datetime)}</div>
          <div>Operator: ${m.operator||''}</div>
          <div>Type: ${m.type||''}</div>
          <div>Location: ${m.location||''}</div>
          <div>HSN: ${fmtInt(m.hsn)} &middot; HSSV: ${fmtInt(m.hssv)}</div>
        </div>`;
      }
    },
    fill: { opacity:0.75 },
  }).render();

  // Chart 6 - HSN histogram
  new ApexCharts(document.getElementById('chart-hsn'), {
    ...baseOpts,
    chart: { ...baseOpts.chart, type:'bar', height:280 },
    series: [{ name:'Visits', data: DATA.charts.hsn_histogram.map(d=>d.count) }],
    xaxis: { categories: DATA.charts.hsn_histogram.map(d=>d.label), labels:{rotate:-45, style:{fontSize:'10px'}} },
    colors: ['#a78bfa'],
    dataLabels: { enabled:false },
    plotOptions: { bar: { borderRadius:2, columnWidth:'80%' } },
  }).render();

  // Chart 7 - CSN vs HSN scatter
  new ApexCharts(document.getElementById('chart-scatter'), {
    ...baseOpts,
    chart: { ...baseOpts.chart, type:'scatter', height:280, zoom:{enabled:true,type:'xy'} },
    series: [{ name:'Visits', data: DATA.charts.csn_vs_hsn.map(d=>({x:d.hsn, y:d.csn, serial:d.serial})) }],
    xaxis: { title:{text:'HSN (hours)'}, tickAmount:8 },
    yaxis: { title:{text:'CSN (cycles)'} },
    colors: ['#6ee7b7'],
    tooltip: {
      custom: ({seriesIndex, dataPointIndex, w}) => {
        const p = w.config.series[seriesIndex].data[dataPointIndex];
        return `<div style="padding:6px 10px;background:#121a2b;border:1px solid #1f2a44;border-radius:6px">
          Serial ${p.serial||''}<br>HSN ${fmtInt(p.x)}, CSN ${fmtInt(p.y)}
        </div>`;
      }
    },
  }).render();

  // ---- Filter dropdowns ----
  const fOperator = document.getElementById('f-operator');
  const fLocation = document.getElementById('f-location');
  const fType = document.getElementById('f-type');
  const fYmin = document.getElementById('f-year-min');
  const fYmax = document.getElementById('f-year-max');
  const fSerial = document.getElementById('f-serial');
  DATA.filters.operators.forEach(v => fOperator.insertAdjacentHTML('beforeend', `<option>${v}</option>`));
  DATA.filters.locations.forEach(v => fLocation.insertAdjacentHTML('beforeend', `<option>${v}</option>`));
  DATA.filters.types.forEach(v => fType.insertAdjacentHTML('beforeend', `<option>${v}</option>`));
  const years = DATA.filters.years;
  if(years.length){
    fYmin.value = years[0]; fYmin.min = years[0]; fYmin.max = years[years.length-1];
    fYmax.value = years[years.length-1]; fYmax.min = years[0]; fYmax.max = years[years.length-1];
  }

  // ---- Tables ----
  const svTbody = document.querySelector('#sv-table tbody');
  const csTbody = document.querySelector('#cs-table tbody');

  // Build "latest per engine" for current status: group by serial, keep latest
  const latestBySerial = new Map();
  DATA.current_status.forEach(r => {
    if(!r.serial_number) return;
    const prev = latestBySerial.get(r.serial_number);
    if(!prev || (r.datetime && (!prev.datetime || r.datetime > prev.datetime))){
      latestBySerial.set(r.serial_number, r);
    }
  });
  const latestCSRows = Array.from(latestBySerial.values()).sort((a,b) => {
    const sa = a.serial_number||'', sb = b.serial_number||'';
    return sa.localeCompare(sb, undefined, {numeric:true});
  });

  let svSortKey = 'datetime', svSortDir = -1;
  let csSortKey = 'serial_number', csSortDir = 1;

  function applyFilters(rows){
    const op = fOperator.value, lo = fLocation.value, ty = fType.value;
    const ymin = parseInt(fYmin.value,10), ymax = parseInt(fYmax.value,10);
    const q = fSerial.value.trim().toLowerCase();
    return rows.filter(r => {
      if(op && r.operator !== op) return false;
      if(lo && r.shop_visit_location !== lo) return false;
      if(ty && r.shop_visit_type !== ty) return false;
      if(!Number.isNaN(ymin) && r.year !== null && r.year < ymin) return false;
      if(!Number.isNaN(ymax) && r.year !== null && r.year > ymax) return false;
      if(q && !(r.serial_number||'').toLowerCase().includes(q)) return false;
      return true;
    });
  }

  function cmp(a,b,k){
    let va=a[k], vb=b[k];
    if(k==='datetime'){ va=va?new Date(va).getTime():0; vb=vb?new Date(vb).getTime():0; }
    if(typeof va==='number' || typeof vb==='number'){
      va = va===null||va===undefined?-Infinity:va;
      vb = vb===null||vb===undefined?-Infinity:vb;
      return va-vb;
    }
    return String(va||'').localeCompare(String(vb||''), undefined, {numeric:true});
  }

  function renderTable(tbody, rows, keys){
    const frag = document.createDocumentFragment();
    rows.forEach(r => {
      const tr = document.createElement('tr');
      keys.forEach(k => {
        const td = document.createElement('td');
        let v = r[k];
        if(k==='datetime') v = shortDate(v);
        else if(['hsn','csn','hssv','cssv'].includes(k)){ td.className='num'; v = fmtInt(v); }
        else if(v===null||v===undefined) v = '';
        td.textContent = v;
        tr.appendChild(td);
      });
      frag.appendChild(tr);
    });
    tbody.innerHTML = '';
    tbody.appendChild(frag);
  }

  const SV_KEYS = ['serial_number','part_number','action_code','datetime','operator','hsn','csn','hssv','cssv','shop_visit_type','shop_visit_location'];
  const CS_KEYS = SV_KEYS;

  function refreshSV(){
    let rows = applyFilters(DATA.shop_visits).slice();
    rows.sort((a,b) => svSortDir * cmp(a,b,svSortKey));
    renderTable(svTbody, rows, SV_KEYS);
    document.getElementById('sv-count').textContent = `Showing ${rows.length.toLocaleString()} of ${DATA.shop_visits.length.toLocaleString()} shop visits`;
    document.getElementById('sv-note').textContent = `${DATA.meta.current_status_rows.toLocaleString()} current-status rows excluded from event table`;
    document.getElementById('pill-sv').textContent = DATA.shop_visits.length.toLocaleString();
    // Mark sorted header
    document.querySelectorAll('#sv-table th').forEach(th => {
      th.classList.remove('sorted','asc');
      if(th.dataset.k === svSortKey){
        th.classList.add('sorted');
        if(svSortDir === 1) th.classList.add('asc');
      }
    });
  }

  function refreshCS(){
    let rows = latestCSRows.slice();
    rows.sort((a,b) => csSortDir * cmp(a,b,csSortKey));
    renderTable(csTbody, rows, CS_KEYS);
    document.getElementById('cs-count').textContent = `${rows.length.toLocaleString()} engines (latest status per engine; ${DATA.current_status.length.toLocaleString()} total snapshot rows)`;
    document.getElementById('pill-cs').textContent = rows.length.toLocaleString();
    document.querySelectorAll('#cs-table th').forEach(th => {
      th.classList.remove('sorted','asc');
      if(th.dataset.k === csSortKey){
        th.classList.add('sorted');
        if(csSortDir === 1) th.classList.add('asc');
      }
    });
  }

  // Wire filters
  [fOperator,fLocation,fType,fYmin,fYmax].forEach(el => el.addEventListener('change', refreshSV));
  fSerial.addEventListener('input', refreshSV);

  // Column sort
  document.querySelectorAll('#sv-table th').forEach(th => th.addEventListener('click', () => {
    const k = th.dataset.k;
    if(svSortKey === k) svSortDir *= -1; else { svSortKey = k; svSortDir = 1; }
    refreshSV();
  }));
  document.querySelectorAll('#cs-table th').forEach(th => th.addEventListener('click', () => {
    const k = th.dataset.k;
    if(csSortKey === k) csSortDir *= -1; else { csSortKey = k; csSortDir = 1; }
    refreshCS();
  }));

  // Tabs
  const tabSV = document.getElementById('tab-sv');
  const tabCS = document.getElementById('tab-cs');
  const panelSV = document.getElementById('panel-sv');
  const panelCS = document.getElementById('panel-cs');
  tabSV.addEventListener('click', () => {
    tabSV.classList.add('active'); tabCS.classList.remove('active');
    panelSV.classList.remove('panel-hidden'); panelCS.classList.add('panel-hidden');
  });
  tabCS.addEventListener('click', () => {
    tabCS.classList.add('active'); tabSV.classList.remove('active');
    panelCS.classList.remove('panel-hidden'); panelSV.classList.add('panel-hidden');
  });

  refreshSV();
  refreshCS();
})();
</script>
</body>
</html>
"""


if __name__ == "__main__":
    build()
