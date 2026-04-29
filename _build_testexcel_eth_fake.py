"""
Build a self-contained HTML visualization for ethiopian_fake_soa.xlsx (SOA).

Runs independently of V6/parser.py - reads the xlsx directly with openpyxl.
Emits V6/TESTEXCEL/Ethiopian_Fake_SOA.html embedding all data as a JSON
constant, rendered with ApexCharts 3.49.0 via CDN.

Structure of the source workbook:
  - Sheet "SOA 26.1.26": rows 1-6 metadata, row 7 header (with \\n in labels),
    rows 8+ five sections separated by banner rows in col A and total rows.
  - Sheets "Offset" and "Payment": 1x1 empty, ignored.
  - Dates are dd/mm/yyyy strings (no datetime cells).
"""

from __future__ import annotations

import json
from datetime import datetime, date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SRC_PATH = Path(
    r"C:\Users\Rutishkrishna\Desktop\RR\RR Powerbi\V6\New info\ethiopian_fake_soa.xlsx"
)
OUT_PATH = Path(
    r"C:\Users\Rutishkrishna\Desktop\RR\RR Powerbi\V6\TESTEXCEL\Ethiopian_Fake_SOA.html"
)

# Today anchor - per metadata row 4 col M, the sheet's self-declared "Today" is 30/01/2026.
# We use that so days_late matches the embedded DAYS LATE column.
TODAY = datetime(2026, 1, 30)

AGING_BUCKETS = [
    "Not due",
    "0-30",
    "31-60",
    "61-90",
    "91-180",
    "181-365",
    "365+",
]

SECTION_KEYWORDS = {
    "credits usable": "Credits Usable",
    "totalcare charges": "TotalCare Charges",
    "customer responsible charges": "Customer Responsible Charges",
    "spare parts charges": "Spare Parts Charges",
    "late payment interest": "Late Payment Interest",
}

TOTAL_MARKERS = {
    "total",
    "overdue",
    "available credits",
    "available credit",
    "total overdue",
}


def norm_header(v: Any) -> str:
    """Collapse literal \\n and extra whitespace in a header cell."""
    if v is None:
        return ""
    s = str(v).replace("\n", " ").replace("\r", " ")
    return " ".join(s.split()).strip()


def aging_bucket(days_late: int) -> str:
    if days_late <= 0:
        return "Not due"
    if days_late <= 30:
        return "0-30"
    if days_late <= 60:
        return "31-60"
    if days_late <= 90:
        return "61-90"
    if days_late <= 180:
        return "91-180"
    if days_late <= 365:
        return "181-365"
    return "365+"


def parse_ddmmyyyy(v: Any) -> datetime | None:
    """Parse dd/mm/yyyy string, or pass through datetime/date."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime.combine(v, datetime.min.time())
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def iso_or_none(dt: datetime | None) -> str | None:
    return dt.strftime("%Y-%m-%d") if dt else None


def ref_to_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        if float(v).is_integer():
            return str(int(v))
        return str(v)
    return str(v)


def infer_status(text: str, comments: str, amount: float, days_late: int) -> str:
    blob = f"{text} {comments}".lower()
    if amount < 0:
        return "Credit"
    if "disput" in blob:
        return "Disputed"
    if "ready for payment" in blob or "paid" in blob:
        return "Ready for Payment"
    if days_late > 0:
        return "Overdue"
    return "Outstanding"


def detect_section(col_a: Any) -> str | None:
    if col_a is None:
        return None
    s = str(col_a).strip().lower()
    for k, v in SECTION_KEYWORDS.items():
        if k in s:
            return v
    return None


def extract() -> dict[str, Any]:
    wb = load_workbook(SRC_PATH, data_only=True)
    ws = wb["SOA 26.1.26"]

    # Metadata rows 1-6
    metadata = {
        "title": ws.cell(row=1, column=1).value,
        "customer_name": ws.cell(row=2, column=2).value,
        "customer_number": ws.cell(row=3, column=2).value,
        "contact_email": ws.cell(row=4, column=2).value,
        "lpi_rate": ws.cell(row=2, column=13).value,
        "today_str": ws.cell(row=4, column=13).value,
        "avg_days_late_reported": ws.cell(row=5, column=11).value,
    }

    # Header row 7 - normalize linebreaks
    headers = [
        norm_header(ws.cell(row=7, column=c).value)
        for c in range(1, ws.max_column + 1)
    ]

    line_items: list[dict[str, Any]] = []
    total_rows: list[dict[str, Any]] = []
    current_section: str | None = None
    date_parse_failures: list[dict[str, Any]] = []

    for r in range(8, ws.max_row + 1):
        col_a = ws.cell(row=r, column=1).value
        col_b = ws.cell(row=r, column=2).value
        col_c = ws.cell(row=r, column=3).value
        col_d = ws.cell(row=r, column=4).value
        col_e = ws.cell(row=r, column=5).value
        col_f = ws.cell(row=r, column=6).value
        col_g = ws.cell(row=r, column=7).value
        col_h = ws.cell(row=r, column=8).value
        col_i = ws.cell(row=r, column=9).value
        col_j = ws.cell(row=r, column=10).value
        col_k = ws.cell(row=r, column=11).value
        col_l = ws.cell(row=r, column=12).value
        col_m = ws.cell(row=r, column=13).value
        col_n = ws.cell(row=r, column=14).value
        col_o = ws.cell(row=r, column=15).value

        # Section banner?
        sec = detect_section(col_a)
        if sec is not None and col_c is None and col_f is None:
            current_section = sec
            continue

        # Total / Overdue marker row in col E?
        if isinstance(col_e, str) and col_e.strip().lower() in TOTAL_MARKERS:
            total_rows.append(
                {
                    "row": r,
                    "section": current_section,
                    "label": col_e.strip(),
                    "value": col_f if isinstance(col_f, (int, float)) else None,
                }
            )
            continue

        # Skip fully blank spacer rows
        if all(
            v is None or v == ""
            for v in (col_a, col_b, col_c, col_d, col_e, col_f, col_g, col_h)
        ):
            continue

        # Data row: require reference (col C) and amount (col F) numeric
        if col_c is None or not isinstance(col_f, (int, float)):
            continue

        doc_dt = parse_ddmmyyyy(col_d)
        due_dt = parse_ddmmyyyy(col_e)

        if col_d not in (None, "") and doc_dt is None:
            date_parse_failures.append({"row": r, "col": "Document Date", "raw": col_d})
        if col_e not in (None, "") and due_dt is None:
            date_parse_failures.append({"row": r, "col": "Net due date", "raw": col_e})

        # Days late - prefer embedded col L if provided (authoritative), else compute
        embedded_days = col_l if isinstance(col_l, (int, float)) else None
        if due_dt:
            computed_days = (TODAY - due_dt).days
        else:
            computed_days = 0
        days_late = int(embedded_days) if embedded_days is not None else computed_days

        amount = float(col_f)
        text = str(col_h or "")
        rr_comments = str(col_j or "")
        status = infer_status(text, rr_comments, amount, days_late)

        line_items.append(
            {
                "row": r,
                "section": current_section or "Unknown",
                "company": ref_to_str(col_a),
                "account": ref_to_str(col_b),
                "reference": ref_to_str(col_c),
                "doc_date": iso_or_none(doc_dt),
                "doc_date_raw": str(col_d) if col_d is not None else "",
                "net_due_date": iso_or_none(due_dt),
                "net_due_date_raw": str(col_e) if col_e is not None else "",
                "amount": amount,
                "currency": str(col_g or ""),
                "text": text,
                "assignment": str(col_i or ""),
                "rr_comments": rr_comments,
                "action_owner": str(col_k or ""),
                "days_late": days_late,
                "aging_bucket": aging_bucket(days_late),
                "eth_comments": str(col_m or ""),
                "eth_po_reference": str(col_n or ""),
                "lpi_cumulated": str(col_o or ""),
                "status": status,
            }
        )

    # Aggregations
    total_count = len(line_items)
    total_amount = sum(i["amount"] for i in line_items)
    total_debits = sum(i["amount"] for i in line_items if i["amount"] > 0)
    total_credits = sum(i["amount"] for i in line_items if i["amount"] < 0)
    total_overdue = sum(
        i["amount"]
        for i in line_items
        if i["amount"] > 0 and i["days_late"] > 0
    )
    overdue_items = [i for i in line_items if i["amount"] > 0 and i["days_late"] > 0]
    avg_days_late = (
        sum(i["days_late"] for i in overdue_items) / len(overdue_items)
        if overdue_items
        else 0
    )

    # Section distribution (count + sum)
    sections: dict[str, dict[str, float]] = {}
    for s in SECTION_KEYWORDS.values():
        sections[s] = {"count": 0, "amount": 0.0, "overdue": 0.0}
    for i in line_items:
        sec = i["section"]
        if sec not in sections:
            sections[sec] = {"count": 0, "amount": 0.0, "overdue": 0.0}
        sections[sec]["count"] += 1
        sections[sec]["amount"] += i["amount"]
        if i["amount"] > 0 and i["days_late"] > 0:
            sections[sec]["overdue"] += i["amount"]

    # Aging distribution
    aging_counts = {b: 0 for b in AGING_BUCKETS}
    aging_amounts = {b: 0.0 for b in AGING_BUCKETS}
    for i in line_items:
        aging_counts[i["aging_bucket"]] += 1
        aging_amounts[i["aging_bucket"]] += i["amount"]

    # Status distribution
    status_counts: dict[str, int] = {}
    status_amounts: dict[str, float] = {}
    for i in line_items:
        status_counts[i["status"]] = status_counts.get(i["status"], 0) + 1
        status_amounts[i["status"]] = status_amounts.get(i["status"], 0.0) + i["amount"]

    # Top 10 overdue by amount (positive, days_late > 0)
    top10_overdue = sorted(
        overdue_items, key=lambda x: x["amount"], reverse=True
    )[:10]

    return {
        "meta": {
            "source_path": str(SRC_PATH),
            "sheet": "SOA 26.1.26",
            "headers": headers,
            "metadata": metadata,
            "row_count": total_count,
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "today": TODAY.strftime("%Y-%m-%d"),
            "date_parse_failures": date_parse_failures,
            "notes": "Synthetic / anonymized SOA test fixture. Do not use for real analysis.",
        },
        "kpis": {
            "line_count": total_count,
            "total_amount": total_amount,
            "total_debits": total_debits,
            "total_credits": total_credits,
            "total_overdue": total_overdue,
            "avg_days_late": avg_days_late,
            "lpi_rate": metadata.get("lpi_rate"),
            "net_balance": total_amount,
        },
        "sections": sections,
        "aging_counts": aging_counts,
        "aging_amounts": aging_amounts,
        "status_counts": status_counts,
        "status_amounts": status_amounts,
        "top10_overdue": top10_overdue,
        "line_items": line_items,
        "total_rows": total_rows,
    }


HTML_TEMPLATE = r"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8" />
<meta name="viewport" content="width=device-width, initial-scale=1" />
<title>Ethiopian (Fake) SOA — Benchmark</title>
<script src="https://cdn.jsdelivr.net/npm/apexcharts@3.49.0/dist/apexcharts.min.js"></script>
<style>
  :root {
    --bg: #0b1220;
    --panel: #111a2e;
    --panel-2: #162241;
    --border: #24324f;
    --text: #e6ecf7;
    --muted: #8b99b8;
    --accent: #5aa2ff;
    --ok: #22c55e;
    --warn: #f59e0b;
    --bad: #ef4444;
    --purple: #a78bfa;
  }
  * { box-sizing: border-box; }
  html, body { margin: 0; padding: 0; background: var(--bg); color: var(--text);
               font: 14px/1.45 "Segoe UI", Roboto, system-ui, -apple-system, Arial, sans-serif; }
  a { color: var(--accent); }
  header { padding: 22px 28px; border-bottom: 1px solid var(--border);
           background: linear-gradient(180deg, #0f1a35, #0b1220); }
  header h1 { margin: 0 0 6px; font-size: 22px; font-weight: 600; letter-spacing: .2px; }
  header .sub { color: var(--muted); font-size: 12.5px; }
  .banner { margin: 10px 0 0; padding: 8px 12px; background: #3a2a14; border: 1px solid #7a5a20;
            border-radius: 6px; color: #f5c26b; font-size: 12.5px; display: inline-block; }
  main { padding: 22px 28px 60px; max-width: 1480px; margin: 0 auto; }
  .meta-grid { display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; margin: 14px 0 4px; }
  .meta-grid .m { background: var(--panel-2); border: 1px solid var(--border); border-radius: 8px;
                  padding: 10px 12px; font-size: 12.5px; }
  .meta-grid .m .l { color: var(--muted); font-size: 11px; text-transform: uppercase; letter-spacing: .6px; }
  .meta-grid .m .v { font-weight: 600; margin-top: 4px; word-break: break-word; }
  .kpis { display: grid; grid-template-columns: repeat(6, minmax(0,1fr)); gap: 14px; margin: 18px 0 22px; }
  .kpi { background: var(--panel); border: 1px solid var(--border); border-radius: 10px; padding: 14px 16px; }
  .kpi .label { color: var(--muted); font-size: 11.5px; text-transform: uppercase; letter-spacing: .8px; }
  .kpi .val { font-size: 20px; font-weight: 600; margin-top: 6px; }
  .kpi .val.small { font-size: 16px; }
  .kpi.ok .val { color: var(--ok); }
  .kpi.warn .val { color: var(--warn); }
  .kpi.bad .val { color: var(--bad); }
  .kpi.acc .val { color: var(--accent); }
  .kpi.purple .val { color: var(--purple); }
  .charts { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 18px; }
  .charts-3 { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 16px; margin-bottom: 18px; }
  .card { background: var(--panel); border: 1px solid var(--border); border-radius: 10px; padding: 14px 16px; }
  .card h2 { margin: 0 0 12px; font-size: 14px; font-weight: 600; color: #cfd8ec; letter-spacing: .3px; }
  .card h2 .hint { color: var(--muted); font-weight: 400; font-size: 12px; margin-left: 8px; }
  details.section { background: var(--panel); border: 1px solid var(--border); border-radius: 10px;
                    margin-bottom: 12px; overflow: hidden; }
  details.section > summary {
    padding: 12px 16px; cursor: pointer; list-style: none; user-select: none;
    display: flex; justify-content: space-between; align-items: center;
    background: linear-gradient(180deg, #162241, #111a2e);
  }
  details.section > summary::-webkit-details-marker { display: none; }
  details.section > summary:hover { background: #1a2850; }
  details.section > summary .title { font-size: 14px; font-weight: 600; }
  details.section > summary .meta { color: var(--muted); font-size: 12.5px; }
  details.section > summary .caret { color: var(--muted); margin-right: 8px; transition: transform .2s; }
  details.section[open] > summary .caret { transform: rotate(90deg); }
  .sec-body { padding: 0; }
  .filters { display: flex; flex-wrap: wrap; gap: 10px; margin: 4px 0 14px; align-items: center; }
  .filters input, .filters select {
    background: var(--panel-2); color: var(--text); border: 1px solid var(--border);
    border-radius: 8px; padding: 8px 10px; font-size: 13px; min-width: 160px;
  }
  .filters input[type=search] { min-width: 280px; }
  .filters .count { color: var(--muted); margin-left: auto; font-size: 12.5px; }
  table { width: 100%; border-collapse: collapse; font-size: 12.5px; }
  th, td { padding: 7px 10px; text-align: left; border-bottom: 1px solid var(--border); vertical-align: top; }
  thead th { background: var(--panel-2); color: #cfd8ec; font-weight: 600; cursor: pointer;
             user-select: none; position: sticky; top: 0; z-index: 1; }
  thead th:hover { background: #1c2a4e; }
  tbody tr:hover { background: #16213e; }
  td.num { text-align: right; font-variant-numeric: tabular-nums; white-space: nowrap; }
  td.ref { font-family: ui-monospace, Consolas, monospace; white-space: nowrap; }
  .pill { display: inline-block; padding: 2px 8px; border-radius: 999px; font-size: 11px; font-weight: 600; }
  .pill.Outstanding { background: #0f2a4a; color: #7ab5ff; }
  .pill.Overdue { background: #3a1414; color: #ff8b8b; }
  .pill.Disputed { background: #3a2a14; color: #f5c26b; }
  .pill.Credit { background: #28124a; color: #c7a3ff; }
  .pill { }
  .pill.ReadyforPayment { background: #0f3b22; color: #5ee89a; }
  .table-wrap { max-height: 540px; overflow: auto; }
  footer { padding: 20px 28px; color: var(--muted); font-size: 12px; border-top: 1px solid var(--border); }
  .desc-trunc { max-width: 340px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; display: inline-block; }
  .sec-total { display: inline-block; padding: 2px 8px; border-radius: 6px; margin-left: 8px;
               font-size: 11.5px; background: #24324f; color: var(--muted); }
  .sec-total.pos { color: #ff8b8b; background: #2a1414; }
  .sec-total.neg { color: #c7a3ff; background: #241240; }
  @media (max-width: 1200px) {
    .kpis { grid-template-columns: repeat(3, 1fr); }
    .charts { grid-template-columns: 1fr; }
    .charts-3 { grid-template-columns: 1fr; }
    .meta-grid { grid-template-columns: repeat(2, 1fr); }
  }
</style>
</head>
<body>
<header>
  <h1>Ethiopian (Fake) — Statement of Account</h1>
  <div class="sub">
    <span id="meta-file"></span> · Sheet <span id="meta-sheet"></span> ·
    <span id="meta-rows"></span> line items
  </div>
  <div class="banner">Synthetic / anonymized SOA test fixture. Do not use for real analysis.</div>
</header>
<main>
  <section class="meta-grid" id="metaGrid"></section>

  <section class="kpis" id="kpis"></section>

  <section class="charts">
    <div class="card">
      <h2>Aging distribution <span class="hint">line count per bucket</span></h2>
      <div id="chartAging"></div>
    </div>
    <div class="card">
      <h2>Section totals <span class="hint">amount in document currency</span></h2>
      <div id="chartSections"></div>
    </div>
  </section>

  <section class="charts-3">
    <div class="card">
      <h2>Balance overview</h2>
      <div id="chartBalance"></div>
    </div>
    <div class="card">
      <h2>Top 10 overdue invoices</h2>
      <div id="chartTop10"></div>
    </div>
    <div class="card">
      <h2>Status mix <span class="hint">inferred per line</span></h2>
      <div id="chartStatus"></div>
    </div>
  </section>

  <section class="card">
    <h2>Global filters</h2>
    <div class="filters">
      <input id="q" type="search" placeholder="Search reference / text / comments / assignment..." />
      <select id="fSection"><option value="">All sections</option></select>
      <select id="fStatus"><option value="">All statuses</option></select>
      <select id="fBucket"><option value="">All aging buckets</option></select>
      <span class="count" id="globalCount"></span>
    </div>
  </section>

  <section id="sections"></section>

</main>
<footer>
  Source: <span id="f-src"></span><br/>
  Sheet: <span id="f-sheet"></span> · Today anchor: <span id="f-today"></span> · Generated <span id="gen-ts"></span><br/>
  <em>Synthetic / anonymized test fixture. Values do not represent real customer activity.</em>
</footer>

<script id="soa-data" type="application/json">__DATA_JSON__</script>
<script>
(function(){
  const D = JSON.parse(document.getElementById('soa-data').textContent);

  const fmtMoney = (v, cur) => {
    const n = Number(v || 0);
    const sign = n < 0 ? '-' : '';
    const body = Math.abs(n).toLocaleString(undefined, {maximumFractionDigits: 2, minimumFractionDigits: 2});
    return sign + (cur ? (cur + ' ') : '$') + body;
  };
  const fmtInt = (v) => Number(v || 0).toLocaleString();
  const fmtPct = (v) => (Number(v || 0) * 100).toFixed(2) + '%';

  // Header/meta
  document.getElementById('meta-file').textContent = D.meta.source_path;
  document.getElementById('meta-sheet').textContent = D.meta.sheet;
  document.getElementById('meta-rows').textContent = D.meta.row_count;
  document.getElementById('gen-ts').textContent = D.meta.generated_at;
  document.getElementById('f-src').textContent = D.meta.source_path;
  document.getElementById('f-sheet').textContent = D.meta.sheet;
  document.getElementById('f-today').textContent = D.meta.today;

  const m = D.meta.metadata;
  document.getElementById('metaGrid').innerHTML = [
    {l:'Customer', v: m.customer_name || '-'},
    {l:'Customer number', v: m.customer_number || '-'},
    {l:'Contact email', v: m.contact_email || '-'},
    {l:'LPI rate / Reported avg days late / Today (sheet)',
      v: `${fmtPct(m.lpi_rate)} · ${m.avg_days_late_reported || '-'}d · ${m.today_str || '-'}`},
  ].map(x => `<div class="m"><div class="l">${x.l}</div><div class="v">${x.v}</div></div>`).join('');

  // KPIs
  const k = D.kpis;
  const kpis = [
    {label:'Line items', val: fmtInt(k.line_count), cls:'acc'},
    {label:'Total overdue (debits)', val: fmtMoney(k.total_overdue), cls:'bad'},
    {label:'Avg days late (overdue)', val: Math.round(k.avg_days_late) + 'd', cls:'warn'},
    {label:'LPI rate', val: fmtPct(k.lpi_rate), cls:'purple'},
    {label:'Total credits', val: fmtMoney(k.total_credits), cls:'ok'},
    {label:'Net balance', val: fmtMoney(k.net_balance), cls:'small'},
  ];
  document.getElementById('kpis').innerHTML = kpis.map(kk =>
    `<div class="kpi ${kk.cls||''}"><div class="label">${kk.label}</div><div class="val ${kk.cls==='small'?'small':''}">${kk.val}</div></div>`
  ).join('');

  // Chart 1 - aging horizontal bar
  const agingOrder = ["Not due","0-30","31-60","61-90","91-180","181-365","365+"];
  new ApexCharts(document.getElementById('chartAging'), {
    chart: {type:'bar', height: 320, toolbar:{show:false}, background:'transparent', foreColor:'#cfd8ec'},
    theme: {mode:'dark'},
    series: [{name:'Line items', data: agingOrder.map(b => D.aging_counts[b] || 0)}],
    plotOptions: {bar: {horizontal: true, borderRadius: 3, distributed: true, dataLabels:{position:'top'}}},
    colors: ['#4ade80','#60a5fa','#fbbf24','#f97316','#ef4444','#b91c1c','#7f1d1d'],
    xaxis: {categories: agingOrder},
    dataLabels: {enabled: true, offsetX: 30, style: {colors:['#e6ecf7']}},
    legend: {show: false},
    grid: {borderColor: '#24324f'},
    tooltip: {y: {formatter: (v, {dataPointIndex}) => `${v} lines · ${fmtMoney(D.aging_amounts[agingOrder[dataPointIndex]])}`}},
  }).render();

  // Chart 2 - section totals bar
  const secOrder = ["Credits Usable","TotalCare Charges","Customer Responsible Charges","Spare Parts Charges","Late Payment Interest"];
  const secNames = Object.keys(D.sections);
  const orderedSecs = secOrder.filter(s => secNames.includes(s)).concat(secNames.filter(s => !secOrder.includes(s)));
  new ApexCharts(document.getElementById('chartSections'), {
    chart: {type:'bar', height: 320, toolbar:{show:false}, background:'transparent', foreColor:'#cfd8ec'},
    theme: {mode:'dark'},
    series: [{name:'Amount', data: orderedSecs.map(s => D.sections[s]?.amount || 0)}],
    plotOptions: {bar: {horizontal: false, borderRadius: 4, distributed: true,
      colors: {ranges: [{from: -1e18, to: 0, color: '#a78bfa'}, {from: 0, to: 1e18, color: '#60a5fa'}]}}},
    xaxis: {categories: orderedSecs, labels: {rotate: -20, style: {fontSize:'11px'}}},
    yaxis: {labels: {formatter: (v) => fmtMoney(v)}},
    dataLabels: {enabled: true, formatter: (v) => fmtMoney(v), style: {fontSize:'10px', colors:['#fff']}},
    legend: {show: false},
    grid: {borderColor: '#24324f'},
    tooltip: {y: {formatter: (v) => fmtMoney(v)}},
  }).render();

  // Chart 3 - balance overview (debits / credits / net)
  new ApexCharts(document.getElementById('chartBalance'), {
    chart: {type:'bar', height: 300, toolbar:{show:false}, background:'transparent', foreColor:'#cfd8ec'},
    theme: {mode:'dark'},
    series: [{name:'Amount', data: [k.total_debits, k.total_credits, k.net_balance]}],
    plotOptions: {bar: {horizontal: false, borderRadius: 4, distributed: true,
      colors: {ranges: [{from: -1e18, to: 0, color: '#a78bfa'}, {from: 0, to: 1e18, color: '#ef4444'}]}}},
    colors: ['#ef4444','#a78bfa','#60a5fa'],
    xaxis: {categories: ['Debits','Credits','Net balance']},
    yaxis: {labels: {formatter: (v) => fmtMoney(v)}},
    dataLabels: {enabled: true, formatter: (v) => fmtMoney(v), style:{colors:['#fff'], fontSize:'11px'}},
    legend: {show: false},
    grid: {borderColor: '#24324f'},
    tooltip: {y: {formatter: (v) => fmtMoney(v)}},
  }).render();

  // Chart 4 - top 10 overdue horizontal
  const t10 = D.top10_overdue;
  new ApexCharts(document.getElementById('chartTop10'), {
    chart: {type:'bar', height: 340, toolbar:{show:false}, background:'transparent', foreColor:'#cfd8ec'},
    theme: {mode:'dark'},
    series: [{name:'Amount', data: t10.map(i => i.amount)}],
    plotOptions: {bar: {horizontal: true, borderRadius: 2, distributed: true}},
    colors: ['#ef4444','#f97316','#f59e0b','#fbbf24','#eab308','#dc2626','#b91c1c','#991b1b','#7f1d1d','#450a0a'],
    xaxis: {categories: t10.map(i => (i.reference || '(blank)') + ' · ' + i.days_late + 'd'),
            labels: {formatter: (v) => fmtMoney(v)}},
    dataLabels: {enabled: true, formatter: (v) => fmtMoney(v), style:{colors:['#fff'], fontSize:'10px'}},
    legend: {show: false},
    grid: {borderColor: '#24324f'},
    tooltip: {y: {formatter: (v) => fmtMoney(v)}},
  }).render();

  // Chart 5 - status donut
  const statusLabels = Object.keys(D.status_counts);
  const statusValues = statusLabels.map(l => D.status_counts[l]);
  const statusColorMap = {
    'Outstanding':'#60a5fa', 'Overdue':'#ef4444',
    'Disputed':'#f59e0b', 'Credit':'#a78bfa', 'Ready for Payment':'#22c55e'
  };
  new ApexCharts(document.getElementById('chartStatus'), {
    chart: {type:'donut', height: 300, background:'transparent', foreColor:'#cfd8ec'},
    theme: {mode:'dark'},
    series: statusValues,
    labels: statusLabels,
    colors: statusLabels.map(l => statusColorMap[l] || '#64748b'),
    legend: {position: 'bottom', fontSize:'11px'},
    dataLabels: {enabled: true, formatter: (v) => v.toFixed(0) + '%'},
    plotOptions: {pie: {donut: {size:'58%', labels: {show:true, total:{show:true, label:'Total', formatter: () => D.kpis.line_count}}}}},
  }).render();

  // Section tables
  const secContainer = document.getElementById('sections');
  const sectionData = {};
  for (const s of orderedSecs) sectionData[s] = [];
  for (const item of D.line_items) {
    (sectionData[item.section] = sectionData[item.section] || []).push(item);
  }

  const columns = [
    {k:'reference', t:'Reference', cls:'ref'},
    {k:'doc_date', t:'Doc Date'},
    {k:'net_due_date', t:'Net Due'},
    {k:'amount', t:'Amount', cls:'num', fmt:(v) => fmtMoney(v)},
    {k:'currency', t:'Cur'},
    {k:'days_late', t:'Days Late', cls:'num'},
    {k:'aging_bucket', t:'Aging'},
    {k:'status', t:'Status', fmt:(v) => `<span class="pill ${v.replace(/\s+/g,'')}">${v}</span>`},
    {k:'text', t:'Description', fmt:(v) => `<span class="desc-trunc" title="${(v||'').replace(/"/g,'&quot;')}">${v||''}</span>`},
    {k:'action_owner', t:'Action Owner'},
    {k:'rr_comments', t:'R-R Comments', fmt:(v) => `<span class="desc-trunc" title="${(v||'').replace(/"/g,'&quot;')}">${v||''}</span>`},
    {k:'eth_comments', t:'ETH Comments', fmt:(v) => `<span class="desc-trunc" title="${(v||'').replace(/"/g,'&quot;')}">${v||''}</span>`},
    {k:'lpi_cumulated', t:'LPI Cum.'},
  ];

  const tableStates = {};  // per section: {sortKey, sortDir}

  function renderSectionTable(sec) {
    const st = tableStates[sec];
    const tbody = document.querySelector(`#tbody-${CSS.escape(sec.replace(/\s+/g,'_'))}`);
    const countEl = document.querySelector(`#count-${CSS.escape(sec.replace(/\s+/g,'_'))}`);
    if (!tbody) return;

    const q = (document.getElementById('q').value || '').trim().toLowerCase();
    const gs = document.getElementById('fSection').value;
    const gst = document.getElementById('fStatus').value;
    const gb = document.getElementById('fBucket').value;

    let rows = (sectionData[sec] || []).slice();
    if (gs && gs !== sec) rows = [];
    if (gst) rows = rows.filter(r => r.status === gst);
    if (gb) rows = rows.filter(r => r.aging_bucket === gb);
    if (q) rows = rows.filter(r => {
      const hay = [r.reference, r.text, r.rr_comments, r.eth_comments, r.assignment, r.status, r.action_owner]
        .filter(Boolean).join(' ').toLowerCase();
      return hay.includes(q);
    });
    rows.sort((a, b) => {
      const va = a[st.sortKey], vb = b[st.sortKey];
      if (va == null && vb == null) return 0;
      if (va == null) return 1;
      if (vb == null) return -1;
      if (typeof va === 'number' && typeof vb === 'number') return (va - vb) * st.sortDir;
      return String(va).localeCompare(String(vb)) * st.sortDir;
    });
    countEl.textContent = rows.length + ' of ' + (sectionData[sec]||[]).length + ' lines';
    tbody.innerHTML = rows.map(r => `
      <tr>${columns.map(c => {
        const val = r[c.k];
        const shown = c.fmt ? c.fmt(val) : (val ?? '');
        return `<td class="${c.cls||''}">${shown}</td>`;
      }).join('')}</tr>
    `).join('');
  }

  function renderAllTables() {
    let globalTotal = 0;
    let globalMatch = 0;
    const q = (document.getElementById('q').value || '').trim().toLowerCase();
    const gs = document.getElementById('fSection').value;
    const gst = document.getElementById('fStatus').value;
    const gb = document.getElementById('fBucket').value;
    for (const sec of Object.keys(sectionData)) {
      const all = sectionData[sec] || [];
      globalTotal += all.length;
      let filtered = all.slice();
      if (gs && gs !== sec) filtered = [];
      if (gst) filtered = filtered.filter(r => r.status === gst);
      if (gb) filtered = filtered.filter(r => r.aging_bucket === gb);
      if (q) filtered = filtered.filter(r => {
        const hay = [r.reference, r.text, r.rr_comments, r.eth_comments, r.assignment, r.status, r.action_owner]
          .filter(Boolean).join(' ').toLowerCase();
        return hay.includes(q);
      });
      globalMatch += filtered.length;
      renderSectionTable(sec);
    }
    document.getElementById('globalCount').textContent = `${globalMatch} / ${globalTotal} lines match filters`;
  }

  // Build HTML for section tables
  secContainer.innerHTML = orderedSecs.map(sec => {
    const items = sectionData[sec] || [];
    const totAmt = items.reduce((a, x) => a + x.amount, 0);
    const cls = totAmt > 0 ? 'pos' : (totAmt < 0 ? 'neg' : '');
    const safe = sec.replace(/\s+/g,'_');
    tableStates[sec] = {sortKey:'days_late', sortDir:-1};
    return `
      <details class="section" ${items.length ? 'open' : ''}>
        <summary>
          <span><span class="caret">▸</span><span class="title">${sec}</span>
            <span class="sec-total ${cls}">${items.length} lines · ${fmtMoney(totAmt)}</span>
          </span>
          <span class="meta"><span id="count-${safe}"></span></span>
        </summary>
        <div class="sec-body">
          <div class="table-wrap">
            <table>
              <thead><tr>
                ${columns.map(c => `<th data-k="${c.k}" data-sec="${sec}" class="${c.cls||''}">${c.t}</th>`).join('')}
              </tr></thead>
              <tbody id="tbody-${safe}"></tbody>
            </table>
          </div>
        </div>
      </details>
    `;
  }).join('');

  // Header click -> sort
  document.querySelectorAll('thead th[data-k]').forEach(th => {
    th.addEventListener('click', () => {
      const k = th.dataset.k;
      const sec = th.dataset.sec;
      const st = tableStates[sec];
      if (st.sortKey === k) st.sortDir = -st.sortDir;
      else {
        st.sortKey = k;
        st.sortDir = (k === 'amount' || k === 'days_late') ? -1 : 1;
      }
      renderSectionTable(sec);
    });
  });

  // Populate global filter dropdowns
  const fSection = document.getElementById('fSection');
  orderedSecs.forEach(s => {
    const o = document.createElement('option'); o.value = s; o.textContent = s; fSection.appendChild(o);
  });
  const fStatus = document.getElementById('fStatus');
  Object.keys(D.status_counts).sort().forEach(s => {
    const o = document.createElement('option'); o.value = s; o.textContent = s; fStatus.appendChild(o);
  });
  const fBucket = document.getElementById('fBucket');
  agingOrder.forEach(b => {
    if ((D.aging_counts[b] || 0) > 0) {
      const o = document.createElement('option'); o.value = b; o.textContent = b; fBucket.appendChild(o);
    }
  });

  ['q','fSection','fStatus','fBucket'].forEach(id =>
    document.getElementById(id).addEventListener('input', renderAllTables));

  renderAllTables();
})();
</script>
</body>
</html>
"""


def build() -> None:
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    data = extract()

    # Safe JSON embed: avoid </script> breakout and ensure dict/datetime serializable
    payload = json.dumps(data, default=str, ensure_ascii=False).replace("</", "<\\/")
    html = HTML_TEMPLATE.replace("__DATA_JSON__", payload)
    OUT_PATH.write_text(html, encoding="utf-8")

    # Report
    k = data["kpis"]
    print("=" * 70)
    print("Ethiopian (Fake) SOA - build report")
    print("=" * 70)
    print(f"Source                : {SRC_PATH}")
    print(f"Output                : {OUT_PATH}")
    print(f"Output size           : {OUT_PATH.stat().st_size:,} bytes")
    print(f"Line items extracted  : {k['line_count']}")
    print("")
    print("Section distribution:")
    for s, v in data["sections"].items():
        print(f"  - {s:<35} {v['count']:>3} lines   amount={v['amount']:>20,.2f}   overdue={v['overdue']:>18,.2f}")
    print("")
    print(f"Total debits          : {k['total_debits']:>20,.2f}")
    print(f"Total credits         : {k['total_credits']:>20,.2f}")
    print(f"Net balance           : {k['net_balance']:>20,.2f}")
    print(f"Total overdue (pos)   : {k['total_overdue']:>20,.2f}")
    print(f"Avg days late (od)    : {k['avg_days_late']:.1f}d")
    print(f"LPI rate              : {k['lpi_rate']}")
    print("")
    print("Aging distribution (count / amount):")
    for b in AGING_BUCKETS:
        print(f"  {b:<10} {data['aging_counts'][b]:>3}   {data['aging_amounts'][b]:>20,.2f}")
    print("")
    print("Status distribution:")
    for s, c in data["status_counts"].items():
        print(f"  {s:<22} {c:>3}   {data['status_amounts'][s]:>20,.2f}")
    print("")
    print(f"Total-rows captured   : {len(data['total_rows'])}")
    for t in data["total_rows"]:
        print(f"  row {t['row']:>3} [{t['section']}] {t['label']:<20} = {t['value']}")
    print("")
    print(f"Date parse failures   : {len(data['meta']['date_parse_failures'])}")
    for f in data["meta"]["date_parse_failures"]:
        print(f"  row {f['row']} {f['col']}: {f['raw']!r}")


if __name__ == "__main__":
    build()
