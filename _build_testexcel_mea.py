"""
Build a self-contained HTML visualization for MEA Profit Opportunities Tracker.

Independent of V6/parser.py. Reads xlsx via openpyxl directly.
Emits JSON payload + ApexCharts visualization as a single HTML file.
"""
from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from datetime import datetime, date
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

SRC = Path(r"C:\Users\Rutishkrishna\Desktop\RR\RR Powerbi\V6\New info\MEA Profit Opportunities Tracker 21.04.xlsx")
OUT_HTML = Path(r"C:\Users\Rutishkrishna\Desktop\RR\RR Powerbi\V6\TESTEXCEL\MEA_Profit_Opportunities_Tracker.html")

# -----------------------------------------------------------------------------
# Header mapping via regex
# -----------------------------------------------------------------------------
HEADER_PATTERNS: list[tuple[str, re.Pattern]] = [
    ("num",                re.compile(r"^\s*#\s*$")),
    ("project",            re.compile(r"^project\s*$", re.I)),
    ("programme",          re.compile(r"^programme\s*$", re.I)),
    ("customer",           re.compile(r"^customer\s*$", re.I)),
    ("region",             re.compile(r"^region\s*$", re.I)),
    ("asks",               re.compile(r"^asks\s*$", re.I)),
    ("type_of_opportunity", re.compile(r"^type of opportunity", re.I)),
    ("levers",             re.compile(r"^levers", re.I)),
    ("priority",           re.compile(r"^priority\s*$", re.I)),
    ("spe_related",        re.compile(r"^spe\s*related", re.I)),
    ("no_of_spe",          re.compile(r"^no[,. ]*of\s*spe", re.I)),
    ("crp_pct",            re.compile(r"^crp\s*%?", re.I)),
    ("external_probability", re.compile(r"^external probability", re.I)),
    ("internal_complexity", re.compile(r"^internal complexity", re.I)),
    ("status",             re.compile(r"^status\s*$", re.I)),
    ("evaluation_level",   re.compile(r"^evaluation level", re.I)),
    ("term_benefit",       re.compile(r"^term benefit\s*$", re.I)),
    ("benefit_2026",       re.compile(r"^2026\s*$")),
    ("benefit_2027",       re.compile(r"^2027\s*$")),
    ("sum_26_27",          re.compile(r"^sum of\s*26/27", re.I)),
]

# Only accept the FIRST occurrence of a repeated header (the ones in the compact 20-col section).
def map_headers(ws, header_row: int) -> dict[str, int]:
    mapped: dict[str, int] = {}
    # Restrict mapping for repeated year columns (2026, 2027) to the first block (<=V/22)
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        text = str(v).strip()
        for key, pat in HEADER_PATTERNS:
            if key in mapped:
                continue
            if pat.match(text):
                # For benefit_2026/benefit_2027/sum we must stay within core block (cols <= 22 / V)
                if key in {"benefit_2026", "benefit_2027", "sum_26_27"} and c > 22:
                    continue
                mapped[key] = c
                break
    return mapped


# -----------------------------------------------------------------------------
# Sheet extraction
# -----------------------------------------------------------------------------
def detect_header_row(ws) -> int:
    """Scan rows 1-20 for the first row matching >= 5 canonical header tokens."""
    canon = ("customer", "region", "priority", "probability", "benefit", "project", "programme", "status")
    best_row = 14
    best_score = 0
    for r in range(1, min(25, ws.max_row + 1)):
        score = 0
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            t = str(v).lower()
            for k in canon:
                if k in t:
                    score += 1
                    break
        if score > best_score:
            best_score = score
            best_row = r
    return best_row if best_score >= 5 else 14


def cell_to_json(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        try:
            return v.isoformat()
        except Exception:
            return str(v)
    if isinstance(v, (int, float)):
        # Coerce floats that are actually ints
        if isinstance(v, float) and v != v:  # NaN
            return None
        return v
    return str(v)


def extract_sheet(ws) -> dict:
    header_row = detect_header_row(ws)
    mapped = map_headers(ws, header_row)
    # All row 14 headers (full) for completeness
    all_headers: list[dict] = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is not None:
            all_headers.append({"col": c, "letter": get_column_letter(c), "name": str(v).strip()})

    # Hidden metadata
    hidden_cols = sorted([k for k, cd in ws.column_dimensions.items() if cd.hidden])
    hidden_rows = sorted([k for k, rd in ws.row_dimensions.items() if rd.hidden])

    rows: list[dict] = []
    start_row = header_row + 1
    for r in range(start_row, ws.max_row + 1):
        # Require at least a project or customer to treat as data
        project = ws.cell(row=r, column=mapped.get("project", 4)).value
        customer = ws.cell(row=r, column=mapped.get("customer", 6)).value
        if not project and not customer:
            continue
        rec: dict = {"row": r}
        for key, col in mapped.items():
            rec[key] = cell_to_json(ws.cell(row=r, column=col).value)
        rows.append(rec)

    # Sums / aggregates using mapped columns
    sum_2026 = 0.0
    sum_2027 = 0.0
    for rec in rows:
        v1 = rec.get("benefit_2026")
        v2 = rec.get("benefit_2027")
        if isinstance(v1, (int, float)):
            sum_2026 += v1
        if isinstance(v2, (int, float)):
            sum_2027 += v2

    return {
        "header_row": header_row,
        "mapped_columns": {k: get_column_letter(v) + f"({v})" for k, v in mapped.items()},
        "all_headers": all_headers,
        "hidden_columns": hidden_cols,
        "hidden_rows_count": len(hidden_rows),
        "row_count_nonblank": len(rows),
        "sum_benefit_2026": round(sum_2026, 4),
        "sum_benefit_2027": round(sum_2027, 4),
        "rows": rows,
    }


# -----------------------------------------------------------------------------
# Timeline / Milestones extraction
# -----------------------------------------------------------------------------
def extract_timeline(ws) -> dict:
    # Header at row 14: cols C=Project, D=Customer, E.. = date columns (one per week)
    header_row = 14
    # Build date lookup for cols >= 5
    date_cols: list[dict] = []
    for c in range(5, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, (datetime, date)):
            date_cols.append({"col": c, "date": v.isoformat()})
        elif isinstance(v, str) and re.match(r"\d{2}/\d{2}/\d{4}", v.strip()):
            # dd/mm/yyyy → iso
            try:
                d = datetime.strptime(v.strip(), "%d/%m/%Y").date()
                date_cols.append({"col": c, "date": d.isoformat()})
            except Exception:
                continue

    col_to_date = {d["col"]: d["date"] for d in date_cols}
    projects: list[dict] = []
    milestone_types: Counter = Counter()

    for r in range(header_row + 1, ws.max_row + 1):
        proj = ws.cell(row=r, column=3).value
        cust = ws.cell(row=r, column=4).value
        if not (proj or cust):
            continue
        milestones: list[dict] = []
        for c, iso in col_to_date.items():
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str):
                name = v.strip()
                milestones.append({"date": iso, "milestone": name})
                milestone_types[name] += 1
        if milestones:
            projects.append({
                "project": str(proj) if proj else None,
                "customer": str(cust) if cust else None,
                "milestones": milestones,
            })

    return {
        "project_count": len(projects),
        "milestone_type_counts": dict(milestone_types),
        "projects": projects,
    }


# -----------------------------------------------------------------------------
# Opps & Threats
# -----------------------------------------------------------------------------
def extract_opps_threats(ws) -> list[dict]:
    # Row 2 headers
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=2, column=c).value
        if v:
            headers[c] = str(v).strip()
    # Rows 4+
    out: list[dict] = []
    for r in range(4, ws.max_row + 1):
        proj = ws.cell(row=r, column=2).value
        opp = ws.cell(row=r, column=5).value
        if not (proj or opp):
            continue
        rec: dict = {}
        for c, h in headers.items():
            rec[h] = cell_to_json(ws.cell(row=r, column=c).value)
        out.append(rec)
    return out


# -----------------------------------------------------------------------------
# Main build
# -----------------------------------------------------------------------------
def main() -> dict:
    print(f"[build] Opening {SRC}")
    wb = openpyxl.load_workbook(SRC, data_only=True)
    print(f"[build] Sheets: {wb.sheetnames}")

    sheets_out: dict[str, dict] = {}
    for sn in ("MEA LOG", "L2", "L3"):
        if sn in wb.sheetnames:
            print(f"[build] Extracting sheet: {sn}")
            sheets_out[sn] = extract_sheet(wb[sn])
            so = sheets_out[sn]
            print(f"  -> header_row={so['header_row']} rows={so['row_count_nonblank']} sum26={so['sum_benefit_2026']} sum27={so['sum_benefit_2027']}")

    timeline = extract_timeline(wb["Timeline"]) if "Timeline" in wb.sheetnames else {"project_count": 0, "projects": []}
    print(f"[build] Timeline: {timeline['project_count']} projects")

    opps_threats = extract_opps_threats(wb["Opps and Threats"]) if "Opps and Threats" in wb.sheetnames else []
    print(f"[build] Opps&Threats: {len(opps_threats)} rows")

    # Extract away-day date (E8) and today's date (S6)
    mea = wb["MEA LOG"]
    away_day = mea.cell(row=8, column=5).value
    today_formula = mea.cell(row=6, column=19).value

    payload = {
        "meta": {
            "source_file": str(SRC),
            "source_basename": SRC.name,
            "generated_at": datetime.now().isoformat(),
            "away_day_date": cell_to_json(away_day),
            "today_value": cell_to_json(today_formula),
            "sheet_names": wb.sheetnames,
        },
        "sheets": sheets_out,
        "timeline": timeline,
        "opps_threats": opps_threats,
    }
    return payload


# -----------------------------------------------------------------------------
# HTML emitter
# -----------------------------------------------------------------------------
HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<title>MEA Profit Opportunities Tracker — Ground Truth</title>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<script src="https://cdn.jsdelivr.net/npm/apexcharts@3.49.0/dist/apexcharts.min.js"></script>
<style>
  :root{
    --bg: #0d1117;
    --panel: #161b22;
    --panel-2: #1c2230;
    --border: #2b3342;
    --text: #e6edf3;
    --muted: #8b96a6;
    --accent: #5b8cff;
    --accent-2: #7ed4b2;
    --high: #ef476f;
    --med:  #ffd166;
    --low:  #06d6a0;
    --alt1: #c77dff;
    --alt2: #ffbe6a;
    --alt3: #4cc9f0;
  }
  *{ box-sizing: border-box; }
  html,body{ margin:0; padding:0; background:var(--bg); color:var(--text);
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
    font-size: 14px; line-height:1.45;
  }
  header.top{
    padding: 22px 28px 14px;
    border-bottom: 1px solid var(--border);
    background: linear-gradient(180deg, #141a24 0%, #0d1117 100%);
  }
  header.top h1{ margin:0 0 4px; font-size: 22px; letter-spacing: .2px; }
  header.top .sub{ color: var(--muted); font-size: 12.5px; }
  main{ padding: 18px 28px 48px; max-width: 1680px; margin: 0 auto; }

  .tabs{ display:flex; gap:6px; margin: 16px 0 10px; }
  .tab{
    padding: 8px 16px; border:1px solid var(--border); background: var(--panel);
    color: var(--muted); border-radius: 6px; cursor: pointer; font-weight: 500;
    transition: all .15s ease;
  }
  .tab:hover{ color:var(--text); border-color:#3b475c; }
  .tab.active{ background: var(--accent); color:#fff; border-color:var(--accent); }

  .kpi-row{ display:grid; grid-template-columns: repeat(auto-fit, minmax(160px,1fr)); gap:12px; margin: 12px 0 18px; }
  .kpi{
    background: var(--panel); border:1px solid var(--border); border-radius: 8px;
    padding: 12px 14px; position:relative; overflow:hidden;
  }
  .kpi .label{ font-size: 11px; color:var(--muted); text-transform: uppercase; letter-spacing: .5px; }
  .kpi .value{ font-size: 22px; font-weight: 700; margin-top:4px; color:var(--text); }
  .kpi .unit{ font-size: 11px; color: var(--muted); margin-left: 4px; font-weight: 400; }
  .kpi.accent .value{ color: var(--accent-2); }

  .grid{ display:grid; gap:14px; }
  .grid.g2{ grid-template-columns: repeat(2, minmax(0,1fr)); }
  .grid.g3{ grid-template-columns: repeat(3, minmax(0,1fr)); }
  @media (max-width: 1100px){ .grid.g2, .grid.g3{ grid-template-columns: 1fr; } }

  .card{
    background: var(--panel); border:1px solid var(--border); border-radius: 8px;
    padding: 14px 16px 10px;
  }
  .card h3{ margin:0 0 8px; font-size: 13px; color:var(--muted); text-transform: uppercase; letter-spacing:.4px; font-weight: 600;}
  .card.chart{ min-height: 310px; }
  .card.chart.tall{ min-height: 420px; }
  .card.chart.xtall{ min-height: 520px; }

  .filter-bar{
    display:flex; flex-wrap: wrap; gap: 10px; margin: 10px 0 12px; align-items: flex-end;
    background: var(--panel); border:1px solid var(--border); border-radius: 8px; padding: 12px;
  }
  .filter-bar .fld{ display:flex; flex-direction:column; gap:4px; }
  .filter-bar label{ font-size: 11px; color: var(--muted); text-transform: uppercase; letter-spacing:.4px; }
  .filter-bar select, .filter-bar input[type="text"]{
    background: var(--panel-2); border:1px solid var(--border); color: var(--text);
    padding: 6px 10px; border-radius: 5px; font-size: 13px; min-width: 160px;
  }
  .filter-bar button{
    background: var(--accent-2); border:none; color:#0d1117; font-weight:600;
    padding: 7px 14px; border-radius: 5px; cursor:pointer;
  }
  .filter-bar button.ghost{ background: transparent; color: var(--muted); border:1px solid var(--border); }

  table.dt{
    width: 100%; border-collapse: collapse; font-size: 12.5px;
  }
  table.dt th, table.dt td{
    border-bottom: 1px solid var(--border);
    padding: 7px 9px; text-align: left; vertical-align: top;
  }
  table.dt th{
    position: sticky; top: 0; background: var(--panel-2); color: var(--muted);
    text-transform: uppercase; font-size: 11px; letter-spacing:.3px;
    cursor: pointer; user-select:none;
  }
  table.dt tbody tr:hover{ background: rgba(91,140,255,0.06); }
  table.dt td.num{ text-align: right; font-variant-numeric: tabular-nums; }
  .pri-badge{ display:inline-block; padding: 2px 8px; border-radius: 999px; font-size: 11px; font-weight: 600; }
  .pri-1{ background: rgba(239,71,111,.18); color: #ff6b8b; }
  .pri-2{ background: rgba(255,209,102,.18); color: #ffd166; }
  .pri-3{ background: rgba(6,214,160,.18); color: #4be3b2; }
  .prob-badge{ display:inline-block; padding: 2px 8px; border-radius: 4px; font-size: 11px; font-weight: 600; }
  .prob-High{ background: rgba(6,214,160,.15); color:#4be3b2; }
  .prob-Med{  background: rgba(255,209,102,.15); color:#ffd166; }
  .prob-Low{  background: rgba(239,71,111,.15); color:#ff6b8b; }

  .table-wrap{
    background: var(--panel); border:1px solid var(--border); border-radius: 8px;
    max-height: 620px; overflow:auto;
  }
  .section-head{ display:flex; align-items:center; justify-content:space-between; margin: 22px 0 6px; }
  .section-head h2{ margin:0; font-size: 15px; letter-spacing:.3px; }
  .section-head .hint{ font-size: 12px; color: var(--muted); }

  .opps-table th, .opps-table td{ font-size: 12px; }
  .footer{ margin-top: 28px; padding-top: 14px; border-top: 1px solid var(--border); color: var(--muted); font-size: 11.5px; display:flex; justify-content:space-between; flex-wrap: wrap; gap: 8px;}

  .unmapped{ color: var(--high); font-size: 11.5px; }
</style>
</head>
<body>
<header class="top">
  <h1>MEA Profit Opportunities Tracker</h1>
  <div class="sub" id="subtitle">loading…</div>
</header>
<main>

  <div class="tabs" id="tabs"></div>

  <div class="kpi-row" id="kpiRow"></div>

  <div class="grid g3">
    <div class="card chart"><h3>Priority Distribution</h3><div id="chPriority"></div></div>
    <div class="card chart"><h3>External Probability</h3><div id="chProb"></div></div>
    <div class="card chart"><h3>Status Funnel</h3><div id="chFunnel"></div></div>
  </div>

  <div class="grid g2" style="margin-top:14px;">
    <div class="card chart tall"><h3>Benefit 2026 vs 2027 — Top 10 Customers (£m)</h3><div id="chBenefitCust"></div></div>
    <div class="card chart tall"><h3>Benefit by Programme × Priority (stacked)</h3><div id="chBenefitProg"></div></div>
  </div>

  <div class="grid g2" style="margin-top:14px;">
    <div class="card chart tall"><h3>Top 10 Customers — Leaderboard (Sum 26+27 £m)</h3><div id="chLeader"></div></div>
    <div class="card chart tall"><h3>Project Milestones Timeline</h3><div id="chTimeline"></div></div>
  </div>

  <div class="section-head"><h2>Opportunities</h2><span class="hint" id="rowcountHint"></span></div>
  <div class="filter-bar">
    <div class="fld"><label>Priority</label>
      <select id="fPriority" multiple size="3" style="min-height:70px"></select>
    </div>
    <div class="fld"><label>Status</label><select id="fStatus"></select></div>
    <div class="fld"><label>Programme</label><select id="fProgramme"></select></div>
    <div class="fld"><label>Probability</label><select id="fProb"></select></div>
    <div class="fld" style="flex:1; min-width:220px"><label>Search</label><input type="text" id="fSearch" placeholder="customer / project / ask…"/></div>
    <button id="btnClear" class="ghost">Clear</button>
  </div>
  <div class="table-wrap">
    <table class="dt" id="mainTable">
      <thead><tr id="mainHead"></tr></thead>
      <tbody id="mainBody"></tbody>
    </table>
  </div>

  <div class="section-head"><h2>Opps &amp; Threats</h2><span class="hint" id="otHint"></span></div>
  <div class="table-wrap">
    <table class="dt opps-table" id="otTable">
      <thead><tr id="otHead"></tr></thead>
      <tbody id="otBody"></tbody>
    </table>
  </div>

  <div class="footer">
    <div id="footLeft"></div>
    <div id="footRight"></div>
  </div>

</main>

<script id="dataPayload" type="application/json">__PAYLOAD__</script>
<script>
const DATA = JSON.parse(document.getElementById('dataPayload').textContent);

// -------------------- state --------------------
const state = {
  activeSheet: 'MEA LOG',
  sortKey: null,
  sortDir: 1,
  filters: { priorities: [], status: '', programme: '', prob: '', search: '' },
};

// -------------------- header + subtitle --------------------
(function initHeader(){
  const m = DATA.meta;
  const counts = Object.entries(DATA.sheets).map(([k,v])=>`${k}=${v.row_count_nonblank}`).join(', ');
  document.getElementById('subtitle').textContent =
    `${m.source_basename} · generated ${new Date(m.generated_at).toLocaleString()} · sheets parsed: ${counts} · away-day ${m.away_day_date||'-'}`;
  document.getElementById('footLeft').textContent = `Source: ${m.source_basename}`;
  document.getElementById('footRight').textContent = `Generated: ${new Date(m.generated_at).toLocaleString()} · Sheets: ${m.sheet_names.join(', ')}`;
})();

// -------------------- tabs --------------------
(function initTabs(){
  const el = document.getElementById('tabs');
  Object.keys(DATA.sheets).forEach(name=>{
    const b = document.createElement('button');
    b.className = 'tab' + (name===state.activeSheet ? ' active':'');
    b.textContent = `${name} (${DATA.sheets[name].row_count_nonblank})`;
    b.dataset.sheet = name;
    b.onclick = () => { state.activeSheet = name; render(); };
    el.appendChild(b);
  });
})();

// -------------------- helpers --------------------
function fmtMoney(v){
  if(v===null||v===undefined||v==='') return '';
  const n = Number(v); if(!isFinite(n)) return '';
  return n.toLocaleString(undefined, { minimumFractionDigits: 2, maximumFractionDigits: 2 });
}
function sumBy(rows, key){
  let s = 0;
  for(const r of rows){ const v = r[key]; if(typeof v === 'number' && isFinite(v)) s += v; }
  return s;
}
function countBy(rows, key){
  const c = {};
  for(const r of rows){ const v = (r[key]===null||r[key]===undefined||r[key]==='') ? '(blank)' : String(r[key]).trim(); c[v]=(c[v]||0)+1; }
  return c;
}
function priorityLabel(p){
  const n = Number(p);
  if(n===1) return 'High';
  if(n===2) return 'Medium';
  if(n===3) return 'Low';
  return '(n/a)';
}
function cleanStatus(s){ return (s||'').toString().trim() || '(blank)'; }

let charts = {};
function renderChart(key, id, opts){
  if(charts[key]){ charts[key].destroy(); }
  const el = document.getElementById(id);
  el.innerHTML = '';
  charts[key] = new ApexCharts(el, opts);
  charts[key].render();
}
function baseChart(extra){
  return Object.assign({
    chart: { fontFamily: 'inherit', toolbar: { show: false }, foreColor: '#c9d1d9', background: 'transparent' },
    theme: { mode: 'dark' },
    grid: { borderColor: '#2b3342', strokeDashArray: 3 },
    dataLabels: { enabled: false },
    tooltip: { theme: 'dark' },
  }, extra);
}

// -------------------- filter options --------------------
function rebuildFilters(rows){
  // Priority multi
  const pSel = document.getElementById('fPriority');
  const priorities = Array.from(new Set(rows.map(r=>r.priority).filter(v=>v!==null&&v!==undefined))).sort();
  pSel.innerHTML = priorities.map(p=>`<option value="${p}">${priorityLabel(p)} (${p})</option>`).join('');
  // status
  const stSel = document.getElementById('fStatus');
  stSel.innerHTML = '<option value="">(all)</option>' +
    Array.from(new Set(rows.map(r=>cleanStatus(r.status)))).sort().map(s=>`<option value="${s}">${s}</option>`).join('');
  // programme
  const pgSel = document.getElementById('fProgramme');
  pgSel.innerHTML = '<option value="">(all)</option>' +
    Array.from(new Set(rows.map(r=>r.programme||'(blank)'))).sort().map(s=>`<option value="${s}">${s}</option>`).join('');
  // probability
  const prSel = document.getElementById('fProb');
  prSel.innerHTML = '<option value="">(all)</option>' +
    Array.from(new Set(rows.map(r=>r.external_probability||'(blank)'))).sort().map(s=>`<option value="${s}">${s}</option>`).join('');
}

function applyFilters(rows){
  const f = state.filters;
  return rows.filter(r=>{
    if(f.priorities.length && !f.priorities.includes(String(r.priority))) return false;
    if(f.status && cleanStatus(r.status)!==f.status) return false;
    if(f.programme && (r.programme||'(blank)')!==f.programme) return false;
    if(f.prob && (r.external_probability||'(blank)')!==f.prob) return false;
    if(f.search){
      const s = f.search.toLowerCase();
      const blob = [r.project, r.customer, r.programme, r.asks, r.levers, r.status, r.type_of_opportunity]
        .filter(Boolean).map(x=>String(x).toLowerCase()).join(' | ');
      if(!blob.includes(s)) return false;
    }
    return true;
  });
}

function wireFilters(){
  const f = state.filters;
  document.getElementById('fPriority').onchange = (e)=>{
    f.priorities = Array.from(e.target.selectedOptions).map(o=>o.value);
    render();
  };
  document.getElementById('fStatus').onchange = (e)=>{ f.status = e.target.value; render(); };
  document.getElementById('fProgramme').onchange = (e)=>{ f.programme = e.target.value; render(); };
  document.getElementById('fProb').onchange = (e)=>{ f.prob = e.target.value; render(); };
  document.getElementById('fSearch').oninput = (e)=>{ f.search = e.target.value.trim(); render(); };
  document.getElementById('btnClear').onclick = ()=>{
    f.priorities=[]; f.status=''; f.programme=''; f.prob=''; f.search='';
    document.getElementById('fPriority').selectedIndex = -1;
    document.getElementById('fStatus').value='';
    document.getElementById('fProgramme').value='';
    document.getElementById('fProb').value='';
    document.getElementById('fSearch').value='';
    render();
  };
}
wireFilters();

// -------------------- KPIs --------------------
function renderKPIs(rows){
  const totalOpps = rows.length;
  const sum26 = sumBy(rows,'benefit_2026');
  const sum27 = sumBy(rows,'benefit_2027');
  const sumTotal = sum26 + sum27;
  const probCounts = rows.reduce((a,r)=>{const k=(r.external_probability||'(blank)'); a[k]=(a[k]||0)+1; return a;},{});
  const probScore = rows.reduce((a,r)=>{
    const p = (r.external_probability||'').toString().toLowerCase();
    if(p.startsWith('high')) a.push(0.9);
    else if(p.startsWith('med')) a.push(0.5);
    else if(p.startsWith('low')) a.push(0.2);
    return a;
  },[]);
  const avgProb = probScore.length ? (probScore.reduce((a,b)=>a+b,0)/probScore.length) : 0;

  const priCounts = rows.reduce((a,r)=>{const k=priorityLabel(r.priority); a[k]=(a[k]||0)+1; return a;},{});

  const kpis = [
    { label: 'Opportunities', value: totalOpps, unit: '' },
    { label: 'Benefit 2026', value: fmtMoney(sum26), unit: '£m', accent: true },
    { label: 'Benefit 2027', value: fmtMoney(sum27), unit: '£m', accent: true },
    { label: 'Sum 26+27', value: fmtMoney(sumTotal), unit: '£m', accent: true },
    { label: 'Avg Probability', value: (avgProb*100).toFixed(0), unit: '%' },
    { label: 'High Priority', value: priCounts['High']||0, unit: '' },
    { label: 'Medium Priority', value: priCounts['Medium']||0, unit: '' },
    { label: 'Low Priority', value: priCounts['Low']||0, unit: '' },
  ];
  document.getElementById('kpiRow').innerHTML = kpis.map(k =>
    `<div class="kpi ${k.accent?'accent':''}"><div class="label">${k.label}</div>
     <div class="value">${k.value}<span class="unit">${k.unit}</span></div></div>`
  ).join('');
}

// -------------------- Charts --------------------
const PRIORITY_COLORS = { 'High':'#ef476f', 'Medium':'#ffd166', 'Low':'#06d6a0', '(n/a)':'#5d6b7e' };

function renderPriorityDonut(rows){
  const c = rows.reduce((a,r)=>{const k=priorityLabel(r.priority); a[k]=(a[k]||0)+1; return a;},{});
  const labels = Object.keys(c);
  const series = Object.values(c);
  const colors = labels.map(l=>PRIORITY_COLORS[l]||'#5b8cff');
  renderChart('priority','chPriority', baseChart({
    series, labels, colors,
    chart: { type:'donut', height: 280, toolbar:{show:false}, foreColor:'#c9d1d9' },
    legend: { position: 'bottom' },
    plotOptions: { pie: { donut: { size: '62%',
      labels: { show: true, total: { show: true, label: 'Total', color: '#c9d1d9' }}}} },
  }));
}
function renderProbBar(rows){
  const order = ['High','Med','Low'];
  const c = rows.reduce((a,r)=>{const k=(r.external_probability||'(blank)'); a[k]=(a[k]||0)+1; return a;},{});
  const keys = [...order.filter(k=>c[k]), ...Object.keys(c).filter(k=>!order.includes(k))];
  const series = [{ name:'Opportunities', data: keys.map(k=>c[k]||0) }];
  const colors = keys.map(k=> k==='High'?'#06d6a0':k==='Med'?'#ffd166':k==='Low'?'#ef476f':'#5b8cff');
  renderChart('prob','chProb', baseChart({
    series,
    chart: { type:'bar', height:280, toolbar:{show:false}, foreColor:'#c9d1d9' },
    xaxis: { categories: keys },
    colors: ['#5b8cff'],
    plotOptions: { bar: { distributed: true, borderRadius: 4, columnWidth: '50%' } },
    legend: { show: false },
    dataLabels: { enabled: true, style: { colors:['#0d1117'] } },
    fill: { colors },
  }));
}
function renderStatusFunnel(rows){
  // Known stages in rough order
  const stageOrder = ['Idea','Hopper','ICT','Negotiations','Contracting','Completed','Cancelled'];
  const c = rows.reduce((a,r)=>{const k=cleanStatus(r.status); a[k]=(a[k]||0)+1; return a;},{});
  const keys = [...stageOrder.filter(s=>Object.keys(c).some(k=>k.toLowerCase().trim()===s.toLowerCase())), ...Object.keys(c).filter(k=>!stageOrder.some(s=>s.toLowerCase()===k.toLowerCase().trim()))];
  // resolve actual matched keys
  const resolved = keys.map(k=>{
    const match = Object.keys(c).find(kk=>kk.toLowerCase().trim()===k.toLowerCase().trim());
    return { name: match || k, count: c[match||k]||0 };
  }).filter(x=>x.count>0);
  const series = [{ name:'Opportunities', data: resolved.map(r=>r.count) }];
  renderChart('funnel','chFunnel', baseChart({
    series,
    chart: { type:'bar', height:280, toolbar:{show:false}, foreColor:'#c9d1d9' },
    plotOptions: { bar: { horizontal: true, distributed: true, borderRadius: 4, barHeight: '65%' } },
    xaxis: { categories: resolved.map(r=>r.name) },
    legend: { show: false },
    dataLabels: { enabled: true, style:{ colors:['#0d1117']} },
    colors: ['#5b8cff','#7ed4b2','#c77dff','#ffbe6a','#ef476f','#06d6a0','#4cc9f0','#8b96a6'],
  }));
}
function renderBenefitCust(rows){
  // Top 10 customers by sum26+sum27
  const agg = {};
  rows.forEach(r=>{
    const k = r.customer || '(blank)';
    if(!agg[k]) agg[k] = { b26:0, b27:0 };
    if(typeof r.benefit_2026==='number') agg[k].b26 += r.benefit_2026;
    if(typeof r.benefit_2027==='number') agg[k].b27 += r.benefit_2027;
  });
  const ranked = Object.entries(agg).map(([k,v])=>({k, total:v.b26+v.b27, b26:v.b26, b27:v.b27}))
    .sort((a,b)=>b.total-a.total).slice(0,10);
  renderChart('benefitCust','chBenefitCust', baseChart({
    chart: { type:'bar', height:380, toolbar:{show:false}, stacked:false, foreColor:'#c9d1d9' },
    series: [
      { name:'2026', data: ranked.map(r=>+r.b26.toFixed(2)) },
      { name:'2027', data: ranked.map(r=>+r.b27.toFixed(2)) },
    ],
    xaxis: { categories: ranked.map(r=>r.k) },
    yaxis: { title:{ text:'£m', style:{color:'#8b96a6'} } },
    colors: ['#5b8cff','#7ed4b2'],
    plotOptions: { bar: { borderRadius: 3, columnWidth: '60%' } },
    legend: { position: 'top' },
  }));
}
function renderBenefitProg(rows){
  const progSet = Array.from(new Set(rows.map(r=>r.programme||'(blank)')));
  const prioLabels = ['High','Medium','Low'];
  const data = prioLabels.map(lab=>{
    return progSet.map(pg=>{
      let s = 0;
      rows.forEach(r=>{
        if((r.programme||'(blank)')===pg && priorityLabel(r.priority)===lab){
          if(typeof r.benefit_2026==='number') s += r.benefit_2026;
          if(typeof r.benefit_2027==='number') s += r.benefit_2027;
        }
      });
      return +s.toFixed(3);
    });
  });
  renderChart('benefitProg','chBenefitProg', baseChart({
    chart: { type:'bar', height:380, stacked:true, toolbar:{show:false}, foreColor:'#c9d1d9' },
    series: prioLabels.map((lab,i)=>({ name: lab+' priority', data: data[i] })),
    xaxis: { categories: progSet },
    colors: ['#ef476f','#ffd166','#06d6a0'],
    plotOptions: { bar: { borderRadius: 3, columnWidth: '55%' } },
    legend: { position: 'top' },
    yaxis: { title:{ text:'Sum 26+27 (£m)', style:{color:'#8b96a6'}} },
  }));
}
function renderLeader(rows){
  const agg = {};
  rows.forEach(r=>{
    const k = r.customer || '(blank)';
    if(!agg[k]) agg[k] = 0;
    if(typeof r.benefit_2026==='number') agg[k] += r.benefit_2026;
    if(typeof r.benefit_2027==='number') agg[k] += r.benefit_2027;
  });
  const ranked = Object.entries(agg).map(([k,v])=>({k,v})).sort((a,b)=>b.v-a.v).slice(0,10);
  renderChart('leader','chLeader', baseChart({
    chart: { type:'bar', height:380, toolbar:{show:false}, foreColor:'#c9d1d9' },
    series: [{ name:'Sum 26+27', data: ranked.map(r=>+r.v.toFixed(3)) }],
    xaxis: { categories: ranked.map(r=>r.k) },
    plotOptions: { bar: { horizontal:true, borderRadius:3, barHeight:'65%', distributed: true } },
    legend: { show: false },
    colors: ['#5b8cff','#7ed4b2','#c77dff','#ffbe6a','#ef476f','#06d6a0','#4cc9f0','#ffd166','#9d4edd','#ff6b6b'],
    dataLabels: { enabled:true, style:{ colors:['#0d1117']}, formatter: v => v.toFixed(2) },
    yaxis: { title:{ text:'£m', style:{color:'#8b96a6'}} },
  }));
}
function renderTimeline(){
  const tl = DATA.timeline || { projects: [] };
  // Build range chart: one series per milestone type; each item = {x: project, y: [start, start+7 days]}
  const typeColors = {
    'Idea Generation':'#5b8cff',
    'Approval to Launch':'#7ed4b2',
    'Strategy Approval':'#c77dff',
    'BE Generated':'#ffbe6a',
    'Approval':'#ef476f',
    'Negotiation Strategy':'#ffd166',
    'Proposal Submitted':'#4cc9f0',
    'Proposal Signed':'#06d6a0',
  };
  const byType = {};
  tl.projects.forEach(p=>{
    p.milestones.forEach(m=>{
      if(!byType[m.milestone]) byType[m.milestone] = [];
      const start = new Date(m.date).getTime();
      byType[m.milestone].push({ x: `${p.project||''} / ${p.customer||''}`, y: [start, start + 7*86400000] });
    });
  });
  const typeOrder = ['Idea Generation','Approval to Launch','Strategy Approval','BE Generated','Approval','Negotiation Strategy','Proposal Submitted','Proposal Signed'];
  const series = typeOrder.filter(t=>byType[t]).map(t=>({ name:t, data: byType[t] }));
  const colors = typeOrder.filter(t=>byType[t]).map(t=>typeColors[t]||'#5b8cff');
  renderChart('timeline','chTimeline', baseChart({
    chart: { type:'rangeBar', height: 480, toolbar:{show:false}, foreColor:'#c9d1d9' },
    series,
    plotOptions: { bar: { horizontal:true, distributed:false, rangeBarGroupRows:true, barHeight:'70%' } },
    xaxis: { type: 'datetime', labels: { style: { colors: '#8b96a6' } } },
    colors,
    legend: { position:'top' },
    tooltip: { theme:'dark', x: { format: 'dd MMM yyyy' } },
  }));
}

// -------------------- Main table --------------------
const COLS = [
  { key:'num', label:'#', num:true },
  { key:'project', label:'Project' },
  { key:'customer', label:'Customer' },
  { key:'programme', label:'Programme' },
  { key:'priority', label:'Priority' },
  { key:'external_probability', label:'Probability' },
  { key:'status', label:'Status' },
  { key:'type_of_opportunity', label:'Type' },
  { key:'benefit_2026', label:'2026 £m', num:true },
  { key:'benefit_2027', label:'2027 £m', num:true },
  { key:'sum_26_27', label:'Sum £m', num:true },
  { key:'term_benefit', label:'Term £m', num:true },
  { key:'next_milestone', label:'Next Milestone' },
];

function nextMilestoneFor(project, customer){
  const tl = DATA.timeline; if(!tl) return '';
  const today = Date.now();
  for(const p of tl.projects){
    if((p.project||'')===(project||'') && (p.customer||'')===(customer||'')){
      // find earliest milestone in future, else latest past
      const future = p.milestones.map(m=>({...m, t: new Date(m.date).getTime()})).filter(m=>m.t>=today).sort((a,b)=>a.t-b.t);
      if(future.length) return `${future[0].milestone} (${future[0].date.slice(0,10)})`;
      const past = p.milestones.map(m=>({...m, t: new Date(m.date).getTime()})).sort((a,b)=>b.t-a.t);
      if(past.length) return `${past[0].milestone} (${past[0].date.slice(0,10)})`;
    }
  }
  return '';
}

function renderTable(rows){
  const head = document.getElementById('mainHead');
  head.innerHTML = COLS.map(c=>`<th data-key="${c.key}">${c.label}${state.sortKey===c.key?(state.sortDir>0?' ↑':' ↓'):''}</th>`).join('');
  head.querySelectorAll('th').forEach(th=>{
    th.onclick = ()=>{
      const k = th.dataset.key;
      if(state.sortKey===k) state.sortDir*=-1; else { state.sortKey=k; state.sortDir=1; }
      render();
    };
  });

  let display = rows.slice();
  if(state.sortKey){
    const dir = state.sortDir;
    display.sort((a,b)=>{
      let va = a[state.sortKey], vb = b[state.sortKey];
      if(state.sortKey==='next_milestone'){ va = a._next_milestone||''; vb = b._next_milestone||''; }
      const na = (va===null||va===undefined||va==='') ? -Infinity : va;
      const nb = (vb===null||vb===undefined||vb==='') ? -Infinity : vb;
      if(typeof na==='number' && typeof nb==='number') return (na-nb)*dir;
      return String(na).localeCompare(String(nb))*dir;
    });
  }

  const body = document.getElementById('mainBody');
  body.innerHTML = display.map(r=>{
    const nm = r._next_milestone || '';
    return `<tr>
      <td class="num">${r.num??''}</td>
      <td>${r.project??''}</td>
      <td>${r.customer??''}</td>
      <td>${r.programme??''}</td>
      <td>${r.priority!==null&&r.priority!==undefined ? `<span class="pri-badge pri-${r.priority}">${priorityLabel(r.priority)}</span>`:''}</td>
      <td>${r.external_probability?`<span class="prob-badge prob-${r.external_probability}">${r.external_probability}</span>`:''}</td>
      <td>${r.status??''}</td>
      <td>${r.type_of_opportunity??''}</td>
      <td class="num">${typeof r.benefit_2026==='number'?fmtMoney(r.benefit_2026):''}</td>
      <td class="num">${typeof r.benefit_2027==='number'?fmtMoney(r.benefit_2027):''}</td>
      <td class="num">${typeof r.sum_26_27==='number'?fmtMoney(r.sum_26_27):''}</td>
      <td class="num">${typeof r.term_benefit==='number'?fmtMoney(r.term_benefit):''}</td>
      <td>${nm}</td>
    </tr>`;
  }).join('');
  document.getElementById('rowcountHint').textContent = `${display.length} / ${DATA.sheets[state.activeSheet].row_count_nonblank} rows · sheet: ${state.activeSheet}`;
}

// -------------------- Opps & Threats table --------------------
function renderOT(){
  const rows = DATA.opps_threats || [];
  const keys = rows.length ? Object.keys(rows[0]) : [];
  document.getElementById('otHead').innerHTML = keys.map(k=>`<th>${k}</th>`).join('');
  document.getElementById('otBody').innerHTML = rows.map(r=>
    `<tr>${keys.map(k=>{
      const v=r[k];
      if(v===null||v===undefined) return '<td></td>';
      if(typeof v==='number') return `<td class="num">${fmtMoney(v)}</td>`;
      if(typeof v==='string' && /^\d{4}-\d{2}-\d{2}T/.test(v)) return `<td>${v.slice(0,10)}</td>`;
      return `<td>${v}</td>`;
    }).join('')}</tr>`
  ).join('');
  document.getElementById('otHint').textContent = `${rows.length} rows (Owners, Due dates, Programme)`;
}
renderOT();

// -------------------- Render --------------------
function render(){
  // tabs
  document.querySelectorAll('.tab').forEach(t=>{
    t.classList.toggle('active', t.dataset.sheet===state.activeSheet);
  });
  const rows = DATA.sheets[state.activeSheet].rows;
  // attach computed next_milestone
  rows.forEach(r=>{ r._next_milestone = nextMilestoneFor(r.project, r.customer); });

  rebuildFilters(rows);
  const filtered = applyFilters(rows);

  renderKPIs(filtered);
  renderPriorityDonut(filtered);
  renderProbBar(filtered);
  renderStatusFunnel(filtered);
  renderBenefitCust(filtered);
  renderBenefitProg(filtered);
  renderLeader(filtered);
  renderTimeline();  // timeline is not filtered by sheet
  renderTable(filtered);
}
render();
</script>
</body>
</html>
"""


def write_html(payload: dict):
    OUT_HTML.parent.mkdir(parents=True, exist_ok=True)
    blob = json.dumps(payload, ensure_ascii=False, default=str)
    # Escape forbidden sequences that would break embedded <script>
    blob = blob.replace("</script", "<\\/script").replace("<!--", "<\\!--")
    html = HTML_TEMPLATE.replace("__PAYLOAD__", blob)
    OUT_HTML.write_text(html, encoding="utf-8")
    print(f"[build] wrote {OUT_HTML} ({OUT_HTML.stat().st_size:,} bytes)")


if __name__ == "__main__":
    payload = main()
    write_html(payload)
    # Summary stats
    print("\n=== SUMMARY ===")
    for sn, s in payload["sheets"].items():
        print(f"  {sn}: rows={s['row_count_nonblank']}  sum26={s['sum_benefit_2026']}  sum27={s['sum_benefit_2027']}")
        missing = [k for k,_ in HEADER_PATTERNS if k not in s['mapped_columns']]
        if missing:
            print(f"    unmapped: {missing}")
    print(f"  Timeline projects: {payload['timeline']['project_count']}")
    print(f"  Opps & Threats rows: {len(payload['opps_threats'])}")
