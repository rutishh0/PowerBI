"""
Build a self-contained HTML visualization for the GLOBAL LOG sheet of
Global_Commercial_Optimisation_Hopper_(v2)_(1).xlsx.

Runs independently of V6/parser.py - reads the xlsx directly with openpyxl.
Emits V6/TESTEXCEL/Global_Hopper.html embedding all data as a JSON constant,
rendered with ApexCharts 3.49.0 via CDN.

HARD CONSTRAINT: only the sheet named exactly "GLOBAL LOG" is read.
All other sheets (COVER, DETAIL_REPORT, EXEC_REPORT, Data Validations,
3+9, COUNT, SUM) are ignored.
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SRC_PATH = Path(
    r"C:\Users\Rutishkrishna\Desktop\RR\RR Powerbi\V6\New info"
    r"\Global_Commercial_Optimisation_Hopper_(v2)_(1).xlsx"
)
OUT_PATH = Path(
    r"C:\Users\Rutishkrishna\Desktop\RR\RR Powerbi\V6\TESTEXCEL\Global_Hopper.html"
)

SHEET_NAME = "GLOBAL LOG"
HEADER_ROW = 5  # 1-based
FIRST_COL = 3   # C
LAST_COL = 21   # U
FIRST_DATA_ROW = 6

# Canonical keys mirror the spec appendix, preserving "Customer " trailing space
# and "Initative" misspelling in raw_header while emitting clean JSON keys.
COLUMNS = [
    ("C",  3, "Region",                    "region",                       "str"),
    ("D",  4, "Customer ",                 "customer",                     "str"),
    ("E",  5, "Engine Value Stream",       "evs",                          "str"),
    ("F",  6, "Top Level EVS",             "top_level_evs",                "str"),
    ("G",  7, "VP/Account Manager Owner",  "vp_owner",                     "str"),
    ("H",  8, "Restructure Type",          "restructure_type",             "str"),
    ("I",  9, "Opportunity Maturity",      "maturity",                     "str"),
    ("J", 10, "Onerous/Non Onerous",       "onerous",                      "str"),
    ("K", 11, "Initative",                 "initiative",                   "str"),
    ("L", 12, "Project Plan Requirements", "project_plan_requirements",    "str"),
    ("M", 13, "Status",                    "status",                       "str"),
    ("N", 14, "Expected year of signature","signature_year",               "int"),
    ("O", 15, "Signature AP",              "signature_ap",                 "str"),
    ("P", 16, "CRP Term Benefit £m",  "crp_term_benefit",             "money"),
    ("Q", 17, "Profit 2026 £m",       "profit_2026",                  "money"),
    ("R", 18, "Profit 2027 £m",       "profit_2027",                  "money"),
    ("S", 19, "Profit 2028 £m",       "profit_2028",                  "money"),
    ("T", 20, "Profit 2029 £m",       "profit_2029",                  "money"),
    ("U", 21, "Profit 2030 £m",       "profit_2030",                  "money"),
]

MONEY_KEYS = [c[3] for c in COLUMNS if c[4] == "money"]
PROFIT_YEAR_KEYS = ["profit_2026", "profit_2027", "profit_2028", "profit_2029", "profit_2030"]

# Pipeline order from spec §3 (misspellings preserved verbatim).
STATUS_ORDER = [
    "Initial idea",
    "ICT formed",
    "Strategy Approved",
    "Financial Modelling Started",
    "Financial Modelling Complete",
    "Financials Approved",
    "Negotations Started",
    "Negotations Concluded",
    "Contracting Started",
    "Contracting Concluded",
]

MATURITY_ORDER = ["Immature", "Mature"]

# Explicit alias map — per spec §5 / §7 (never fuzzy-dedupe without review).
VP_NORMALIZE = {
    "Dan Hector": "Dan Hector",
    "Daniel Hector": "Dan Hector",
    "Nick Chadwick": "Nick Chadwick",
    "Nicholas Chadwick": "Nick Chadwick",
}


def as_numeric(v: Any) -> float | None:
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def as_text(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v)
    return s if s != "" else None


def normalize_vp(raw: str | None) -> str | None:
    if not raw:
        return None
    stripped = raw.strip()
    if stripped in VP_NORMALIZE:
        return VP_NORMALIZE[stripped]
    # Handle slash-joined duos — keep primary owner, but present normalized
    if "/" in stripped:
        primary = stripped.split("/")[0].strip()
        if primary in VP_NORMALIZE:
            return VP_NORMALIZE[primary] + " (+co-lead)"
        return primary + " (+co-lead)"
    return stripped


def extract() -> dict[str, Any]:
    wb = load_workbook(SRC_PATH, data_only=True)

    # Exact match required per directive; fallback only on stripped upper.
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        match = next(
            (s for s in wb.sheetnames if s.strip().upper() == SHEET_NAME),
            None,
        )
        if match is None:
            raise RuntimeError(
                f"Sheet 'GLOBAL LOG' not found. Available sheets: {wb.sheetnames}"
            )
        ws = wb[match]

    # Header validation fingerprint (spec §9 item 3)
    header_region = ws.cell(row=HEADER_ROW, column=FIRST_COL).value
    header_last = ws.cell(row=HEADER_ROW, column=LAST_COL).value
    assert header_region == "Region", (
        f"Header row mismatch: col C row 5 = {header_region!r}, expected 'Region'"
    )
    assert isinstance(header_last, str) and header_last.startswith("Profit 2030"), (
        f"Header row mismatch: col U row 5 = {header_last!r}, expected 'Profit 2030 £m'"
    )

    raw_headers = [
        ws.cell(row=HEADER_ROW, column=c).value
        for c in range(FIRST_COL, LAST_COL + 1)
    ]

    rows: list[dict[str, Any]] = []

    # Use max_row to be safe, but also trust that orphan row 129 exists.
    # auto_filter ref is C5:U129, so honor that hint.
    last_row_candidate = ws.max_row
    auto_filter_hint = 129
    scan_last = max(last_row_candidate, auto_filter_hint)

    for r in range(FIRST_DATA_ROW, scan_last + 1):
        raw_cells = {
            col_letter: ws.cell(row=r, column=col_idx).value
            for col_letter, col_idx, _hdr, _key, _typ in COLUMNS
        }

        # Skip fully blank rows — but DO NOT stop iterating (orphan at r=129).
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in raw_cells.values()):
            continue

        item: dict[str, Any] = {"row": r}
        notes: dict[str, str] = {}

        for col_letter, col_idx, header_text, key, typ in COLUMNS:
            v = raw_cells[col_letter]
            if typ == "str":
                item[key] = as_text(v)
            elif typ == "int":
                if isinstance(v, bool):
                    item[key] = None
                elif isinstance(v, (int, float)):
                    item[key] = int(v)
                else:
                    item[key] = None
                    text = as_text(v)
                    if text is not None:
                        notes[key] = text
            elif typ == "money":
                num = as_numeric(v)
                if num is not None:
                    item[key] = num
                else:
                    item[key] = None
                    text = as_text(v)
                    if text is not None:
                        notes[key] = text

        # Normalize customer (strip) but keep raw
        raw_customer = item.get("customer")
        item["raw_customer"] = raw_customer
        item["customer"] = raw_customer.strip() if isinstance(raw_customer, str) else None

        # Normalize VP owner
        raw_vp = item.get("vp_owner")
        item["raw_vp_owner"] = raw_vp
        item["vp_owner_normalized"] = normalize_vp(raw_vp)

        # Sum profit 26-30 (numeric only) for convenience
        profit_sum = 0.0
        for k in PROFIT_YEAR_KEYS:
            val = item.get(k)
            if isinstance(val, (int, float)):
                profit_sum += float(val)
        item["profit_total_26_30"] = profit_sum

        if notes:
            item["notes"] = notes

        rows.append(item)

    # ---- Aggregations ----
    total_count = len(rows)

    def safe_sum(items, key):
        return sum(float(i[key]) for i in items if isinstance(i.get(key), (int, float)))

    total_crp = safe_sum(rows, "crp_term_benefit")
    total_profit_26_30 = sum(safe_sum(rows, k) for k in PROFIT_YEAR_KEYS)

    unique_customers = sorted({i["customer"] for i in rows if i["customer"]})
    unique_evs = sorted({i["evs"] for i in rows if i["evs"]})
    unique_vp = sorted({
        i["vp_owner_normalized"] for i in rows if i["vp_owner_normalized"]
    })
    unique_regions = sorted({i["region"] for i in rows if i["region"]})
    unique_restructure = sorted({
        i["restructure_type"] for i in rows if i["restructure_type"]
    })
    unique_status = sorted({i["status"] for i in rows if i["status"]})
    unique_maturity = sorted({i["maturity"] for i in rows if i["maturity"]})
    unique_onerous = sorted({i["onerous"] for i in rows if i["onerous"]})

    # Region × Restructure Type CRP
    region_restructure: dict[str, dict[str, float]] = {}
    for i in rows:
        reg = i["region"] or "(blank)"
        rt = i["restructure_type"] or "(blank)"
        val = i.get("crp_term_benefit")
        if not isinstance(val, (int, float)):
            continue
        region_restructure.setdefault(reg, {}).setdefault(rt, 0.0)
        region_restructure[reg][rt] += float(val)

    # Status count (ordered per pipeline)
    status_counts: dict[str, int] = {s: 0 for s in STATUS_ORDER}
    for i in rows:
        s = i["status"]
        if s in status_counts:
            status_counts[s] += 1
        elif s:
            status_counts[s] = status_counts.get(s, 0) + 1

    # Maturity counts
    maturity_counts: dict[str, int] = {m: 0 for m in MATURITY_ORDER}
    for i in rows:
        m = i["maturity"]
        if m:
            maturity_counts[m] = maturity_counts.get(m, 0) + 1

    # EVS treemap (profit 26-30)
    evs_profit: dict[str, float] = {}
    for i in rows:
        e = i["evs"] or "(blank)"
        evs_profit[e] = evs_profit.get(e, 0.0) + float(i["profit_total_26_30"] or 0.0)

    # VP leaderboard
    vp_benefit: dict[str, float] = {}
    vp_count: dict[str, int] = {}
    for i in rows:
        vp = i["vp_owner_normalized"]
        if not vp:
            continue
        vp_count[vp] = vp_count.get(vp, 0) + 1
        val = i.get("crp_term_benefit")
        if isinstance(val, (int, float)):
            vp_benefit[vp] = vp_benefit.get(vp, 0.0) + float(val)
    vp_leaderboard = sorted(
        ({"vp": k, "benefit": v, "count": vp_count.get(k, 0)} for k, v in vp_benefit.items()),
        key=lambda x: x["benefit"],
        reverse=True,
    )[:10]

    # Signature year
    year_counts: dict[str, int] = {}
    for i in rows:
        y = i["signature_year"]
        k = str(int(y)) if isinstance(y, int) else "(blank)"
        year_counts[k] = year_counts.get(k, 0) + 1

    # Onerous donut
    onerous_counts: dict[str, int] = {}
    for i in rows:
        key = i["onerous"] or "(blank)"
        onerous_counts[key] = onerous_counts.get(key, 0) + 1

    # Profit by year line
    profit_by_year = [
        {"year": 2026, "total": safe_sum(rows, "profit_2026")},
        {"year": 2027, "total": safe_sum(rows, "profit_2027")},
        {"year": 2028, "total": safe_sum(rows, "profit_2028")},
        {"year": 2029, "total": safe_sum(rows, "profit_2029")},
        {"year": 2030, "total": safe_sum(rows, "profit_2030")},
    ]

    # Region count
    region_counts: dict[str, int] = {}
    for i in rows:
        key = i["region"] or "(blank)"
        region_counts[key] = region_counts.get(key, 0) + 1

    # Top-5 EVS by total benefit (for final report)
    evs_benefit: dict[str, float] = {}
    for i in rows:
        e = i["evs"] or "(blank)"
        val = i.get("crp_term_benefit")
        if isinstance(val, (int, float)):
            evs_benefit[e] = evs_benefit.get(e, 0.0) + float(val)

    # Rows with placeholder notes
    notes_rows = [
        {
            "row": i["row"],
            "customer": i["customer"],
            "notes": i["notes"],
        }
        for i in rows
        if "notes" in i
    ]

    # Provenance — last row inspection
    last_row_capture = rows[-1] if rows else None

    return {
        "meta": {
            "source_path": str(SRC_PATH),
            "sheet": SHEET_NAME,
            "raw_headers": raw_headers,
            "row_count": total_count,
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "last_row": last_row_capture["row"] if last_row_capture else None,
            "last_customer": last_row_capture["customer"] if last_row_capture else None,
            "currency": "GBP (millions, £m)",
        },
        "kpis": {
            "opportunities": total_count,
            "total_crp": total_crp,
            "total_profit_26_30": total_profit_26_30,
            "unique_customers": len(unique_customers),
            "unique_evs": len(unique_evs),
            "unique_vp_owners": len(unique_vp),
        },
        "filters": {
            "regions": unique_regions,
            "restructure_types": unique_restructure,
            "statuses": unique_status,
            "vp_owners": unique_vp,
            "maturities": unique_maturity,
            "onerous": unique_onerous,
        },
        "region_counts": region_counts,
        "region_restructure": region_restructure,
        "status_counts": status_counts,
        "status_order": STATUS_ORDER,
        "maturity_counts": maturity_counts,
        "evs_profit": evs_profit,
        "evs_benefit": evs_benefit,
        "vp_leaderboard": vp_leaderboard,
        "year_counts": year_counts,
        "onerous_counts": onerous_counts,
        "profit_by_year": profit_by_year,
        "notes_rows": notes_rows,
        "rows": rows,
    }


HTML_TEMPLATE = r"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8" />
<meta name="viewport" content="width=device-width, initial-scale=1" />
<title>Global Commercial Optimisation Hopper — GLOBAL LOG</title>
<script src="https://cdn.jsdelivr.net/npm/apexcharts@3.49.0/dist/apexcharts.min.js"></script>
<style>
  :root {
    --bg: #0b1220;
    --panel: #111a2e;
    --panel-2: #162241;
    --border: #24324f;
    --text: #e6ecf7;
    --muted: #8b99b8;
    --accent: #5aa2ff;
    --ok: #22c55e;
    --warn: #f59e0b;
    --bad: #ef4444;
    --purple: #a78bfa;
  }
  * { box-sizing: border-box; }
  html, body { margin: 0; padding: 0; background: var(--bg); color: var(--text);
               font: 14px/1.45 "Segoe UI", Roboto, system-ui, -apple-system, Arial, sans-serif; }
  a { color: var(--accent); }
  header { padding: 22px 28px; border-bottom: 1px solid var(--border);
           background: linear-gradient(180deg, #0f1a35, #0b1220); }
  header h1 { margin: 0 0 6px; font-size: 22px; font-weight: 600; letter-spacing: .2px; }
  header .sub { color: var(--muted); font-size: 12.5px; }
  main { padding: 22px 28px 60px; max-width: 1500px; margin: 0 auto; }
  .kpis { display: grid; grid-template-columns: repeat(6, minmax(0,1fr)); gap: 14px; margin: 18px 0 22px; }
  .kpi { background: var(--panel); border: 1px solid var(--border); border-radius: 10px; padding: 14px 16px; }
  .kpi .label { color: var(--muted); font-size: 11.5px; text-transform: uppercase; letter-spacing: .8px; }
  .kpi .val { font-size: 20px; font-weight: 600; margin-top: 6px; }
  .kpi .val.small { font-size: 16px; }
  .kpi.ok .val { color: var(--ok); }
  .kpi.warn .val { color: var(--warn); }
  .kpi.bad .val { color: var(--bad); }
  .kpi.accent .val { color: var(--accent); }
  .row-2 { display: grid; grid-template-columns: 2fr 1fr; gap: 16px; margin-bottom: 18px; }
  .row-3 { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 16px; margin-bottom: 18px; }
  .row-2b { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 18px; }
  .card { background: var(--panel); border: 1px solid var(--border); border-radius: 10px; padding: 14px 16px; }
  .card h2 { margin: 0 0 12px; font-size: 14px; font-weight: 600; color: #cfd8ec; letter-spacing: .3px; }
  .card h2 .hint { color: var(--muted); font-weight: 400; font-size: 12px; margin-left: 8px; }
  .filters { display: flex; flex-wrap: wrap; gap: 10px; margin: 4px 0 14px; align-items: center; }
  .filters input, .filters select {
    background: var(--panel-2); color: var(--text); border: 1px solid var(--border);
    border-radius: 8px; padding: 8px 10px; font-size: 13px; min-width: 140px;
  }
  .filters input[type=search] { min-width: 260px; }
  .filters button {
    background: var(--panel-2); color: var(--text); border: 1px solid var(--border);
    border-radius: 8px; padding: 8px 14px; font-size: 13px; cursor: pointer;
  }
  .filters button:hover { background: #1c2a4e; }
  .filters .count { color: var(--muted); margin-left: auto; font-size: 12.5px; }
  table { width: 100%; border-collapse: collapse; font-size: 12px; }
  th, td { padding: 7px 9px; text-align: left; border-bottom: 1px solid var(--border);
           vertical-align: top; }
  thead th { background: var(--panel-2); color: #cfd8ec; font-weight: 600; cursor: pointer;
             user-select: none; position: sticky; top: 0; font-size: 11.5px; white-space: nowrap; }
  thead th:hover { background: #1c2a4e; }
  tbody tr:hover { background: #16213e; }
  td.num { text-align: right; font-variant-numeric: tabular-nums; white-space: nowrap; }
  td.mono { font-family: ui-monospace, Consolas, monospace; font-size: 11px; }
  .pill { display: inline-block; padding: 2px 8px; border-radius: 999px; font-size: 11px;
          font-weight: 600; background: #0f2a4a; color: #7ab5ff; white-space: nowrap; }
  .pill.region { background: #12263f; color: #7ab5ff; }
  .pill.status-initial { background: #222c42; color: #8b99b8; }
  .pill.status-active { background: #0f3b22; color: #5ee89a; }
  .pill.status-nego { background: #3a2a14; color: #f5c26b; }
  .pill.status-contract { background: #28124a; color: #c7a3ff; }
  .pill.onerous { background: #3a1414; color: #ff8b8b; }
  .pill.non-onerous { background: #0f2a1f; color: #6be0a5; }
  .table-wrap { max-height: 620px; overflow: auto; border: 1px solid var(--border);
                border-radius: 10px; }
  .desc-trunc { max-width: 280px; overflow: hidden; text-overflow: ellipsis;
                white-space: nowrap; display: inline-block; }
  .notes-panel { background: #281414; border: 1px solid #5a1f1f; border-radius: 10px;
                 padding: 14px 16px; margin-bottom: 18px; }
  .notes-panel h2 { margin: 0 0 10px; font-size: 14px; font-weight: 600;
                    color: #ffb4b4; letter-spacing: .3px; }
  .notes-panel .hint { color: #c27a7a; font-weight: 400; font-size: 12px; margin-left: 8px; }
  .notes-list { display: grid; grid-template-columns: repeat(auto-fill, minmax(320px,1fr));
                gap: 10px; }
  .note-item { background: #1f1212; border: 1px solid #5a1f1f; border-radius: 8px;
               padding: 10px 12px; font-size: 12.5px; }
  .note-item .cust { font-weight: 600; color: #ffb4b4; }
  .note-item .pair { color: #c27a7a; font-size: 11.5px; margin-top: 3px;
                     font-family: ui-monospace, Consolas, monospace; }
  footer { padding: 20px 28px; color: var(--muted); font-size: 12px;
           border-top: 1px solid var(--border); }
  @media (max-width: 1200px) {
    .kpis { grid-template-columns: repeat(3, 1fr); }
    .row-2, .row-2b, .row-3 { grid-template-columns: 1fr; }
  }
</style>
</head>
<body>
<header>
  <h1>Global Commercial Optimisation Hopper — GLOBAL LOG</h1>
  <div class="sub">
    <span id="meta-file"></span> · sheet <b>GLOBAL LOG</b> ·
    <span id="meta-rows"></span> opportunities · <span id="meta-currency"></span>
  </div>
</header>
<main>
  <section class="kpis" id="kpis"></section>

  <section class="row-2">
    <div class="card">
      <h2>CRP Term Benefit by Region × Restructure Type
          <span class="hint">stacked — £m</span></h2>
      <div id="chartRegionStack"></div>
    </div>
    <div class="card">
      <h2>Status funnel <span class="hint">pipeline order — count</span></h2>
      <div id="chartStatus"></div>
    </div>
  </section>

  <section class="row-2">
    <div class="card">
      <h2>Profit 2026–2030 by Engine Value Stream
          <span class="hint">treemap — sum £m</span></h2>
      <div id="chartEvsTree"></div>
    </div>
    <div class="card">
      <h2>Top 10 VP Owners <span class="hint">by CRP Term Benefit, £m</span></h2>
      <div id="chartVpLeader"></div>
    </div>
  </section>

  <section class="row-3">
    <div class="card">
      <h2>Expected year of signature</h2>
      <div id="chartYear"></div>
    </div>
    <div class="card">
      <h2>Onerous vs Non Onerous <span class="hint">opportunity count</span></h2>
      <div id="chartOnerous"></div>
    </div>
    <div class="card">
      <h2>Opportunity maturity <span class="hint">funnel</span></h2>
      <div id="chartMaturity"></div>
    </div>
  </section>

  <section class="row-2b">
    <div class="card">
      <h2>Profit by year <span class="hint">numeric-only sum £m</span></h2>
      <div id="chartProfitYear"></div>
    </div>
    <div class="card">
      <h2>Region mix <span class="hint">opportunity count</span></h2>
      <div id="chartRegionMix"></div>
    </div>
  </section>

  <section class="notes-panel" id="notesPanel" style="display:none">
    <h2>Incomplete data <span class="hint">rows with TBD / tbc / "Confirm with …" placeholders</span></h2>
    <div class="notes-list" id="notesList"></div>
  </section>

  <section class="card">
    <h2>GLOBAL LOG data table <span class="hint">all 19 columns · filter and sort</span></h2>
    <div class="filters">
      <input id="q" type="search" placeholder="Search customer, initiative, owner..." />
      <select id="fRegion"><option value="">All regions</option></select>
      <select id="fRestructure"><option value="">All restructure types</option></select>
      <select id="fStatus"><option value="">All statuses</option></select>
      <select id="fVp"><option value="">All VP owners</option></select>
      <select id="fMaturity"><option value="">All maturity</option></select>
      <select id="fOnerous"><option value="">All onerous</option></select>
      <button id="btnReset" type="button">Reset</button>
      <span class="count" id="tableCount"></span>
    </div>
    <div class="table-wrap">
      <table id="tbl">
        <thead>
          <tr>
            <th data-k="row" class="num">Row</th>
            <th data-k="region">Region</th>
            <th data-k="customer">Customer</th>
            <th data-k="evs">EVS</th>
            <th data-k="top_level_evs">Top EVS</th>
            <th data-k="vp_owner_normalized">VP Owner</th>
            <th data-k="restructure_type">Restructure</th>
            <th data-k="maturity">Maturity</th>
            <th data-k="onerous">Onerous</th>
            <th data-k="initiative">Initiative</th>
            <th data-k="project_plan_requirements">Plan Req</th>
            <th data-k="status">Status</th>
            <th data-k="signature_year" class="num">Sig Year</th>
            <th data-k="signature_ap">AP</th>
            <th data-k="crp_term_benefit" class="num">CRP £m</th>
            <th data-k="profit_2026" class="num">2026</th>
            <th data-k="profit_2027" class="num">2027</th>
            <th data-k="profit_2028" class="num">2028</th>
            <th data-k="profit_2029" class="num">2029</th>
            <th data-k="profit_2030" class="num">2030</th>
          </tr>
        </thead>
        <tbody id="tbody"></tbody>
      </table>
    </div>
  </section>
</main>
<footer>
  Generated <span id="gen-ts"></span> ·
  Source <span id="f-src"></span> ·
  Last captured row <span id="f-last"></span> (<span id="f-last-cust"></span>) ·
  <b>Uses GLOBAL LOG sheet only — other sheets (COVER, DETAIL_REPORT, EXEC_REPORT,
     Data Validations, 3+9, COUNT, SUM) intentionally ignored per user directive.</b>
</footer>

<script id="hopper-data" type="application/json">__DATA_JSON__</script>
<script>
(function(){
  const D = JSON.parse(document.getElementById('hopper-data').textContent);

  const fmtM = (v) => {
    if (v == null || isNaN(v)) return '';
    const n = Number(v);
    const abs = Math.abs(n);
    const digits = abs >= 100 ? 0 : abs >= 10 ? 1 : 2;
    return (n < 0 ? '-' : '') + '£' + abs.toLocaleString(undefined,
      {minimumFractionDigits: digits, maximumFractionDigits: digits}) + 'm';
  };
  const fmtInt = (v) => Number(v || 0).toLocaleString();
  const fmtPlain = (v) => {
    if (v == null || isNaN(v)) return '';
    const n = Number(v);
    return n.toLocaleString(undefined, {maximumFractionDigits: 2});
  };

  // Meta
  document.getElementById('meta-file').textContent = D.meta.source_path;
  document.getElementById('meta-rows').textContent = D.meta.row_count;
  document.getElementById('meta-currency').textContent = D.meta.currency;
  document.getElementById('gen-ts').textContent = D.meta.generated_at;
  document.getElementById('f-src').textContent = D.meta.source_path;
  document.getElementById('f-last').textContent = D.meta.last_row;
  document.getElementById('f-last-cust').textContent = D.meta.last_customer || '(unknown)';

  // KPIs
  const kpis = [
    {label:'Total opportunities', val:fmtInt(D.kpis.opportunities), cls:'accent'},
    {label:'Total CRP Term Benefit', val:fmtM(D.kpis.total_crp), cls:'ok'},
    {label:'Profit 2026–2030 sum', val:fmtM(D.kpis.total_profit_26_30), cls:'ok'},
    {label:'Unique customers', val:fmtInt(D.kpis.unique_customers)},
    {label:'Unique EVS', val:fmtInt(D.kpis.unique_evs)},
    {label:'Unique VP owners', val:fmtInt(D.kpis.unique_vp_owners)},
  ];
  document.getElementById('kpis').innerHTML = kpis.map(k =>
    `<div class="kpi ${k.cls||''}"><div class="label">${k.label}</div>
     <div class="val ${k.cls==='small'?'small':''}">${k.val}</div></div>`
  ).join('');

  // ---------------- Chart 1: Region × Restructure Type stacked bar ----------------
  const regions = Object.keys(D.region_restructure).sort();
  const allRestructure = Array.from(new Set(
    regions.flatMap(r => Object.keys(D.region_restructure[r]))
  )).sort();
  const seriesStack = allRestructure.map(rt => ({
    name: rt,
    data: regions.map(r => Number((D.region_restructure[r]||{})[rt] || 0).toFixed(2) * 1),
  }));
  new ApexCharts(document.getElementById('chartRegionStack'), {
    chart: {type:'bar', stacked: true, height: 360, toolbar:{show:false},
            background:'transparent', foreColor:'#cfd8ec'},
    theme: {mode:'dark'},
    series: seriesStack,
    xaxis: {categories: regions, labels: {style:{colors:'#cfd8ec'}}},
    yaxis: {labels: {formatter: v => fmtM(v)}},
    plotOptions: {bar: {borderRadius: 3, columnWidth: '55%'}},
    colors: ['#60a5fa','#f59e0b','#a78bfa','#22c55e','#ef4444','#64748b'],
    dataLabels: {enabled: false},
    legend: {position: 'bottom', horizontalAlign: 'left'},
    grid: {borderColor: '#24324f'},
    tooltip: {y: {formatter: v => fmtM(v)}},
  }).render();

  // ---------------- Chart 2: Status funnel ----------------
  const statusOrdered = D.status_order.filter(s => D.status_counts[s] != null);
  const extraStatus = Object.keys(D.status_counts).filter(s => !D.status_order.includes(s));
  const statusKeys = statusOrdered.concat(extraStatus);
  new ApexCharts(document.getElementById('chartStatus'), {
    chart: {type:'bar', height: 360, toolbar:{show:false},
            background:'transparent', foreColor:'#cfd8ec'},
    theme: {mode:'dark'},
    series: [{name:'Opportunities', data: statusKeys.map(s => D.status_counts[s] || 0)}],
    plotOptions: {bar: {horizontal: true, borderRadius: 3, distributed: true,
                         barHeight: '85%'}},
    xaxis: {categories: statusKeys, labels: {style:{colors:'#cfd8ec'}}},
    colors: ['#475569','#64748b','#60a5fa','#3b82f6','#8b5cf6','#a78bfa',
             '#f59e0b','#f97316','#22c55e','#16a34a'],
    dataLabels: {enabled: true, style: {colors: ['#fff']}},
    legend: {show: false},
    grid: {borderColor: '#24324f'},
  }).render();

  // ---------------- Chart 3: EVS treemap (profit 26-30) ----------------
  const evsData = Object.entries(D.evs_profit)
    .filter(([k,v]) => v > 0)
    .sort((a,b) => b[1]-a[1])
    .map(([k,v]) => ({x: k, y: Number(v.toFixed(2))}));
  new ApexCharts(document.getElementById('chartEvsTree'), {
    chart: {type:'treemap', height: 360, toolbar:{show:false},
            background:'transparent', foreColor:'#cfd8ec'},
    theme: {mode:'dark'},
    series: [{data: evsData}],
    colors: ['#60a5fa','#a78bfa','#22c55e','#f59e0b','#ef4444',
             '#8b5cf6','#06b6d4','#f97316','#eab308','#84cc16'],
    plotOptions: {treemap: {distributed: true, enableShades: false}},
    dataLabels: {enabled: true, style: {fontSize:'12px'},
                 formatter: (t, op) => [t, fmtM(op.value)]},
    tooltip: {y: {formatter: v => fmtM(v)}},
    legend: {show: false},
  }).render();

  // ---------------- Chart 4: VP leaderboard ----------------
  new ApexCharts(document.getElementById('chartVpLeader'), {
    chart: {type:'bar', height: 360, toolbar:{show:false},
            background:'transparent', foreColor:'#cfd8ec'},
    theme: {mode:'dark'},
    series: [{name:'CRP Term Benefit', data: D.vp_leaderboard.map(v => Number(v.benefit.toFixed(2)))}],
    plotOptions: {bar: {horizontal: true, borderRadius: 3, distributed: true,
                         barHeight: '75%'}},
    xaxis: {categories: D.vp_leaderboard.map(v => v.vp + ' ('+v.count+')'),
            labels: {formatter: v => fmtM(v), style:{colors:'#cfd8ec'}}},
    colors: ['#60a5fa','#a78bfa','#22c55e','#f59e0b','#ef4444',
             '#06b6d4','#8b5cf6','#eab308','#f97316','#84cc16'],
    dataLabels: {enabled: true, formatter: v => fmtM(v), style:{colors:['#fff']}},
    legend: {show: false},
    grid: {borderColor: '#24324f'},
    tooltip: {y: {formatter: v => fmtM(v)}},
  }).render();

  // ---------------- Chart 5: Year column ----------------
  const yearsSorted = Object.keys(D.year_counts).sort((a,b) => {
    if (a === '(blank)') return 1;
    if (b === '(blank)') return -1;
    return Number(a) - Number(b);
  });
  new ApexCharts(document.getElementById('chartYear'), {
    chart: {type:'bar', height: 280, toolbar:{show:false},
            background:'transparent', foreColor:'#cfd8ec'},
    theme: {mode:'dark'},
    series: [{name:'Opportunities', data: yearsSorted.map(k => D.year_counts[k] || 0)}],
    xaxis: {categories: yearsSorted, labels: {style:{colors:'#cfd8ec'}}},
    plotOptions: {bar: {borderRadius: 3, columnWidth: '45%', distributed: true}},
    colors: ['#60a5fa','#22c55e','#f59e0b','#64748b'],
    dataLabels: {enabled: true, style: {colors:['#fff']}},
    legend: {show: false},
    grid: {borderColor: '#24324f'},
  }).render();

  // ---------------- Chart 6: Onerous donut ----------------
  const oKeys = Object.keys(D.onerous_counts);
  new ApexCharts(document.getElementById('chartOnerous'), {
    chart: {type:'donut', height: 280,
            background:'transparent', foreColor:'#cfd8ec'},
    theme: {mode:'dark'},
    series: oKeys.map(k => D.onerous_counts[k]),
    labels: oKeys,
    colors: oKeys.map(k => {
      if (k === 'Onerous Contract') return '#ef4444';
      if (k === 'Not Onerous') return '#22c55e';
      return '#64748b';
    }),
    legend: {position: 'bottom'},
    dataLabels: {enabled: true},
    plotOptions: {pie: {donut: {size:'60%'}}},
  }).render();

  // ---------------- Chart 7 (was 8): Maturity funnel ----------------
  const mKeys = Object.keys(D.maturity_counts).filter(k => D.maturity_counts[k] > 0);
  new ApexCharts(document.getElementById('chartMaturity'), {
    chart: {type:'bar', height: 280, toolbar:{show:false},
            background:'transparent', foreColor:'#cfd8ec'},
    theme: {mode:'dark'},
    series: [{name:'Opportunities', data: mKeys.map(k => D.maturity_counts[k])}],
    plotOptions: {bar: {horizontal: true, borderRadius: 3, distributed: true,
                         barHeight: '60%'}},
    xaxis: {categories: mKeys, labels: {style:{colors:'#cfd8ec'}}},
    colors: ['#f59e0b','#22c55e'],
    dataLabels: {enabled: true, style: {colors:['#fff']}},
    legend: {show: false},
    grid: {borderColor: '#24324f'},
  }).render();

  // ---------------- Chart: Profit by year line ----------------
  new ApexCharts(document.getElementById('chartProfitYear'), {
    chart: {type:'area', height: 280, toolbar:{show:false},
            background:'transparent', foreColor:'#cfd8ec'},
    theme: {mode:'dark'},
    series: [{name:'Profit', data: D.profit_by_year.map(p => Number(p.total.toFixed(2)))}],
    xaxis: {categories: D.profit_by_year.map(p => String(p.year))},
    yaxis: {labels: {formatter: v => fmtM(v)}},
    colors: ['#60a5fa'],
    stroke: {curve:'smooth', width: 2},
    fill: {type:'gradient', gradient:{shadeIntensity:1, opacityFrom:0.4, opacityTo:0.05}},
    dataLabels: {enabled: true, formatter: v => fmtM(v), style:{colors:['#fff']}},
    grid: {borderColor: '#24324f'},
    tooltip: {y: {formatter: v => fmtM(v)}},
  }).render();

  // ---------------- Chart: Region mix (donut) ----------------
  const rKeys = Object.keys(D.region_counts).sort((a,b) => D.region_counts[b] - D.region_counts[a]);
  new ApexCharts(document.getElementById('chartRegionMix'), {
    chart: {type:'donut', height: 280,
            background:'transparent', foreColor:'#cfd8ec'},
    theme: {mode:'dark'},
    series: rKeys.map(k => D.region_counts[k]),
    labels: rKeys,
    colors: ['#60a5fa','#a78bfa','#22c55e','#f59e0b','#ef4444','#64748b'],
    legend: {position: 'bottom'},
    dataLabels: {enabled: true},
    plotOptions: {pie: {donut: {size:'55%'}}},
  }).render();

  // ---------------- Notes panel ----------------
  if (D.notes_rows.length > 0) {
    document.getElementById('notesPanel').style.display = 'block';
    document.getElementById('notesList').innerHTML = D.notes_rows.map(n => {
      const pairs = Object.entries(n.notes).map(([k,v]) =>
        `<div class="pair">${k}: ${String(v).replace(/</g,'&lt;')}</div>`
      ).join('');
      return `<div class="note-item">
        <div class="cust">${n.customer || '(no customer)'} · row ${n.row}</div>
        ${pairs}
      </div>`;
    }).join('');
  }

  // ---------------- Table with filters ----------------
  const tbody = document.getElementById('tbody');
  const qEl = document.getElementById('q');
  const fRegion = document.getElementById('fRegion');
  const fRestructure = document.getElementById('fRestructure');
  const fStatus = document.getElementById('fStatus');
  const fVp = document.getElementById('fVp');
  const fMaturity = document.getElementById('fMaturity');
  const fOnerous = document.getElementById('fOnerous');
  const btnReset = document.getElementById('btnReset');
  const countEl = document.getElementById('tableCount');

  const fillSelect = (el, items) => {
    items.forEach(v => {
      if (!v) return;
      const o = document.createElement('option');
      o.value = v; o.textContent = v;
      el.appendChild(o);
    });
  };
  fillSelect(fRegion, D.filters.regions);
  fillSelect(fRestructure, D.filters.restructure_types);
  fillSelect(fStatus, D.filters.statuses);
  fillSelect(fVp, D.filters.vp_owners);
  fillSelect(fMaturity, D.filters.maturities);
  fillSelect(fOnerous, D.filters.onerous);

  let sortKey = 'crp_term_benefit';
  let sortDir = -1;

  const statusClass = (s) => {
    if (!s) return '';
    if (s.startsWith('Initial') || s.startsWith('ICT')) return 'status-initial';
    if (s.startsWith('Strategy') || s.startsWith('Financial')) return 'status-active';
    if (s.startsWith('Negotation') || s.startsWith('Negotiation')) return 'status-nego';
    if (s.startsWith('Contract')) return 'status-contract';
    return '';
  };

  const onerousClass = (s) => {
    if (!s) return '';
    if (s === 'Onerous Contract') return 'onerous';
    if (s === 'Not Onerous') return 'non-onerous';
    return '';
  };

  const truncate = (s, n) => {
    if (s == null) return '';
    s = String(s).replace(/\n/g, ' ');
    return s.length > n ? s.slice(0, n-1) + '…' : s;
  };

  function escapeHtml(s) {
    if (s == null) return '';
    return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;')
                    .replace(/"/g,'&quot;');
  }

  function render() {
    const q = qEl.value.trim().toLowerCase();
    const fR = fRegion.value;
    const fRt = fRestructure.value;
    const fS = fStatus.value;
    const fV = fVp.value;
    const fM = fMaturity.value;
    const fO = fOnerous.value;

    let rows = D.rows.filter(i => {
      if (fR && i.region !== fR) return false;
      if (fRt && i.restructure_type !== fRt) return false;
      if (fS && i.status !== fS) return false;
      if (fV && i.vp_owner_normalized !== fV) return false;
      if (fM && i.maturity !== fM) return false;
      if (fO && i.onerous !== fO) return false;
      if (q) {
        const hay = [
          i.customer, i.initiative, i.vp_owner_normalized, i.raw_vp_owner,
          i.evs, i.top_level_evs, i.region, i.restructure_type, i.status,
          i.project_plan_requirements, i.signature_ap
        ].filter(Boolean).join(' ').toLowerCase();
        if (!hay.includes(q)) return false;
      }
      return true;
    });

    rows.sort((a, b) => {
      const va = a[sortKey], vb = b[sortKey];
      if (va == null && vb == null) return 0;
      if (va == null) return 1;
      if (vb == null) return -1;
      if (typeof va === 'number' && typeof vb === 'number') return (va - vb) * sortDir;
      return String(va).localeCompare(String(vb)) * sortDir;
    });

    countEl.textContent = rows.length + ' of ' + D.rows.length + ' opportunities';

    tbody.innerHTML = rows.map(i => `
      <tr>
        <td class="num mono">${i.row}</td>
        <td><span class="pill region">${escapeHtml(i.region || '')}</span></td>
        <td>${escapeHtml(i.customer || '')}</td>
        <td>${escapeHtml(i.evs || '')}</td>
        <td>${escapeHtml(i.top_level_evs || '')}</td>
        <td>${escapeHtml(i.vp_owner_normalized || '')}</td>
        <td>${escapeHtml(i.restructure_type || '')}</td>
        <td>${escapeHtml(i.maturity || '')}</td>
        <td>${i.onerous ? `<span class="pill ${onerousClass(i.onerous)}">${escapeHtml(i.onerous)}</span>` : ''}</td>
        <td title="${escapeHtml(i.initiative || '')}">${escapeHtml(truncate(i.initiative, 60))}</td>
        <td>${escapeHtml(i.project_plan_requirements || '')}</td>
        <td>${i.status ? `<span class="pill ${statusClass(i.status)}">${escapeHtml(i.status)}</span>` : ''}</td>
        <td class="num">${i.signature_year != null ? i.signature_year : ''}</td>
        <td>${escapeHtml(i.signature_ap || '')}</td>
        <td class="num">${i.crp_term_benefit != null ? fmtPlain(i.crp_term_benefit)
                                                      : (i.notes && i.notes.crp_term_benefit
                                                         ? '<span style="color:#ffb4b4">'+escapeHtml(i.notes.crp_term_benefit)+'</span>'
                                                         : '')}</td>
        <td class="num">${i.profit_2026 != null ? fmtPlain(i.profit_2026)
                                                : (i.notes && i.notes.profit_2026
                                                   ? '<span style="color:#ffb4b4">'+escapeHtml(i.notes.profit_2026)+'</span>'
                                                   : '')}</td>
        <td class="num">${i.profit_2027 != null ? fmtPlain(i.profit_2027)
                                                : (i.notes && i.notes.profit_2027
                                                   ? '<span style="color:#ffb4b4">'+escapeHtml(i.notes.profit_2027)+'</span>'
                                                   : '')}</td>
        <td class="num">${i.profit_2028 != null ? fmtPlain(i.profit_2028)
                                                : (i.notes && i.notes.profit_2028
                                                   ? '<span style="color:#ffb4b4">'+escapeHtml(i.notes.profit_2028)+'</span>'
                                                   : '')}</td>
        <td class="num">${i.profit_2029 != null ? fmtPlain(i.profit_2029)
                                                : (i.notes && i.notes.profit_2029
                                                   ? '<span style="color:#ffb4b4">'+escapeHtml(i.notes.profit_2029)+'</span>'
                                                   : '')}</td>
        <td class="num">${i.profit_2030 != null ? fmtPlain(i.profit_2030)
                                                : (i.notes && i.notes.profit_2030
                                                   ? '<span style="color:#ffb4b4">'+escapeHtml(i.notes.profit_2030)+'</span>'
                                                   : '')}</td>
      </tr>
    `).join('');
  }

  document.querySelectorAll('thead th').forEach(th => {
    th.addEventListener('click', () => {
      const k = th.dataset.k;
      if (!k) return;
      if (sortKey === k) sortDir = -sortDir;
      else {
        sortKey = k;
        const numericCols = ['row','signature_year','crp_term_benefit',
                             'profit_2026','profit_2027','profit_2028',
                             'profit_2029','profit_2030'];
        sortDir = numericCols.includes(k) ? -1 : 1;
      }
      render();
    });
  });

  [qEl, fRegion, fRestructure, fStatus, fVp, fMaturity, fOnerous]
    .forEach(el => el.addEventListener('input', render));

  btnReset.addEventListener('click', () => {
    qEl.value = '';
    [fRegion, fRestructure, fStatus, fVp, fMaturity, fOnerous].forEach(el => el.value = '');
    render();
  });

  render();
})();
</script>
</body>
</html>
"""


def build() -> None:
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    data = extract()
    payload = json.dumps(data, default=str, ensure_ascii=False).replace("</", "<\\/")
    html = HTML_TEMPLATE.replace("__DATA_JSON__", payload)
    OUT_PATH.write_text(html, encoding="utf-8")

    # Console report
    kpis = data["kpis"]
    rows = data["rows"]
    region_counts = data["region_counts"]
    evs_benefit = data["evs_benefit"]
    vp_lb = data["vp_leaderboard"]
    notes_rows = data["notes_rows"]

    print(f"[OK] Opportunities        : {kpis['opportunities']}")
    print(f"[OK] Last row captured    : row {data['meta']['last_row']} "
          f"(customer={data['meta']['last_customer']!r})")
    print(f"[OK] Total CRP Term £m    : {kpis['total_crp']:.2f}")
    print(f"[OK] Profit 2026-2030 £m  : {kpis['total_profit_26_30']:.2f}")
    print(f"[OK] Unique customers     : {kpis['unique_customers']}")
    print(f"[OK] Unique EVS           : {kpis['unique_evs']}")
    print(f"[OK] Unique VP (norm)     : {kpis['unique_vp_owners']}")
    print("[OK] Top 5 regions by count:")
    for reg, cnt in sorted(region_counts.items(), key=lambda x: -x[1])[:5]:
        print(f"        {reg:18s} {cnt}")
    print("[OK] Top 5 EVS by CRP benefit (£m):")
    for evs, val in sorted(evs_benefit.items(), key=lambda x: -x[1])[:5]:
        print(f"        {evs:18s} {val:9.2f}")
    print("[OK] Top 5 VP owners by CRP benefit (£m):")
    for vp in vp_lb[:5]:
        print(f"        {vp['vp']:35s} {vp['benefit']:9.2f}  (n={vp['count']})")
    print(f"[OK] Rows with placeholder notes: {len(notes_rows)}")
    for n in notes_rows[:3]:
        pairs = ", ".join(f"{k}={v!r}" for k, v in n["notes"].items())
        print(f"        row {n['row']:>3} · {n['customer']!r} · {pairs}")
    print(f"[OK] Output               : {OUT_PATH}")
    print(f"[OK] Size                 : {OUT_PATH.stat().st_size:,} bytes")


if __name__ == "__main__":
    build()
