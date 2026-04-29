"""
Build a self-contained HTML visualization for the Hopper v2 ORIGINAL file
(NOT the (1) variant): only the GLOBAL LOG sheet is parsed.

Emits V6/TESTEXCEL/Global_Hopper_v2_original.html with data embedded as JSON
and rendered with ApexCharts 3.49.0 via CDN.

Ground truth expectations (per _hopper_global_log_spec.md + direct audit):
- Sheet 'GLOBAL LOG' exists; 7 sheets total in workbook
- Header row = 5, data row range = 6..128 (103 data rows, ~107 hidden)
- No orphan Uganda row at 129 (that only exists in the (1) variant)
- 19 real columns C..U, same schema as the (1) variant
- Mixed-type money columns P..U: numeric when possible, preserve string placeholders as notes

This script is independent of V6/parser.py.
"""

from __future__ import annotations

import html
import json
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SRC_PATH = Path(
    r"C:\Users\Rutishkrishna\Desktop\RR\RR Powerbi\V6\New info\Global Commercial Optimisation Hopper (v2).xlsx"
)
OUT_PATH = Path(
    r"C:\Users\Rutishkrishna\Desktop\RR\RR Powerbi\V6\TESTEXCEL\Global_Hopper_v2_original.html"
)

SHEET_NAME = "GLOBAL LOG"
HEADER_ROW = 5
DATA_FIRST_ROW = 6
COL_FIRST = 3   # C
COL_LAST = 21   # U

# Canonical column set (col letter, 1-based index, verbatim header, json key, is_money)
COLUMNS = [
    ("C",  3, "Region",                     "region",                False),
    ("D",  4, "Customer ",                  "customer",              False),  # trailing space in sheet
    ("E",  5, "Engine Value Stream",        "evs",                   False),
    ("F",  6, "Top Level EVS",              "top_level_evs",         False),
    ("G",  7, "VP/Account Manager Owner",   "vp_owner",              False),
    ("H",  8, "Restructure Type",           "restructure_type",      False),
    ("I",  9, "Opportunity Maturity",       "opportunity_maturity",  False),
    ("J", 10, "Onerous/Non Onerous",        "onerous",               False),
    ("K", 11, "Initative",                  "initiative",            False),  # sic
    ("L", 12, "Project Plan Requirements",  "project_plan_req",      False),
    ("M", 13, "Status",                     "status",                False),
    ("N", 14, "Expected year of signature", "year_of_signature",     False),
    ("O", 15, "Signature AP",               "signature_ap",          False),
    ("P", 16, "CRP Term Benefit £m",   "crp_benefit",           True),
    ("Q", 17, "Profit 2026 £m",        "profit_2026",           True),
    ("R", 18, "Profit 2027 £m",        "profit_2027",           True),
    ("S", 19, "Profit 2028 £m",        "profit_2028",           True),
    ("T", 20, "Profit 2029 £m",        "profit_2029",           True),
    ("U", 21, "Profit 2030 £m",        "profit_2030",           True),
]

MONEY_KEYS = [k for (_, _, _, k, is_money) in COLUMNS if is_money]
PROFIT_YEAR_KEYS = ["profit_2026", "profit_2027", "profit_2028", "profit_2029", "profit_2030"]

# Canonical pipeline ordering for Status funnel (misspellings preserved)
STATUS_ORDER = [
    "Initial idea",
    "ICT formed",
    "Strategy Approved",
    "Financial Modelling Started",
    "Financial Modelling Complete",
    "Financials Approved",
    "Negotations Started",
    "Negotations Concluded",
    "Contracting Started",
    "Contracting Concluded",
]

# VP/Account Manager alias map -> canonical name
VP_ALIAS = {
    "Dan Hector": "Daniel Hector",
    "Daniel Hector": "Daniel Hector",
    "Nick Chadwick": "Nicholas Chadwick",
    "Nicholas Chadwick": "Nicholas Chadwick",
}


def canonical_vp(name: str | None) -> str:
    if name is None:
        return ""
    base = str(name).strip()
    if base == "":
        return ""
    # strip trailing-space duplicate bug
    return VP_ALIAS.get(base, base)


def coerce_year(v: Any) -> int | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            i = int(v)
            if 2000 <= i <= 2100:
                return i
        except Exception:
            return None
    s = str(v).strip()
    try:
        i = int(float(s))
        if 2000 <= i <= 2100:
            return i
    except Exception:
        return None
    return None


def coerce_money(v: Any) -> tuple[float | None, str | None]:
    """
    Returns (numeric_value, note_text).
    If cell is numeric -> (float, None).
    If cell is a non-numeric string (TBD / tbc / Confirm with X) -> (None, string).
    Blank -> (None, None).
    """
    if v is None:
        return None, None
    if isinstance(v, bool):
        return None, None
    if isinstance(v, (int, float)):
        return float(v), None
    s = str(v).strip()
    if s == "":
        return None, None
    # Try numeric conversion (handles cases where the cell stored a numeric string)
    try:
        return float(s.replace(",", "")), None
    except Exception:
        return None, s


def cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def extract() -> dict[str, Any]:
    wb = load_workbook(SRC_PATH, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(
            f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}"
        )
    all_sheets = list(wb.sheetnames)
    ws = wb[SHEET_NAME]

    # Sanity-check header row
    assert ws.cell(row=HEADER_ROW, column=3).value == "Region", (
        f"Header row check failed: C5 = {ws.cell(row=HEADER_ROW, column=3).value!r}"
    )
    assert ws.cell(row=HEADER_ROW, column=21).value == "Profit 2030 £m", (
        f"Header row check failed: U5 = {ws.cell(row=HEADER_ROW, column=21).value!r}"
    )

    # Snapshot hidden rows
    hidden_rows = {
        r for r, rd in ws.row_dimensions.items() if rd.hidden
    }

    rows: list[dict[str, Any]] = []

    for r in range(DATA_FIRST_ROW, ws.max_row + 1):
        raw = [ws.cell(row=r, column=c).value for c in range(COL_FIRST, COL_LAST + 1)]
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in raw):
            continue

        rec: dict[str, Any] = {"row_index": r, "row_hidden": r in hidden_rows}
        notes: dict[str, str] = {}

        for i, (_letter, col_idx, _header, key, is_money) in enumerate(COLUMNS):
            v = raw[col_idx - COL_FIRST]
            if key == "year_of_signature":
                rec[key] = coerce_year(v)
            elif is_money:
                num, note = coerce_money(v)
                rec[key] = num
                if note is not None:
                    notes[key] = note
            else:
                s = "" if v is None else str(v)
                if key == "customer":
                    rec["customer_raw"] = s
                    rec[key] = s.strip()
                elif key == "vp_owner":
                    rec["vp_owner_raw"] = s
                    rec[key] = canonical_vp(s)
                else:
                    rec[key] = s.strip() if isinstance(v, str) else (v if v is not None else "")
                    if isinstance(rec[key], str):
                        rec[key] = rec[key]

        # Force string types (not None, not blank-fill) for categorical keys we use
        for k in ["region", "evs", "top_level_evs", "restructure_type", "opportunity_maturity",
                  "onerous", "status", "signature_ap", "project_plan_req", "initiative",
                  "customer", "vp_owner"]:
            if not isinstance(rec.get(k), str):
                rec[k] = "" if rec.get(k) is None else str(rec[k])

        rec["notes"] = notes
        rows.append(rec)

    # ---- Aggregations ----

    total_count = len(rows)
    hidden_in_source = sum(1 for r in rows if r["row_hidden"])

    def num_sum(key: str) -> float:
        return round(sum((r[key] or 0.0) for r in rows if isinstance(r.get(key), (int, float))), 4)

    crp_total = num_sum("crp_benefit")
    profit_totals = {k: num_sum(k) for k in PROFIT_YEAR_KEYS}
    profit_grand = round(sum(profit_totals.values()), 4)

    unique_customers = sorted({r["customer"] for r in rows if r["customer"]})
    unique_evs = sorted({r["evs"] for r in rows if r["evs"]})
    unique_vp = sorted({r["vp_owner"] for r in rows if r["vp_owner"]})

    # Region counts
    region_counts = Counter(r["region"] for r in rows if r["region"])

    # Region x Restructure Type -> sum CRP benefit
    regions = sorted({r["region"] for r in rows if r["region"]})
    rtypes = sorted({r["restructure_type"] for r in rows if r["restructure_type"]}) or ["(blank)"]
    region_rtype_crp: dict[str, dict[str, float]] = {
        rt: {rg: 0.0 for rg in regions} for rt in rtypes
    }
    for r in rows:
        rg = r["region"]
        rt = r["restructure_type"] or "(blank)"
        val = r.get("crp_benefit")
        if not rg or not isinstance(val, (int, float)):
            continue
        region_rtype_crp.setdefault(rt, {rg2: 0.0 for rg2 in regions})
        region_rtype_crp[rt].setdefault(rg, 0.0)
        region_rtype_crp[rt][rg] += val

    # Status funnel (canonical order)
    status_counter = Counter(r["status"] for r in rows if r["status"])
    status_funnel = [
        {"status": s, "count": status_counter.get(s, 0)} for s in STATUS_ORDER
    ]
    other_statuses = [
        {"status": s, "count": c}
        for s, c in sorted(status_counter.items(), key=lambda x: -x[1])
        if s not in STATUS_ORDER
    ]

    # Treemap: Profit 2026-2030 by EVS (sum of numeric only)
    evs_profit: dict[str, float] = defaultdict(float)
    for r in rows:
        evs = r["evs"] or "(blank)"
        total = 0.0
        has_any = False
        for k in PROFIT_YEAR_KEYS:
            v = r.get(k)
            if isinstance(v, (int, float)):
                total += v
                has_any = True
        if has_any:
            evs_profit[evs] += total
    evs_profit_list = sorted(
        ({"evs": k, "profit": round(v, 4)} for k, v in evs_profit.items()),
        key=lambda x: -x["profit"],
    )

    # VP leaderboard (top 10 by CRP benefit, fallback count)
    vp_agg: dict[str, dict[str, float]] = defaultdict(lambda: {"crp": 0.0, "count": 0})
    for r in rows:
        owner = r["vp_owner"]
        if not owner:
            continue
        vp_agg[owner]["count"] += 1
        v = r.get("crp_benefit")
        if isinstance(v, (int, float)):
            vp_agg[owner]["crp"] += v
    vp_leaderboard = sorted(
        ({"owner": k, "crp": round(v["crp"], 4), "count": int(v["count"])} for k, v in vp_agg.items()),
        key=lambda x: (-x["crp"], -x["count"]),
    )[:10]

    # Year of signature column chart (integer bucket + blank)
    year_counter: Counter = Counter()
    blank_year = 0
    for r in rows:
        y = r.get("year_of_signature")
        if isinstance(y, int):
            year_counter[y] += 1
        else:
            blank_year += 1
    year_series = [
        {"year": str(y), "count": year_counter[y]} for y in sorted(year_counter)
    ]
    if blank_year:
        year_series.append({"year": "Blank", "count": blank_year})

    # Onerous donut
    onerous_counter = Counter(r["onerous"] if r["onerous"] else "(blank)" for r in rows)
    onerous_series = [
        {"label": k, "count": v} for k, v in sorted(onerous_counter.items(), key=lambda x: -x[1])
    ]

    # Profit line (aggregate sums per year)
    profit_line = [
        {"year": k.replace("profit_", ""), "profit": round(profit_totals[k], 4)}
        for k in PROFIT_YEAR_KEYS
    ]

    # Maturity funnel
    maturity_counter = Counter(r["opportunity_maturity"] if r["opportunity_maturity"] else "(blank)" for r in rows)
    maturity_funnel = [
        {"maturity": k, "count": v} for k, v in sorted(maturity_counter.items(), key=lambda x: -x[1])
    ]

    # Notes panel rows
    note_rows = [
        {
            "row_index": r["row_index"],
            "customer": r["customer"],
            "region": r["region"],
            "evs": r["evs"],
            "notes": r["notes"],
        }
        for r in rows
        if r["notes"]
    ]

    return {
        "meta": {
            "source_path": str(SRC_PATH),
            "source_file_name": SRC_PATH.name,
            "sheet": SHEET_NAME,
            "all_sheets": all_sheets,
            "header_row": HEADER_ROW,
            "data_first_row": DATA_FIRST_ROW,
            "row_count": total_count,
            "hidden_row_count": len(hidden_rows),
            "hidden_in_source_parsed": hidden_in_source,
            "ws_max_row": ws.max_row,
            "ws_max_col": ws.max_column,
            "currency": "GBP (£m)",
            "variant": "v2 original (NOT the (1) variant — 103 rows, no Uganda orphan, no 3+9 sheet)",
        },
        "kpis": {
            "total_opportunities": total_count,
            "crp_total": round(crp_total, 4),
            "profit_totals": {k: round(v, 4) for k, v in profit_totals.items()},
            "profit_grand_total": profit_grand,
            "unique_customers": len(unique_customers),
            "unique_evs": len(unique_evs),
            "unique_vp_owners": len(unique_vp),
        },
        "regions": regions,
        "restructure_types": rtypes,
        "region_counts": dict(region_counts),
        "region_rtype_crp": region_rtype_crp,
        "status_funnel": status_funnel,
        "status_other": other_statuses,
        "evs_profit": evs_profit_list,
        "vp_leaderboard": vp_leaderboard,
        "year_series": year_series,
        "onerous_series": onerous_series,
        "profit_line": profit_line,
        "maturity_funnel": maturity_funnel,
        "note_rows": note_rows,
        "rows": rows,
    }


# ---------------------- HTML rendering ----------------------

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8" />
<meta name="viewport" content="width=device-width,initial-scale=1" />
<title>Global Hopper (v2 original) — GLOBAL LOG visualization</title>
<script src="https://cdn.jsdelivr.net/npm/apexcharts@3.49.0/dist/apexcharts.min.js"></script>
<style>
  :root {
    --bg: #0f1115;
    --panel: #161922;
    --panel-2: #1d2130;
    --border: #262b3b;
    --text: #e6e8ee;
    --muted: #9aa2b1;
    --accent: #4aa3ff;
    --accent-2: #ff6b9a;
    --ok: #4caf8a;
    --warn: #e6b85c;
    --bad: #e57373;
  }
  * { box-sizing: border-box; }
  body {
    margin: 0;
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
    background: var(--bg);
    color: var(--text);
    line-height: 1.45;
  }
  header {
    padding: 22px 28px;
    border-bottom: 1px solid var(--border);
    background: linear-gradient(180deg, #1a1f2c 0%, #0f1115 100%);
  }
  h1 { margin: 0 0 4px 0; font-size: 20px; font-weight: 600; }
  .meta { color: var(--muted); font-size: 12.5px; }
  .meta code { background: #0b0d13; padding: 1px 6px; border-radius: 4px; color: #c9d2e3; }
  main { padding: 20px 28px 60px 28px; display: grid; gap: 20px; }
  .grid {
    display: grid;
    gap: 16px;
  }
  .kpis { grid-template-columns: repeat(auto-fit, minmax(160px, 1fr)); }
  .kpi {
    background: var(--panel);
    border: 1px solid var(--border);
    border-radius: 10px;
    padding: 14px 16px;
  }
  .kpi .label { color: var(--muted); font-size: 11.5px; text-transform: uppercase; letter-spacing: 0.04em; }
  .kpi .value { font-size: 22px; font-weight: 600; margin-top: 4px; }
  .kpi .sub { color: var(--muted); font-size: 11.5px; margin-top: 2px; }

  .charts { grid-template-columns: repeat(auto-fit, minmax(420px, 1fr)); }
  .card {
    background: var(--panel);
    border: 1px solid var(--border);
    border-radius: 10px;
    padding: 14px 16px;
  }
  .card h3 {
    margin: 0 0 10px 0;
    font-size: 14px;
    font-weight: 600;
    color: #dfe4f0;
  }
  .card .sub { color: var(--muted); font-size: 11.5px; margin-bottom: 8px; }

  .filters { display: flex; gap: 10px; flex-wrap: wrap; align-items: flex-end; }
  .filters label { display: flex; flex-direction: column; font-size: 11.5px; color: var(--muted); }
  .filters select, .filters input {
    background: var(--panel-2);
    color: var(--text);
    border: 1px solid var(--border);
    border-radius: 6px;
    padding: 6px 8px;
    font-size: 12.5px;
    min-width: 140px;
  }
  .filters button {
    background: var(--accent);
    color: #0b0d13;
    border: none;
    border-radius: 6px;
    padding: 8px 14px;
    font-weight: 600;
    cursor: pointer;
  }
  .filters button.ghost {
    background: transparent;
    color: var(--text);
    border: 1px solid var(--border);
  }

  table.data {
    width: 100%;
    border-collapse: collapse;
    font-size: 12px;
  }
  table.data th, table.data td {
    padding: 6px 8px;
    border-bottom: 1px solid var(--border);
    text-align: left;
    vertical-align: top;
  }
  table.data th {
    position: sticky;
    top: 0;
    background: var(--panel-2);
    cursor: pointer;
    user-select: none;
    white-space: nowrap;
  }
  table.data th.sort-asc::after { content: " \\2191"; color: var(--accent); }
  table.data th.sort-desc::after { content: " \\2193"; color: var(--accent); }
  table.data td.num { text-align: right; font-variant-numeric: tabular-nums; }
  table.data tr:hover td { background: #1a1f2c; }

  .badge {
    display: inline-block;
    padding: 1px 6px;
    border-radius: 10px;
    font-size: 10.5px;
    font-weight: 600;
    letter-spacing: 0.02em;
  }
  .badge.hidden { background: #3a2b4d; color: #d3b5ff; border: 1px solid #553c75; }
  .badge.note   { background: #3a2f1c; color: #e6b85c; border: 1px solid #5b4a2a; }

  .scroll-x { overflow-x: auto; max-height: 560px; overflow-y: auto; border: 1px solid var(--border); border-radius: 8px; }

  .notes-panel ul { padding-left: 18px; margin: 6px 0; }
  .notes-panel li { margin-bottom: 4px; font-size: 12.5px; }

  footer { color: var(--muted); font-size: 11.5px; padding: 16px 28px 30px 28px; border-top: 1px solid var(--border); }
</style>
</head>
<body>
<header>
  <h1>Global Hopper (v2 original)</h1>
  <div class="meta">
    Sheet parsed: <code>__SHEET__</code> &nbsp;·&nbsp;
    <b>__ROW_COUNT__</b> opportunities &nbsp;·&nbsp;
    Currency: <code>__CURRENCY__</code> &nbsp;·&nbsp;
    <span title="Rows marked hidden in the source workbook. Still parsed.">Hidden rows in source: <b>__HIDDEN_COUNT__</b></span>
  </div>
  <div class="meta" style="margin-top:4px;">
    Source: <code>__SRC_PATH__</code> &nbsp;·&nbsp; Workbook sheets: <code>__ALL_SHEETS__</code>
  </div>
  <div class="meta" style="margin-top:4px;">__VARIANT_NOTE__</div>
</header>

<main>

  <section>
    <div class="grid kpis" id="kpis"></div>
  </section>

  <section class="card">
    <h3>Filter bar</h3>
    <div class="filters">
      <label>Region<select id="f-region"><option value="">(all)</option></select></label>
      <label>Engine Value Stream<select id="f-evs"><option value="">(all)</option></select></label>
      <label>Restructure Type<select id="f-rtype"><option value="">(all)</option></select></label>
      <label>Status<select id="f-status"><option value="">(all)</option></select></label>
      <label>VP/Owner<select id="f-vp"><option value="">(all)</option></select></label>
      <label>Text search<input id="f-search" type="search" placeholder="customer, initiative..." /></label>
      <button id="f-apply">Apply</button>
      <button id="f-reset" class="ghost">Reset</button>
      <span class="sub" id="f-count" style="margin-left:auto;color:var(--muted);"></span>
    </div>
  </section>

  <section class="grid charts">
    <div class="card"><h3>CRP Term Benefit by Region &times; Restructure Type</h3><div class="sub">Stacked bar, &pound;m</div><div id="chart-region-rtype"></div></div>
    <div class="card"><h3>Status funnel</h3><div class="sub">Pipeline ordering preserved (misspellings included)</div><div id="chart-status"></div></div>
    <div class="card"><h3>Profit 2026-2030 by Engine Value Stream</h3><div class="sub">Treemap, &pound;m (numeric cells only)</div><div id="chart-evs-treemap"></div></div>
    <div class="card"><h3>VP/Owner leaderboard (top 10)</h3><div class="sub">Sum of CRP Term Benefit &pound;m</div><div id="chart-vp"></div></div>
    <div class="card"><h3>Expected year of signature</h3><div class="sub">Counts incl. blank bucket</div><div id="chart-year"></div></div>
    <div class="card"><h3>Onerous / Non-Onerous mix</h3><div class="sub">Donut</div><div id="chart-onerous"></div></div>
    <div class="card"><h3>Profit trajectory 2026-2030</h3><div class="sub">Line, total &pound;m per year</div><div id="chart-profit-line"></div></div>
    <div class="card"><h3>Opportunity Maturity</h3><div class="sub">Funnel / horizontal bar</div><div id="chart-maturity"></div></div>
  </section>

  <section class="card notes-panel">
    <h3>Placeholder / TBD / tbc rows</h3>
    <div class="sub">Rows where one or more money columns held a non-numeric placeholder in the source sheet.</div>
    <div id="notes-list"></div>
  </section>

  <section class="card">
    <h3>All opportunities (19 columns, sortable, filterable)</h3>
    <div class="sub">Rows with the <span class="badge hidden">hidden</span> badge are marked hidden in the source workbook by openpyxl; they are still parsed.</div>
    <div class="scroll-x">
      <table class="data" id="data-table"></table>
    </div>
  </section>

</main>

<footer>
  Built from <code>__SRC_PATH__</code>, sheet <code>__SHEET__</code> only, via direct openpyxl read (independent of V6/parser.py).
  Charts: ApexCharts 3.49.0 via jsDelivr CDN. Self-contained HTML — data embedded as JSON literal below.
</footer>

<script id="payload" type="application/json">__PAYLOAD__</script>
<script>
(function(){
  const DATA = JSON.parse(document.getElementById('payload').textContent);
  const $ = (s) => document.querySelector(s);
  const fmtM = (v) => (v==null||isNaN(v)) ? '' : (Math.round(v*100)/100).toLocaleString('en-GB', {maximumFractionDigits:2});
  const fmtInt = (v) => (v==null||isNaN(v)) ? '' : Math.round(v).toLocaleString('en-GB');

  const MONEY_KEYS = ['crp_benefit','profit_2026','profit_2027','profit_2028','profit_2029','profit_2030'];
  const COLS = [
    {key:'row_index', label:'Row', type:'num'},
    {key:'region', label:'Region'},
    {key:'customer', label:'Customer'},
    {key:'evs', label:'Engine Value Stream'},
    {key:'top_level_evs', label:'Top Level EVS'},
    {key:'vp_owner', label:'VP/Owner (canonical)'},
    {key:'restructure_type', label:'Restructure Type'},
    {key:'opportunity_maturity', label:'Maturity'},
    {key:'onerous', label:'Onerous'},
    {key:'initiative', label:'Initative'},
    {key:'project_plan_req', label:'Project Plan Req'},
    {key:'status', label:'Status'},
    {key:'year_of_signature', label:'Year of Sig', type:'num'},
    {key:'signature_ap', label:'Sig AP'},
    {key:'crp_benefit', label:'CRP Term Benefit £m', type:'money'},
    {key:'profit_2026', label:'Profit 2026 £m', type:'money'},
    {key:'profit_2027', label:'Profit 2027 £m', type:'money'},
    {key:'profit_2028', label:'Profit 2028 £m', type:'money'},
    {key:'profit_2029', label:'Profit 2029 £m', type:'money'},
    {key:'profit_2030', label:'Profit 2030 £m', type:'money'},
  ];

  // ----- KPIs -----
  function renderKPIs(){
    const k = DATA.kpis;
    const items = [
      {label:'Total opportunities', value: fmtInt(k.total_opportunities)},
      {label:'CRP Term Benefit', value: '£' + fmtM(k.crp_total) + 'm'},
      {label:'Profit 2026-2030', value: '£' + fmtM(k.profit_grand_total) + 'm'},
      {label:'Unique customers', value: fmtInt(k.unique_customers)},
      {label:'Unique EVS', value: fmtInt(k.unique_evs)},
      {label:'Unique VP owners', value: fmtInt(k.unique_vp_owners)},
    ];
    $('#kpis').innerHTML = items.map(i =>
      `<div class="kpi"><div class="label">${i.label}</div><div class="value">${i.value}</div></div>`
    ).join('');
  }

  // ----- Filters -----
  function uniqueVals(key){
    const s = new Set();
    DATA.rows.forEach(r => { if (r[key]) s.add(r[key]); });
    return [...s].sort();
  }
  function fillSelect(id, values){
    const sel = $(id);
    values.forEach(v => {
      const opt = document.createElement('option');
      opt.value = v; opt.textContent = v;
      sel.appendChild(opt);
    });
  }
  function currentFilter(r){
    const region = $('#f-region').value;
    const evs = $('#f-evs').value;
    const rtype = $('#f-rtype').value;
    const status = $('#f-status').value;
    const vp = $('#f-vp').value;
    const q = ($('#f-search').value || '').trim().toLowerCase();
    if (region && r.region !== region) return false;
    if (evs && r.evs !== evs) return false;
    if (rtype && r.restructure_type !== rtype) return false;
    if (status && r.status !== status) return false;
    if (vp && r.vp_owner !== vp) return false;
    if (q) {
      const hay = [r.customer, r.initiative, r.status, r.evs, r.vp_owner].join(' ').toLowerCase();
      if (hay.indexOf(q) === -1) return false;
    }
    return true;
  }
  function applyFilters(){
    const filtered = DATA.rows.filter(currentFilter);
    $('#f-count').textContent = filtered.length + ' of ' + DATA.rows.length + ' rows';
    renderTable(filtered);
  }

  // ----- Table -----
  let sortKey = 'row_index', sortDir = 1;
  function renderTable(rows){
    const table = $('#data-table');
    const theadHtml = '<thead><tr>' + COLS.map(c => {
      const cls = (c.key === sortKey) ? (sortDir > 0 ? ' class="sort-asc"' : ' class="sort-desc"') : '';
      return `<th data-key="${c.key}"${cls}>${c.label}</th>`;
    }).join('') + '<th>Flags</th></tr></thead>';
    const data = [...rows].sort((a,b) => {
      const va = a[sortKey], vb = b[sortKey];
      if (va == null && vb == null) return 0;
      if (va == null) return 1;
      if (vb == null) return -1;
      if (typeof va === 'number' && typeof vb === 'number') return (va - vb) * sortDir;
      return String(va).localeCompare(String(vb)) * sortDir;
    });
    const tbodyHtml = '<tbody>' + data.map(r => {
      const cells = COLS.map(c => {
        const v = r[c.key];
        if (c.type === 'money') {
          const cls = 'num';
          if (v == null) {
            const note = (r.notes && r.notes[c.key]) ? r.notes[c.key] : '';
            return `<td class="${cls}" title="${note ? note.replace(/"/g,'&quot;') : ''}">${note ? '<span class="badge note">'+escapeHtml(note)+'</span>' : ''}</td>`;
          }
          return `<td class="${cls}">${fmtM(v)}</td>`;
        }
        if (c.type === 'num') {
          return `<td class="num">${v==null?'':fmtInt(v)}</td>`;
        }
        return `<td>${escapeHtml(v==null?'':String(v))}</td>`;
      }).join('');
      const flags = [];
      if (r.row_hidden) flags.push('<span class="badge hidden">hidden in source</span>');
      if (r.notes && Object.keys(r.notes).length) flags.push('<span class="badge note">'+Object.keys(r.notes).length+' note(s)</span>');
      return `<tr>${cells}<td>${flags.join(' ')}</td></tr>`;
    }).join('') + '</tbody>';
    table.innerHTML = theadHtml + tbodyHtml;
    table.querySelectorAll('th[data-key]').forEach(th => {
      th.addEventListener('click', () => {
        const k = th.getAttribute('data-key');
        if (sortKey === k) sortDir = -sortDir; else { sortKey = k; sortDir = 1; }
        applyFilters();
      });
    });
  }
  function escapeHtml(s){
    return s.replace(/[&<>"']/g, ch => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[ch]));
  }

  // ----- Charts -----
  const commonOpts = {
    chart: {
      foreColor: '#cfd4e0',
      toolbar: { show: false },
      fontFamily: 'inherit',
    },
    grid: { borderColor: '#262b3b' },
    theme: { mode: 'dark' },
    tooltip: { theme: 'dark' },
    dataLabels: { enabled: false },
  };
  function mergeOpts(...objs){
    return objs.reduce((acc, o) => {
      for (const k in o) {
        if (o[k] && typeof o[k] === 'object' && !Array.isArray(o[k])) {
          acc[k] = Object.assign({}, acc[k] || {}, o[k]);
        } else {
          acc[k] = o[k];
        }
      }
      return acc;
    }, {});
  }

  function renderStackedRegionRtype(){
    const regions = DATA.regions;
    const rtypes = DATA.restructure_types;
    const matrix = DATA.region_rtype_crp;
    const series = rtypes.map(rt => ({
      name: rt,
      data: regions.map(rg => Math.round((matrix[rt] && matrix[rt][rg] || 0) * 100) / 100),
    }));
    const opts = mergeOpts(commonOpts, {
      chart: { type: 'bar', stacked: true, height: 340 },
      series: series,
      xaxis: { categories: regions },
      yaxis: { title: { text: '£m' } },
      plotOptions: { bar: { borderRadius: 2 } },
      legend: { position: 'bottom' },
    });
    new ApexCharts($('#chart-region-rtype'), opts).render();
  }

  function renderStatusFunnel(){
    const stages = DATA.status_funnel.concat(DATA.status_other);
    const opts = mergeOpts(commonOpts, {
      chart: { type: 'bar', height: 340 },
      series: [{ name: 'Opportunities', data: stages.map(s => s.count) }],
      xaxis: { categories: stages.map(s => s.status) },
      plotOptions: { bar: { horizontal: true, borderRadius: 2 } },
      colors: ['#4aa3ff'],
    });
    new ApexCharts($('#chart-status'), opts).render();
  }

  function renderEVSTreemap(){
    const data = DATA.evs_profit.map(d => ({ x: d.evs, y: d.profit }));
    const opts = mergeOpts(commonOpts, {
      chart: { type: 'treemap', height: 340 },
      series: [{ data }],
      plotOptions: { treemap: { distributed: true, enableShades: true } },
      dataLabels: { enabled: true, style: { fontSize: '11px' } },
    });
    new ApexCharts($('#chart-evs-treemap'), opts).render();
  }

  function renderVPLeaderboard(){
    const rows = DATA.vp_leaderboard;
    const opts = mergeOpts(commonOpts, {
      chart: { type: 'bar', height: 340 },
      series: [{ name: 'CRP Term Benefit £m', data: rows.map(r => Math.round(r.crp*100)/100) }],
      xaxis: { categories: rows.map(r => r.owner + ' (' + r.count + ')') },
      plotOptions: { bar: { horizontal: true, borderRadius: 2 } },
      colors: ['#ff6b9a'],
    });
    new ApexCharts($('#chart-vp'), opts).render();
  }

  function renderYearChart(){
    const rows = DATA.year_series;
    const opts = mergeOpts(commonOpts, {
      chart: { type: 'bar', height: 280 },
      series: [{ name: 'Opportunities', data: rows.map(r => r.count) }],
      xaxis: { categories: rows.map(r => r.year) },
      plotOptions: { bar: { borderRadius: 3, columnWidth: '50%' } },
      colors: ['#4caf8a'],
    });
    new ApexCharts($('#chart-year'), opts).render();
  }

  function renderOnerous(){
    const rows = DATA.onerous_series;
    const opts = mergeOpts(commonOpts, {
      chart: { type: 'donut', height: 280 },
      series: rows.map(r => r.count),
      labels: rows.map(r => r.label),
      legend: { position: 'bottom' },
    });
    new ApexCharts($('#chart-onerous'), opts).render();
  }

  function renderProfitLine(){
    const rows = DATA.profit_line;
    const opts = mergeOpts(commonOpts, {
      chart: { type: 'line', height: 280 },
      series: [{ name: 'Profit £m', data: rows.map(r => r.profit) }],
      xaxis: { categories: rows.map(r => r.year) },
      stroke: { curve: 'smooth', width: 3 },
      markers: { size: 5 },
      colors: ['#e6b85c'],
      yaxis: { title: { text: '£m' } },
    });
    new ApexCharts($('#chart-profit-line'), opts).render();
  }

  function renderMaturityFunnel(){
    const rows = DATA.maturity_funnel;
    const opts = mergeOpts(commonOpts, {
      chart: { type: 'bar', height: 280 },
      series: [{ name: 'Opportunities', data: rows.map(r => r.count) }],
      xaxis: { categories: rows.map(r => r.maturity) },
      plotOptions: { bar: { horizontal: true, borderRadius: 2 } },
      colors: ['#4aa3ff'],
    });
    new ApexCharts($('#chart-maturity'), opts).render();
  }

  // ----- Notes panel -----
  function renderNotes(){
    const container = $('#notes-list');
    if (!DATA.note_rows.length) {
      container.innerHTML = '<p class="sub">No placeholder / TBD / tbc rows detected.</p>';
      return;
    }
    const html = DATA.note_rows.map(n => {
      const parts = Object.entries(n.notes).map(([k, v]) =>
        `<span class="badge note" title="${escapeHtml(v)}">${k}: ${escapeHtml(v)}</span>`
      ).join(' ');
      return `<li><code>row ${n.row_index}</code> &nbsp; <b>${escapeHtml(n.customer || '(no customer)')}</b> &mdash; ${escapeHtml(n.region)} / ${escapeHtml(n.evs)} &nbsp; ${parts}</li>`;
    }).join('');
    container.innerHTML = '<ul>' + html + '</ul>';
  }

  // ----- Boot -----
  function boot(){
    renderKPIs();
    fillSelect('#f-region', uniqueVals('region'));
    fillSelect('#f-evs', uniqueVals('evs'));
    fillSelect('#f-rtype', uniqueVals('restructure_type'));
    fillSelect('#f-status', uniqueVals('status'));
    fillSelect('#f-vp', uniqueVals('vp_owner'));
    $('#f-apply').addEventListener('click', applyFilters);
    $('#f-reset').addEventListener('click', () => {
      ['#f-region','#f-evs','#f-rtype','#f-status','#f-vp'].forEach(s => { $(s).value = ''; });
      $('#f-search').value = '';
      applyFilters();
    });
    $('#f-search').addEventListener('keydown', (e) => { if (e.key === 'Enter') applyFilters(); });
    applyFilters();

    renderStackedRegionRtype();
    renderStatusFunnel();
    renderEVSTreemap();
    renderVPLeaderboard();
    renderYearChart();
    renderOnerous();
    renderProfitLine();
    renderMaturityFunnel();
    renderNotes();
  }
  boot();
})();
</script>
</body>
</html>
"""


def render_html(payload: dict[str, Any]) -> str:
    meta = payload["meta"]
    js_payload = json.dumps(payload, ensure_ascii=False, allow_nan=False)
    # Safe escape for embedding inside <script type="application/json">
    js_payload = js_payload.replace("</", "<\\/")
    out = HTML_TEMPLATE
    out = out.replace("__SHEET__", html.escape(meta["sheet"]))
    out = out.replace("__ROW_COUNT__", str(meta["row_count"]))
    out = out.replace("__CURRENCY__", html.escape(meta["currency"]))
    out = out.replace("__HIDDEN_COUNT__", str(meta["hidden_row_count"]))
    out = out.replace("__SRC_PATH__", html.escape(meta["source_path"]))
    out = out.replace("__ALL_SHEETS__", html.escape(", ".join(meta["all_sheets"])))
    out = out.replace("__VARIANT_NOTE__", html.escape(meta["variant"]))
    out = out.replace("__PAYLOAD__", js_payload)
    return out


def main() -> None:
    payload = extract()
    html_out = render_html(payload)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(html_out, encoding="utf-8")

    # -- Report --
    m = payload["meta"]
    k = payload["kpis"]
    rc = Counter(r["region"] for r in payload["rows"] if r["region"])
    top_evs = payload["evs_profit"][:5]
    print("=" * 60)
    print("GLOBAL HOPPER v2 ORIGINAL — build report")
    print("=" * 60)
    print(f"Source       : {m['source_path']}")
    print(f"Output       : {OUT_PATH}")
    print(f"Sheet parsed : {m['sheet']} (all sheets: {m['all_sheets']})")
    print(f"Rows parsed  : {m['row_count']}")
    print(f"Hidden rows  : {m['hidden_row_count']}  (captured in-row: {m['hidden_in_source_parsed']})")
    print(f"CRP total    : £{k['crp_total']:.2f}m")
    print(f"Profit 26-30 : £{k['profit_grand_total']:.2f}m")
    print("Region counts:")
    for rg, n in rc.most_common():
        print(f"  {rg:<16} {n}")
    print("Top EVS by profit 26-30:")
    for e in top_evs:
        print(f"  {e['evs']:<16} £{e['profit']:.2f}m")


if __name__ == "__main__":
    main()
