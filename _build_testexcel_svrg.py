"""
Builds C:/Users/Rutishkrishna/Desktop/RR/RR Powerbi/V6/TESTEXCEL/SVRG_MASTER.html
Standalone openpyxl extractor for the Trent 900 SVRG Master workbook.

Rules:
 - Skip chartsheets.
 - Clip phantom widths.
 - Clamp dates to 1970-2075.
 - Convert Excel error strings (#N/A, #VALUE!, #REF!, etc.) to None.
 - Emit JSON embedded in a self-contained HTML page.
"""
from __future__ import annotations
import json
import datetime as dt
from pathlib import Path
import openpyxl
from openpyxl.chartsheet import Chartsheet

SRC = Path(r"C:\Users\Rutishkrishna\Desktop\RR\RR Powerbi\V6\New info\VERSION 2 Enhanced SVRG MASTER FILE (version 1).xlsb LOCAL.xlsx")
OUT_DIR = Path(r"C:\Users\Rutishkrishna\Desktop\RR\RR Powerbi\V6\TESTEXCEL")
OUT_HTML = OUT_DIR / "SVRG_MASTER.html"

ERROR_STRINGS = {"#N/A", "#VALUE!", "#REF!", "#NAME?", "#DIV/0!", "#NULL!", "#NUM!", "#GETTING_DATA"}
DATE_MIN = dt.date(1970, 1, 1)
DATE_MAX = dt.date(2075, 12, 31)


def clean(val):
    """Normalize one cell value: strip error strings, clamp dates, return JSON-safe."""
    if val is None:
        return None
    if isinstance(val, str):
        s = val.strip()
        if not s or s in ERROR_STRINGS:
            return None
        return s
    if isinstance(val, dt.datetime):
        d = val.date()
        if d < DATE_MIN or d > DATE_MAX:
            return None
        # emit ISO8601
        return val.isoformat()
    if isinstance(val, dt.date):
        if val < DATE_MIN or val > DATE_MAX:
            return None
        return val.isoformat()
    if isinstance(val, dt.time):
        return None  # time-only cells are noise in this workbook
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        # NaN / inf check
        if isinstance(val, float):
            if val != val or val in (float("inf"), float("-inf")):
                return None
        return val
    # fallback
    return str(val)


def clean_date_only(val):
    """Return ISO date string (no time) or None. Used for aggregation by year."""
    if val is None:
        return None
    if isinstance(val, dt.datetime):
        d = val.date()
    elif isinstance(val, dt.date):
        d = val
    else:
        return None
    if d < DATE_MIN or d > DATE_MAX:
        return None
    return d.isoformat()


def year_of(iso):
    if not iso:
        return None
    try:
        return int(iso[:4])
    except Exception:
        return None


def last_nonempty_col(ws, header_row, hard_cap=512):
    """Find last non-empty col in `header_row` to clip phantom widths."""
    max_c = min(ws.max_column, hard_cap)
    row = next(ws.iter_rows(min_row=header_row, max_row=header_row,
                            min_col=1, max_col=max_c, values_only=True))
    last = 0
    for i, v in enumerate(row, 1):
        if v is not None and (not isinstance(v, str) or v.strip()):
            last = i
    return max(last, 1)


# -------- Extractors ----------------------------------------------------------

def extract_qualified_engines(ws):
    """Merges two blocks:
      - Primary table at cols A-R (header row 6): engine_family, delivery_date, qualifications, etc.
      - Asset# table at cols U-V (header row 6): canonical 161-engine list with Asset# + ESN.
    Returns the canonical 161-engine list joined with metadata by ESN.
    """
    header_row = 6
    # Build metadata lookup from primary table
    meta = {}
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row,
                            min_col=1, max_col=18, values_only=True):
        vals = [clean(c) for c in row]
        esn = vals[3] if len(vals) > 3 else None
        if esn is None or esn == "" or esn == 0:
            continue
        meta[str(esn)] = {
            "engine_family": str(vals[2]) if len(vals) > 2 and vals[2] is not None else None,
            "delivery_date": vals[4] if len(vals) > 4 and isinstance(vals[4], str) else None,
            "improved_hptb_from_new": vals[5] if len(vals) > 5 else None,
            "final_fix_from_new": vals[6] if len(vals) > 6 else None,
            "first_improved_hptb_date": vals[7] if len(vals) > 7 else None,
            "improved_auto_derate_date": vals[9] if len(vals) > 9 else None,
            "qualified_flag": vals[12] if len(vals) > 12 else None,
            "qualified_date": vals[15] if len(vals) > 15 else None,
            "qualified_year": vals[16] if len(vals) > 16 else None,
        }

    # Walk canonical Asset# table at cols T-U (20, 21 1-indexed): T=Asset#, U=ESN
    items = []
    seen = set()
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row,
                            min_col=20, max_col=21, values_only=True):
        asset = clean(row[0])
        esn = clean(row[1])
        if asset is None and esn is None:
            continue
        # ESN of 0 or None is a filler row
        if esn in (None, 0, "0"):
            continue
        esn_str = str(esn)
        if esn_str in seen:
            continue
        seen.add(esn_str)
        m = meta.get(esn_str, {})
        qualified_flag = m.get("qualified_flag")
        qdate = m.get("qualified_date")
        qyear = m.get("qualified_year")
        # Prefer the date/year presence — Excel uses these as the truth-source when the
        # "Is Engine a Qualified Engine..." flag is #N/A.
        if isinstance(qualified_flag, str) and qualified_flag.lower().startswith("yes"):
            status = "Qualified"
        elif isinstance(qualified_flag, str) and qualified_flag.lower().startswith("no"):
            status = "Not qualified"
        elif qdate or qyear:
            status = "Qualified"
        else:
            status = "Pending / Unknown"
        items.append({
            "asset_id": str(asset) if asset is not None else None,
            "engine_serial": esn_str,
            "engine_family": m.get("engine_family"),
            "rating": m.get("engine_family"),
            "delivery_date": m.get("delivery_date"),
            "improved_hptb_from_new": m.get("improved_hptb_from_new"),
            "final_fix_from_new": m.get("final_fix_from_new"),
            "first_improved_hptb_date": m.get("first_improved_hptb_date"),
            "improved_auto_derate_date": m.get("improved_auto_derate_date"),
            "qualified_flag": qualified_flag,
            "qualified_date": m.get("qualified_date"),
            "qualified_year": m.get("qualified_year"),
            "operator": "EMIRATES",
            "status": status,
        })
    return [], items


def extract_qualified_svs(ws):
    """Header row 6, row 7 is 'Actual shop visits' label (skip), data row 8+."""
    header_row = 6
    width = min(last_nonempty_col(ws, header_row), 16)
    headers = [clean(c) for c in next(ws.iter_rows(
        min_row=header_row, max_row=header_row, min_col=1, max_col=width, values_only=True))]
    items = []
    for row in ws.iter_rows(min_row=header_row + 2, max_row=ws.max_row,
                            min_col=1, max_col=width, values_only=True):
        vals = [clean(c) for c in row]
        # Asset#, Engine, Date of engine removal, CSN, Year of SV, Cause, Qualified?, OSP, HPTB?, Month Qualified, Qualified HPTB, Comments, RSV#
        asset_id = vals[0] if len(vals) > 0 else None
        engine = vals[1] if len(vals) > 1 else None
        if engine is None and asset_id is None:
            continue
        # Skip marker rows
        if isinstance(asset_id, str) and asset_id.strip().lower().startswith("actual shop visit"):
            continue
        sv_date = vals[2] if len(vals) > 2 else None
        items.append({
            "asset_id": asset_id,
            "engine_serial": str(engine) if engine is not None else None,
            "sv_date": sv_date if isinstance(sv_date, str) else None,
            "csn": vals[3] if len(vals) > 3 else None,
            "sv_year": vals[4] if len(vals) > 4 else None,
            "cause": vals[5] if len(vals) > 5 else None,
            "qualified_cause": vals[6] if len(vals) > 6 else None,
            "osp_250k_received": vals[7] if len(vals) > 7 else None,
            "hptb_driven": vals[8] if len(vals) > 8 else None,
            "qualified_month": vals[9] if len(vals) > 9 else None,
            "qualified_hptb_sv": vals[10] if len(vals) > 10 else None,
            "comments": vals[11] if len(vals) > 11 else None,
            "rsv_num": vals[12] if len(vals) > 12 else None,
            "operator": "EMIRATES",
            "sv_type": vals[5] if len(vals) > 5 else None,
            "sv_location": None,
        })
    return headers, items


def extract_qualified_efh(ws):
    """Year headers in row 4 (col D onwards), engine serials col C row 6 onwards.
    Emit per-engine annual totals.
    """
    width = min(last_nonempty_col(ws, 5), 200)
    year_row = [c for c in next(ws.iter_rows(min_row=4, max_row=4,
                                             min_col=4, max_col=width,
                                             values_only=True))]
    # year_row is list of years aligned to columns (col D -> index 0, etc.)
    # Walk engine rows (row 6+)
    per_engine_year = {}  # (esn, year) -> total efh
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row,
                            min_col=3, max_col=width, values_only=True):
        esn = row[0]
        if esn is None:
            continue
        esn_cleaned = clean(esn)
        if esn_cleaned is None:
            continue
        esn_str = str(esn_cleaned)
        # columns D.. map to year_row
        for i, year in enumerate(year_row):
            if year is None:
                continue
            try:
                y = int(year)
            except (ValueError, TypeError):
                continue
            if y < 1970 or y > 2075:
                continue
            v = row[i + 1] if (i + 1) < len(row) else None
            v = clean(v)
            if not isinstance(v, (int, float)):
                continue
            key = (esn_str, y)
            per_engine_year[key] = per_engine_year.get(key, 0.0) + float(v)

    items = [{"engine_serial": esn, "year": yr, "efh": round(v, 2)}
             for (esn, yr), v in sorted(per_engine_year.items())]
    return [], items


def extract_hours_cycles_input(ws):
    """Header row 1, data row 2+. Emit row_count, sample 500 most recent."""
    width = min(last_nonempty_col(ws, 1), 16)
    headers = [clean(c) for c in next(ws.iter_rows(
        min_row=1, max_row=1, min_col=1, max_col=width, values_only=True))]
    all_rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                            min_col=1, max_col=width, values_only=True):
        vals = [clean(c) for c in row]
        if all(v is None or v == "" for v in vals):
            continue
        all_rows.append({
            "operator": vals[0],
            "engine_family": vals[1],
            "engine_serial": str(vals[2]) if vals[2] is not None else None,
            "year": vals[3],
            "month": vals[4],
            "hours": vals[5],
            "cycles": vals[6] if len(vals) > 6 else None,
        })
    # Sample: keep 500 most recent (by year desc, month desc)
    row_count = len(all_rows)

    def _to_int(v):
        if isinstance(v, (int, float)):
            return int(v)
        if isinstance(v, str):
            try:
                return int(v)
            except ValueError:
                return 0
        return 0

    def sort_key(r):
        return (_to_int(r.get("year")), _to_int(r.get("month")))

    sample = sorted(all_rows, key=sort_key, reverse=True)[:500]
    return headers, {"row_count": row_count, "sample": sample, "all_rows_count": row_count}


def extract_claims_summary(ws):
    """Header row 7: DATE, YEAR, CREDIT NOTE REFERENCE, GUARANTEE, CREDIT NOTE VALUE, CUMULATIVE CLAIM VALUE."""
    header_row = 7
    width = min(last_nonempty_col(ws, header_row), 12)
    headers = [clean(c) for c in next(ws.iter_rows(
        min_row=header_row, max_row=header_row, min_col=1, max_col=width, values_only=True))]
    items = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row,
                            min_col=1, max_col=width, values_only=True):
        vals = [clean(c) for c in row]
        # col B = DATE, C = YEAR, D = CREDIT NOTE REF, E = GUARANTEE, F = CN VALUE, G = CUMULATIVE
        date_v = vals[1] if len(vals) > 1 else None
        credit_note_ref = vals[3] if len(vals) > 3 else None
        guarantee = vals[4] if len(vals) > 4 else None
        cn_val = vals[5] if len(vals) > 5 else None
        # Require at least one meaningful field
        if all(v in (None, "", 0) for v in (date_v, credit_note_ref, guarantee, cn_val)):
            continue
        items.append({
            "date": date_v,
            "year": vals[2] if len(vals) > 2 else None,
            "credit_note_ref": credit_note_ref,
            "guarantee": guarantee,
            "credit_note_value": cn_val,
            "cumulative_value": vals[6] if len(vals) > 6 else None,
        })
    return headers, items


def extract_event_entry(ws):
    """Header row 5, data row 6+. Cols: A=ABTO flag, B=Date, C=Engine Serial, D=A/C, E=TSN, F=CSN, G=Cause, ..."""
    header_row = 5
    width = min(last_nonempty_col(ws, header_row), 22)
    headers = [clean(c) for c in next(ws.iter_rows(
        min_row=header_row, max_row=header_row, min_col=1, max_col=width, values_only=True))]
    items = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row,
                            min_col=1, max_col=width, values_only=True):
        vals = [clean(c) for c in row]
        date_v = vals[1] if len(vals) > 1 else None
        esn = vals[2] if len(vals) > 2 else None
        cause = vals[6] if len(vals) > 6 else None
        if date_v is None and esn is None and cause is None:
            continue
        items.append({
            "flag": vals[0] if len(vals) > 0 else None,
            "date": date_v,
            "engine_serial": str(esn) if esn is not None else None,
            "ac_position": vals[3] if len(vals) > 3 else None,
            "tsn": vals[4] if len(vals) > 4 else None,
            "csn": vals[5] if len(vals) > 5 else None,
            "cause": cause,
            "emirates_input": vals[9] if len(vals) > 9 else None,
            "emirates_justification": vals[10] if len(vals) > 10 else None,
            "rr_input": vals[11] if len(vals) > 11 else None,
            "rr_justification": vals[12] if len(vals) > 12 else None,
            "coverage": vals[13] if len(vals) > 13 else None,
            "comments": vals[14] if len(vals) > 14 else None,
        })
    return headers, items


def extract_svrg_esvrg(ws):
    """Metric rows with year columns (E..S = years 1..15 = 2017..2031).
    Row 5: year labels (2017..2031). Rows 6+: metric rows where col B is description.
    """
    # years live in row 5, cols E..S (idx 4..18)
    year_row = [c for c in next(ws.iter_rows(min_row=5, max_row=5,
                                             min_col=5, max_col=19, values_only=True))]
    years = [int(y) if isinstance(y, (int, float)) else None for y in year_row]
    metrics = []
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row,
                            min_col=2, max_col=19, values_only=True):
        desc = clean(row[0])  # col B
        if not desc or (isinstance(desc, str) and len(desc.strip()) < 3):
            continue
        # skip section headers like "Shop Visit Rate Guarantee (SVRG)" that have no numeric data
        vals = [clean(v) for v in row[3:18]]  # cols E..S
        numeric_vals = [v for v in vals if isinstance(v, (int, float))]
        if not numeric_vals:
            continue
        metrics.append({
            "description": str(desc)[:200],
            "code": clean(row[1]),  # col C e.g. 'A' 'B' 'C'
            "calculation": clean(row[2]),  # col D
            "values_by_year": [{"year": y, "value": v} for y, v in zip(years, vals) if y is not None],
        })
    return [], metrics


def extract_secondary_summary(ws, max_width=40, sample_rows=20):
    """Generic summary sheet — find a sensible header row, emit name + row_count + column_headers + sample."""
    # Find first row with 3+ non-empty string cells — call that the header
    header_row = 1
    for ri in range(1, min(ws.max_row, 20) + 1):
        row = next(ws.iter_rows(min_row=ri, max_row=ri, min_col=1,
                                max_col=min(ws.max_column, max_width), values_only=True))
        non_empty_strs = sum(1 for v in row if isinstance(v, str) and v.strip())
        if non_empty_strs >= 3:
            header_row = ri
            break
    width = min(last_nonempty_col(ws, header_row, hard_cap=max_width), max_width)
    headers = [clean(c) for c in next(ws.iter_rows(
        min_row=header_row, max_row=header_row, min_col=1, max_col=width, values_only=True))]
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1,
                            max_row=min(ws.max_row, header_row + sample_rows),
                            min_col=1, max_col=width, values_only=True):
        vals = [clean(c) for c in row]
        if all(v is None or v == "" for v in vals):
            continue
        rows.append(vals)
    # Data row count (rough — anything non-empty below header)
    data_row_count = 0
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row,
                            min_col=1, max_col=width, values_only=True):
        if any(v is not None and (not isinstance(v, str) or v.strip()) for v in row):
            data_row_count += 1
    return {
        "header_row": header_row,
        "column_headers": headers,
        "sample_rows": rows,
        "data_row_count": data_row_count,
    }


# -------- Main ----------------------------------------------------------------

def main():
    wb = openpyxl.load_workbook(SRC, data_only=True, read_only=False)
    sheets_meta = []
    parse_errors = []

    data = {
        "source_file": SRC.name,
        "sheet_count": len(wb.sheetnames),
        "visible_sheets": [],
        "hidden_sheets": [],
        "chartsheets": [],
        "qualified_engines": [],
        "qualified_svs": [],
        "qualified_efh": [],  # per-engine per-year
        "hours_cycles": {"row_count": 0, "sample": []},
        "claims": [],
        "events": [],
        "svrg_esvrg_metrics": [],
        "secondary_summaries": {},
    }

    for name in wb.sheetnames:
        ws = wb[name]
        if isinstance(ws, Chartsheet):
            data["chartsheets"].append(name)
            sheets_meta.append({"name": name, "kind": "chartsheet"})
            continue
        meta = {"name": name, "kind": "worksheet",
                "state": ws.sheet_state,
                "rows": ws.max_row, "cols": ws.max_column}
        sheets_meta.append(meta)
        if ws.sheet_state == "visible":
            data["visible_sheets"].append(name)
        else:
            data["hidden_sheets"].append(name)

    # ---- structured extracts
    try:
        _, data["qualified_engines"] = extract_qualified_engines(wb["QUALIFIED ENGINES"])
    except Exception as e:
        parse_errors.append(f"QUALIFIED ENGINES: {e!r}")

    try:
        _, data["qualified_svs"] = extract_qualified_svs(wb["QUALIFIED SVs"])
    except Exception as e:
        parse_errors.append(f"QUALIFIED SVs: {e!r}")

    try:
        _, data["qualified_efh"] = extract_qualified_efh(wb["QUALIFIED EFH"])
    except Exception as e:
        parse_errors.append(f"QUALIFIED EFH: {e!r}")

    try:
        _, hc = extract_hours_cycles_input(wb["HOURS&CYCLES INPUT"])
        data["hours_cycles"] = hc
    except Exception as e:
        parse_errors.append(f"HOURS&CYCLES INPUT: {e!r}")

    try:
        _, data["claims"] = extract_claims_summary(wb["CLAIMS SUMMARY"])
    except Exception as e:
        parse_errors.append(f"CLAIMS SUMMARY: {e!r}")

    try:
        _, data["events"] = extract_event_entry(wb["EVENT ENTRY"])
    except Exception as e:
        parse_errors.append(f"EVENT ENTRY: {e!r}")

    try:
        _, data["svrg_esvrg_metrics"] = extract_svrg_esvrg(wb["SVRG+ESVRG"])
    except Exception as e:
        parse_errors.append(f"SVRG+ESVRG: {e!r}")

    # Secondary summary sheets
    secondary_names = [
        "RATE BASED SUMMARY", "DI SUMMARY", "HPTB&VANE SUMMARY", "ELMB SUMMARY",
        "EMISSIONS SUMMARY", "WEIGHT SUMMARY", "OIL  ", "OIL SUMMARY",
        "FBR&UNCONT FAIL", "TGT DETERIORATION", "EFH AND REV DEPS",
        "CHARTS", "EHPTB MEASURE", "2024 Expected SV's", "ASSUMPTIONS",
    ]
    for sname in secondary_names:
        if sname not in wb.sheetnames:
            continue
        try:
            ws = wb[sname]
            if isinstance(ws, Chartsheet):
                continue
            data["secondary_summaries"][sname] = extract_secondary_summary(ws)
        except Exception as e:
            parse_errors.append(f"{sname}: {e!r}")

    # Derived aggregates for UI convenience
    # engine family distribution
    from collections import Counter
    family_counts = Counter()
    for e in data["qualified_engines"]:
        family_counts[e.get("engine_family") or "Unknown"] += 1
    data["engine_family_counts"] = dict(family_counts)

    status_counts = Counter()
    for e in data["qualified_engines"]:
        status_counts[e.get("status") or "Unknown"] += 1
    data["engine_status_counts"] = dict(status_counts)

    # SVs per year
    svs_per_year = Counter()
    for sv in data["qualified_svs"]:
        y = sv.get("sv_year")
        if isinstance(y, (int, float)):
            svs_per_year[int(y)] += 1
        elif isinstance(sv.get("sv_date"), str):
            yr = year_of(sv["sv_date"])
            if yr:
                svs_per_year[yr] += 1
    data["svs_per_year"] = dict(sorted(svs_per_year.items()))

    # EFH per engine (total)
    efh_per_engine = {}
    for it in data["qualified_efh"]:
        efh_per_engine[it["engine_serial"]] = efh_per_engine.get(it["engine_serial"], 0.0) + it["efh"]
    data["efh_per_engine"] = {k: round(v, 2) for k, v in efh_per_engine.items()}

    # EFH per year (fleet total)
    efh_per_year = {}
    for it in data["qualified_efh"]:
        efh_per_year[it["year"]] = efh_per_year.get(it["year"], 0.0) + it["efh"]
    data["efh_per_year"] = {str(k): round(v, 2) for k, v in sorted(efh_per_year.items())}

    # Claims per year
    claims_per_year = Counter()
    total_claim_value = 0.0
    for c in data["claims"]:
        y = c.get("year")
        if isinstance(y, (int, float)):
            claims_per_year[int(y)] += 1
        elif isinstance(c.get("date"), str):
            yr = year_of(c["date"])
            if yr:
                claims_per_year[yr] += 1
        v = c.get("credit_note_value")
        if isinstance(v, (int, float)):
            total_claim_value += v
    data["claims_per_year"] = dict(sorted(claims_per_year.items()))
    data["total_claim_value"] = round(total_claim_value, 2)

    # Events per year
    events_per_year = Counter()
    for ev in data["events"]:
        d = ev.get("date")
        if isinstance(d, str):
            yr = year_of(d)
            if yr:
                events_per_year[yr] += 1
    data["events_per_year"] = dict(sorted(events_per_year.items()))

    data["sheets_meta"] = sheets_meta
    data["parse_errors"] = parse_errors

    # ---- report
    print(f"\n=== EXTRACTION SUMMARY ===")
    print(f"Sheets: {len(sheets_meta)} | visible={len(data['visible_sheets'])} hidden={len(data['hidden_sheets'])} chart={len(data['chartsheets'])}")
    print(f"QUALIFIED ENGINES: {len(data['qualified_engines'])}")
    print(f"QUALIFIED SVs:     {len(data['qualified_svs'])}")
    print(f"QUALIFIED EFH items (engine x year): {len(data['qualified_efh'])}")
    print(f"HOURS&CYCLES INPUT row_count: {data['hours_cycles']['row_count']} (sampled {len(data['hours_cycles']['sample'])})")
    print(f"CLAIMS SUMMARY: {len(data['claims'])} (total value {data['total_claim_value']})")
    print(f"EVENT ENTRY:   {len(data['events'])}")
    print(f"SVRG+ESVRG metrics: {len(data['svrg_esvrg_metrics'])}")
    print(f"Engine family counts: {data['engine_family_counts']}")
    print(f"SVs per year: {data['svs_per_year']}")
    print(f"Claims per year: {data['claims_per_year']}")
    print(f"Events per year: {data['events_per_year']}")
    print(f"EFH per year: {data['efh_per_year']}")
    if parse_errors:
        print(f"PARSE ERRORS:")
        for pe in parse_errors:
            print(f"  - {pe}")

    # ---- write HTML
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    html = render_html(data)
    OUT_HTML.write_text(html, encoding="utf-8")
    print(f"\nWrote: {OUT_HTML}  ({OUT_HTML.stat().st_size:,} bytes)")


# -------- HTML renderer -------------------------------------------------------

def render_html(data):
    payload = json.dumps(data, ensure_ascii=False, default=str)
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>SVRG Master — Trent 900 Guarantee Administration (Emirates)</title>
<script src="https://cdn.jsdelivr.net/npm/apexcharts@3.49.0/dist/apexcharts.min.js"></script>
<style>
  :root {{
    --bg: #0b0d12;
    --panel: #12151c;
    --panel2: #181c25;
    --border: #242a36;
    --text: #e6e9ef;
    --muted: #9aa3b2;
    --accent: #7c5cff;
    --accent2: #22d3ee;
    --green: #34d399;
    --amber: #fbbf24;
    --red: #f87171;
  }}
  * {{ box-sizing: border-box; }}
  html, body {{ margin: 0; background: var(--bg); color: var(--text); font: 14px/1.5 -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Arial, sans-serif; }}
  header {{
    padding: 18px 24px; border-bottom: 1px solid var(--border);
    background: linear-gradient(180deg, #1a1f2b 0%, #0d1017 100%);
    display: flex; align-items: center; justify-content: space-between;
  }}
  header h1 {{ margin: 0; font-size: 18px; font-weight: 600; letter-spacing: .2px; }}
  header .meta {{ color: var(--muted); font-size: 12px; }}
  .filters {{
    padding: 12px 24px; border-bottom: 1px solid var(--border);
    display: flex; gap: 12px; flex-wrap: wrap; background: var(--panel);
    position: sticky; top: 0; z-index: 40;
  }}
  .filters label {{ font-size: 12px; color: var(--muted); display: block; margin-bottom: 4px; }}
  .filters select, .filters input {{
    background: var(--panel2); border: 1px solid var(--border); color: var(--text);
    padding: 6px 10px; border-radius: 6px; font: inherit; min-width: 140px;
  }}
  .tabs {{ display: flex; gap: 4px; border-bottom: 1px solid var(--border); padding: 0 24px; background: var(--panel); }}
  .tab {{
    padding: 10px 16px; cursor: pointer; color: var(--muted); border-bottom: 2px solid transparent;
    font-size: 13px; letter-spacing: .2px;
  }}
  .tab.active {{ color: var(--text); border-bottom-color: var(--accent); }}
  .tab:hover {{ color: var(--text); }}
  .pane {{ display: none; padding: 20px 24px; }}
  .pane.active {{ display: block; }}
  .kpis {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px,1fr)); gap: 12px; margin-bottom: 16px; }}
  .kpi {{
    background: var(--panel); border: 1px solid var(--border);
    border-radius: 10px; padding: 14px 16px;
  }}
  .kpi .label {{ color: var(--muted); font-size: 12px; text-transform: uppercase; letter-spacing: .4px; }}
  .kpi .value {{ font-size: 22px; font-weight: 600; margin-top: 4px; }}
  .kpi .sub {{ color: var(--muted); font-size: 11px; margin-top: 2px; }}
  .grid {{ display: grid; gap: 16px; }}
  .grid.two {{ grid-template-columns: 1fr 1fr; }}
  .grid.three {{ grid-template-columns: 1fr 1fr 1fr; }}
  @media (max-width: 1100px) {{ .grid.two, .grid.three {{ grid-template-columns: 1fr; }} }}
  .card {{ background: var(--panel); border: 1px solid var(--border); border-radius: 10px; padding: 14px 16px; }}
  .card h3 {{ margin: 0 0 8px 0; font-size: 14px; font-weight: 600; color: var(--text); }}
  .card .hint {{ color: var(--muted); font-size: 11px; margin-bottom: 8px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
  table th, table td {{ padding: 6px 8px; border-bottom: 1px solid var(--border); text-align: left; vertical-align: top; }}
  table th {{
    position: sticky; top: 0; background: var(--panel2); color: var(--muted); font-weight: 600;
    text-transform: uppercase; letter-spacing: .3px; font-size: 11px; cursor: pointer; user-select: none;
  }}
  table th.sorted-asc::after {{ content: " ▲"; color: var(--accent); }}
  table th.sorted-desc::after {{ content: " ▼"; color: var(--accent); }}
  table tbody tr:hover {{ background: rgba(124,92,255,0.06); }}
  .table-wrap {{ max-height: 420px; overflow: auto; border: 1px solid var(--border); border-radius: 8px; }}
  .pill {{ display: inline-block; padding: 2px 8px; border-radius: 12px; font-size: 10px;
    background: var(--panel2); color: var(--muted); border: 1px solid var(--border); }}
  .pill.good {{ color: #065f46; background: #34d399; border-color: #10b981; }}
  .pill.bad  {{ color: #7f1d1d; background: #fca5a5; border-color: #ef4444; }}
  .search {{ background: var(--panel2); border: 1px solid var(--border); color: var(--text);
    padding: 6px 10px; border-radius: 6px; font: inherit; width: 100%; margin-bottom: 8px; }}
  footer {{ padding: 16px 24px; color: var(--muted); font-size: 11px; border-top: 1px solid var(--border); }}
  .secondary-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(280px, 1fr)); gap: 12px; }}
  .summary-card {{ background: var(--panel); border: 1px solid var(--border); border-radius: 8px; padding: 12px; }}
  .summary-card h4 {{ margin: 0 0 4px 0; font-size: 13px; }}
  .summary-card .cnt {{ color: var(--accent2); font-size: 20px; font-weight: 600; }}
  .summary-card details {{ margin-top: 8px; }}
  .summary-card summary {{ cursor: pointer; color: var(--muted); font-size: 12px; }}
</style>
</head>
<body>
<header>
  <div>
    <h1>SVRG Master — Trent 900 Guarantee Administration</h1>
    <div class="meta">Emirates · Shop Visit Rate Guarantee + Enhanced SVRG + Enhanced HPTB Measure</div>
  </div>
  <div class="meta" id="source-meta"></div>
</header>

<div class="filters">
  <div><label>Engine Family</label><select id="f-family"><option value="">All</option></select></div>
  <div><label>Engine Status</label><select id="f-status"><option value="">All</option></select></div>
  <div><label>Engine Serial</label><select id="f-engine"><option value="">All</option></select></div>
  <div><label>Year from</label><input type="number" id="f-year-from" placeholder="e.g. 2017" /></div>
  <div><label>Year to</label><input type="number" id="f-year-to" placeholder="e.g. 2031" /></div>
  <div style="align-self:end"><button id="f-reset" style="background:var(--panel2);border:1px solid var(--border);color:var(--text);padding:7px 14px;border-radius:6px;cursor:pointer">Reset</button></div>
</div>

<div class="tabs">
  <div class="tab active" data-tab="fleet">Engine Fleet</div>
  <div class="tab" data-tab="svs">Shop Visit History</div>
  <div class="tab" data-tab="efh">Flight Hours</div>
  <div class="tab" data-tab="claims">Claims &amp; Events</div>
  <div class="tab" data-tab="svrg">SVRG+ESVRG Metrics</div>
  <div class="tab" data-tab="secondary">Secondary Summaries</div>
</div>

<div class="pane active" id="pane-fleet">
  <div class="kpis" id="kpi-fleet"></div>
  <div class="grid two">
    <div class="card"><h3>Engine Family Distribution</h3><div id="ch-family"></div></div>
    <div class="card"><h3>Engine Age (Years Since Delivery)</h3><div id="ch-age"></div></div>
  </div>
  <div class="card" style="margin-top:16px">
    <h3>Engine Fleet List</h3>
    <input type="text" class="search" id="search-engines" placeholder="Search engine serial, family, or status..." />
    <div class="table-wrap"><table id="tbl-engines"></table></div>
  </div>
</div>

<div class="pane" id="pane-svs">
  <div class="kpis" id="kpi-svs"></div>
  <div class="grid two">
    <div class="card"><h3>Shop Visits per Year</h3><div id="ch-svs-year"></div></div>
    <div class="card"><h3>Shop Visit Cause Breakdown</h3><div id="ch-svs-cause"></div></div>
  </div>
  <div class="card" style="margin-top:16px">
    <h3>Engine SV Timeline</h3>
    <div class="hint">Scatter of CSN vs SV date, per engine serial.</div>
    <div id="ch-svs-scatter"></div>
  </div>
  <div class="card" style="margin-top:16px">
    <h3>Shop Visit List</h3>
    <input type="text" class="search" id="search-svs" placeholder="Search engine serial, cause, comments..." />
    <div class="table-wrap"><table id="tbl-svs"></table></div>
  </div>
</div>

<div class="pane" id="pane-efh">
  <div class="kpis" id="kpi-efh"></div>
  <div class="grid two">
    <div class="card"><h3>Fleet EFH per Year</h3><div id="ch-efh-year"></div></div>
    <div class="card"><h3>Top 10 Engines by EFH</h3><div id="ch-efh-top"></div></div>
  </div>
  <div class="card" style="margin-top:16px">
    <h3>Per-Engine Annual EFH</h3>
    <input type="text" class="search" id="search-efh" placeholder="Search engine serial..." />
    <div class="table-wrap"><table id="tbl-efh"></table></div>
  </div>
</div>

<div class="pane" id="pane-claims">
  <div class="kpis" id="kpi-claims"></div>
  <div class="grid two">
    <div class="card"><h3>Claims per Year</h3><div id="ch-claims-year"></div></div>
    <div class="card"><h3>Events per Year</h3><div id="ch-events-year"></div></div>
  </div>
  <div class="card" style="margin-top:16px">
    <h3>Claims</h3>
    <div class="table-wrap"><table id="tbl-claims"></table></div>
  </div>
  <div class="card" style="margin-top:16px">
    <h3>Events</h3>
    <input type="text" class="search" id="search-events" placeholder="Search cause, ESN, AC position..." />
    <div class="table-wrap"><table id="tbl-events"></table></div>
  </div>
</div>

<div class="pane" id="pane-svrg">
  <div class="card"><h3>SVRG + ESVRG Metrics</h3>
    <div class="hint">Per-year values for each metric row in the SVRG+ESVRG calculation sheet.</div>
    <div class="table-wrap"><table id="tbl-svrg"></table></div>
  </div>
</div>

<div class="pane" id="pane-secondary">
  <div class="card"><h3>Hidden / Secondary Summary Sheets</h3>
    <div class="hint">Sample rows + row counts. Many are hidden in the workbook but retained for reference.</div>
    <div class="secondary-grid" id="secondary-grid"></div>
  </div>
</div>

<footer id="footer-meta"></footer>

<script id="svrg-data" type="application/json">{payload}</script>

<script>
(function() {{
  const DATA = JSON.parse(document.getElementById('svrg-data').textContent);
  window.SVRG = DATA;

  // ------- utils
  const num = v => (typeof v === 'number' && isFinite(v)) ? v : null;
  const fmtNum = n => n == null ? '—' : Math.round(n).toLocaleString();
  const fmtFloat = (n, d=1) => n == null ? '—' : n.toLocaleString(undefined, {{maximumFractionDigits: d, minimumFractionDigits: d}});
  const fmtMoney = n => n == null ? '—' : '£' + Math.round(n).toLocaleString();
  const fmtDate = s => !s ? '—' : String(s).slice(0,10);
  const yearOf = s => (!s || typeof s !== 'string') ? null : parseInt(s.slice(0,4));

  // ------- filter state
  const filters = {{ family: '', status: '', engine: '', yearFrom: null, yearTo: null }};

  function getEngineSet() {{
    // Intersection of engines that pass current filters
    const set = new Set();
    DATA.qualified_engines.forEach(e => {{
      if (filters.family && e.engine_family !== filters.family) return;
      if (filters.status && e.status !== filters.status) return;
      if (filters.engine && e.engine_serial !== filters.engine) return;
      set.add(e.engine_serial);
    }});
    return set;
  }}

  function inYear(y) {{
    if (y == null) return true;
    if (filters.yearFrom != null && y < filters.yearFrom) return false;
    if (filters.yearTo != null && y > filters.yearTo) return false;
    return true;
  }}

  // ------- filter bar setup
  function uniq(arr) {{ return Array.from(new Set(arr.filter(x => x != null && x !== ''))).sort(); }}
  const families = uniq(DATA.qualified_engines.map(e => e.engine_family));
  const statuses = uniq(DATA.qualified_engines.map(e => e.status));
  const engines  = uniq(DATA.qualified_engines.map(e => e.engine_serial));

  function populateSelect(id, options) {{
    const sel = document.getElementById(id);
    options.forEach(o => {{
      const opt = document.createElement('option');
      opt.value = o; opt.textContent = o;
      sel.appendChild(opt);
    }});
  }}
  populateSelect('f-family', families);
  populateSelect('f-status', statuses);
  populateSelect('f-engine', engines);

  ['f-family','f-status','f-engine','f-year-from','f-year-to'].forEach(id => {{
    document.getElementById(id).addEventListener('input', e => {{
      if (id === 'f-family') filters.family = e.target.value;
      if (id === 'f-status') filters.status = e.target.value;
      if (id === 'f-engine') filters.engine = e.target.value;
      if (id === 'f-year-from') filters.yearFrom = e.target.value ? parseInt(e.target.value) : null;
      if (id === 'f-year-to') filters.yearTo = e.target.value ? parseInt(e.target.value) : null;
      renderAll();
    }});
  }});
  document.getElementById('f-reset').addEventListener('click', () => {{
    filters.family = filters.status = filters.engine = '';
    filters.yearFrom = filters.yearTo = null;
    ['f-family','f-status','f-engine','f-year-from','f-year-to'].forEach(id => document.getElementById(id).value = '');
    renderAll();
  }});

  // ------- tabs
  document.querySelectorAll('.tab').forEach(t => {{
    t.addEventListener('click', () => {{
      document.querySelectorAll('.tab').forEach(x => x.classList.remove('active'));
      document.querySelectorAll('.pane').forEach(x => x.classList.remove('active'));
      t.classList.add('active');
      document.getElementById('pane-' + t.dataset.tab).classList.add('active');
      setTimeout(renderAll, 50);
    }});
  }});

  // ------- KPIs
  function kpi(label, value, sub) {{
    return `<div class="kpi"><div class="label">${{label}}</div><div class="value">${{value}}</div>${{sub ? `<div class="sub">${{sub}}</div>` : ''}}</div>`;
  }}

  function renderKPIs() {{
    const esnSet = getEngineSet();
    const engines = DATA.qualified_engines.filter(e => esnSet.has(e.engine_serial));
    const svs = DATA.qualified_svs.filter(sv => esnSet.has(sv.engine_serial) && inYear(num(sv.sv_year)));
    const efh = DATA.qualified_efh.filter(r => esnSet.has(r.engine_serial) && inYear(r.year));
    const claims = DATA.claims.filter(c => inYear(num(c.year) || yearOf(c.date)));
    const events = DATA.events.filter(ev => esnSet.has(ev.engine_serial) && inYear(yearOf(ev.date)));

    // Fleet
    const qualified = engines.filter(e => e.status && e.status.toLowerCase().startsWith('qualified')).length;
    document.getElementById('kpi-fleet').innerHTML =
      kpi('Total Engines', engines.length, `${{families.length}} families`) +
      kpi('Qualified Engines', qualified, `${{Math.round(100*qualified/Math.max(engines.length,1))}}% of fleet`) +
      kpi('Engine Families', uniq(engines.map(e => e.engine_family)).length) +
      kpi('Earliest Delivery', (() => {{ const d = engines.map(e=>e.delivery_date).filter(Boolean).sort()[0]; return fmtDate(d); }})());

    // SVs
    const svYears = svs.map(s => num(s.sv_year) || yearOf(s.sv_date)).filter(Boolean);
    const svEsn = uniq(svs.map(s => s.engine_serial));
    const hptb = svs.filter(s => typeof s.hptb_driven === 'string' && s.hptb_driven.toLowerCase().startsWith('yes')).length;
    document.getElementById('kpi-svs').innerHTML =
      kpi('Total Shop Visits', svs.length) +
      kpi('Unique Engines', svEsn.length) +
      kpi('HPTB-Driven SVs', hptb) +
      kpi('Date Range', svYears.length ? `${{Math.min(...svYears)}} — ${{Math.max(...svYears)}}` : '—');

    // EFH
    const totalEfh = efh.reduce((s,r) => s + (r.efh||0), 0);
    const byEng = {{}};
    efh.forEach(r => byEng[r.engine_serial] = (byEng[r.engine_serial]||0) + r.efh);
    const maxEng = Object.entries(byEng).sort((a,b) => b[1]-a[1])[0];
    const engCount = Object.keys(byEng).length;
    document.getElementById('kpi-efh').innerHTML =
      kpi('Total Fleet EFH', fmtNum(totalEfh)) +
      kpi('Avg EFH / Engine', fmtNum(totalEfh / Math.max(engCount,1))) +
      kpi('Highest-Hour Engine', maxEng ? maxEng[0] : '—', maxEng ? `${{fmtNum(maxEng[1])}} hrs` : '') +
      kpi('Engines w/ EFH Data', engCount);

    // Claims & Events
    const claimValue = claims.reduce((s,c) => s + (num(c.credit_note_value) || 0), 0);
    document.getElementById('kpi-claims').innerHTML =
      kpi('Total Claims', claims.length) +
      kpi('Total Claim Value', fmtMoney(claimValue)) +
      kpi('Total Events', events.length) +
      kpi('Unique Event ESNs', uniq(events.map(e=>e.engine_serial)).length);
  }}

  // ------- charts
  const chartRegistry = {{}};
  function renderChart(id, options) {{
    if (chartRegistry[id]) {{ chartRegistry[id].destroy(); }}
    const el = document.getElementById(id);
    if (!el) return;
    chartRegistry[id] = new ApexCharts(el, options);
    chartRegistry[id].render();
  }}

  const APEX_COMMON = {{
    chart: {{ fontFamily: 'inherit', foreColor: '#9aa3b2', toolbar: {{ show: false }}, background: 'transparent' }},
    theme: {{ mode: 'dark' }},
    grid: {{ borderColor: '#242a36', strokeDashArray: 3 }},
    tooltip: {{ theme: 'dark' }},
  }};

  function renderCharts() {{
    const esnSet = getEngineSet();
    const engines = DATA.qualified_engines.filter(e => esnSet.has(e.engine_serial));
    const svs = DATA.qualified_svs.filter(sv => esnSet.has(sv.engine_serial) && inYear(num(sv.sv_year)));
    const efh = DATA.qualified_efh.filter(r => esnSet.has(r.engine_serial) && inYear(r.year));
    const events = DATA.events.filter(ev => esnSet.has(ev.engine_serial) && inYear(yearOf(ev.date)));
    const claims = DATA.claims.filter(c => inYear(num(c.year) || yearOf(c.date)));

    // --- family donut
    const famCounts = {{}};
    engines.forEach(e => {{
      const k = e.engine_family || 'Unknown';
      famCounts[k] = (famCounts[k]||0) + 1;
    }});
    renderChart('ch-family', {{
      ...APEX_COMMON,
      chart: {{ ...APEX_COMMON.chart, type: 'donut', height: 280 }},
      series: Object.values(famCounts),
      labels: Object.keys(famCounts),
      colors: ['#7c5cff','#22d3ee','#34d399','#fbbf24','#f87171','#c084fc'],
      legend: {{ position: 'bottom' }},
    }});

    // --- age histogram
    const thisYear = new Date().getFullYear();
    const ageBuckets = {{}};
    engines.forEach(e => {{
      const y = yearOf(e.delivery_date);
      if (!y) return;
      const age = thisYear - y;
      const bucket = Math.floor(age/2)*2;
      ageBuckets[bucket] = (ageBuckets[bucket]||0) + 1;
    }});
    const ageKeys = Object.keys(ageBuckets).map(Number).sort((a,b)=>a-b);
    renderChart('ch-age', {{
      ...APEX_COMMON,
      chart: {{ ...APEX_COMMON.chart, type: 'bar', height: 280 }},
      series: [{{ name: 'Engines', data: ageKeys.map(k => ageBuckets[k]) }}],
      xaxis: {{ categories: ageKeys.map(k => k + '–' + (k+1) + ' yrs'), title: {{ text: 'Age bucket' }} }},
      yaxis: {{ title: {{ text: 'Engines' }} }},
      colors: ['#22d3ee'],
      plotOptions: {{ bar: {{ columnWidth: '55%', borderRadius: 4 }} }},
      dataLabels: {{ enabled: false }},
    }});

    // --- SVs per year
    const sy = {{}};
    svs.forEach(s => {{ const y = num(s.sv_year) || yearOf(s.sv_date); if (y) sy[y] = (sy[y]||0)+1; }});
    const syKeys = Object.keys(sy).map(Number).sort((a,b)=>a-b);
    renderChart('ch-svs-year', {{
      ...APEX_COMMON,
      chart: {{ ...APEX_COMMON.chart, type: 'bar', height: 300 }},
      series: [{{ name: 'Shop Visits', data: syKeys.map(k => sy[k]) }}],
      xaxis: {{ categories: syKeys }},
      colors: ['#7c5cff'],
      dataLabels: {{ enabled: true, style: {{ colors: ['#e6e9ef'] }} }},
      plotOptions: {{ bar: {{ columnWidth: '55%', borderRadius: 4 }} }},
    }});

    // --- SV cause donut
    const causeCounts = {{}};
    svs.forEach(s => {{
      const c = (s.cause || 'Unknown').toString().slice(0, 30);
      causeCounts[c] = (causeCounts[c]||0) + 1;
    }});
    renderChart('ch-svs-cause', {{
      ...APEX_COMMON,
      chart: {{ ...APEX_COMMON.chart, type: 'donut', height: 300 }},
      series: Object.values(causeCounts),
      labels: Object.keys(causeCounts),
      legend: {{ position: 'bottom' }},
      colors: ['#7c5cff','#22d3ee','#34d399','#fbbf24','#f87171','#c084fc','#fb923c','#60a5fa'],
    }});

    // --- SV scatter
    const series = {{}};
    svs.forEach(s => {{
      if (!s.sv_date || !s.engine_serial) return;
      const y = yearOf(s.sv_date); if (!y) return;
      const esn = s.engine_serial;
      if (!series[esn]) series[esn] = [];
      series[esn].push({{ x: new Date(s.sv_date).getTime(), y: parseInt(s.engine_serial) || 0, z: num(s.csn) || 10 }});
    }});
    const scatterSeries = Object.entries(series).slice(0, 60).map(([esn, pts]) => ({{ name: esn, data: pts }}));
    renderChart('ch-svs-scatter', {{
      ...APEX_COMMON,
      chart: {{ ...APEX_COMMON.chart, type: 'scatter', height: 400, zoom: {{ enabled: true, type: 'xy' }} }},
      series: scatterSeries,
      xaxis: {{ type: 'datetime', labels: {{ format: 'yyyy' }}, title: {{ text: 'Shop Visit Date' }} }},
      yaxis: {{ title: {{ text: 'Engine Serial' }}, labels: {{ formatter: v => Math.round(v).toString() }} }},
      legend: {{ show: false }},
      markers: {{ size: 6 }},
    }});

    // --- EFH per year (fleet)
    const efhY = {{}};
    efh.forEach(r => {{ efhY[r.year] = (efhY[r.year]||0) + r.efh; }});
    const efhKeys = Object.keys(efhY).map(Number).sort((a,b)=>a-b);
    renderChart('ch-efh-year', {{
      ...APEX_COMMON,
      chart: {{ ...APEX_COMMON.chart, type: 'area', height: 300 }},
      series: [{{ name: 'Fleet EFH', data: efhKeys.map(k => Math.round(efhY[k])) }}],
      xaxis: {{ categories: efhKeys }},
      colors: ['#22d3ee'],
      stroke: {{ curve: 'smooth', width: 2 }},
      fill: {{ type: 'gradient', gradient: {{ shadeIntensity: 0.6, opacityFrom: 0.4, opacityTo: 0.05 }} }},
      dataLabels: {{ enabled: false }},
    }});

    // --- Top 10 engines by EFH
    const byEng = {{}};
    efh.forEach(r => {{ byEng[r.engine_serial] = (byEng[r.engine_serial]||0) + r.efh; }});
    const top10 = Object.entries(byEng).sort((a,b)=>b[1]-a[1]).slice(0,10);
    renderChart('ch-efh-top', {{
      ...APEX_COMMON,
      chart: {{ ...APEX_COMMON.chart, type: 'bar', height: 300 }},
      series: [{{ name: 'EFH', data: top10.map(([,v]) => Math.round(v)) }}],
      xaxis: {{ categories: top10.map(([k]) => k) }},
      plotOptions: {{ bar: {{ horizontal: true, borderRadius: 4, barHeight: '65%' }} }},
      colors: ['#34d399'],
      dataLabels: {{ enabled: true, style: {{ colors: ['#0b0d12'] }}, formatter: v => v.toLocaleString() }},
    }});

    // --- claims per year
    const cy = {{}};
    claims.forEach(c => {{ const y = num(c.year) || yearOf(c.date); if (y) cy[y] = (cy[y]||0) + 1; }});
    const cyKeys = Object.keys(cy).map(Number).sort((a,b)=>a-b);
    renderChart('ch-claims-year', {{
      ...APEX_COMMON,
      chart: {{ ...APEX_COMMON.chart, type: 'bar', height: 300 }},
      series: [{{ name: 'Claims', data: cyKeys.map(k => cy[k]) }}],
      xaxis: {{ categories: cyKeys.length ? cyKeys : ['No claims'] }},
      colors: ['#fbbf24'],
      plotOptions: {{ bar: {{ columnWidth: '55%', borderRadius: 4 }} }},
    }});

    // --- events per year
    const ey = {{}};
    events.forEach(e => {{ const y = yearOf(e.date); if (y) ey[y] = (ey[y]||0) + 1; }});
    const eyKeys = Object.keys(ey).map(Number).sort((a,b)=>a-b);
    renderChart('ch-events-year', {{
      ...APEX_COMMON,
      chart: {{ ...APEX_COMMON.chart, type: 'bar', height: 300 }},
      series: [{{ name: 'Events', data: eyKeys.map(k => ey[k]) }}],
      xaxis: {{ categories: eyKeys.length ? eyKeys : ['No events'] }},
      colors: ['#f87171'],
      plotOptions: {{ bar: {{ columnWidth: '55%', borderRadius: 4 }} }},
    }});
  }}

  // ------- Tables (sortable, filterable)
  function buildTable(tblId, rows, columns, searchId) {{
    const el = document.getElementById(tblId);
    el.innerHTML = '';
    const thead = document.createElement('thead');
    const tr = document.createElement('tr');
    columns.forEach((c, i) => {{
      const th = document.createElement('th');
      th.textContent = c.label;
      th.dataset.col = i;
      th.addEventListener('click', () => {{
        const cur = th.classList.contains('sorted-asc') ? 'asc' : (th.classList.contains('sorted-desc') ? 'desc' : null);
        thead.querySelectorAll('th').forEach(x => x.classList.remove('sorted-asc','sorted-desc'));
        const next = cur === 'asc' ? 'desc' : 'asc';
        th.classList.add('sorted-' + next);
        const key = c.key;
        const mul = next === 'asc' ? 1 : -1;
        rows.sort((a,b) => {{
          const av = a[key], bv = b[key];
          if (av == null && bv == null) return 0;
          if (av == null) return 1; if (bv == null) return -1;
          if (typeof av === 'number' && typeof bv === 'number') return (av-bv)*mul;
          return String(av).localeCompare(String(bv)) * mul;
        }});
        renderBody();
      }});
      tr.appendChild(th);
    }});
    thead.appendChild(tr);
    const tbody = document.createElement('tbody');
    el.appendChild(thead); el.appendChild(tbody);
    let searchTerm = '';
    function renderBody() {{
      tbody.innerHTML = '';
      const filtered = searchTerm ?
        rows.filter(r => columns.some(c => String(r[c.key] ?? '').toLowerCase().includes(searchTerm))) : rows;
      const slice = filtered.slice(0, 2000);
      slice.forEach(r => {{
        const tr = document.createElement('tr');
        columns.forEach(c => {{
          const td = document.createElement('td');
          let v = r[c.key];
          if (c.format) v = c.format(v, r);
          td.innerHTML = v == null ? '—' : String(v);
          tr.appendChild(td);
        }});
        tbody.appendChild(tr);
      }});
      if (filtered.length > slice.length) {{
        const tr = document.createElement('tr');
        const td = document.createElement('td');
        td.colSpan = columns.length;
        td.style.color = 'var(--muted)'; td.style.textAlign = 'center';
        td.textContent = `… and ${{filtered.length - slice.length}} more rows (truncated for performance)`;
        tr.appendChild(td); tbody.appendChild(tr);
      }}
    }}
    renderBody();
    if (searchId) {{
      const s = document.getElementById(searchId);
      if (s) {{
        s.value = '';
        s.oninput = e => {{ searchTerm = e.target.value.toLowerCase(); renderBody(); }};
      }}
    }}
    return {{ rerender: renderBody }};
  }}

  function renderTables() {{
    const esnSet = getEngineSet();
    const engines = DATA.qualified_engines.filter(e => esnSet.has(e.engine_serial));
    const svs = DATA.qualified_svs.filter(sv => esnSet.has(sv.engine_serial) && inYear(num(sv.sv_year)));
    const efh = DATA.qualified_efh.filter(r => esnSet.has(r.engine_serial) && inYear(r.year));
    const events = DATA.events.filter(ev => esnSet.has(ev.engine_serial) && inYear(yearOf(ev.date)));
    const claims = DATA.claims.filter(c => inYear(num(c.year) || yearOf(c.date)));

    buildTable('tbl-engines', engines, [
      {{ label: 'Asset#', key: 'asset_id' }},
      {{ label: 'ESN', key: 'engine_serial' }},
      {{ label: 'Family', key: 'engine_family' }},
      {{ label: 'Delivery', key: 'delivery_date', format: fmtDate }},
      {{ label: 'Status', key: 'status' }},
      {{ label: 'Improved HPTB', key: 'improved_hptb_from_new' }},
      {{ label: 'Final Fix', key: 'final_fix_from_new' }},
      {{ label: 'Qualified Date', key: 'qualified_date', format: fmtDate }},
      {{ label: 'Qualified Year', key: 'qualified_year' }},
    ], 'search-engines');

    buildTable('tbl-svs', svs, [
      {{ label: 'Asset#', key: 'asset_id' }},
      {{ label: 'ESN', key: 'engine_serial' }},
      {{ label: 'SV Date', key: 'sv_date', format: fmtDate }},
      {{ label: 'CSN', key: 'csn', format: v => v==null?'—':Math.round(v).toLocaleString() }},
      {{ label: 'Year', key: 'sv_year' }},
      {{ label: 'Cause', key: 'cause' }},
      {{ label: 'Qualified?', key: 'qualified_cause' }},
      {{ label: 'HPTB-Driven', key: 'hptb_driven' }},
      {{ label: 'Qualified HPTB SV', key: 'qualified_hptb_sv' }},
      {{ label: 'Comments', key: 'comments' }},
    ], 'search-svs');

    buildTable('tbl-efh', efh, [
      {{ label: 'Engine', key: 'engine_serial' }},
      {{ label: 'Year', key: 'year' }},
      {{ label: 'EFH', key: 'efh', format: v => v==null?'—':fmtFloat(v,1) }},
    ], 'search-efh');

    buildTable('tbl-claims', claims, [
      {{ label: 'Date', key: 'date', format: fmtDate }},
      {{ label: 'Year', key: 'year' }},
      {{ label: 'Credit Note Ref', key: 'credit_note_ref' }},
      {{ label: 'Guarantee', key: 'guarantee' }},
      {{ label: 'CN Value', key: 'credit_note_value', format: fmtMoney }},
      {{ label: 'Cumulative', key: 'cumulative_value', format: fmtMoney }},
    ]);

    buildTable('tbl-events', events, [
      {{ label: 'Flag', key: 'flag' }},
      {{ label: 'Date', key: 'date', format: fmtDate }},
      {{ label: 'ESN', key: 'engine_serial' }},
      {{ label: 'AC/Pos', key: 'ac_position' }},
      {{ label: 'Cause', key: 'cause' }},
      {{ label: 'Emirates Input', key: 'emirates_input' }},
      {{ label: 'RR Input', key: 'rr_input' }},
      {{ label: 'RR Justification', key: 'rr_justification' }},
      {{ label: 'Coverage', key: 'coverage' }},
    ], 'search-events');

    // SVRG+ESVRG: expand into flat (metric × year) rows
    const svrgRows = [];
    DATA.svrg_esvrg_metrics.forEach(m => {{
      m.values_by_year.forEach(vy => {{
        svrgRows.push({{
          description: m.description, code: m.code, calculation: m.calculation,
          year: vy.year, value: vy.value
        }});
      }});
    }});
    buildTable('tbl-svrg', svrgRows, [
      {{ label: 'Metric', key: 'description' }},
      {{ label: 'Code', key: 'code' }},
      {{ label: 'Formula', key: 'calculation' }},
      {{ label: 'Year', key: 'year' }},
      {{ label: 'Value', key: 'value', format: v => v==null?'—':(typeof v==='number'?fmtFloat(v,2):String(v)) }},
    ]);
  }}

  function renderSecondary() {{
    const grid = document.getElementById('secondary-grid');
    grid.innerHTML = '';
    Object.entries(DATA.secondary_summaries).forEach(([name, s]) => {{
      const card = document.createElement('div');
      card.className = 'summary-card';
      card.innerHTML = `<h4>${{name}}</h4><div class="cnt">${{s.data_row_count}} rows</div>
        <div class="hint">Header row ${{s.header_row}}, ${{(s.column_headers||[]).filter(Boolean).length}} columns</div>
        <details><summary>Preview sample (${{(s.sample_rows||[]).length}} rows)</summary>
          <div style="overflow:auto;max-height:220px;margin-top:6px"><table style="font-size:11px">
            <thead><tr>${{(s.column_headers||[]).map(h => `<th>${{h==null?'—':String(h).slice(0,40)}}</th>`).join('')}}</tr></thead>
            <tbody>${{(s.sample_rows||[]).map(r =>
              `<tr>${{r.map(c => `<td>${{c==null?'—':(typeof c==='string'?c.slice(0,80):String(c))}}</td>`).join('')}}</tr>`
            ).join('')}}</tbody>
          </table></div>
        </details>`;
      grid.appendChild(card);
    }});
  }}

  // ------- footer + meta
  document.getElementById('source-meta').innerHTML =
    `<div>Source: ${{DATA.source_file}}</div>` +
    `<div>${{DATA.sheet_count}} sheets (${{DATA.visible_sheets.length}} visible, ${{DATA.hidden_sheets.length}} hidden, ${{DATA.chartsheets.length}} chartsheets)</div>`;
  document.getElementById('footer-meta').innerHTML =
    `Engines: ${{DATA.qualified_engines.length}} · SVs: ${{DATA.qualified_svs.length}} · EFH items: ${{DATA.qualified_efh.length}} · H&amp;C rows: ${{DATA.hours_cycles.row_count.toLocaleString()}} · Claims: ${{DATA.claims.length}} · Events: ${{DATA.events.length}}`;

  function renderAll() {{
    renderKPIs();
    renderCharts();
    renderTables();
    renderSecondary();
  }}
  renderAll();
}})();
</script>
</body>
</html>
"""


if __name__ == "__main__":
    main()
