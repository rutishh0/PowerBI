"""Independent xlsx reconciliation.
Reads source xlsx files with openpyxl directly and computes ground-truth values
to compare against embedded JSON payloads (saved in _payloads/)."""
import os, json, sys
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.abspath(__file__))
PARENT = os.path.dirname(ROOT)  # V6
NEWINFO = os.path.join(PARENT, 'New info')
PAYLOADS = os.path.join(ROOT, '_payloads')

def load_payload(name):
    with open(os.path.join(PAYLOADS, name + '.json'), encoding='utf-8') as fp:
        return json.load(fp)

def num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace(',', '').replace('£', '').replace('$', '').replace('€', '')
    if s.startswith('(') and s.endswith(')'):
        s = '-' + s[1:-1]
    try: return float(s)
    except: return None

results = []

def log(section, key, truth, html_val, status, note=''):
    results.append((section, key, truth, html_val, status, note))

# ------------------------------------------------------------------
# 1) ETH_SOA (ETH SOA 30.1.26.xlsx)
# ------------------------------------------------------------------
def verify_eth_soa():
    sect = 'ETH_SOA_30_1_26'
    path = os.path.join(NEWINFO, 'ETH SOA 30.1.26.xlsx')
    wb = load_workbook(path, data_only=True)
    # Expectation: 94 items, net balance -33,588,435.28, 5 sections, CRC item count 40
    p = load_payload('ETH_SOA_30_1_26')
    # flat_items count
    flat_n = len(p.get('flat_items', []))
    # Sections in payload
    sections = p.get('sections', {})
    # Each section may be dict with items
    section_item_counts = {}
    total_items = 0
    for s, body in sections.items():
        items = body.get('items') if isinstance(body, dict) else body
        if isinstance(items, list):
            section_item_counts[s] = len(items)
            total_items += len(items)
        elif isinstance(body, list):
            section_item_counts[s] = len(body)
            total_items += len(body)
    # Net balance from payload summary
    summary = p.get('summary', {})
    html_net_balance = summary.get('net_balance') or summary.get('sheet_balance') or summary.get('balance_total')
    # Try sheet_totals
    if html_net_balance is None:
        st = p.get('sheet_totals', {})
        html_net_balance = st.get('net_balance') or st.get('balance') or st.get('grand_total')
    # TRUTH - parse xlsx directly
    # Sheet structure varies; collect rows where Amount col is numeric and not a header/subtotal
    truth_total_rows = 0
    truth_net = 0.0
    truth_sections = set()
    for sh in wb.sheetnames:
        ws = wb[sh]
        # Find header row by scanning for "Amount" header
        header_row = None
        header_map = {}
        for r in range(1, min(ws.max_row, 30)+1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 30)+1)]
            low = [str(v).lower() if v is not None else '' for v in row_vals]
            if 'amount' in low or any('amount' in v for v in low):
                header_row = r
                for c, v in enumerate(row_vals, start=1):
                    if v is not None:
                        header_map[str(v).strip().lower()] = c
                break
        if not header_row:
            continue
        amt_col = None
        for k, c in header_map.items():
            if k == 'amount' or 'amount' == k.strip():
                amt_col = c; break
        if amt_col is None:
            for k, c in header_map.items():
                if 'amount' in k:
                    amt_col = c; break
        if amt_col is None:
            continue
        # Section column candidate
        sec_col = header_map.get('section') or header_map.get('charges') or header_map.get('debit type')
        rows_with_amount = 0
        for r in range(header_row+1, ws.max_row+1):
            v = ws.cell(row=r, column=amt_col).value
            nv = num(v)
            # Skip subtotals: any row where any cell mentions 'total'
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 30)+1)]
            row_has_total = any(isinstance(x, str) and 'total' in x.lower() for x in row_vals)
            if nv is not None and not row_has_total:
                rows_with_amount += 1
                truth_net += nv
                if sec_col:
                    sv = ws.cell(row=r, column=sec_col).value
                    if sv: truth_sections.add(str(sv).strip())
        truth_total_rows += rows_with_amount
    print(f'[ETH_SOA] TRUTH: rows={truth_total_rows} net={truth_net:,.2f} sections={sorted(truth_sections)}')
    print(f'[ETH_SOA] HTML:  flat_items={flat_n} total_from_sections={total_items} section_names={list(section_item_counts.keys())}')
    print(f'[ETH_SOA] HTML:  sheet_totals={p.get("sheet_totals")}')
    print(f'[ETH_SOA] HTML:  summary keys={list(summary.keys())}')
    print(f'[ETH_SOA] HTML:  section_item_counts={section_item_counts}')
    # Expected 94 items and 5 sections
    log(sect, 'flat_items_count', 94, flat_n, 'MATCH' if flat_n == 94 else 'MISMATCH')
    log(sect, 'num_sections', 5, len(section_item_counts), 'MATCH' if len(section_item_counts) == 5 else 'MISMATCH')
    # Net balance: expected -33,588,435.28
    expected_net = -33588435.28
    # Figure out html net
    sht = p.get('sheet_totals') or {}
    html_net = None
    for k in ('net_balance','balance','grand_total','total','net','running_balance'):
        if k in sht and isinstance(sht[k], (int,float)):
            html_net = sht[k]; break
    # Also check summary
    for k in ('net_balance','balance','grand_total','total'):
        if html_net is None and k in summary and isinstance(summary[k], (int,float)):
            html_net = summary[k]
    log(sect, 'net_balance(expected)', expected_net, html_net,
        'MATCH' if html_net is not None and abs(html_net - expected_net) < 1 else 'MISMATCH')
    log(sect, 'truth_net_from_xlsx', truth_net, expected_net,
        'WITHIN_TOLERANCE' if abs(truth_net - expected_net) < abs(expected_net)*0.01 else 'CHECK')
    # CRC item count == 40
    crc_count = None
    for k, v in section_item_counts.items():
        if 'customer responsible' in k.lower() or 'crc' in k.lower():
            crc_count = v
    log(sect, 'CRC_items(expected 40)', 40, crc_count, 'MATCH' if crc_count == 40 else 'MISMATCH')

# ------------------------------------------------------------------
# 2) EPI (EPI 16.02.xlsx)
# ------------------------------------------------------------------
def verify_epi():
    sect = 'EPI'
    path = os.path.join(NEWINFO, 'EPI 16.02.xlsx')
    wb = load_workbook(path, data_only=True)
    p = load_payload('EPI_16_02')
    invoices = p.get('invoices', [])
    subtotals = p.get('subtotals', [])
    aging_counts = p.get('aging_counts', {})
    # Count leading-zero refs in invoices
    leading_zero = 0
    for inv in invoices:
        ref = inv.get('invoice') or inv.get('ref') or inv.get('reference') or inv.get('doc_no') or inv.get('number') or ''
        s = str(ref)
        if s.startswith('0'):
            leading_zero += 1
    # Truth: scan xlsx directly
    truth_rows = 0
    truth_total = 0.0
    truth_subtotals = 0
    truth_leading_zero = 0
    for sh in wb.sheetnames:
        ws = wb[sh]
        header_row = None
        header_map = {}
        for r in range(1, min(ws.max_row, 20)+1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 30)+1)]
            low = [str(v).lower() if v is not None else '' for v in row_vals]
            if any('amount' in v for v in low) and any(('invoice' in v or 'reference' in v or 'document' in v) for v in low):
                header_row = r
                for c, v in enumerate(row_vals, start=1):
                    if v is not None:
                        header_map[str(v).strip().lower()] = c
                break
        if not header_row:
            continue
        amt_col = None
        ref_col = None
        for k, c in header_map.items():
            if k == 'amount' or k.endswith('amount'): amt_col = c
            if 'invoice' in k or 'reference' in k or 'document' in k: ref_col = c
        if amt_col is None: continue
        for r in range(header_row+1, ws.max_row+1):
            v = ws.cell(row=r, column=amt_col).value
            nv = num(v)
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 30)+1)]
            row_has_total = any(isinstance(x, str) and 'total' in x.lower() for x in row_vals)
            if nv is not None:
                if row_has_total:
                    truth_subtotals += 1
                else:
                    truth_rows += 1
                    truth_total += nv
                    if ref_col:
                        rv = ws.cell(row=r, column=ref_col).value
                        if rv is not None and str(rv).startswith('0'):
                            truth_leading_zero += 1
    print(f'[EPI] TRUTH: rows={truth_rows} subtotals={truth_subtotals} total={truth_total:,.2f} leading_zero={truth_leading_zero}')
    print(f'[EPI] HTML:  invoices={len(invoices)} subtotals={len(subtotals)} aging_buckets={list(aging_counts.keys())}')
    print(f'[EPI] HTML:  sum_amounts={sum(num(i.get("amount")) or 0 for i in invoices):,.2f}')
    log(sect, 'invoice_rows(expected 106)', 106, len(invoices), 'MATCH' if len(invoices)==106 else 'MISMATCH')
    log(sect, 'subtotals(expected 3)', 3, len(subtotals), 'MATCH' if len(subtotals)==3 else 'MISMATCH')
    # 7 aging buckets populated
    populated = sum(1 for k,v in aging_counts.items() if v and v>0)
    log(sect, 'aging_buckets_populated(>=7)', '>=7', populated, 'MATCH' if populated>=7 else 'MISMATCH_or_LESS')
    log(sect, 'TRUTH vs HTML invoice rows', truth_rows, len(invoices),
        'MATCH' if truth_rows == len(invoices) else 'MISMATCH')

# ------------------------------------------------------------------
# 3) MEA tracker
# ------------------------------------------------------------------
def verify_mea():
    sect = 'MEA'
    path = os.path.join(NEWINFO, 'MEA Profit Opportunities Tracker 21.04.xlsx')
    wb = load_workbook(path, data_only=True)
    p = load_payload('MEA_Profit_Opportunities_Tracker')
    sheets = p.get('sheets', {})
    for k, v in sheets.items():
        rows = v.get('rows') or v.get('data') or []
        print(f'[MEA] HTML sheet={k!r} row_count_nonblank={v.get("row_count_nonblank")} rows_len={len(rows)} benefit_2026={v.get("benefit_2026")} benefit_2027={v.get("benefit_2027")}')
    # TRUTH: Look at each sheet directly
    for sh in wb.sheetnames:
        ws = wb[sh]
        # header row = first non-empty row
        header_row = None
        for r in range(1, min(ws.max_row, 10)+1):
            vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column,50)+1)]
            if sum(1 for v in vals if v not in (None, '', ' ')) >= 3:
                header_row = r; break
        if not header_row:
            continue
        headers = {}
        for c in range(1, ws.max_column+1):
            v = ws.cell(row=header_row, column=c).value
            if v is not None:
                headers[str(v).strip().lower()] = c
        nonblank = 0
        b26_tot = 0.0
        b27_tot = 0.0
        b26_col = None; b27_col = None
        for k, c in headers.items():
            if ('benefit' in k and '2026' in k) or ('2026' in k and 'benefit' in k): b26_col = c
            if ('benefit' in k and '2027' in k) or ('2027' in k and 'benefit' in k): b27_col = c
        for r in range(header_row+1, ws.max_row+1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
            if sum(1 for v in row_vals if v not in (None, '', ' ')) == 0:
                continue
            # Skip total rows
            if any(isinstance(x,str) and 'total' in x.lower() for x in row_vals):
                continue
            nonblank += 1
            if b26_col:
                n = num(ws.cell(row=r, column=b26_col).value)
                if n is not None: b26_tot += n
            if b27_col:
                n = num(ws.cell(row=r, column=b27_col).value)
                if n is not None: b27_tot += n
        print(f'[MEA] TRUTH sheet={sh!r} rows_nonblank={nonblank} benefit_2026={b26_tot:,.2f} benefit_2027={b27_tot:,.2f}')
    log(sect, 'sheet_MEA_LOG_rows(expected 45)', 45, sheets.get('MEA LOG', {}).get('row_count_nonblank'),
        'MATCH' if sheets.get('MEA LOG', {}).get('row_count_nonblank') == 45 else 'MISMATCH')
    log(sect, 'sheet_L2_rows(expected 45)', 45, sheets.get('L2', {}).get('row_count_nonblank'),
        'MATCH' if sheets.get('L2', {}).get('row_count_nonblank') == 45 else 'MISMATCH')
    log(sect, 'sheet_L3_rows(expected 45)', 45, sheets.get('L3', {}).get('row_count_nonblank'),
        'MATCH' if sheets.get('L3', {}).get('row_count_nonblank') == 45 else 'MISMATCH')

# ------------------------------------------------------------------
# 4) Shop Visit
# ------------------------------------------------------------------
def verify_shop_visit():
    sect = 'SHOP_VISIT'
    import glob
    candidates = glob.glob(os.path.join(NEWINFO, 'SV008RV08_Trent 900 Shop Visit History*.xlsx'))
    assert candidates, 'shop visit xlsx not found'
    path = candidates[0]
    wb = load_workbook(path, data_only=True)
    p = load_payload('Trent_900_Shop_Visit_History')
    sv = p.get('shop_visits', [])
    cs = p.get('current_status', [])
    print(f'[SHOP_VISIT] HTML shop_visits={len(sv)} current_status={len(cs)} total={len(sv)+len(cs)}')
    # TRUTH: iterate workbook
    for sh in wb.sheetnames:
        ws = wb[sh]
        nonblank = 0
        for r in range(2, ws.max_row+1):
            if sum(1 for c in range(1, ws.max_column+1)
                   if ws.cell(row=r, column=c).value not in (None, '', ' ')) > 0:
                nonblank += 1
        print(f'[SHOP_VISIT] TRUTH sheet={sh!r} rows_below_header={nonblank}')
    log(sect, 'shop_visits(expected 1332)', 1332, len(sv),
        'MATCH' if len(sv)==1332 else 'MISMATCH')
    log(sect, 'current_status(expected 680)', 680, len(cs),
        'MATCH' if len(cs)==680 else 'MISMATCH')
    log(sect, 'total(expected 2012)', 2012, len(sv)+len(cs),
        'MATCH' if len(sv)+len(cs)==2012 else 'MISMATCH')

# ------------------------------------------------------------------
# 5) SVRG
# ------------------------------------------------------------------
def verify_svrg():
    sect = 'SVRG'
    import glob
    candidates = glob.glob(os.path.join(NEWINFO, 'VERSION 2 Enhanced SVRG MASTER*.xlsx'))
    assert candidates
    path = candidates[0]
    wb = load_workbook(path, data_only=True)
    p = load_payload('SVRG_MASTER')
    engines = p.get('qualified_engines', [])
    svs = p.get('qualified_svs', [])
    claims = p.get('claims', [])
    total_claim_value = p.get('total_claim_value')
    print(f'[SVRG] HTML engines={len(engines)} svs={len(svs)} claims={len(claims)} total_claim_value={total_claim_value}')
    # Filter placeholder engines (no ESN or ESN null)
    real_engines = [e for e in engines if e.get('engine_serial') and str(e.get('engine_serial')).strip()]
    print(f'[SVRG] HTML engines_with_ESN={len(real_engines)}')
    # Total fleet EFH
    efh = p.get('qualified_efh', []) or p.get('efh_per_engine', {})
    total_efh_html = None
    if isinstance(efh, list):
        total_efh_html = sum(num(e.get('efh')) or 0 for e in efh)
    elif isinstance(efh, dict):
        total_efh_html = sum(num(v) or 0 for v in efh.values())
    print(f'[SVRG] HTML total_fleet_EFH (approx sum)={total_efh_html}')
    # TRUTH: scan sheets for H&C
    for sh in wb.sheetnames[:5]:
        ws = wb[sh]
        print(f'[SVRG] SHEET {sh!r} max_row={ws.max_row}')
    # H&C sheet
    for sh in wb.sheetnames:
        if 'h&c' in sh.lower() or 'hours' in sh.lower() or 'cycles' in sh.lower():
            ws = wb[sh]
            print(f'[SVRG] H&C-like sheet={sh!r} max_row={ws.max_row}')
    # Log results
    log(sect, 'engines(expected 151)', 151, len(engines),
        'MATCH' if len(engines)==151 else 'MISMATCH')
    log(sect, 'svs(expected 215)', 215, len(svs),
        'MATCH' if len(svs)==215 else 'MISMATCH')
    # Total fleet EFH expected ~4,359,187
    if total_efh_html is not None:
        diff = abs(total_efh_html - 4359187)/4359187
        log(sect, 'total_fleet_EFH(~4,359,187)', 4359187, f'{total_efh_html:.0f}',
            'WITHIN_TOL' if diff < 0.005 else f'OFF_BY_{diff*100:.2f}%')

# ------------------------------------------------------------------
# 6) Global Hopper (v2 (1)) - canonical
# ------------------------------------------------------------------
def verify_hopper():
    sect = 'HOPPER_v2_1'
    path = os.path.join(NEWINFO, 'Global_Commercial_Optimisation_Hopper_(v2)_(1).xlsx')
    wb = load_workbook(path, data_only=True)
    p = load_payload('Global_Hopper')
    rows = p.get('rows', [])
    kpis = p.get('kpis', {})
    meta = p.get('meta', {})
    print(f'[HOPPER] HTML rows={len(rows)} kpis_keys={list(kpis.keys())}')
    # Last opp
    if rows:
        last = rows[-1]
        print(f'[HOPPER] HTML last_row customer/operator fields:', {k:v for k,v in last.items() if v and len(str(v))<50})
    # CRP total
    crp_total = None
    for k in ('crp_term_benefit_total_m', 'total_crp_term_benefit_m', 'total_crp_term_benefit', 'crp_total_m'):
        if k in kpis: crp_total = kpis[k]; break
    if crp_total is None:
        # sum manually
        crp_total = 0.0
        for r in rows:
            for k in ('crp_term_benefit_m','crp_term_benefit','crp_benefit_m'):
                n = num(r.get(k))
                if n is not None:
                    crp_total += n; break
    print(f'[HOPPER] HTML CRP_total(m)={crp_total}')
    print(f'[HOPPER] HTML sheets_parsed={meta.get("sheets_parsed") or meta.get("sheets")}')
    log(sect, 'opp_count(expected 109)', 109, len(rows),
        'MATCH' if len(rows)==109 else 'MISMATCH')
    # Check last is Ugandan Airlines
    last_text = json.dumps(rows[-1]) if rows else ''
    log(sect, 'last_row_contains_Ugandan', 'Ugandan Airlines', 'Ugandan' in last_text,
        'MATCH' if 'Ugandan' in last_text else 'MISMATCH')
    # CRP total expected 3069.21
    if crp_total is not None:
        diff = abs(crp_total - 3069.21)
        log(sect, 'CRP_total(expected 3069.21m)', 3069.21, f'{crp_total:.2f}',
            'MATCH' if diff < 1.0 else 'MISMATCH')
    # sheets_parsed only GLOBAL LOG
    sp = meta.get('sheets_parsed') or meta.get('sheets') or []
    only_global_log = (isinstance(sp, list) and len(sp) == 1 and 'GLOBAL LOG' in str(sp[0]).upper()) \
                       or (isinstance(sp, str) and sp.upper() == 'GLOBAL LOG')
    log(sect, 'sheets_parsed == [GLOBAL LOG]', 'GLOBAL LOG only', sp,
        'MATCH' if only_global_log else 'CHECK')

# ------------------------------------------------------------------
# 7) Global Hopper v2 original
# ------------------------------------------------------------------
def verify_hopper_orig():
    sect = 'HOPPER_v2_orig'
    path = os.path.join(NEWINFO, 'Global Commercial Optimisation Hopper (v2).xlsx')
    wb = load_workbook(path, data_only=True)
    p = load_payload('Global_Hopper_v2_original')
    rows = p.get('rows', [])
    meta = p.get('meta', {})
    note_rows = p.get('note_rows', [])
    print(f'[HOPPER_ORIG] HTML rows={len(rows)} note_rows={len(note_rows)}')
    # hidden_rows flagged
    hidden_flag = meta.get('hidden_rows_flagged') or meta.get('hidden_rows')
    print(f'[HOPPER_ORIG] HTML hidden rows meta={hidden_flag}')
    # no 3+9 rows - check for sub-section markers
    # Look for text "3+9" anywhere
    as_text = json.dumps(p)
    has_3_plus_9 = '3+9' in as_text
    print(f'[HOPPER_ORIG] HTML contains_3+9?={has_3_plus_9}')
    log(sect, 'opp_count(expected 103)', 103, len(rows),
        'MATCH' if len(rows)==103 else 'MISMATCH')
    log(sect, 'hidden_rows_flagged(expected 107)', 107, hidden_flag,
        'MATCH' if hidden_flag == 107 else 'CHECK')
    log(sect, 'no_3+9 markers', False, has_3_plus_9,
        'MATCH' if not has_3_plus_9 else 'MISMATCH')

# ------------------------------------------------------------------
# 8) Rutish
# ------------------------------------------------------------------
def verify_rutish():
    sect = 'RUTISH'
    # Could be at V6 root or in New info
    import glob
    cand = glob.glob(os.path.join(NEWINFO, 'Rutish_Airways_Statement_of_Account.xlsx')) + \
           glob.glob(os.path.join(PARENT, 'Rutish_Airways_Statement_of_Account.xlsx'))
    path = cand[0]
    wb = load_workbook(path, data_only=True)
    p = load_payload('Rutish_Airways_SOA')
    sections = p.get('sections', {})
    all_items = p.get('all_items', [])
    meta = p.get('meta', {})
    print(f'[RUTISH] HTML sections={list(sections.keys())} all_items={len(all_items)}')
    for s, items in sections.items():
        n = len(items) if isinstance(items, list) else (len(items.get('items', [])) if isinstance(items, dict) else 0)
        print(f'[RUTISH]   section={s!r} items={n}')
    print(f'[RUTISH] HTML meta.customer_name={meta.get("customer_name") or meta.get("metadata",{}).get("customer_name")}')
    expected_total = 3+35+25+12  # 75
    log(sect, 'section_count(expected 4 non-empty)', 4, len([s for s in sections if sections[s]]),
        'MATCH' if len([s for s in sections if sections[s]])==4 else 'CHECK')
    log(sect, 'total_items(expected 75)', 75, len(all_items),
        'MATCH' if len(all_items)==75 else 'MISMATCH')
    cname = meta.get('customer_name') or meta.get('metadata',{}).get('customer_name') or ''
    log(sect, 'customer_name contains Rutish', 'Rutish Airways', cname,
        'MATCH' if 'Rutish' in str(cname) else 'MISMATCH')

# ------------------------------------------------------------------
# 9) Ethiopian Fake
# ------------------------------------------------------------------
def verify_eth_fake():
    sect = 'ETH_FAKE'
    import glob
    cand = glob.glob(os.path.join(PARENT, 'ethiopian_fake_soa.xlsx')) + glob.glob(os.path.join(NEWINFO, 'ethiopian_fake_soa.xlsx'))
    path = cand[0]
    wb = load_workbook(path, data_only=True)
    p = load_payload('Ethiopian_Fake_SOA')
    line_items = p.get('line_items', [])
    sections = p.get('sections', {})
    aging_counts = p.get('aging_counts', {})
    populated = sum(1 for k,v in aging_counts.items() if v and v>0)
    print(f'[ETH_FAKE] HTML line_items={len(line_items)} sections={list(sections.keys())} aging_populated={populated}')
    log(sect, 'line_items(expected 82)', 82, len(line_items),
        'MATCH' if len(line_items)==82 else 'MISMATCH')
    log(sect, 'sections(expected 5)', 5, len(sections),
        'MATCH' if len(sections)==5 else 'MISMATCH')
    log(sect, 'aging_buckets populated (expected 7)', 7, populated,
        'MATCH' if populated >= 7 else 'CHECK')

# ============ RUN ============
for fn in [verify_eth_soa, verify_epi, verify_mea, verify_shop_visit,
           verify_svrg, verify_hopper, verify_hopper_orig, verify_rutish, verify_eth_fake]:
    print(f'\n---- {fn.__name__} ----')
    try:
        fn()
    except Exception as e:
        import traceback
        traceback.print_exc()
        log(fn.__name__, 'RUN', 'exception', str(e), 'FAIL')

# Save results
with open(os.path.join(ROOT, '_verify_results.json'), 'w', encoding='utf-8') as fp:
    json.dump(results, fp, indent=2, default=str)
print('\n===== SUMMARY =====')
for r in results:
    print(r)
